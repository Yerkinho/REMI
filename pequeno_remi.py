# ══════════════════════════════════════════════════════════════════════
#  PEQUEÑO REMI  v10.4  —  Asistente IA · Remuneraciones & Control Gestión
#  Autor: Yerko  |  Empresa: Área Remuneraciones Chile
#
#  pip install anthropic google-generativeai openai mistralai groq
#             customtkinter pillow openpyxl requests beautifulsoup4
#             duckduckgo-search plyer psutil SpeechRecognition pyaudio
# ══════════════════════════════════════════════════════════════════════
REMI_VERSION = "10.4"

import tkinter as tk
from tkinter import filedialog, messagebox
import threading, json, datetime, os, subprocess, sys, io, tempfile
import webbrowser, glob, base64, time, re, math
import urllib.request, urllib.parse, urllib.error
from pathlib import Path
import logging
import warnings

# ── Suprimir FutureWarning/DeprecationWarning de paquetes externos ─────────────
warnings.filterwarnings("ignore", category=FutureWarning)

# ══════════════════════════════════════════════════════════════════════
#  AUTO-ACTUALIZACIÓN  —  verifica GitHub al iniciar (no bloquea UI)
# ══════════════════════════════════════════════════════════════════════
# Para habilitar: sube pequeno_remi.py a tu repositorio GitHub y
# configura REMI_UPDATE_URL con la URL raw del archivo.
# Ejemplo:
#   REMI_UPDATE_URL = "https://raw.githubusercontent.com/TU_USUARIO/remi/main/pequeno_remi.py"
# Si dejas la URL vacía (""), la auto-actualización queda desactivada.
REMI_UPDATE_URL = "https://raw.githubusercontent.com/Yerkinho/REMI/main/pequeno_remi.py"   # ← pega aquí tu URL raw de GitHub

def _remi_check_update(callback_ui=None):
    """
    Verifica en segundo plano si hay una versión más nueva en GitHub.
    callback_ui(tipo, msg): función que se llama en el hilo principal con el resultado.
      tipo: "ok" | "update" | "error"
    """
    if not REMI_UPDATE_URL:
        return  # Auto-update desactivado

    def _run():
        try:
            import urllib.request, re as _re
            req = urllib.request.Request(
                REMI_UPDATE_URL,
                headers={"User-Agent": f"REMI/{REMI_VERSION}",
                         "Cache-Control": "no-cache"}
            )
            with urllib.request.urlopen(req, timeout=8) as r:
                remote_src = r.read().decode("utf-8", errors="replace")

            # Extraer versión del archivo remoto
            m = _re.search(r'REMI_VERSION\s*=\s*"([^"]+)"', remote_src)
            if not m:
                return  # No se encontró versión → no hacer nada

            remote_ver = m.group(1).strip()
            local_ver  = REMI_VERSION.strip()

            def _ver_tuple(v):
                try: return tuple(int(x) for x in v.split("."))
                except: return (0,)

            if _ver_tuple(remote_ver) <= _ver_tuple(local_ver):
                # Ya tenemos la versión más reciente
                if callback_ui:
                    callback_ui("ok", f"REMI v{local_ver} — actualizado ✓")
                return

            # Hay una versión más nueva
            if callback_ui:
                callback_ui("update", f"Nueva versión disponible: v{remote_ver} → di 'actualizar remi'")

        except Exception as _e:
            # Silencioso — no interrumpir la app por errores de red
            pass

    import threading
    threading.Thread(target=_run, daemon=True).start()


def _remi_aplicar_update(callback_ui=None):
    """
    Descarga la nueva versión, hace backup del archivo actual y lo reemplaza.
    Luego reinicia REMI automáticamente.
    """
    if not REMI_UPDATE_URL:
        if callback_ui: callback_ui("error", "URL de actualización no configurada.")
        return

    def _run():
        try:
            import urllib.request, shutil, subprocess, re as _re
            current_file = Path(__file__).resolve()
            backup_file  = current_file.with_suffix(f".v{REMI_VERSION}.bak")

            # Descargar nueva versión
            if callback_ui: callback_ui("ok", "⬇️ Descargando actualización…")
            req = urllib.request.Request(
                REMI_UPDATE_URL,
                headers={"User-Agent": f"REMI/{REMI_VERSION}"}
            )
            with urllib.request.urlopen(req, timeout=30) as r:
                new_src = r.read()

            # Verificar que es válido (tiene REMI_VERSION)
            if b'REMI_VERSION' not in new_src:
                if callback_ui: callback_ui("error", "El archivo descargado no parece válido.")
                return

            # Extraer nueva versión
            m = _re.search(rb'REMI_VERSION\s*=\s*b?"([^"]+)"', new_src)
            new_ver = m.group(1).decode() if m else "?"

            # Backup del archivo actual
            shutil.copy2(current_file, backup_file)

            # Escribir nuevo archivo
            current_file.write_bytes(new_src)

            if callback_ui:
                callback_ui("ok", f"✅ REMI v{new_ver} instalado. Reiniciando…")

            # Reiniciar REMI
            import time; time.sleep(1.5)
            subprocess.Popen(
                [sys.executable, str(current_file)],
                creationflags=0x08000000 if sys.platform=="win32" else 0
            )
            # Cerrar la instancia actual
            import os; os._exit(0)

        except Exception as e:
            if callback_ui: callback_ui("error", f"Error al actualizar: {e}")

    import threading
    threading.Thread(target=_run, daemon=True).start()

warnings.filterwarnings("ignore", category=DeprecationWarning)


def _bootstrap_msg(title: str, body: str, cmd: str | None = None):
    """Mensaje de arranque seleccionable + botón Copiar (no depende de customtkinter)."""
    try:
        root = tk.Tk()
        root.title(title)
        root.geometry("640x300")
        root.configure(bg="#111111")

        frm = tk.Frame(root, bg="#111111")
        frm.pack(fill="both", expand=True)

        txt = tk.Text(frm, wrap="word", bg="#111111", fg="#e6e6e6",
                      insertbackground="#e6e6e6", relief="flat")
        txt.pack(fill="both", expand=True, padx=12, pady=12)
        txt.insert("1.0", body.strip() + "\n")
        if cmd:
            txt.insert("end", "\nComando:\n")
            txt.insert("end", cmd.strip() + "\n")
        txt.configure(state="disabled")

        bar = tk.Frame(frm, bg="#111111")
        bar.pack(fill="x", padx=12, pady=(0, 12))

        def _copy():
            if not cmd:
                return
            try:
                root.clipboard_clear()
                root.clipboard_append(cmd)
                btn.configure(text="Copiado")
                root.after(1800, lambda: btn.configure(text="Copiar comando"))
            except Exception:
                pass

        btn = tk.Button(bar, text="Copiar comando", command=_copy,
                        bg="#2d6cdf", fg="white", bd=0, padx=10, pady=6)
        btn.pack(side="left")

        tk.Button(bar, text="Cerrar", command=root.destroy,
                  bg="#333333", fg="white", bd=0, padx=10, pady=6).pack(side="right")

        root.mainloop()
    except Exception:
        # Último recurso: imprimir
        print(title)
        print(body)
        if cmd:
            print("Comando:")
            print(cmd)

try:
    import customtkinter as ctk
except ModuleNotFoundError:
    _bootstrap_msg(
        "Pequeño Remi — dependencia faltante",
        "Falta la librería 'customtkinter'.\n\nInstálala y vuelve a abrir Remi.",
        "python -m pip install customtkinter",
    )
    sys.exit(1)

# ── CTK PRIMERO — antes de cualquier widget ────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

# ── PIL ────────────────────────────────────────────────────────────────
try:    from PIL import Image, ImageTk; PIL_OK = True
except: PIL_OK = False

# Intento configuración automática de Tesseract/pytesseract si está disponible
try:
    import pytesseract
    import shutil
    tpath = shutil.which('tesseract')
    if tpath:
        try:
            pytesseract.pytesseract.tesseract_cmd = tpath
        except Exception:
            pass
    # Si existe instalación en Program Files, preferirla
    possible = [r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"]
    for p in possible:
        if os.path.exists(p):
            try:
                pytesseract.pytesseract.tesseract_cmd = p
            except Exception:
                pass
            break

    # configurar TESSDATA_PREFIX si existe carpeta de tessdata en user local
    user_tess = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Tesseract', 'tessdata')
    pf_tess = os.path.join('C:\\Program Files\\Tesseract-OCR', 'tessdata')
    if os.path.isdir(user_tess):
        os.environ.setdefault('TESSDATA_PREFIX', user_tess)
    elif os.path.isdir(pf_tess):
        os.environ.setdefault('TESSDATA_PREFIX', pf_tess)
except Exception:
    pass

# ── IAs ───────────────────────────────────────────────────────────────
try:    import anthropic;                     CLAUDE  = True
except: CLAUDE  = False

# Gemini: intentar primero la nueva librería google.genai, luego la antigua
genai = None
GEMINI = False
genai_new = None
GENAI_NEW = False

try:
    import importlib
    genai_new = importlib.import_module("google.genai")
    GENAI_NEW = True
    GEMINI    = True   # nueva librería también activa el flag Gemini
except Exception:
    GENAI_NEW = False

if not GENAI_NEW:
    try:
        # Silenciar FutureWarning del paquete deprecado antes de importarlo
        import warnings as _w_gem
        with _w_gem.catch_warnings():
            _w_gem.simplefilter("ignore")
            import google.generativeai as genai
        GEMINI = True
    except Exception:
        GEMINI = False
try:    import openai;                        OPENAI  = True
except: OPENAI  = False
try:    from mistralai import Mistral;        MISTRAL = True
except: MISTRAL = False
try:    import groq as _groq_lib;            GROQ    = True
except: GROQ    = False

# ── Extras ────────────────────────────────────────────────────────────
try:    from duckduckgo_search import DDGS;   WEB   = True
except: WEB   = False
try:    import speech_recognition as sr;      VOZ   = True
except: VOZ   = False
try:    from plyer import notification;       NOTIF = True
except: NOTIF = False
try:    import psutil;                        SYS   = True
except: SYS   = False

# ── Auto-instalar dependencias críticas si faltan ──────────────────────
def _ensure_pkg(pkg, import_name=None):
    """Instala un paquete si no está disponible."""
    import_name = import_name or pkg
    try:
        __import__(import_name)
        return True
    except ImportError:
        try:
            import subprocess as _sp
            _flags = getattr(_sp, "CREATE_NO_WINDOW", 0)
            _sp.run([sys.executable, "-m", "pip", "install", pkg, "-q",
                     "--disable-pip-version-check"],
                    capture_output=True, timeout=120,
                    creationflags=_flags)
            __import__(import_name)
            return True
        except Exception:
            return False

_ensure_pkg("openpyxl")
_ensure_pkg("pandas")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, GradientFill
    from openpyxl.utils import get_column_letter, column_index_from_string
    try:
        from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY, FORMAT_NUMBER_COMMA_SEPARATED1
    except ImportError:
        FORMAT_DATE_DDMMYY = "DD/MM/YY"
        FORMAT_NUMBER_COMMA_SEPARATED1 = "#,##0.00"
    XL = True
except Exception:
    XL = False
try:    import requests as _rq;               REQ   = True
except: REQ   = False

# ══════════════════════════════════════════════════════════════════════
#  CONSTANTES Y RUTAS
# ══════════════════════════════════════════════════════════════════════
def _safe_mkdir(p: Path) -> Path | None:
    try:
        p.mkdir(parents=True, exist_ok=True)
        return p
    except Exception:
        return None

def _descargas_dir() -> Path:
    """Carpeta de Descargas del usuario (Windows / Linux / macOS)."""
    # 1. Variable de entorno USERPROFILE (Windows)
    up = os.environ.get("USERPROFILE", "")
    if up:
        d = Path(up) / "Downloads"
        if d.exists(): return d
        d2 = Path(up) / "Descargas"
        if d2.exists(): return d2
    # 2. Path.home() universal
    for name in ("Downloads", "Descargas"):
        d = Path.home() / name
        if d.exists(): return d
    # 3. Fallback: crear Downloads en home
    fallback = Path.home() / "Downloads"
    try: fallback.mkdir(parents=True, exist_ok=True)
    except Exception: pass
    return fallback

DESCARGAS = _descargas_dir()

def _base_dir() -> Path:
    """Directorio de datos portable.

    Orden:
    1) env REMI_BASE
    2) Escritorio del usuario actual
    3) LOCALAPPDATA
    4) carpeta actual
    """
    env = os.environ.get("REMI_BASE", "").strip().strip('"')
    cands: list[Path] = []
    if env:
        cands.append(Path(env))
    cands.append(DESCARGAS / "PequeñoRemi")
    cands.append(Path.home() / "Desktop" / "PequeñoRemi")
    cands.append(Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / "PequeñoRemi")
    cands.append(Path.cwd() / "PequeñoRemiData")

    for c in cands:
        ok = _safe_mkdir(c)
        if ok:
            return ok
    return Path.cwd()

BASE = _base_dir()

# Logging: archivo en la carpeta de datos de Remi
try:
    LOG_FILE = BASE / "remi.log"
    logger = logging.getLogger("remi")
    if not logger.handlers:
        logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        logger.addHandler(fh)
        ch = logging.StreamHandler()
        ch.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
        logger.addHandler(ch)
except Exception:
    logger = logging.getLogger("remi")

F = {k: BASE / v for k, v in {
    "cfg":"config.json", "hist":"historial.json", "tema":"temas.json", "conv_sessions":"conversaciones.json",
    "alarm":"alarmas.json", "notas":"notas.json", "proc":"procesos.json",
    "kpi":"kpi.json", "gantt":"gantt.json",
    "usuarios":"usuarios.json", "acceso_log":"acceso_log.json",
}.items()}

# ══════════════════════════════════════════════════════════════════════
#  REMI REGISTRO — Carpetas automáticas por tipo de proceso
#
#  Estructura generada en BASE/Registro/:
#    Registro/
#      Macros/          ← archivos .bas / .vba generados
#      Excel/           ← archivos .xlsx creados o modificados
#      Modificaciones/  ← copias backup antes de editar
#      Scripts/         ← código Python ejecutado
#      Informes/        ← reportes y outputs de texto
#      Finiquitos/      ← matrices y consolidados
#      log_registro.json ← bitácora de todo lo registrado
# ══════════════════════════════════════════════════════════════════════
class RemiRegistro:
    """
    Sistema de registro organizado por carpetas.
    Cada vez que REMI genera o modifica un archivo lo copia aquí
    con nombre de fecha+hora para no perder nada y tener trazabilidad.
    """
    TIPOS = {
        "macro":         "Macros",
        "excel":         "Excel",
        "backup":        "Modificaciones",
        "script":        "Scripts",
        "informe":       "Informes",
        "finiquito":     "Finiquitos",
    }
    _EXT_TIPO = {
        ".bas":  "macro",
        ".vba":  "macro",
        ".bas":  "macro",
        ".py":   "script",
        ".txt":  "informe",
        ".md":   "informe",
        ".xlsx": "excel",
        ".xlsm": "excel",
        ".xls":  "excel",
        ".csv":  "informe",
    }

    def __init__(self):
        self._raiz = BASE / "Registro"
        self._log_f = self._raiz / "log_registro.json"
        # Crear carpetas base
        for nombre in self.TIPOS.values():
            _safe_mkdir(self._raiz / nombre)

    def _carpeta(self, tipo: str) -> Path:
        """Devuelve la carpeta correspondiente al tipo, creándola si no existe."""
        nombre = self.TIPOS.get(tipo, "Informes")
        p = self._raiz / nombre
        _safe_mkdir(p)
        return p

    def _slug_fecha(self) -> str:
        return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    def guardar(self, contenido: str, tipo: str, nombre_base: str,
                extension: str = ".txt", descripcion: str = "") -> Path | None:
        """
        Guarda un contenido de texto en la carpeta correspondiente.
        Retorna la ruta del archivo guardado, o None si falló.
        tipo: 'macro' | 'excel' | 'script' | 'informe' | 'finiquito' | 'backup'
        """
        try:
            carpeta = self._carpeta(tipo)
            # Nombre seguro: sin caracteres inválidos
            base_safe = re.sub(r'[\\/:*?"<>|]', "_", nombre_base)[:40].strip("_. ")
            fname = f"{self._slug_fecha()}_{base_safe}{extension}"
            dest  = carpeta / fname
            dest.write_text(contenido, encoding="utf-8", errors="replace")
            self._bitacora(tipo, str(dest), descripcion)
            logger.info("RemiRegistro: %s guardado → %s", tipo, dest)
            return dest
        except Exception as e:
            logger.warning("RemiRegistro.guardar error: %s", e)
            return None

    def copiar_archivo(self, ruta_origen: str, tipo: str,
                       descripcion: str = "") -> Path | None:
        """
        Copia un archivo existente a la carpeta de registro.
        Ideal para hacer backup antes de modificar un Excel.
        """
        try:
            import shutil as _sh
            carpeta = self._carpeta(tipo)
            nombre  = Path(ruta_origen).name
            ext     = Path(ruta_origen).suffix.lower()
            base    = Path(ruta_origen).stem
            base_safe = re.sub(r'[\\/:*?"<>|]', "_", base)[:40].strip("_. ")
            fname   = f"{self._slug_fecha()}_{base_safe}{ext}"
            dest    = carpeta / fname
            _sh.copy2(ruta_origen, dest)
            self._bitacora(tipo, str(dest), descripcion or f"Copia de {nombre}")
            logger.info("RemiRegistro: copia de '%s' → %s", nombre, dest)
            return dest
        except Exception as e:
            logger.warning("RemiRegistro.copiar_archivo error: %s", e)
            return None

    def detectar_tipo(self, ruta: str) -> str:
        """Detecta el tipo de registro por extensión del archivo."""
        ext = Path(ruta).suffix.lower()
        return self._EXT_TIPO.get(ext, "informe")

    def _bitacora(self, tipo: str, ruta: str, desc: str):
        """Agrega entrada al log JSON de registro."""
        try:
            entradas = []
            if self._log_f.exists():
                try:
                    entradas = json.loads(self._log_f.read_text(encoding="utf-8"))
                    if not isinstance(entradas, list):
                        entradas = []
                except Exception:
                    entradas = []
            entradas.append({
                "fecha":       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "tipo":        tipo,
                "archivo":     Path(ruta).name,
                "ruta":        ruta,
                "descripcion": desc or "",
            })
            # Mantener solo los últimos 500 registros
            entradas = entradas[-500:]
            self._log_f.write_text(
                json.dumps(entradas, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )
        except Exception as e:
            logger.warning("RemiRegistro._bitacora error: %s", e)

    def listar(self, tipo: str | None = None, n: int = 20) -> list[dict]:
        """Retorna los últimos N registros (todos o filtrado por tipo)."""
        try:
            if not self._log_f.exists():
                return []
            entradas = json.loads(self._log_f.read_text(encoding="utf-8"))
            if not isinstance(entradas, list):
                return []
            if tipo:
                entradas = [e for e in entradas if e.get("tipo") == tipo]
            return list(reversed(entradas))[:n]
        except Exception:
            return []

    def resumen_texto(self, n: int = 10) -> str:
        """Genera un resumen de texto de los últimos registros para mostrar en chat."""
        entradas = self.listar(n=n)
        if not entradas:
            return "📂 Registro vacío — todavía no hay archivos guardados."
        lineas = ["**📂 Registro de archivos generados por REMI**\n"]
        for e in entradas:
            ic = {"macro":"📜","excel":"📊","script":"🐍","informe":"📄",
                  "backup":"🔒","finiquito":"📋"}.get(e.get("tipo",""), "📁")
            lineas.append(
                f"{ic} **{e['archivo']}**\n"
                f"   `{e['fecha']}` · {e.get('descripcion','')}"
            )
        lineas.append(f"\n_Carpeta: `{self._raiz}`_")
        return "\n".join(lineas)

    def abrir_carpeta(self, tipo: str | None = None):
        """Abre en el explorador la carpeta de registro (o una subcarpeta)."""
        try:
            ruta = self._carpeta(tipo) if tipo else self._raiz
            import subprocess as _sp
            _sp.Popen(f'explorer "{ruta}"',
                      creationflags=getattr(_sp, "CREATE_NO_WINDOW", 0))
        except Exception as e:
            logger.warning("RemiRegistro.abrir_carpeta: %s", e)


# Instancia global
REGISTRO = RemiRegistro()


def _registro_auto_macro(codigo_bas: str, nombre: str = "macro") -> Path | None:
    """Registra automáticamente un bloque de código VBA/BAS."""
    return REGISTRO.guardar(codigo_bas, "macro", nombre, ".bas",
                             f"Macro generada por REMI")

def _registro_auto_script(codigo_py: str, nombre: str = "script") -> Path | None:
    """Registra automáticamente un bloque de código Python antes de ejecutarlo."""
    return REGISTRO.guardar(codigo_py, "script", nombre, ".py",
                             "Script ejecutado por REMI")

def _registro_auto_excel(ruta_xlsx: str, descripcion: str = "") -> Path | None:
    """Registra (copia) un archivo Excel generado o modificado."""
    if os.path.exists(ruta_xlsx):
        return REGISTRO.copiar_archivo(ruta_xlsx, "excel",
                                        descripcion or "Excel generado por REMI")
    return None

def _registro_backup_excel(ruta_xlsx: str) -> Path | None:
    """Hace backup de un Excel ANTES de modificarlo."""
    if os.path.exists(ruta_xlsx):
        return REGISTRO.copiar_archivo(ruta_xlsx, "backup",
                                        f"Backup previo a modificación")
    return None

# ── GANTT: Tareas del mes ────────────────────────────────────────────
GANTT_TASKS_DEFAULT = [
    {"id":1,  "name":"Capturas con macro",     "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-21","avance":0.0},
    {"id":2,  "name":"Datos Softland",         "resp":"Ignacio",       "inicio":"2026-02-02","fin":"2026-02-20","avance":0.0},
    {"id":3,  "name":"Dot Auto",               "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-02","avance":1.0},
    {"id":4,  "name":"Dot diarias",            "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-28","avance":0.44},
    {"id":5,  "name":"Envios auto",            "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-02","avance":1.0},
    {"id":6,  "name":"GeoVictoria",            "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-13","avance":1.0},
    {"id":7,  "name":"Libro nuevo",            "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-03","avance":1.0},
    {"id":8,  "name":"Libros rem",             "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-03","avance":1.0},
    {"id":9,  "name":"Prevención",             "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-13","avance":0.1},
    {"id":10, "name":"Selección",              "resp":"Yerko",         "inicio":"2026-02-02","fin":"2026-02-25","avance":0.1},
    {"id":11, "name":"Libros Rem (sem1)",      "resp":"Yerko",         "inicio":"2026-02-03","fin":"2026-02-09","avance":1.0},
    {"id":12, "name":"Funes (Isapre y AFP)",   "resp":"Yerko",         "inicio":"2026-02-05","fin":"2026-02-18","avance":0.0},
    {"id":13, "name":"Licencias auto",         "resp":"Yerko",         "inicio":"2026-02-05","fin":"2026-02-05","avance":1.0},
    {"id":14, "name":"Liquidaciones BUK",      "resp":"Yerko",         "inicio":"2026-02-11","fin":"2026-02-12","avance":0.0},
    {"id":15, "name":"Liquidaciones Xiaomi",   "resp":"Yerko",         "inicio":"2026-02-12","fin":"2026-02-13","avance":0.0},
    {"id":16, "name":"Proyecto Bono Vacas",    "resp":"Ignacio/Yerko", "inicio":"2026-02-13","fin":"2026-02-13","avance":1.0},
    {"id":17, "name":"Libros Rem (final)",     "resp":"Yerko",         "inicio":"2026-02-15","fin":"2026-02-25","avance":0.0},
]

def gantt_load():
    d = _r(F["gantt"])
    if isinstance(d, dict) and "tasks" in d and isinstance(d["tasks"], list) and d["tasks"]:
        return d
    return {"tasks": [t.copy() for t in GANTT_TASKS_DEFAULT], "checkin": {}}

def gantt_save(data):
    _w(F["gantt"], data)

FONT  = "Segoe UI"
FMONO = "Consolas"

M_FAST  = "claude-haiku-4-5-20251001"
M_SMART = "claude-sonnet-4-6"
OLLAMA_URL = "http://localhost:11434"

RUTAS = {
    "SLD": r"X:\REMUNERACIONES\REM2026-1\Rem 2026-02\Comisiones\SLD",
    "PDV": r"X:\REMUNERACIONES\REM2026-1\Rem 2026-02\Comisiones\PDV",
    "EOS": r"X:\REMUNERACIONES\REM2026-1\Rem 2026-02\Comisiones\EOS",
    "DOT": r"X:\DOTACIONES_DIARIAS",
    "INF": r"X:\INFORMES",
}

# Calendario remuneraciones Chile
CAL_REM = {
    1:  "Dotación diaria",
    5:  "Corte previred empleados",
    10: "Pago AFP/previsión",
    15: "Corte comisiones PDV",
    20: "Cierre comisiones SLD",
    25: "⚠ Pre-cierre mes — verifica comisiones",
    28: "🔴 CIERRE COMISIONES",
    30: "Informe dotación mensual",
}

PROV_NOMBRES = {"claude":"Claude","gemini":"Gemini","openai":"OpenAI",
                "mistral":"Mistral","ollama":"Ollama","groq":"Groq (gratis)","ninguno":"Sin IA"}
PROV_COLOR   = {"claude":"#cc785c","gemini":"#4285f4","openai":"#10a37f",
                "mistral":"#f55036","ollama":"#c084fc","groq":"#f5a623","ninguno":"#666666"}
PROV_ICON    = {"claude":"◆","gemini":"◉","openai":"◈","mistral":"◇","ollama":"◎","groq":"⚡","ninguno":"○"}

# ══════════════════════════════════════════════════════════════════════
#  PERSISTENCIA
# ══════════════════════════════════════════════════════════════════════
def _r(f):
    try:
        if f.exists():
            with open(f,"r",encoding="utf-8") as fp: return json.load(fp)
    except Exception as _e:
        logging.warning("_r(%s): %s", f, _e)
    return {}

def _rl(f):
    d = _r(f); return d if isinstance(d,list) else []

def _w(f,d):
    try:
        with open(f,"w",encoding="utf-8") as fp: json.dump(d,fp,ensure_ascii=False,indent=2)
    except Exception as _e:
        logging.warning("_w(%s): %s", f, _e)

# ══════════════════════════════════════════════════════════════════════
#  TEMA
# ══════════════════════════════════════════════════════════════════════
# ── Paletas de color REMI v10 — renovadas para mayor comodidad visual ──
PALETAS = {
    # ── Matrix: basado en el prototipo RemiUI — fondo #141619, verde moderno
    "Matrix": {
        "bg":"#141619","sb":"#101214","card":"#1c2027","inp":"#1a1d24",
        "a1":"#3fb950","a2":"#52c768","tx":"#eaecef","t2":"#8b949e",
        "br":"#2a2e35","bub_u":"#2a2e35","bub_b":"#101214","acc":"#3fb950",
        "err":"#f85149","ok":"#3fb950","warn":"#e3b341","code":"#0d1117",
        "link":"#58a6ff","idea":"#e3b341",
    },
    # ── Moderno: gris azulado premium — suave para largas sesiones
    "Moderno": {
        "bg":"#16181d","sb":"#11131a","card":"#1e2128","inp":"#191c23",
        "a1":"#4d8ef0","a2":"#6ba3f5","tx":"#e8eaf0","t2":"#828a9e",
        "br":"#282d3a","bub_u":"#21253060","bub_b":"#11131a","acc":"#4d8ef0",
        "err":"#f07070","ok":"#52c78a","warn":"#e8b84b","code":"#0e1018",
        "link":"#4d8ef0","idea":"#e8b84b",
    },
    # ── Crepúsculo: violeta apagado — tranquilo, sin cansancio
    "Crepúsculo": {
        "bg":"#17151e","sb":"#110f18","card":"#201d2c","inp":"#1a1826",
        "a1":"#9b7cf0","a2":"#b89af5","tx":"#e6e2f5","t2":"#7a72a0",
        "br":"#2d2840","bub_u":"#231f34","bub_b":"#110f18","acc":"#9b7cf0",
        "err":"#e86c6c","ok":"#5aba88","warn":"#d4a84b","code":"#0d0b14",
        "link":"#7cb4f5","idea":"#d4a84b",
    },
    # ── Océano: azul profundo renovado — alto contraste sin agresividad
    "Océano": {
        "bg":"#0e1219","sb":"#0a0e15","card":"#141c28","inp":"#10161f",
        "a1":"#3d9be0","a2":"#5db5f5","tx":"#d0dcf0","t2":"#5878a0",
        "br":"#1c2a3e","bub_u":"#141e30","bub_b":"#0a0e15","acc":"#3d9be0",
        "err":"#e05858","ok":"#44b87a","warn":"#c8a040","code":"#080c14",
        "link":"#5db5f5","idea":"#c8a040",
    },
    # ── Claude: cálido ámbar — conservado y mejorado
    "Claude": {
        "bg":"#1a1813","sb":"#141108","card":"#21201a","inp":"#1c1a12",
        "a1":"#cc7a48","a2":"#e09468","tx":"#ede8de","t2":"#8a7d6a",
        "br":"#302e24","bub_u":"#211d15","bub_b":"#141108","acc":"#cc7a48",
        "err":"#c45848","ok":"#5a9458","warn":"#c49838","code":"#0f0d09",
        "link":"#7aa8e0","idea":"#c49838",
    },
    # ── Niebla: azul-gris suave — el más descansado para ojos sensibles
    "Niebla": {
        "bg":"#181c22","sb":"#12151c","card":"#1e2230","inp":"#161920",
        "a1":"#7a9ecb","a2":"#9abddf","tx":"#d8e0ee","t2":"#6878a0",
        "br":"#262c3a","bub_u":"#1a2030","bub_b":"#12151c","acc":"#7a9ecb",
        "err":"#c85858","ok":"#4da870","warn":"#c49040","code":"#0c0f18",
        "link":"#9abddf","idea":"#c49040",
    },
    # ── Corporativo: azul ejecutivo — ligeramente renovado
    "Corporativo": {
        "bg":"#13151b","sb":"#0f1117","card":"#1c1f28","inp":"#161820",
        "a1":"#4d85cc","a2":"#6fa3e0","tx":"#e2e8f4","t2":"#7080a8",
        "br":"#282e3e","bub_u":"#181e30","bub_b":"#0f1117","acc":"#4d85cc",
        "err":"#d95c5c","ok":"#4db877","warn":"#d4974a","code":"#0b0d14",
        "link":"#6fa3e0","idea":"#d4974a",
    },
}

def cargar_tema():
    d = _r(F["tema"])
    if isinstance(d,dict) and "bg" in d and len(d) >= 10: return d
    return PALETAS["Matrix"].copy()

def guardar_tema(p): _w(F["tema"], p)
T = cargar_tema()

def cargar_cfg():
    d = _r(F["cfg"])
    if not isinstance(d,dict): d = {}
    defs = {
        "api_key":"","gemini_key":"","openai_key":"","mistral_key":"","mistral_model":"mistral-small-latest","groq_key":"",
        "gemini_model":"gemini-2.5-flash",
        "modelo":M_FAST,"proveedor":"claude",
        "ollama_url":OLLAMA_URL,"ollama_model":"qwen2.5-coder:7b",
        "web_auto":True,"multi_agent":False,"nombre_usuario":"Yerko","server_url":"http://127.0.0.1:8765","server_token":"remi-piloto-2026","server_activo":False,"graph_tenant":"","graph_client":"","graph_secret":"","graph_user_email":"","daemon_activo":False,"daemon_intervalo":30,"onedrive_dotacion":"","onedrive_liquidaciones":"","onedrive_comisiones_sld":"","onedrive_comisiones_pdv":"","onedrive_comisiones_eos":"","finiquitos_ruta":"","finiquitos_madre":"MATRIZ_GENERAL_FINIQUITOS.xlsx","finiquitos_ruta_od":"","finiquitos_hoja":"Finiquitos",
    }
    for k,v in defs.items(): d.setdefault(k,v)
    # Permitir sobreescritura por variables de entorno (útil para .env o CI)
    # Variables soportadas: GEMINI_KEY, API_KEY (Claude), OPENAI_KEY, MISTRAL_KEY, GROQ_KEY, PROVIDER
    try:
        env_map = {
            "GEMINI_KEY":"gemini_key",
            "API_KEY":"api_key",
            "OPENAI_KEY":"openai_key",
            "MISTRAL_KEY":"mistral_key",
            "GROQ_KEY":"groq_key",
            "PROVIDER":"proveedor",
        }
        for ev, kk in env_map.items():
            v = os.environ.get(ev)
            if v:
                d[kk] = v
    except Exception:
        pass
    return d

def guardar_cfg(c): _w(F["cfg"],c)
CFG = cargar_cfg()

# ══════════════════════════════════════════════════════════════════════
#  GESTIÓN DE PROVEEDORES IA
#  — Singleton por proveedor, cooldown de error, cadena automática
# ══════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════
#  AI ENGINE — Rotación automática de proveedores (patrón AIEngine)
#  Basado en el prototipo: cooldown escalonado + contador de errores
#  Integrado con la cadena existente: claude→gemini→openai→mistral→groq→ollama
# ══════════════════════════════════════════════════════════════════════
_ERR_PROV   = {}   # nombre → {"ts": float, "count": int}
_PROV_CACHE = {}   # singleton de clientes

# Cooldown escalonado: 1er error=60s, 2do=300s, 3ro=900s, 4to+=2700s
_COOLDOWN_TIERS = [60, 300, 900, 2700]

def _prov_libre(p: str) -> bool:
    """True si el proveedor ya superó su período de cooldown."""
    err = _ERR_PROV.get(p)
    if not err:
        return True
    count  = err.get("count", 1)
    tier   = min(count - 1, len(_COOLDOWN_TIERS) - 1)
    cd_seg = _COOLDOWN_TIERS[tier]
    return (time.time() - err["ts"]) > cd_seg

def _marcar_error(p: str):
    """Registra un error en el proveedor e incrementa su nivel de cooldown."""
    prev  = _ERR_PROV.get(p, {})
    count = prev.get("count", 0) + 1
    _ERR_PROV[p] = {"ts": time.time(), "count": count}
    _PROV_CACHE.pop(p, None)   # invalidar singleton
    tier   = min(count - 1, len(_COOLDOWN_TIERS) - 1)
    cd_seg = _COOLDOWN_TIERS[tier]
    logger.warning("AIEngine: %s marcado con error #%d — cooldown %ds", p, count, cd_seg)

def _marcar_ok(p: str):
    """Resetea el contador de errores cuando una llamada tiene éxito."""
    if p in _ERR_PROV:
        _ERR_PROV.pop(p, None)

def cooldown_info(p: str) -> str:
    """Retorna texto legible sobre el estado de cooldown de un proveedor."""
    err = _ERR_PROV.get(p)
    if not err:
        return "✅ libre"
    count  = err.get("count", 1)
    tier   = min(count - 1, len(_COOLDOWN_TIERS) - 1)
    cd_seg = _COOLDOWN_TIERS[tier]
    restante = max(0, cd_seg - (time.time() - err["ts"]))
    if restante <= 0:
        return "✅ libre (cooldown expirado)"
    mins = int(restante // 60)
    segs = int(restante % 60)
    return f"⏳ {mins}m {segs}s  (error #{count})"

def siguiente_prov_libre() -> str | None:
    """Retorna el próximo proveedor configurado y sin cooldown activo."""
    c = cadena_proveedores()
    return c[0] if c else None

def _get_claude():
    k = CFG.get("api_key","").strip()
    if not CLAUDE or not k.startswith("sk-"): return None
    if "claude" not in _PROV_CACHE:
        _PROV_CACHE["claude"] = anthropic.Anthropic(api_key=k)
    return _PROV_CACHE["claude"]

def _get_gemini():
    global GEMINI, genai
    if not GEMINI:
        try:
            import google.generativeai as genai
            GEMINI = True
        except:
            return None
    k = CFG.get("gemini_key","").strip()
    if not GEMINI or not k:
        return None

    if "gemini" not in _PROV_CACHE:
        # Registrar el módulo de Gemini (antiguo o nuevo) en el caché de proveedores.
        try:
            if GENAI_NEW and genai_new is not None:
                try:
                    genai_new.configure(api_key=k)
                except Exception:
                    pass
                _PROV_CACHE["gemini"] = genai_new
            else:
                try:
                    genai.configure(api_key=k)
                except Exception:
                    pass
                _PROV_CACHE["gemini"] = genai
        except Exception:
            return None
    return _PROV_CACHE["gemini"]


def _gemini_extraer_texto(candidatos):
    try:
        partes_txt = []
        for cand in candidatos:
            content = getattr(cand, "content", None)
            parts = getattr(content, "parts", None) or []
            for part in parts:
                t = getattr(part, "text", None)
                if t:
                    partes_txt.append(t)
        return "".join(partes_txt)
    except Exception:
        return ""


def _gemini_finish_reason(resp):
    try:
        c0 = (getattr(resp, "candidates", None) or [None])[0]
        return getattr(c0, "finish_reason", None)
    except Exception:
        return None

def _get_openai():
    k = CFG.get("openai_key","").strip()
    if not OPENAI or not k.startswith("sk-"): return None
    if "openai" not in _PROV_CACHE:
        _PROV_CACHE["openai"] = openai.OpenAI(api_key=k)
    return _PROV_CACHE["openai"]

def _get_mistral():
    k = CFG.get("mistral_key","").strip()
    if not MISTRAL or not k: return None
    if "mistral" not in _PROV_CACHE:
        _PROV_CACHE["mistral"] = Mistral(api_key=k)
    return _PROV_CACHE["mistral"]

def _get_groq():
    k = CFG.get("groq_key","").strip()
    if not GROQ or not k: return None
    if "groq" not in _PROV_CACHE:
        _PROV_CACHE["groq"] = _groq_lib.Groq(api_key=k)
    return _PROV_CACHE["groq"]

GROQ_MODELOS = [
    "llama-3.3-70b-versatile",               # ✅ Principal — mejor calidad, 32k ctx
    "meta-llama/llama-4-scout-17b-16e-instruct",  # ✅ Llama 4 Scout — multimodal, rápido
    "openai/gpt-oss-120b",                   # ✅ GPT-OSS 120B — alta calidad
    "llama-3.1-8b-instant",                  # ✅ Rápido, bueno para respuestas simples
    # DEPRECADOS — eliminados:
    # "mixtral-8x7b-32768"  → deprecado marzo 2025
    # "gemma2-9b-it"        → deprecado agosto 2025 → reemplazado por llama-3.1-8b-instant
]

# Cache para no congelar UI con chequeos de red
_OLLAMA_OK_CACHE = {"ts": 0.0, "ok": False}

def _get_ollama_ok():
    if not REQ: return False
    now = time.time()
    if (now - _OLLAMA_OK_CACHE.get("ts", 0.0)) < 12.0:
        return bool(_OLLAMA_OK_CACHE.get("ok", False))
    try:
        r = _rq.get(f"{CFG.get('ollama_url',OLLAMA_URL)}/api/tags", timeout=0.35)
        ok = (r.status_code == 200)
        _OLLAMA_OK_CACHE.update({"ts": now, "ok": ok})
        return ok
    except:
        _OLLAMA_OK_CACHE.update({"ts": now, "ok": False})
        return False

def _prov_configurado(p: str) -> bool:
    p = (p or "").lower().strip()
    if p == "claude":
        return bool(CLAUDE and CFG.get("api_key", "").strip().startswith("sk-"))
    if p == "gemini":
        return bool(GEMINI and CFG.get("gemini_key", "").strip())
    if p == "openai":
        return bool(OPENAI and CFG.get("openai_key", "").strip().startswith("sk-"))
    if p == "mistral":
        return bool(MISTRAL and CFG.get("mistral_key", "").strip())
    if p == "groq":
        return bool(GROQ and CFG.get("groq_key", "").strip())
    if p == "ollama":
        # Ollama no usa key; solo validamos "vive" con cache.
        return _get_ollama_ok()
    return False

_PROV_CHECKS = {
    "claude":  _get_claude,
    "gemini":  _get_gemini,
    "openai":  _get_openai,
    "mistral": _get_mistral,
    "groq":    _get_groq,
    "ollama":  _get_ollama_ok,
}

def cadena_proveedores():
    pref  = CFG.get("proveedor","claude")
    todos = ["claude","gemini","openai","mistral","groq","ollama"]
    orden = [pref] + [p for p in todos if p != pref]
    return [p for p in orden if _prov_libre(p) and _PROV_CHECKS.get(p,lambda:False)()]

def prov_activo():
    c = cadena_proveedores(); return c[0] if c else "ninguno"

def invalidar_cache_prov():
    _PROV_CACHE.clear()

# ══════════════════════════════════════════════════════════════════════
#  SISTEMA PROMPT — rico en contexto de remuneraciones Chile
# ══════════════════════════════════════════════════════════════════════
def sistema_prompt(extra_ctx=""):
    hoy  = datetime.datetime.now()
    dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
    meses = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    dia_s = dias[hoy.weekday()]
    nombre = CFG.get("nombre_usuario","Yerko")

    # Contexto calendario
    alertas = []
    for d, msg in CAL_REM.items():
        if abs(hoy.day - d) <= 1: alertas.append(f"HOY/PROX: {msg}")

    # Últimos registros
    hist = _rl(F["hist"])
    ult  = "; ".join(x.get("resumen","")[:60] for x in hist[-3:]) or "—"

    # Procesos aprendidos relevantes
    procs = _rl(F["proc"])
    proc_ctx = ""
    if procs:
        proc_ctx = "\nPROCESOS APRENDIDOS: " + " | ".join(
            f"[{p.get('nombre','')}→{p.get('resumen','')[:50]}]" for p in procs[-4:])

    alert_str = " | ALERTAS: " + " · ".join(alertas) if alertas else ""

    # ── Contexto de macros y procesos en memoria ──────────────────────────
    macros_ctx = ""
    try:
        from pathlib import Path as _P
        _macro_dir = BASE / "Registro" / "Macros"
        if _macro_dir.exists():
            _macros = sorted(_macro_dir.glob("*.bas"), key=lambda x: x.stat().st_mtime, reverse=True)[:5]
            if _macros:
                macros_ctx = "\nMACROS RECIENTES: " + " | ".join(m.stem for m in _macros)
    except Exception:
        pass

    return (
        f"Eres REMI, asistente IA experto en Remuneraciones, Control de Gestión y automatización "
        f"con Excel/VBA/Python. Trabajas junto a {nombre} en Remuneraciones Chile.\n"
        f"Fecha: {dia_s} {hoy.strftime('%d/%m/%Y')} ({meses[hoy.month]} {hoy.year}).{alert_str}"

        f"\n\n━━ INFRAESTRUCTURA ━━"
        f"\nSISTEMAS: Softland ERP · Excel 365 (OneDrive/SharePoint) · SAP consultas · PowerQuery · Power Automate"
        f"\nIAS ACTIVAS: {', '.join([p.upper() for p in ['gemini','mistral','groq','claude','openai'] if _prov_configurado(p)])}"
        f"\nGRAPH API: {'☁️ Activa · ' + CFG.get('graph_user_email','') if _graph_creds_ok() else '⚙️ Sin configurar (falta Email OD)'}\n"
        f"\n  Dotación     : {CFG.get('onedrive_dotacion') or '—'}"
        f"\n  Liquidaciones: {CFG.get('onedrive_liquidaciones') or '—'}"
        f"\n  Comisiones   : SLD={CFG.get('onedrive_comisiones_sld') or '—'} "
        f"PDV={CFG.get('onedrive_comisiones_pdv') or '—'} EOS={CFG.get('onedrive_comisiones_eos') or '—'}"
        f"\n  Finiquitos   : {CFG.get('finiquitos_ruta_od') or '—'} · Madre: {CFG.get('finiquitos_madre') or 'MATRIZ_GENERAL_FINIQUITOS.xlsx'}"

        f"\n\n━━ CONOCIMIENTO EXCEL/VBA/365 ━━"
        f"\nEXCEL CLÁSICO (openpyxl/xlwings): siempre engine='openpyxl', data_only=True al leer, autofit, estilos."
        f"\nEXCEL 365 / SHAREPOINT (Graph API):"
        f"\n  · Leer rangos: GET /workbook/worksheets/{{hoja}}/usedRange"
        f"\n  · Escribir: PATCH /workbook/worksheets/{{hoja}}/range(address='A1:Z100')"
        f"\n  · Fórmulas dinámicas 365: XLOOKUP, FILTER, UNIQUE, SORT, SEQUENCE, LET, LAMBDA"
        f"\n  · Co-autoría: no abrir con openpyxl mientras alguien lo edita en 365"
        f"\n  · Power Query (M): para transformaciones ETL sin VBA"
        f"\nVBA AVANZADO — EXCEL ESCRITORIO:"
        f"\n  ESTRUCTURA OBLIGATORIA DE TODA MACRO:"
        f"\n    Option Explicit"
        f"\n    Sub NombreMacro()"
        f"\n      On Error GoTo ErrHandler"
        f"\n      Application.ScreenUpdating = False : Application.EnableEvents = False"
        f"\n      ' ... código ..."
        f"\n      GoTo FinMacro"
        f"\n    ErrHandler:"
        f"\n      MsgBox Err.Description, vbCritical, \"Error en \" & Application.Caller"
        f"\n    FinMacro:"
        f"\n      Application.ScreenUpdating = True : Application.EnableEvents = True"
        f"\n      Set obj = Nothing  ' liberar todos los objetos"
        f"\n    End Sub"
        f"\n  PATRONES DE CALIDAD:"
        f"\n    · With/End With para múltiples propiedades del mismo objeto"
        f"\n    · Dictionary (Scripting.Dictionary) para búsquedas O(1)"
        f"\n    · Arrays dinámicos: ReDim Preserve solo al final del loop"
        f"\n    · NEVER usar .Select/.Activate — siempre referencia directa: ws.Range(\"A1\")"
        f"\n    · LastRow = ws.Cells(ws.Rows.Count,1).End(xlUp).Row"
        f"\n    · Abrir archivos: Workbooks.Open con UpdateLinks:=False, ReadOnly si solo lectura"
        f"\n  INTEGRACIÓN ENTRE MÓDULOS:"
        f"\n    · Si hay macros previas en Registro, REFERENCIARLAS y extenderlas (no duplicar)"
        f"\n    · Proponer encadenamiento: Macro1 llama Macro2 → resultado cohesivo"
        f"\n    · Si piden mejorar la macro: leer el .bas del Registro, proponer diff claro"
        f"\n    · Cuando la macro maneja múltiples hojas: usar nombres de hoja como constantes"
        f"\n    · Para reportes: generar en hoja nueva con nombre fecha, no pisar datos"
        f"\n  SOFTLAND / BASES DE DATOS:"
        f"\n    · ADODB.Connection para queries directos al ERP desde VBA"
        f"\n    · ConnString = 'Provider=SQLOLEDB;Data Source=SERVER;...'"  
        f"\n    · Siempre .Close y Set = Nothing para conexiones"
        f"\n  EXCEL 365 DESDE VBA:"
        f"\n    · ActiveWorkbook.Queries para gestionar Power Query M"
        f"\n    · Refresh Power Query: ThisWorkbook.Connections('Nombre').Refresh"
        f"\n    · Para SharePoint: usa Graph API (Python) en lugar de VBA para archivos remotos"

        f"\n\n━━ PROCESOS REMUNERACIONES ━━"
        f"\nLEGISLACIÓN: Código Trabajo · DT Chile · AFC · AFP · INP · Mutual · Previred"
        f"\nCIERRES: Previred día 10 · Sueldos último hábil · Comisiones día 28 · Dotación lunes/viernes"
        f"\nFLUJOS: Dotación diaria→semanal→mensual · Comisiones SLD/PDV/EOS · Liquidaciones · Finiquitos · Contratos · Libro Rem"

        f"\n\n━━ MEMORIA ACTIVA ━━"
        f"\nHISTORIAL: {ult}"
        f"{proc_ctx}{macros_ctx}"

        f"\n\n━━ REGLAS DE RESPUESTA ━━"
        f"\n  1. Español chileno, tutea a {nombre}. Tono: colega experto, directo. Sin 'claro que sí', 'por supuesto'"
        f"\n  2. Código SIEMPRE completo en bloques ```vba / ```python / ```m (nunca truncar)"
        f"\n  3. VBA: Option Explicit + On Error + comentarios en español + liberar objetos (Set x=Nothing)"
        f"\n  4. Python Excel: SIEMPRE engine='openpyxl', data_only=True, guardar en Downloads"
        f"\n  5. Excel 365: para leer/escribir en SharePoint, usa Graph API (ya configurada)"
        f"\n  6. Cálculos legales: cita artículo del Código del Trabajo"
        f"\n  7. INFERENCIA ACTIVA: si el mensaje es vago → infiere la intención, propón algo concreto"
        f"\n     Ej: 'el cierre' → guía de cierre mes; 'la macro de comisiones' → mejorar última macro guardada"
        f"\n     Ej: 'algo no cuadra' → pide la diferencia y el Excel; 'el gantt' → abre el panel"
        f"\n  8. ENCADENAMIENTO: si generas una macro, ofrece: (a) probarla, (b) guardarla en Registro, (c) mejora sugerida"
        f"\n  9. CONTEXTO IMPLÍCITO: 'ese excel' / 'lo que hiciste' / 'el último' → archivo más reciente del contexto"
        f"\n  10. RAZONAMIENTO: para código/cálculo complejo → una línea '📌 Enfoque: …' antes del código"
        f"\n  11. PROACTIVIDAD: al final de cada respuesta con código, menciona 1 riesgo o mejora brevemente"
        f"\n  12. MACROS VBA — REGLA DE ORO: siempre generar el módulo COMPLETO listo para pegar en el Editor VBA."
        f"\n      · Incluir TODAS las subrutinas relacionadas en el mismo bloque de código"
        f"\n      · Si la tarea tiene pasos (leer→procesar→escribir→notificar): codificar TODOS los pasos"
        f"\n      · Inferir nombres de hojas/columnas del contexto; si hay ambigüedad, usar constantes Const fáciles de cambiar"
        f"\n      · Al final del código VBA, incluir un comentario '\' CÓMO USAR: ...' con 2-3 líneas de instrucción"
        f"\n      · Si hay una macro en Registro relacionada, mencionarla: 'Detecté [NombreMacro] — la integro aquí'"
        f"\n      · Proponer siempre: (a) cómo probarla, (b) una mejora futura puntual"
        f"\n  13. EXCEL 365: cuando hay una URL SharePoint → ofrecerla como enlace directo + opción de leer con Graph API"
        f"\n  14. DATOS REALES: para dotación/comisiones/finiquitos → usa los comandos de Remi (Graph API), no inventes"
        f"\n  15. FINIQUITOS: si preguntan por un analista → sugiere '📋 Matriz Finiquitos' para filtrar"
        f"{extra_ctx}"
    )

# ══════════════════════════════════════════════════════════════════════
#  BÚSQUEDA WEB
# ══════════════════════════════════════════════════════════════════════
def buscar_web(q, n=4):
    if not WEB: return "duckduckgo-search no instalado"
    try:
        with DDGS() as d:
            rs = list(d.text(q, max_results=n))
        if not rs: return "Sin resultados"
        return "\n\n".join(f"【{r['title']}】\n{r['body'][:300]}\n{r['href']}" for r in rs)
    except Exception as e: return f"Error búsqueda: {e}"


# ══════════════════════════════════════════════════════════════════════
#  CLIENTE REMI SERVER
# ══════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════
#  VALIDACIÓN Y PARCHEO DE CÓDIGO ANTES DE EJECUTAR
# ══════════════════════════════════════════════════════════════════════
def _prevalidar_codigo(code):
    """
    Revisa el código Python ANTES de ejecutarlo y devuelve lista de errores/avisos.
    Verifica: sintaxis, imports disponibles, rutas de archivo, patrones peligrosos.
    """
    errores = []
    py = sys.executable

    # 1. Sintaxis Python
    try:
        compile(code, "<remi>", "exec")
    except SyntaxError as e:
        errores.append(f"Sintaxis inválida en línea {e.lineno}: {e.msg}")
        return errores  # sin sintaxis no tiene sentido seguir

    # 2. Imports faltantes
    import_names = re.findall(r'^\s*(?:import|from)\s+([\w]+)', code, re.MULTILINE)
    PKG_MAP = {"openpyxl":"openpyxl","pandas":"pandas","numpy":"numpy",
               "bs4":"beautifulsoup4","PIL":"pillow","cv2":"opencv-python",
               "sklearn":"scikit-learn","requests":"requests"}
    for mod in import_names:
        try:
            __import__(mod)
        except ImportError:
            pkg = PKG_MAP.get(mod, mod)
            errores.append(f"Módulo '{mod}' no instalado → se instalará '{pkg}' automáticamente")

    # 3. pandas to_excel sin engine
    if "to_excel" in code and "engine=" not in code:
        errores.append("pandas to_excel() sin engine= → se agrega engine='openpyxl' automáticamente")

    # 4. load_workbook sin data_only (puede leer fórmulas como None)
    if "load_workbook" in code and "data_only" not in code:
        errores.append("load_workbook() sin data_only=True → puede leer fórmulas en vez de valores")

    # 5. Rutas de archivo que no existen
    rutas = re.findall(r'[A-Za-z]:[/\\][^\n"\'<>|*?\s]{3,}', code, re.I)
    for ruta in rutas:
        ruta_c = ruta.strip("\"',();r")
        # Solo verificar si parece ruta de lectura (load, open, read)
        if any(k in code for k in ["load_workbook(","open(","read_excel("]):
            carpeta = os.path.dirname(ruta_c)
            if carpeta and not os.path.exists(carpeta):
                errores.append(f"Carpeta no accesible: '{carpeta}' (¿red desconectada?)")

    # 6. Operaciones potencialmente destructivas
    if re.search(r'\bos\.remove\b|\bshutil\.rmtree\b|\bos\.unlink\b', code):
        errores.append("El código elimina archivos — revisa que sea intencional")

    return errores

def _parchear_codigo(code):
    """
    Aplica correcciones automáticas al código para evitar errores conocidos.
    Retorna el código corregido listo para ejecutar.
    """
    fixed = code

    # 1. pandas to_excel sin engine → agregar openpyxl
    if "to_excel" in fixed and "engine=" not in fixed:
        fixed = re.sub(
            r'\.to_excel\(([^)]+)\)',
            lambda m: f".to_excel({m.group(1).rstrip()}, engine='openpyxl')"
            if "engine=" not in m.group(1) else m.group(0),
            fixed
        )
        # fallback DOTALL para multilínea
        if "engine=" not in fixed and "to_excel" in fixed:
            fixed = re.sub(
                r'\.to_excel\((.*?)\)',
                lambda m: f".to_excel({m.group(1).rstrip()}, engine='openpyxl')"
                if "engine=" not in m.group(1) else m.group(0),
                fixed, flags=re.DOTALL
            )

    # 2. load_workbook sin data_only
    if "load_workbook" in fixed and "data_only" not in fixed:
        fixed = re.sub(
            r'load_workbook\(([^)]+)\)',
            lambda m: f"load_workbook({m.group(1).rstrip()}, data_only=True)"
            if "data_only" not in m.group(1) else m.group(0),
            fixed
        )

    # 3. Rutas con backslash en strings sin prefijo → convertir a raw strings
    # Patrón de 2 grupos: (prefijo)(contenido) — la comilla se infiere del match
    def _to_raw(m):
        pre, content = m.group(1), m.group(2)
        q = '"' if '"' in m.group(0) else "'"
        if pre.lower() in ("r","f","b","rb","br","fr","rf","u"):
            return m.group(0)
        return f'r{q}{content}{q}'

    # 3b. Redirigir rutas genéricas de Desktop/Users → Descargas del usuario
    #     Solo aplica a rutas que apuntan a C:\Users\\<usuario>\Desktop o similares
    desc_str = str(DESCARGAS).replace('\\', '\\\\')
    def _redir_ruta(m):
        full = m.group(0)
        pre  = m.group(1)
        path = m.group(2)
        q    = '"' if '"' in full else "'"
        # Solo redirigir si apunta a Desktop o a C:\Users\<user> (rutas genéricas del código)
        if re.search(r'[Uu]suario|[Dd]esktop|[Ee]scritorio|[Uu]ser', path):
            fname = re.split(r'[/\\]', path)[-1]  # solo el nombre del archivo
            new_path = desc_str + '\\\\' + fname
            pfx = pre if pre.lower() in ('r','f','b') else 'r'
            return f"{pfx}{q}{new_path}{q}"
        return full
    fixed = re.sub(r'([rRfFbBuU]?)"([A-Za-z]:[\\][^"\n]+\.xlsx?)"', _redir_ruta, fixed)
    fixed = re.sub(r"([rRfFbBuU]?)'([A-Za-z]:[\\][^'\n]+\.xlsx?)'", _redir_ruta, fixed)
    # Strings con comilla doble
    fixed = re.sub(r'([rRfFbBuU]?)"([A-Za-z]:[\\][^"\n]+)"',
                   _to_raw, fixed)
    # Strings con comilla simple
    fixed = re.sub(r"([rRfFbBuU]?)'([A-Za-z]:[\\][^'\n]+)'",
                   _to_raw, fixed)
    # 4. Inyectar imports faltantes automáticamente
    # Si el código usa un módulo stdlib pero no lo importa, lo agrega al inicio.
    STDLIB_AUTO = {
        "os":        (r'\bos\.', r'^\s*import\s+os\b'),
        "sys":       (r'\bsys\.', r'^\s*import\s+sys\b'),
        "re":        (r'\bre\.', r'^\s*import\s+re\b'),
        "json":      (r'\bjson\.', r'^\s*import\s+json\b'),
        "datetime":  (r'\bdatetime\.', r'^\s*import\s+datetime\b|from\s+datetime\s+import'),
        "pathlib":   (r'\bPath\s*\(|pathlib\.', r'^\s*from\s+pathlib|^\s*import\s+pathlib'),
        "shutil":    (r'\bshutil\.', r'^\s*import\s+shutil\b'),
        "glob":      (r'\bglob\.', r'^\s*import\s+glob\b'),
        "time":      (r'\btime\.sleep\b|\btime\.time\b', r'^\s*import\s+time\b'),
        "math":      (r'\bmath\.', r'^\s*import\s+math\b'),
        "random":    (r'\brandom\.', r'^\s*import\s+random\b'),
        "io":        (r'\bio\.BytesIO\b|\bio\.StringIO\b', r'^\s*import\s+io\b'),
        "tempfile":  (r'\btempfile\.', r'^\s*import\s+tempfile\b'),
        "subprocess":(r'\bsubprocess\.', r'^\s*import\s+subprocess\b'),
        "threading": (r'\bthreading\.', r'^\s*import\s+threading\b'),
        "csv":       (r'\bcsv\.', r'^\s*import\s+csv\b'),
        "copy":      (r'\bcopy\.', r'^\s*import\s+copy\b'),
    }
    missing_imports = []
    for mod, (use_pat, imp_pat) in STDLIB_AUTO.items():
        uses_it    = bool(re.search(use_pat, fixed))
        imports_it = bool(re.search(imp_pat, fixed, re.MULTILINE))
        if uses_it and not imports_it:
            missing_imports.append(f"import {mod}")
    if missing_imports:
        first_imp = re.search(r'^(import |from )', fixed, re.MULTILINE)
        if first_imp:
            fixed = fixed[:first_imp.start()] + "\n".join(missing_imports) + "\n" + fixed[first_imp.start():]
        else:
            fixed = "\n".join(missing_imports) + "\n" + fixed

    return fixed

# ══════════════════════════════════════════════════════════════════════
#  EXCEL — lectura, modificación, comparación
# ══════════════════════════════════════════════════════════════════════
def leer_excel_contexto(ruta, max_filas=40):
    if not XL: return ""
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        partes = [f"── Excel: {os.path.basename(ruta)} ──"]
        for nh in wb.sheetnames[:4]:
            ws  = wb[nh]
            mf  = min(ws.max_row or 0, max_filas)
            rows = []
            for row in ws.iter_rows(max_row=mf, values_only=True):
                r = [str(v) if v is not None else "" for v in row]
                if any(r): rows.append(" │ ".join(r))
            partes.append(f"[{nh}  {ws.max_row or 0} filas]\n" + "\n".join(rows[:max_filas]))
        wb.close()
        return "\n\n".join(partes)[:4000]
    except Exception as e: return f"Error leyendo Excel: {e}"

def extraer_rutas_excel(txt):
    rutas = []
    for m in re.finditer(r'[A-Za-z]:[/\\][^\n"\'<>|*?]+\.xlsx?', txt, re.I):
        r = m.group(0).strip("\"' ")
        if os.path.exists(r) and r not in rutas: rutas.append(r)
    return rutas

def _xl_autofit(ws):
    """Autoajusta el ancho de todas las columnas."""
    try:
        for col_cells in ws.columns:
            # openpyxl puede retornar Cell sola si hay 1 sola columna
            if not isinstance(col_cells, (list, tuple)):
                col_cells = (col_cells,)
            if not col_cells:
                continue
            ml = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 3, 55)
    except Exception:
        pass  # autofit es cosmético, nunca debe romper el flujo

def _xl_estilo_tabla(ws, fila_header=1, color_header="1F6FBF"):
    """Aplica estilo de tabla profesional: encabezado azul, bordes, autofit."""
    try:
        fill_h = PatternFill("solid", fgColor=color_header)
        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ws[fila_header] puede ser Cell sola si hay 1 sola columna
        fila = ws[fila_header]
        if not isinstance(fila, (list, tuple)):
            fila = (fila,)
        for cell in fila:
            cell.fill = fill_h
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        fill_alt = PatternFill("solid", fgColor="EEF2F8")
        max_row = ws.max_row or 0
        if max_row > fila_header:
            for i, row in enumerate(ws.iter_rows(min_row=fila_header+1, max_row=max_row), 1):
                for cell in row:
                    cell.border = border
                    if i % 2 == 0:
                        cell.fill = fill_alt

        _xl_autofit(ws)
        ws.freeze_panes = f"A{fila_header+1}"
        if (ws.max_row or 0) > 1:
            ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass  # estilo es cosmético, nunca debe romper el flujo

def _xl_valor(v):
    """Convierte un valor a tipo nativo para openpyxl."""
    if v is None or v == "": return None
    s = str(v).strip()
    # Número
    try:
        if "." in s: return float(s.replace(",", "."))
        return int(s)
    except (ValueError, AttributeError): pass
    # Fecha dd/mm/yyyy
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try: return datetime.datetime.strptime(s, fmt)
        except ValueError: pass
    return s

def crear_excel_desde_datos(ruta, datos_hojas, estilo=True):
    """
    Crea (o sobreescribe) un Excel con datos reales.
    datos_hojas: dict { nombre_hoja: {'headers': [...], 'rows': [[...], ...]} }
               o list of lists (tabla simple)
    """
    if not XL: return "❌ Instala openpyxl"
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # quitar hoja vacía default

        if isinstance(datos_hojas, list):
            # Lista de listas → hoja única "Datos"
            datos_hojas = {"Datos": {"rows": datos_hojas}}

        for nombre_hoja, contenido in datos_hojas.items():
            ws = wb.create_sheet(title=str(nombre_hoja)[:31])
            headers = contenido.get("headers", [])
            rows    = contenido.get("rows", [])

            fila_ini = 1
            if headers:
                ws.append([str(h) for h in headers])
                fila_ini = 2

            for row in rows:
                ws.append([_xl_valor(v) for v in row])

            if estilo and ws.max_row >= 1:
                _xl_estilo_tabla(ws, fila_header=fila_ini)

        wb.save(ruta)
        sheets = list(datos_hojas.keys())
        total  = sum(len(c.get("rows",[])) for c in datos_hojas.values())
        # ── Registrar Excel creado ───────────────────────────────────────
        try:
            _registro_auto_excel(ruta, f"Excel creado: {', '.join(sheets)}")
        except Exception:
            pass
        return f"✅ Excel creado: {os.path.basename(ruta)}\n   Hojas: {', '.join(sheets)} | {total} filas de datos"
    except Exception as e:
        return f"❌ Error creando Excel: {e}"

def escribir_celdas(ruta, cambios, hoja=None):
    """
    Escribe valores en celdas específicas.
    cambios: dict { "A1": valor, "B2": valor, ... }
             o list [ (fila, col, valor), ... ]
    """
    if not XL: return "❌ Instala openpyxl"
    try:
        # Backup si el archivo existe antes de modificar
        if os.path.exists(ruta):
            try: _registro_backup_excel(ruta)
            except Exception: pass
        wb = openpyxl.load_workbook(ruta) if os.path.exists(ruta) else openpyxl.Workbook()
        ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
        hechos = []
        if isinstance(cambios, dict):
            for ref, val in cambios.items():
                ws[ref] = _xl_valor(val)
                hechos.append(f"{ref}={val}")
        elif isinstance(cambios, list):
            for fila, col, val in cambios:
                ws.cell(row=int(fila), column=int(col), value=_xl_valor(val))
                hechos.append(f"F{fila}C{col}={val}")
        wb.save(ruta)
        try: _registro_auto_excel(ruta, f"{len(hechos)} celdas escritas")
        except Exception: pass
        return f"✅ {len(hechos)} celdas escritas en {os.path.basename(ruta)}"
    except Exception as e:
        return f"❌ Error escribiendo celdas: {e}"

def modificar_excel(ruta, inst):
    if not XL: return "❌ Instala openpyxl"
    if not os.path.exists(ruta): return f"❌ Archivo no encontrado: {ruta}"
    # ── Backup automático antes de modificar ────────────────────────────
    try:
        _registro_backup_excel(ruta)
    except Exception:
        pass
    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        t  = inst.lower(); ok = []

        COLORES = {"naranja":"FF8C00","rojo":"C0392B","azul":"1F6FBF","verde":"1E8449",
                   "amarillo":"F4D03F","gris":"808080","morado":"7D3C98","celeste":"2E86C1"}

        # ── Pintar encabezados ──────────────────────────────────────────
        if any(p in t for p in ["color encabezad","pinta encabezad","encabezad"]):
            col_hex = next((v for k,v in COLORES.items() if k in t), "1F6FBF")
            fill = PatternFill("solid", fgColor=col_hex)
            for cell in (ws[1] if isinstance(ws[1], (list,tuple)) else (ws[1],)):
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")
            ok.append(f"Encabezados pintados ({col_hex})")

        # ── Autoajuste ──────────────────────────────────────────────────
        if any(p in t for p in ["autoajust","ajusta column","ancho column"]):
            _xl_autofit(ws); ok.append("Columnas autoajustadas")

        # ── Congelar ────────────────────────────────────────────────────
        if any(p in t for p in ["congela","inmoviliza","bloquea fila"]):
            ws.freeze_panes = "A2"; ok.append("Fila 1 congelada")

        # ── Negrita ─────────────────────────────────────────────────────
        if any(p in t for p in ["negrita encabez","bold encabez"]):
            for cell in (ws[1] if isinstance(ws[1], (list,tuple)) else (ws[1],)): cell.font = Font(bold=True)
            ok.append("Encabezados en negrita")

        # ── Bordes ──────────────────────────────────────────────────────
        if any(p in t for p in ["borde","tabla","format"]):
            thin = Side(style="thin")
            border = Border(left=thin,right=thin,top=thin,bottom=thin)
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 200)):
                for cell in row:
                    if cell.value is not None: cell.border = border
            ok.append("Bordes aplicados")

        # ── Filtros ─────────────────────────────────────────────────────
        if any(p in t for p in ["filtro","autofilter"]):
            if ws.dimensions:  # None si hoja vacía
                ws.auto_filter.ref = ws.dimensions
            ok.append("Filtros activados")

        # ── Estilo tabla completo ────────────────────────────────────────
        if any(p in t for p in ["estilo tabla","formato tabla","tabla profesional","tabla completa"]):
            _xl_estilo_tabla(ws); ok.append("Estilo tabla aplicado")

        # ── Agregar fila con datos entre comillas ────────────────────────
        if any(p in t for p in ["agrega fila","nueva fila","insertar fila","agregar fila"]):
            datos_q  = re.findall(r'"([^"]+)"', inst)
            datos_sb = re.findall(r'\[([^\]]+)\]', inst)
            if datos_q:
                ws.append([_xl_valor(v) for v in datos_q])
                ok.append(f"Fila agregada: {datos_q}")
            elif datos_sb:
                vals = [v.strip().strip("'\"") for v in datos_sb[0].split(",")]
                ws.append([_xl_valor(v) for v in vals])
                ok.append(f"Fila agregada: {vals}")

        # ── Actualizar celda específica ──────────────────────────────────
        # Ej: "actualiza A3 con 75" / "escribe 100 en B5"
        m_cel = re.search(r'\b([A-Z]{1,2}\d{1,4})\b.*?(\d+(?:\.\d+)?)', inst, re.I)
        if m_cel and any(p in t for p in ["actualiza","escribe","pon","ingresa","cambia","modifica celda","celda"]):
            ref = m_cel.group(1).upper()
            val = _xl_valor(m_cel.group(2))
            ws[ref] = val
            ok.append(f"Celda {ref} = {val}")

        # ── Agregar columna ──────────────────────────────────────────────
        if any(p in t for p in ["agrega columna","nueva columna","añade columna"]):
            nombre_col = re.search(r'(?:columna|column)[:\s]+["\']?([^"\']+)["\']?', inst, re.I)
            if nombre_col:
                new_col = (ws.max_column or 0) + 1
                ws.cell(row=1, column=new_col, value=nombre_col.group(1).strip())
                ok.append(f"Columna '{nombre_col.group(1).strip()}' agregada")

        wb.save(ruta)
        return "✅ " + ("; ".join(ok) if ok else "Archivo guardado.")
    except Exception as e: return f"❌ Error: {e}"
def comparar_excel(ruta1, ruta2):
    if not XL: return "❌ Instala openpyxl"
    try:
        wb1 = openpyxl.load_workbook(ruta1, read_only=True, data_only=True)
        wb2 = openpyxl.load_workbook(ruta2, read_only=True, data_only=True)
        ws1, ws2 = wb1.active, wb2.active
        f1, f2   = ws1.max_row or 0, ws2.max_row or 0
        c1, c2   = ws1.max_column or 0, ws2.max_column or 0
        h1 = [str(c.value or "") for c in ws1[1]]
        h2 = [str(c.value or "") for c in ws2[1]]
        nuevas = [h for h in h2 if h and h not in h1]
        elim   = [h for h in h1 if h and h not in h2]
        difs = 0
        for ri in range(2, min(f1, f2, 201)):
            for v1, v2 in zip([c.value for c in ws1[ri]], [c.value for c in ws2[ri]]):
                if str(v1 or "") != str(v2 or ""): difs += 1
        wb1.close(); wb2.close()
        res = [
            f"**{os.path.basename(ruta1)}** vs **{os.path.basename(ruta2)}**",
            f"Filas: {f1} → {f2}  |  Columnas: {c1} → {c2}",
        ]
        if nuevas: res.append(f"➕ Columnas nuevas: {', '.join(nuevas[:6])}")
        if elim:   res.append(f"➖ Columnas eliminadas: {', '.join(elim[:6])}")
        if difs:   res.append(f"🔍 **{difs} celdas distintas** (primeras 200 filas)")
        else:      res.append("✅ Contenido idéntico en las primeras 200 filas")
        return "\n".join(res)
    except Exception as e: return f"❌ Error comparando: {e}"

# ══════════════════════════════════════════════════════════════════════
#  DASHBOARD FINIQUITOS — Excel visual con xlsxwriter (gráficos + KPIs)
#  Origen: módulo crear_dashboard_finiquitos del prototipo AIEngine
# ══════════════════════════════════════════════════════════════════════
def crear_dashboard_finiquitos(datos_lista: list[dict],
                                output_path: str | None = None) -> str:
    """
    Genera un Excel visual con dashboard de finiquitos:
      · Hoja 'Data_Dashboard': resumen por Centro de Costo + % distribución
      · Hoja 'Dashboard Visual': KPIs (Total monto / Total procesados) +
        gráfico de torta por Centro de Costo
      · Hoja 'Detalle': todos los registros originales con estilos

    Parámetros:
        datos_lista: lista de dicts con keys recomendadas:
            'Encargado', 'Centro Costo', 'Monto', 'Rut', 'Nombre', 'Estado'
        output_path: ruta de salida. Si None, guarda en Descargas.

    Retorna: ruta del archivo generado o mensaje de error.
    """
    # ── Instalación automática de xlsxwriter si falta ────────────────
    try:
        import xlsxwriter as _xlw
    except ImportError:
        try:
            import subprocess as _sp
            _sp.run([sys.executable, "-m", "pip", "install", "xlsxwriter", "-q",
                     "--disable-pip-version-check"],
                    capture_output=True, timeout=120,
                    creationflags=getattr(_sp, "CREATE_NO_WINDOW", 0))
            import xlsxwriter as _xlw
        except Exception as e:
            return f"❌ No se pudo instalar xlsxwriter: {e}"
    try:
        import pandas as _pd
    except ImportError:
        return "❌ Instala pandas: pip install pandas"

    if not datos_lista:
        return "❌ No hay datos para generar el dashboard."

    # ── Ruta de salida ────────────────────────────────────────────────
    if not output_path:
        ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(DESCARGAS / f"Dashboard_Finiquitos_{ts}.xlsx")

    try:
        df = _pd.DataFrame(datos_lista)

        # Normalizar nombres de columnas (flexible)
        col_map = {}
        for col in df.columns:
            cl = col.lower().replace(" ","_").replace("é","e").replace("ó","o")
            if "centro" in cl or "cc" == cl:        col_map[col] = "Centro Costo"
            elif "monto" in cl or "costo" in cl:    col_map[col] = "Monto"
            elif "rut" in cl:                        col_map[col] = "Rut"
            elif "encargado" in cl or "analista" in cl: col_map[col] = "Encargado"
            elif "nombre" in cl:                     col_map[col] = "Nombre"
            elif "estado" in cl:                     col_map[col] = "Estado"
        df = df.rename(columns=col_map)

        for req in ["Centro Costo", "Monto"]:
            if req not in df.columns:
                df[req] = "Sin dato" if req == "Centro Costo" else 0

        df["Monto"] = _pd.to_numeric(df["Monto"], errors="coerce").fillna(0)

        # ── Resumen por Centro de Costo ───────────────────────────────
        resumen = df.groupby("Centro Costo").agg(
            Monto=("Monto", "sum"),
            Cantidad=("Monto", "count"),
        ).reset_index()
        total_monto = df["Monto"].sum()
        resumen["%"] = (resumen["Monto"] / total_monto * 100).round(1) if total_monto else 0
        resumen = resumen.sort_values("Monto", ascending=False)

        # ── Crear workbook con xlsxwriter ─────────────────────────────
        wb = _xlw.Workbook(output_path)

        # Paleta corporativa (usa colores del tema Matrix de REMI)
        C_HDR  = "#1F4E78"   # azul oscuro encabezados
        C_ACC  = "#2E75B6"   # azul acento
        C_OK   = "#1E8449"   # verde positivo
        C_ALT  = "#EBF5FB"   # filas alternas

        fmt_hdr   = wb.add_format({"bg_color":C_HDR,"font_color":"white","bold":True,
                                    "border":1,"align":"center","valign":"vcenter"})
        fmt_money = wb.add_format({"num_format":"$#,##0","bold":True,"font_color":C_HDR,
                                    "border":1})
        fmt_pct   = wb.add_format({"num_format":"0.0%","border":1,"align":"center"})
        fmt_num   = wb.add_format({"num_format":"#,##0","border":1,"align":"center"})
        fmt_txt   = wb.add_format({"border":1})
        fmt_alt   = wb.add_format({"border":1,"bg_color":C_ALT})
        fmt_kpi_t = wb.add_format({"bg_color":C_ACC,"font_color":"white","bold":True,
                                    "font_size":11,"align":"center","valign":"vcenter","border":1})
        fmt_kpi_v = wb.add_format({"num_format":"$#,##0","bold":True,"font_size":14,
                                    "font_color":C_OK,"align":"center","valign":"vcenter",
                                    "border":1})
        fmt_kpi_n = wb.add_format({"bold":True,"font_size":14,"font_color":C_HDR,
                                    "align":"center","valign":"vcenter","border":1})

        # ── Hoja 1: Data_Dashboard (datos para el gráfico) ───────────
        ws_data = wb.add_worksheet("Data_Dashboard")
        headers = ["Centro Costo","Monto","Cantidad","%"]
        for ci, h in enumerate(headers):
            ws_data.write(0, ci, h, fmt_hdr)
        for ri, row in resumen.iterrows():
            alt = fmt_alt if (ri % 2) else fmt_txt
            ws_data.write(ri+1, 0, row["Centro Costo"], alt)
            ws_data.write(ri+1, 1, row["Monto"],        fmt_money)
            ws_data.write(ri+1, 2, int(row["Cantidad"]),fmt_num)
            ws_data.write(ri+1, 3, row["%"]/100,        fmt_pct)
        ws_data.set_column("A:A", 24)
        ws_data.set_column("B:B", 16)
        ws_data.set_column("C:D", 12)

        # ── Hoja 2: Dashboard Visual (KPIs + gráfico) ────────────────
        ws_dash = wb.add_worksheet("Dashboard Visual")
        ws_dash.set_tab_color(C_ACC)
        ws_dash.hide_gridlines(2)

        # Título
        fmt_title = wb.add_format({"bold":True,"font_size":16,"font_color":C_HDR,
                                    "bottom":2,"bottom_color":C_ACC})
        ts_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        ws_dash.merge_range("B2:G2", f"DASHBOARD FINIQUITOS — {ts_str}", fmt_title)

        # KPI 1: Total monto
        ws_dash.merge_range("B4:C4", "TOTAL COSTO FINIQUITOS", fmt_kpi_t)
        ws_dash.merge_range("B5:C5", total_monto, fmt_kpi_v)
        ws_dash.set_row(4, 30); ws_dash.set_row(5, 36)

        # KPI 2: Total casos
        total_casos = len(df)
        ws_dash.merge_range("E4:F4", "TOTAL PROCESADOS", fmt_kpi_t)
        ws_dash.merge_range("E5:F5", total_casos, fmt_kpi_n)

        # KPI 3: Promedio por caso
        prom = total_monto / total_casos if total_casos else 0
        ws_dash.merge_range("B7:C7", "MONTO PROMEDIO / CASO", fmt_kpi_t)
        ws_dash.merge_range("B8:C8", prom, fmt_kpi_v)
        ws_dash.set_row(7, 30); ws_dash.set_row(8, 36)

        # KPI 4: Centros de costo involucrados
        n_cc = resumen["Centro Costo"].nunique()
        ws_dash.merge_range("E7:F7", "CENTROS DE COSTO", fmt_kpi_t)
        ws_dash.merge_range("E8:F8", n_cc, fmt_kpi_n)

        # ── Gráfico torta — distribución por Centro de Costo ─────────
        chart_pie = wb.add_chart({"type": "pie"})
        max_row = len(resumen) + 1
        chart_pie.add_series({
            "name":       "Distribución de costos",
            "categories": f"=Data_Dashboard!$A$2:$A${max_row}",
            "values":     f"=Data_Dashboard!$B$2:$B${max_row}",
            "data_labels":{"percentage": True, "category": True,
                            "position": "outside_end", "font": {"size": 9}},
        })
        chart_pie.set_title({"name": "Impacto por Centro de Costo"})
        chart_pie.set_style(10)
        chart_pie.set_size({"width": 480, "height": 320})
        ws_dash.insert_chart("B10", chart_pie)

        # ── Gráfico barras — top centros de costo ────────────────────
        chart_bar = wb.add_chart({"type": "bar"})
        chart_bar.add_series({
            "name":       "Monto por CC",
            "categories": f"=Data_Dashboard!$A$2:$A${max_row}",
            "values":     f"=Data_Dashboard!$B$2:$B${max_row}",
            "fill":       {"color": C_ACC},
            "data_labels":{"value": True, "num_format": "$#,##0",
                            "font": {"size": 8}},
        })
        chart_bar.set_title({"name": "Monto por Centro de Costo"})
        chart_bar.set_style(11)
        chart_bar.set_size({"width": 420, "height": 300})
        ws_dash.insert_chart("I10", chart_bar)

        ws_dash.set_column("A:A", 3)
        ws_dash.set_column("B:G", 14)

        # ── Hoja 3: Detalle completo ──────────────────────────────────
        ws_det = wb.add_worksheet("Detalle")
        det_cols = [c for c in df.columns if c != "index"]
        for ci, h in enumerate(det_cols):
            ws_det.write(0, ci, h, fmt_hdr)
        for ri, row in df.iterrows():
            alt = fmt_alt if (ri % 2) else fmt_txt
            for ci, col in enumerate(det_cols):
                v = row[col]
                if col == "Monto":
                    ws_det.write(ri+1, ci, float(v) if v else 0, fmt_money)
                else:
                    ws_det.write(ri+1, ci, str(v) if v is not None else "", alt)
        ws_det.set_column("A:Z", 18)
        ws_det.autofilter(0, 0, len(df), len(det_cols)-1)
        ws_det.freeze_panes(1, 0)

        wb.close()

        # ── Registrar en REMI ─────────────────────────────────────────
        try:
            _registro_auto_excel(output_path, "Dashboard Finiquitos generado")
        except Exception:
            pass

        nombre = os.path.basename(output_path)
        return (
            "✅ **Dashboard generado:** `" + nombre + "`\n\n"
            "  · Casos: **" + str(total_casos) + "**\n"
            "  · Monto total: **$" + f"{total_monto:,.0f}" + "**\n"
            "  · Centros de costo: **" + str(n_cc) + "**\n"
            "  · Hojas: Data_Dashboard · Dashboard Visual · Detalle\n\n"
            "Guardado en: `" + output_path + "`"
        )

    except Exception as e:
        import traceback
        return (
            "❌ Error generando dashboard: " + str(e) + "\n\n"
            "```\n" + traceback.format_exc()[-600:] + "\n```"
        )


# ══════════════════════════════════════════════════════════════════════
#  VOZ
# ══════════════════════════════════════════════════════════════════════
def escuchar():
    if not VOZ: return None, "SpeechRecognition no instalado"
    try:
        r = sr.Recognizer()
        with sr.Microphone() as src:
            r.adjust_for_ambient_noise(src, duration=0.5)
            audio = r.listen(src, timeout=8, phrase_time_limit=18)
        return r.recognize_google(audio, language="es-CL"), None
    except sr.WaitTimeoutError: return None, "No escuché nada (timeout)."
    except Exception as e: return None, f"Error voz: {e}"

# ══════════════════════════════════════════════════════════════════════
#  NOTIFICACIONES Y ALARMAS
# ══════════════════════════════════════════════════════════════════════
def notificar(titulo, msg):
    if NOTIF:
        try: notification.notify(title=titulo, message=msg, timeout=10)
        except: pass

class GestorAlarmas:
    def __init__(self, cb):
        threading.Thread(target=self._loop, args=(cb,), daemon=True).start()

    def _loop(self, cb):
        while True:
            try:
                now = datetime.datetime.now().strftime("%H:%M")
                al  = _rl(F["alarm"]); ch = False
                for a in al:
                    if a.get("activa") and a.get("hora") == now:
                        a["activa"] = False; ch = True
                        cb(a.get("mensaje","Alarma"))
                if ch: _w(F["alarm"], al)
            except: pass
            time.sleep(28)

# ══════════════════════════════════════════════════════════════════════
#  MONITOR DE PROCESOS
# ══════════════════════════════════════════════════════════════════════
class MonitorProcesos:
    def __init__(self):
        self._activo=False; self._nombre=""; self._eventos=[]; self._snaps={}

    def iniciar(self, nombre, carpetas):
        self._nombre=nombre; self._activo=True; self._eventos=[]; self._snaps={}
        validas = [c for c in carpetas if os.path.exists(c)]
        for c in validas:
            for f in glob.glob(os.path.join(c,"*.xlsx"))+glob.glob(os.path.join(c,"*.xlsm")):
                self._snaps[f] = os.path.getmtime(f)
        threading.Thread(target=self._loop, daemon=True).start()
        return f"🔴 Grabando proceso **'{nombre}'** — vigilando {len(validas)} carpetas"

    def detener(self):
        self._activo = False
        if not self._eventos: return "Detenido sin cambios detectados."
        proc = {
            "nombre": self._nombre,
            "fecha":  datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "eventos":self._eventos[:30],
            "resumen":f"{len(self._eventos)} cambios en " + ", ".join(
                list({e["archivo"] for e in self._eventos[:6]})),
            "archivos":list({e["archivo"] for e in self._eventos}),
        }
        ps = _rl(F["proc"]); ps.append(proc); _w(F["proc"], ps)
        return f"✅ Proceso **'{self._nombre}'** guardado con {len(self._eventos)} eventos"

    def _loop(self):
        while self._activo:
            time.sleep(8)
            try:
                for c in (c for c in RUTAS.values() if os.path.exists(c)):
                    for f in glob.glob(os.path.join(c,"*.xlsx"))+glob.glob(os.path.join(c,"*.xlsm")):
                        mt = os.path.getmtime(f)
                        if f not in self._snaps:
                            self._snaps[f]=mt
                            self._eventos.append({"tipo":"creado","archivo":os.path.basename(f)})
                        elif mt > self._snaps[f]+2:
                            self._snaps[f]=mt
                            self._eventos.append({"tipo":"modificado","archivo":os.path.basename(f)})
            except: pass

    @property
    def grabando(self): return self._activo
    @property
    def nombre(self): return self._nombre

MONITOR = MonitorProcesos()

# ══════════════════════════════════════════════════════════════════════
#  AVATAR — Robot corporativo serio (no caricaturesco)
#  Dibujado a cualquier tamaño vía factor escala
# ══════════════════════════════════════════════════════════════════════
def dibujar_remi(cv, sz, blink=False, habla=False):
    """
    Logo REMI v10 — Hexágono tech + nodo central + pulso de actividad.
    Inspirado en el prototipo RemiUI: limpio, corporativo, sin caricatura.
    Adapta colores al tema activo. Funciona desde 32px hasta 128px.
    """
    import math as _math
    cv.delete("all")
    cx = sz / 2
    cy = sz / 2
    k  = sz / 80.0

    def s(v):  return v * k
    def si(v): return int(v * k)

    # ── Paleta del tema activo ────────────────────────────────
    C_BG    = T.get("sb",   "#101214")
    C_HEX   = T.get("card", "#1c2027")      # relleno hexágono
    C_RING  = T.get("a1",   "#3fb950")      # borde hexágono / acento
    C_RING2 = T.get("a2",   "#52c768")      # acento secundario
    C_DARK  = T.get("bg",   "#141619")      # sombra
    C_LED   = T.get("ok",   "#3fb950")      # LED activo
    C_LED_H = T.get("warn", "#e3b341")      # LED hablando
    C_TXT   = T.get("tx",   "#eaecef")      # texto / nodo

    cv.configure(bg=C_BG)

    # ── Función auxiliar: hexágono por puntos ─────────────────
    def hex_pts(cx, cy, r, rot=0):
        pts = []
        for i in range(6):
            ang = _math.radians(60 * i + rot)
            pts += [cx + r * _math.cos(ang), cy + r * _math.sin(ang)]
        return pts

    R_OUT = s(30)   # radio exterior hexágono
    R_MID = s(24)   # radio medio (relleno)
    R_IN  = s(17)   # radio interior (zona logo)
    R_NRO = s(5)    # nodo central

    # ── Sombra sutil (offset 2px) ─────────────────────────────
    sh = s(1.5)
    cv.create_polygon(hex_pts(cx+sh, cy+sh, R_OUT, rot=30),
                      fill=C_DARK, outline="", smooth=False)

    # ── Hexágono exterior — borde acento ──────────────────────
    cv.create_polygon(hex_pts(cx, cy, R_OUT, rot=30),
                      fill=C_RING, outline="", smooth=False)

    # ── Hexágono relleno interior ─────────────────────────────
    cv.create_polygon(hex_pts(cx, cy, R_MID, rot=30),
                      fill=C_HEX, outline="", smooth=False)

    # ── 6 líneas de "circuito" desde el centro a cada vértice ─
    # Solo 3 de los 6 para no saturar (cada 2do)
    for i in [0, 2, 4]:
        ang = _math.radians(60 * i + 30)
        x1 = cx + s(6)  * _math.cos(ang)
        y1 = cy + s(6)  * _math.sin(ang)
        x2 = cx + s(20) * _math.cos(ang)
        y2 = cy + s(20) * _math.sin(ang)
        cv.create_line(x1, y1, x2, y2,
                       fill=C_RING, width=max(1, si(0.8)), capstyle="round")
        # Nodo pequeño al final
        nr = s(1.6)
        cv.create_oval(x2-nr, y2-nr, x2+nr, y2+nr, fill=C_RING2, outline="")

    # ── Nodo central (círculo con acento) ─────────────────────
    cv.create_oval(cx-R_NRO, cy-R_NRO, cx+R_NRO, cy+R_NRO,
                   fill=C_RING, outline="")
    cv.create_oval(cx-s(2.8), cy-s(2.8), cx+s(2.8), cy+s(2.8),
                   fill=C_HEX, outline="")

    # ── Texto "R" centrado — identidad REMI ───────────────────
    fz = max(8, si(14))
    if not blink:
        cv.create_text(cx, cy,
                       text="R",
                       fill=C_RING2,
                       font=("Segoe UI", fz, "bold"))
    else:
        # Parpadeo: punto pulsante en lugar de R
        pr = s(3)
        cv.create_oval(cx-pr, cy-pr, cx+pr, cy+pr,
                       fill=C_RING2, outline="")

    # ── Indicador de estado (LED abajo derecha) ───────────────
    led_x = cx + s(18)
    led_y = cy + s(18)
    led_r = s(3.5)
    led_c = C_LED_H if habla else C_LED
    # Halo suave
    cv.create_oval(led_x-led_r-s(1.5), led_y-led_r-s(1.5),
                   led_x+led_r+s(1.5), led_y+led_r+s(1.5),
                   fill=C_DARK, outline="")
    cv.create_oval(led_x-led_r, led_y-led_r,
                   led_x+led_r, led_y+led_r,
                   fill=led_c, outline="")
    # Brillo interno
    bx, by, br = led_x - s(1), led_y - s(1), s(1.2)
    cv.create_oval(bx-br, by-br, bx+br, by+br,
                   fill="white", outline="")

    # ── Pulso animado cuando habla (arco giratorio) ───────────
    if habla:
        arc_r = s(22)
        cv.create_arc(cx-arc_r, cy-arc_r, cx+arc_r, cy+arc_r,
                      start=30, extent=80,
                      style="arc", outline=C_RING2,
                      width=max(1, si(1.2)))


# ══════════════════════════════════════════════════════════════════════
#  WIDGET: BLOQUE DE CÓDIGO
# ══════════════════════════════════════════════════════════════════════
class BloqueCode(tk.Frame):
    LANG_COLOR = {
        "vba":"#d4a520","python":"#4584b6","sql":"#f08040","powershell":"#3a6ba8",
        "bash":"#50a050","json":"#8888f0","xml":"#c07050","javascript":"#f0d040",
        "batch":"#a0a0a0","vbscript":"#d0a030",
    }
    def __init__(self, parent, lang, code):
        super().__init__(parent, bg=T["code"], bd=0,
                         highlightthickness=1, highlightbackground=T["br"])
        lang_c = self.LANG_COLOR.get(lang.lower(), T["a2"])
        lbl    = lang.upper() or "CÓDIGO"
        n_lin  = code.count("\n") + 1

        # ── Header ──
        bar = tk.Frame(self, bg=T["br"]); bar.pack(fill="x")
        # Badge lenguaje
        tk.Label(bar, text=f" {lbl} ", font=(FMONO,8,"bold"),
                 bg=lang_c, fg=T["bg"], padx=4, pady=2).pack(side="left")
        tk.Label(bar, text=f" {n_lin} líneas", font=(FONT,8),
                 bg=T["br"], fg=T["t2"]).pack(side="left", padx=6)
        self._copybtn = tk.Button(
            bar, text="  📋  Copiar", font=(FONT,8),
            bg=T["br"], fg=T["t2"], bd=0, padx=8, pady=3,
            activebackground=T["ok"], activeforeground="white",
            cursor="hand2", relief="flat",
            command=lambda: self._copiar(code))
        self._copybtn.pack(side="right")

        # ── Botón Ejecutar (solo Python) ──
        if lang.lower() == "python":
            self._runbtn = tk.Button(
                bar, text="  ▶  Ejecutar", font=(FONT,8),
                bg=T["br"], fg="#4aaa70", bd=0, padx=8, pady=3,
                activebackground="#1e4a1e", activeforeground="#4aaa70",
                cursor="hand2", relief="flat",
                command=lambda: self._ejecutar(code))
            self._runbtn.pack(side="right")

        # ── Botón Abrir en Excel si hay ruta xlsx en el código ──
        rutas_xl = re.findall(r'[A-Za-z]:[/\\][^\n"\' <>|*?\s]+\.xlsx?', code, re.I)
        rutas_xl = [r.strip("\"'(),;") for r in rutas_xl]
        if rutas_xl:
            ruta_xl = rutas_xl[0]
            self._excel_btn = tk.Button(
                bar, text="  📂  Abrir en Excel", font=(FONT,8),
                bg=T["br"], fg=T["a2"], bd=0, padx=8, pady=3,
                activebackground=T["a1"], activeforeground="white",
                cursor="hand2", relief="flat",
                command=lambda r=ruta_xl: self._abrir_excel(r))
            self._excel_btn.pack(side="right")

        # ── Área código ──
        h = min(n_lin + 1, 22)
        self._txt = tk.Text(self, font=(FMONO, 10), bg=T["code"], fg="#e6edf3",
                            wrap="none", relief="flat", padx=12, pady=8, bd=0,
                            height=h, selectbackground=T["a1"],
                            insertbackground="white", cursor="arrow",
                            yscrollcommand=lambda *a: None)
        self._txt.insert("1.0", code)
        self._txt.configure(state="disabled")
        self._txt.pack(fill="x")

        # Scrollbar horizontal solo si líneas largas
        if max((len(l) for l in code.split("\n")), default=0) > 88:
            sb = tk.Scrollbar(self, orient="horizontal", command=self._txt.xview,
                              bg=T["br"], troughcolor=T["bg"])
            self._txt.configure(xscrollcommand=sb.set); sb.pack(fill="x")

    def _copiar(self, code):
        self.clipboard_clear(); self.clipboard_append(code)
        self._copybtn.configure(text="  ✅  Copiado", fg=T.get("ok","#4aaa70"))
        self.after(2200, lambda: self._copybtn.configure(text="  📋  Copiar", fg=T["t2"]))

    def _abrir_excel(self, ruta):
        """Abre el archivo Excel con la aplicación del sistema."""
        try:
            os.startfile(ruta)
        except Exception:
            try:
                subprocess.Popen(["start", "", ruta], shell=True)
            except Exception as e:
                tk.messagebox.showerror("Error", f"No se pudo abrir:\n{ruta}\n\n{e}")

    def _ejecutar(self, code):
        """Valida y ejecuta el bloque Python."""
        # ── Registrar script antes de ejecutar ───────────────────────────
        try:
            _registro_auto_script(code, "script_ejecutado")
        except Exception:
            pass
        try:
            # ── Pre-validar y parchear ────────────────────────────────────
            try:
                errores = _prevalidar_codigo(code)
            except Exception:
                errores = []
            try:
                code = _parchear_codigo(code)
            except Exception:
                pass  # si el parcheo falla, ejecutar el código original

            try:
                self._runbtn.configure(text="  ⏳  Ejecutando…", fg=T.get("warn","#d49040"), state="disabled")
            except: pass

            # Área de output (crear si no existe)
            if not hasattr(self, "_out_txt"):
                self._out_txt = tk.Text(
                    self, font=(FMONO, 9), bg="#080c10", fg="#a0e0a0",
                    wrap="word", relief="flat", padx=8, pady=6, bd=0,
                    height=6, state="disabled")
                self._out_txt.pack(fill="x")

            self._out_txt.configure(state="normal")
            self._out_txt.delete("1.0", "end")
            # Mostrar advertencias pre-ejecución si las hay
            if errores:
                aviso = "⚠ Correcciones aplicadas:\n" + "\n".join(f"  • {e}" for e in errores) + "\n\nEjecutando…\n"
            else:
                aviso = "Ejecutando…\n"
            self._out_txt.insert("1.0", aviso)
            self._out_txt.configure(state="disabled")
        except Exception as _init_err:
            # Si falla la inicialización UI, al menos intentar ejecutar
            try: self._runbtn.configure(text="  ⏳  Ejecutando…", state="disabled")
            except: pass

        def _run():
            # ── Resolver Python correcto ─────────────────────────────────
            py = sys.executable
            try:
                app_dir = Path(sys.executable).resolve().parent
                for base in [app_dir, app_dir.parent, Path.cwd()]:
                    for venv_rel in [".venv/Scripts/python.exe",
                                     ".venv/bin/python",
                                     "venv/Scripts/python.exe",
                                     "venv/bin/python"]:
                        candidate = base / venv_rel
                        if candidate.exists():
                            py = str(candidate); break
                    else:
                        continue
                    break
            except Exception:
                pass

            # ── Helper: verificar si un módulo existe en 'py' ────────────
            def _mod_ok(mod):
                """Comprueba en el Python ejecutor (no en el de Remi)."""
                try:
                    r = subprocess.run([py, "-c", f"import {mod}"],
                                       capture_output=True, timeout=15)
                    return r.returncode == 0
                except Exception:
                    return False

            # ── Helper: instalar con feedback al usuario ──────────────────
            def _instalar(pkg, display_name=None):
                dn = display_name or pkg
                def _msg(txt):
                    try:
                        self._out_txt.configure(state="normal")
                        self._out_txt.delete("1.0", "end")
                        self._out_txt.insert("1.0", txt)
                        self._out_txt.configure(state="disabled")
                    except Exception: pass
                self.after(0, _msg, f"📦 Instalando {dn}…\n(esto puede tomar 1-2 minutos la primera vez)")
                try:
                    r = subprocess.run(
                        [py, "-m", "pip", "install", pkg, "-q",
                         "--disable-pip-version-check"],
                        capture_output=True, timeout=180,
                        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
                    if r.returncode != 0:
                        err_txt = r.stderr.decode(errors="replace").strip()
                        logging.warning("pip install %s falló: %s", pkg, err_txt)
                        return False
                    return True
                except Exception as _pip_err:
                    logging.warning("pip install %s: %s", pkg, _pip_err)
                    return False

            # ── Pre-instalar dependencias detectadas en el código ─────────
            PKG_MAP = {
                "openpyxl":   "openpyxl",
                "pandas":     "pandas",
                "numpy":      "numpy",
                "requests":   "requests",
                "bs4":        "beautifulsoup4",
                "PIL":        "pillow",
                "cv2":        "opencv-python",
                "sklearn":    "scikit-learn",
                "matplotlib": "matplotlib",
                "seaborn":    "seaborn",
                "xlrd":       "xlrd",
                "xlwt":       "xlwt",
            }
            # Siempre asegurar openpyxl y pandas en el Python ejecutor
            for _pkg in ["openpyxl", "pandas"]:
                if not _mod_ok(_pkg):
                    _instalar(_pkg)

            import_names = re.findall(r'^\s*(?:import|from)\s+([\w]+)', code, re.MULTILINE)
            faltantes = []
            for mod in import_names:
                if not _mod_ok(mod):
                    pkg = PKG_MAP.get(mod, mod)
                    ok = _instalar(pkg, display_name=mod)
                    if not ok:
                        faltantes.append(mod)

            # Si quedaron módulos sin instalar, informar claramente
            if faltantes:
                output = (
                    f"❌ No se pudo instalar automáticamente:\n"
                    + "\n".join(f"  • {m}" for m in faltantes)
                    + "\n\nInstálalos manualmente en la terminal:\n"
                    + "  " + "; ".join(f"pip install {PKG_MAP.get(m,m)}" for m in faltantes)
                )
                def _show_err():
                    try:
                        self._out_txt.configure(state="normal")
                        self._out_txt.delete("1.0","end")
                        self._out_txt.insert("1.0", output)
                        self._out_txt.configure(state="disabled", fg="#f08080")
                        self._runbtn.configure(text="  ▶  Ejecutar", fg="#4aaa70", state="normal")
                    except: pass
                self.after(0, _show_err)
                return

            code_fixed = code  # parcheo ya aplicado

            try:
                result = subprocess.run(
                    [py, "-c", code_fixed],
                    capture_output=True, text=True, timeout=60,
                    encoding="utf-8", errors="replace",
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0)
                )
                out = result.stdout.strip()
                err = result.stderr.strip()

                # Filtrar ruido de pip/warnings
                noise = ["WARNING","DeprecationWarning","FutureWarning",
                         "pip","notice","UserWarning","InsecureRequest"]
                err_lines = [l for l in err.splitlines()
                             if l.strip() and not any(x in l for x in noise)]
                err_clean = "\n".join(err_lines)

                if result.returncode == 0:
                    if out:
                        output = f"✅ Output:\n{out}"
                    else:
                        output = "✅ Ejecutado correctamente."
                    if err_clean:
                        output += f"\n\n⚠️ Avisos:\n{err_clean}"
                else:
                    # Mensaje de error más amigable
                    if "ModuleNotFoundError" in err_clean:
                        mod_falt = re.search(r"No module named '([\w\.]+)'", err_clean)
                        mod_n = mod_falt.group(1) if mod_falt else "desconocido"
                        output = (f"❌ Falta el módulo **{mod_n}**\n\n"
                                  f"Intenta instalar con:\n  pip install {mod_n}\n\n"
                                  f"Error completo:\n{err_clean}")
                    elif "FileNotFoundError" in err_clean:
                        output = (f"❌ Archivo no encontrado\n\n"
                                  f"Verifica que la ruta existe y la unidad de red está conectada.\n\n"
                                  f"{err_clean}")
                    elif "PermissionError" in err_clean:
                        output = (f"❌ Sin permisos de acceso\n\n"
                                  f"El archivo puede estar abierto en Excel. Ciérralo e intenta de nuevo.\n\n"
                                  f"{err_clean}")
                    else:
                        output = f"❌ Error:\n{err_clean}" if err_clean else f"❌ Falló (código {result.returncode})"
            except subprocess.TimeoutExpired:
                output = "⏱ Timeout: el script tardó más de 60 segundos."
            except Exception as e:
                output = f"❌ Error al ejecutar: {e}"

            def _upd():
                try:
                    self._out_txt.configure(state="normal")
                    self._out_txt.delete("1.0", "end")
                    self._out_txt.insert("1.0", output)
                    lines = output.count("\n") + 1
                    color = "#a0e0a0" if "✅" in output else "#f08080"
                    self._out_txt.configure(height=min(lines + 1, 20),
                                            state="disabled", fg=color)
                    self._runbtn.configure(text="  ▶  Ejecutar", fg="#4aaa70", state="normal")
                    # ── Si el código creó un xlsx, ofrecer botón de apertura ──
                    if "✅" in output:
                        rutas_creadas = re.findall(
                            r'[A-Za-z]:[/\\][^\n"\'<>|*?\s]+\.xlsx?', code_fixed, re.I)
                        rutas_creadas = [r.strip("\"'(),;") for r in rutas_creadas]
                        rutas_ok = [r for r in rutas_creadas if os.path.exists(r)]
                        if rutas_ok and not hasattr(self, "_open_btn"):
                            ruta_xl = rutas_ok[0]
                            self._open_btn = tk.Button(
                                self,
                                text=f"  📂  Abrir en Excel: {os.path.basename(ruta_xl)}",
                                font=(FONT, 8), bg=T["br"], fg=T["a2"],
                                bd=0, padx=10, pady=4,
                                activebackground=T["a1"], activeforeground="white",
                                cursor="hand2", relief="flat",
                                command=lambda r=ruta_xl: self._abrir_excel(r))
                            self._open_btn.pack(fill="x", pady=(2, 0))
                except: pass
            try:
                self.after(0, _upd)
            except: pass

        threading.Thread(target=_run, daemon=True).start()


# ══════════════════════════════════════════════════════════════════════
#  WIDGET: BURBUJA DE MENSAJE
# ══════════════════════════════════════════════════════════════════════
class Burbuja(tk.Frame):
    """
    Burbuja de chat. Soporta:
    - Texto con **bold**, listas, párrafos
    - Bloques ```código```
    - Streaming (update incremental, re-render final)
    - Imagen adjunta
    """
    # ── Patrones para detección en _render_texto ──────────────────────
    _URL_PAT = re.compile(r"(https?://[^\s]+)")
    _IDEA_KW = re.compile(
        r"(💡|📋|macro|vba|matriz|sharepoint|onedrive|365|"
        r"inferencia|consolidar|liquidaci|comisi\w+|dotaci\w+|pivot|powerquery)",
        re.IGNORECASE,
    )
    TIPOS = {
        "user":   ("bub_u", "a2",  "a2"),
        "bot":    ("bub_b", "acc", "acc"),
        "system": ("card",  "br",  "t2"),
        "error":  ("card",  "err", "err"),
        "ok":     ("card",  "ok",  "ok"),
        "warn":   ("card",  "warn","warn"),
    }

    def __init__(self, parent, texto="", tipo="bot", img_path=None):
        bg_k, border_k, label_k = self.TIPOS.get(tipo, self.TIPOS["bot"])
        super().__init__(parent, bg=T[bg_k],
                         highlightthickness=1, highlightbackground=T[border_k], bd=0)
        self._bg       = T[bg_k]
        self._tipo     = tipo
        self._stream_l = None   # último Label para streaming (siempre el más reciente)
        # ── Borde redondeado visual (estilo BurbujaChat del prototipo) ─
        # Padding extra en burbujas user para simular corner_radius=15
        self._corner_px = 10   # pixeles de "radio" simulado con padding

        # ── Avatar / ícono ──
        if tipo == "bot":
            av = tk.Canvas(self, width=22, height=22, bg=T[bg_k], highlightthickness=0)
            av.pack(side="left", anchor="n", padx=(10,0), pady=10)
            # Mini avatar circular — refleja el estilo del avatar principal
            a1 = T.get("a1","#4d85cc"); bg_ = T.get("bg","#13151b")
            av.create_oval(1, 1, 21, 21, fill=a1, outline="")
            av.create_oval(3, 3, 19, 19, fill=T.get("card","#1c1f28"), outline="")
            # Ojos
            av.create_oval(6, 8, 9, 11,  fill=a1, outline="")
            av.create_oval(13, 8, 16, 11, fill=a1, outline="")
            # Sonrisa (arco)
            av.create_arc(6, 11, 16, 18, start=200, extent=140,
                          style="arc", outline=a1, width=1)
            ctn = tk.Frame(self, bg=T[bg_k]); ctn.pack(side="left", fill="both",
                                                          expand=True, padx=(4,12), pady=4)
        elif tipo == "user":
            # Usuario: margen izquierdo amplio → alineación visual a la derecha
            # (patrón BurbujaChat emisor="User": side=right + corner_radius)
            ctn = tk.Frame(self, bg=T[bg_k]); ctn.pack(fill="both", expand=True,
                                                          padx=(60,14), pady=5)
        else:
            ctn = tk.Frame(self, bg=T[bg_k]); ctn.pack(fill="both", expand=True, padx=12, pady=4)

        self._ctn = ctn
        self._wrap_labels: list[tk.Label] = []

        # ── Header compacto ──
        hdr = tk.Frame(ctn, bg=T[bg_k]); hdr.pack(fill="x", pady=(4,2))
        LABELS = {"user": CFG.get("nombre_usuario","Yerko"),
                  "bot": "Remi", "system": "Sistema", "error": "Error", "ok": "OK", "warn": "Aviso"}
        tk.Label(hdr, text=LABELS.get(tipo,"Remi"), font=(FONT,9,"bold"),
                 bg=T[bg_k], fg=T[label_k]).pack(side="left")
        tk.Label(hdr, text=datetime.datetime.now().strftime("%H:%M"),
                 font=(FONT,8), bg=T[bg_k], fg=T["t2"]).pack(side="right")

        # ── Imagen adjunta ──
        if img_path and PIL_OK:
            try:
                im = Image.open(img_path); im.thumbnail((300,220))
                ph = ImageTk.PhotoImage(im)
                l  = tk.Label(ctn, image=ph, bg=T[bg_k]); l.image = ph
                l.pack(anchor="w", pady=4)
            except: pass

        # ── Contenido ──
        self._render(texto)

        # Ajustar wrap al ancho real (ocupar mejor el espacio)
        try:
            self.bind("<Configure>", self._reflow)
            self._reflow()
        except Exception:
            pass

    def _wrap_px(self) -> int:
        try:
            w = int(self.winfo_width() or 0)
        except Exception:
            w = 0
        # Márgenes aprox: sidebar + padding interno
        return max(420, w - 120)

    def _reflow(self, _e=None):
        try:
            wp = self._wrap_px()
            for l in self._wrap_labels:
                try:
                    l.configure(wraplength=wp)
                except Exception:
                    pass
        except Exception:
            pass

    def _render(self, texto):
        if not texto: return
        pat = re.compile(r'```(\w*)\n?(.*?)```', re.DOTALL)
        cur = 0
        for m in pat.finditer(texto):
            if m.start() > cur:
                self._render_texto(texto[cur:m.start()])
            BloqueCode(self._ctn, m.group(1).strip(), m.group(2).strip()).pack(
                fill="x", pady=(4,6))
            cur = m.end()
        if cur < len(texto):
            self._render_texto(texto[cur:])

    def _render_texto(self, texto):
        for linea in texto.split("\n"):
            stripped = linea.strip()
            if not stripped:
                tk.Frame(self._ctn, bg=self._bg, height=4).pack(fill="x"); continue

            # Listas con viñeta
            if stripped.startswith(("- ","• ","* ")):
                stripped = "  ·  " + stripped[2:]
                linea = stripped

            # Procesar bold **...**
            partes = re.split(r'\*\*(.+?)\*\*', linea)
            if len(partes) == 1:
                l = tk.Label(self._ctn, text=linea, font=(FONT,11),
                             bg=self._bg, fg=T["tx"],
                             wraplength=self._wrap_px(), justify="left", anchor="w")
                l.pack(anchor="w", fill="x", padx=2, pady=1)
                self._stream_l = l   # SIEMPRE actualizar (fix del bug de bold)
                self._wrap_labels.append(l)
            else:
                f = tk.Frame(self._ctn, bg=self._bg); f.pack(anchor="w", fill="x", pady=1)
                for i, p in enumerate(partes):
                    if not p: continue
                    lbl = tk.Label(f, text=p,
                                   font=(FONT,11,"bold") if i%2==1 else (FONT,11),
                                   bg=self._bg,
                                   fg=T["a2"] if i%2==1 else T["tx"],
                                   anchor="w")
                    lbl.pack(side="left")
                    self._stream_l = lbl   # SIEMPRE actualizar
                    self._wrap_labels.append(lbl)

    # ── Streaming ──
    def stream_upd(self, txt):
        """Actualiza el último label con texto+cursor mientras llega el stream."""
        if self._stream_l:
            try:
                # Mostrar una cola del texto (no solo la última línea), así no parece "cortado".
                lines = txt.split("\n")
                tail = "\n".join(lines[-10:]) if len(lines) > 10 else txt
                tail = tail[-2200:]  # límite razonable para un Label
                self._stream_l.configure(text=tail + "▌")
            except: pass

    def stream_fin(self, txt):
        """Re-render completo con el texto final."""
        # Destruir todo excepto el header (primer hijo de _ctn)
        hijos = list(self._ctn.winfo_children())
        for w in hijos[1:]:
            try: w.destroy()
            except: pass
        self._stream_l = None
        self._wrap_labels = []
        try:
            self._render(txt)
        except Exception:
            # Fallback: si el parser de markdown/código falla, al menos mostrar el texto completo.
            l = tk.Label(self._ctn, text=txt, font=(FONT,11),
                         bg=self._bg, fg=T["tx"],
                         wraplength=self._wrap_px(), justify="left", anchor="w")
            l.pack(anchor="w", fill="x", padx=2, pady=1)
            self._stream_l = l
            self._wrap_labels.append(l)
        # Flash de borde para indicar que terminó
        try:
            self.configure(highlightbackground=T["ok"])
            self.after(800, lambda: (
                self.configure(highlightbackground=T[self.TIPOS.get(self._tipo,("","acc",""))[1]])
                if self.winfo_exists() else None
            ))
        except: pass


# ══════════════════════════════════════════════════════════════════════
#  CHIPS DE ACCESO RÁPIDO — acciones frecuentes de remuneraciones
# ══════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════
#  REMI DAEMON — Consolidación automática de Excel (cada N minutos)
# ══════════════════════════════════════════════════════════════════════

#  MÓDULO GRAPH API — OneDrive / Excel 365 (directo, sin localhost)
#  Credenciales: CFG["graph_tenant"], CFG["graph_client"], CFG["graph_secret"]

# Leer credenciales desde CFG (se actualizan sin reiniciar)
def _graph_creds():
    """Retorna (tenant, client, secret). Todos deben estar para considerar configurado."""
    return (
        CFG.get("graph_tenant","").strip(),
        CFG.get("graph_client","").strip(),
        CFG.get("graph_secret","").strip(),
    )

def _graph_creds_ok() -> bool:
    """True si Graph API está completamente configurado (incluye email de usuario)."""
    tid, cid, sec = _graph_creds()
    email = CFG.get("graph_user_email","").strip()
    return bool(tid and cid and sec and email and "@" in email)

import threading as _threading_lock
_graph_token_cache: dict = {"token": None, "expires": 0}
_graph_token_lock = _threading_lock.Lock()
_wb_sessions_lock = _threading_lock.Lock()

def graph_token() -> str:
    """
    Obtiene (o renueva) el access token de Microsoft Graph.
    Usa Client Credentials flow — no requiere usuario interactivo.
    Thread-safe mediante _graph_token_lock.
    """
    import urllib.request as _ureq, urllib.parse as _up, json as _j

    with _graph_token_lock:
        _tok = _graph_token_cache["token"]
        _exp = _graph_token_cache["expires"]
        if _tok and time.time() < _exp - 60:
            return _tok

    _tid, _cid, _sec = _graph_creds()
    if not all([_tid, _cid, _sec]):
        raise RuntimeError(
            "Graph API no configurado. Ve a Configuracion > Microsoft Graph API."
        )
    if not CFG.get("graph_user_email","").strip():
        raise RuntimeError(
            "Falta el Email OD en Configuracion > Graph API. "
            "Agrega el email corporativo del propietario del OneDrive."
        )

    url  = f"https://login.microsoftonline.com/{_tid}/oauth2/v2.0/token"
    body = _up.urlencode({
        "grant_type":    "client_credentials",
        "client_id":     _cid,
        "client_secret": _sec,
        "scope":         "https://graph.microsoft.com/.default",
    }).encode("utf-8")
    req = _ureq.Request(url, data=body,
                        headers={"Content-Type": "application/x-www-form-urlencoded"})
    try:
        with _ureq.urlopen(req, timeout=15) as r:
            d = _j.loads(r.read())
    except _ureq.error.HTTPError as _he:
        _body = {}
        try: _body = _j.loads(_he.read())
        except Exception: pass
        _desc = _body.get("error_description", "")
        if "AADSTS700016" in _desc:
            raise RuntimeError("Client ID no reconocido en este Tenant.") from None
        if "AADSTS7000215" in _desc:
            raise RuntimeError("Client Secret invalido o expirado.") from None
        if "AADSTS90002" in _desc:
            raise RuntimeError("Tenant ID no encontrado (usa el ID de directorio).") from None
        raise RuntimeError(f"OAuth error {_he.code}: {_desc[:200]}") from None

    _new_token = d.get("access_token", "")
    _expires_in = int(d.get("expires_in", 3600))
    with _graph_token_lock:
        _graph_token_cache["token"]   = _new_token
        _graph_token_cache["expires"] = time.time() + _expires_in
    logger.info("Graph API: token renovado (expira en %ds)", _expires_in)
    return _new_token


def graph_get(path: str, extra_headers: dict | None = None) -> dict:
    """GET a Microsoft Graph API con soporte de paginación y extra headers."""
    import urllib.request, urllib.error, json as _j
    token = graph_token()
    hdrs  = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    if extra_headers:
        hdrs.update(extra_headers)

    all_values: list = []
    next_url: str | None = f"https://graph.microsoft.com/v1.0{path}"

    while next_url:
        req = urllib.request.Request(next_url, headers=hdrs)
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                data = _j.loads(r.read())
        except urllib.error.HTTPError as e:
            body = ""
            try:
                err_json = _j.loads(e.read())
                body = err_json.get("error", {}).get("message", "")
                code = err_json.get("error", {}).get("code", "")
                if code:
                    body = f"[{code}] {body}"
            except Exception:
                pass
            if e.code == 401:
                # Token expirado — limpiar caché con Lock y reintentar una vez
                with _graph_token_lock:
                    _graph_token_cache["token"] = None
                    _graph_token_cache["expires"] = 0
                try:
                    token2 = graph_token()
                    hdrs["Authorization"] = f"Bearer {token2}"
                    req2 = urllib.request.Request(next_url, headers=hdrs)
                    with urllib.request.urlopen(req2, timeout=30) as r2:
                        data = _j.loads(r2.read())
                    # Continuar con data del reintento
                    next_url = data.get("@odata.nextLink")
                    if "value" in data:
                        all_values.extend(data["value"])
                        if not next_url:
                            if all_values:
                                data["value"] = all_values
                            return data
                        continue
                    return data
                except Exception:
                    pass
            raise RuntimeError(f"Graph HTTP {e.code}: {body or path}") from None

        # Paginación: acumular si hay @odata.nextLink
        next_url = data.get("@odata.nextLink")
        if "value" in data:
            all_values.extend(data["value"])
        else:
            return data   # respuesta sin paginación (ej: usedRange)

    # Devolver con todos los valores acumulados
    if all_values:
        data["value"] = all_values
    return data


def graph_patch(path: str, body: dict, extra_headers: dict | None = None) -> dict:
    """PATCH a Microsoft Graph API (actualizar celdas Excel). Soporta sesión de workbook."""
    import urllib.request, urllib.error, json as _j
    token = graph_token()
    data  = _j.dumps(body, default=str).encode("utf-8")
    hdrs  = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
    }
    if extra_headers:
        hdrs.update(extra_headers)
    req = urllib.request.Request(
        f"https://graph.microsoft.com/v1.0{path}",
        data=data, method="PATCH", headers=hdrs
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            try: return _j.loads(r.read())
            except Exception: return {"ok": True}
    except urllib.error.HTTPError as e:
        if e.code == 401:
            with _graph_token_lock:
                _graph_token_cache["token"] = None
                _graph_token_cache["expires"] = 0
            try:
                token2 = graph_token()
                hdrs["Authorization"] = f"Bearer {token2}"
                req2 = urllib.request.Request(
                    f"https://graph.microsoft.com/v1.0{path}",
                    data=data, method="PATCH", headers=hdrs)
                with urllib.request.urlopen(req2, timeout=60) as r2:
                    try: return _j.loads(r2.read())
                    except Exception: return {"ok": True}
            except Exception:
                pass
        err_txt = ""
        try:
            ej = _j.loads(e.read())
            code = ej.get("error",{}).get("code","")
            msg  = ej.get("error",{}).get("message","")[:250]
            err_txt = f"[{code}] {msg}" if code else msg
        except Exception: pass
        raise RuntimeError(f"Graph PATCH {e.code}: {err_txt or path}") from None


def graph_post(path: str, body: dict, extra_headers: dict | None = None) -> dict:
    """POST a Microsoft Graph API. Soporta extra_headers (ej: Workbook-Session-Id)."""
    import urllib.request, urllib.error, json as _j
    token = graph_token()
    data  = _j.dumps(body, default=str).encode("utf-8")
    hdrs  = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
    }
    if extra_headers:
        hdrs.update(extra_headers)
    req = urllib.request.Request(
        f"https://graph.microsoft.com/v1.0{path}",
        data=data, method="POST", headers=hdrs
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            try: return _j.loads(r.read())
            except Exception: return {"ok": True}
    except urllib.error.HTTPError as e:
        if e.code == 401:
            with _graph_token_lock:
                _graph_token_cache["token"] = None
                _graph_token_cache["expires"] = 0
            try:
                token2 = graph_token()
                hdrs["Authorization"] = f"Bearer {token2}"
                req2 = urllib.request.Request(
                    f"https://graph.microsoft.com/v1.0{path}",
                    data=data, method="POST", headers=hdrs)
                with urllib.request.urlopen(req2, timeout=60) as r2:
                    try: return _j.loads(r2.read())
                    except Exception: return {"ok": True}
            except Exception:
                pass
        err_txt = ""
        try:
            ej = _j.loads(e.read())
            code = ej.get("error",{}).get("code","")
            msg  = ej.get("error",{}).get("message","")[:250]
            err_txt = f"[{code}] {msg}" if code else msg
        except Exception: pass
        raise RuntimeError(f"Graph POST {e.code}: {err_txt or path}") from None


# ── Helper: email del usuario OneDrive configurado ───────────────────
def _graph_user() -> str:
    """
    Retorna el email/UPN del usuario OneDrive.
    Con Client Credentials NUNCA se puede usar /me/ — debe ser /users/{email}/.
    """
    email = CFG.get("graph_user_email", "").strip()
    if not email:
        raise RuntimeError(
            "Falta el Email del usuario OneDrive en Configuración > Graph API.\n"
            "Agrega el email corporativo (ej: remuneraciones@empresa.cl) para que "
            "REMI pueda acceder a los archivos con permisos de aplicación."
        )
    return email


def _drive_root() -> str:
    """
    Retorna el prefijo correcto del drive para Client Credentials.
    /users/{email}/drive — NO usa /me/drive (eso requiere usuario delegado).
    Soporta también formato 'user@email.com:/ruta' para multi-usuario.
    """
    return f"/users/{_graph_user()}/drive"


def _drive_path(ruta_onedrive: str) -> str:
    """
    Convierte una ruta OneDrive a la URL de Graph API.

    Formatos aceptados:
      '/RRHH/Dotacion.xlsx'              → /users/{email}/drive/root:/RRHH/Dotacion.xlsx:
      'otro@emp.cl:/Carpeta/Arch.xlsx'   → /users/otro@emp.cl/drive/root:/Carpeta/Arch.xlsx:

    NOTA: Con Client Credentials NO se usa /me/drive.
    """
    import urllib.parse as _up
    ruta = ruta_onedrive.strip()

    # Formato 'user@email:/ruta' — usuario explícito
    if ":" in ruta and "@" in ruta.split(":")[0]:
        user, path = ruta.split(":", 1)
        encoded = "/".join(_up.quote(seg, safe="") for seg in path.strip("/").split("/") if seg)
        return f"/users/{user.strip()}/drive/root:/{encoded}:"

    # Ruta normal — usa el usuario configurado
    encoded = "/".join(_up.quote(seg, safe="") for seg in ruta.strip("/").split("/") if seg)
    return f"{_drive_root()}/root:/{encoded}:"


# ── Caché de sesiones de workbook (para escritura) ───────────────────
_WB_SESSIONS: dict[str, str] = {}   # ruta_onedrive → sessionId (thread-safe via lock)

def _wb_session(drive_p: str) -> str | None:
    """
    Crea (o reutiliza) una sesión de workbook para operaciones de escritura.
    Las sesiones permiten escritura transaccional y evitan conflictos de co-edición.
    Retorna el sessionId o None si falla (en cuyo caso se opera sin sesión).
    """
    with _wb_sessions_lock:
        if drive_p in _WB_SESSIONS:
            return _WB_SESSIONS[drive_p]
    try:
        resp = graph_post(f"{drive_p}/workbook/createSession",
                          {"persistChanges": True})
        sid = resp.get("id") or resp.get("sessionId")
        if sid:
            with _wb_sessions_lock:
                _WB_SESSIONS[drive_p] = sid
        return sid
    except Exception:
        return None   # continuar sin sesión (solo lectura o archivo no abierto)

def _wb_session_headers(drive_p: str) -> dict:
    """Retorna headers con Workbook-Session-Id si existe sesión activa (thread-safe)."""
    with _wb_sessions_lock:
        sid = _WB_SESSIONS.get(drive_p)
    if sid:
        return {"Workbook-Session-Id": sid}
    return {}

def _wb_close_session(drive_p: str):
    """Cierra y elimina la sesión del workbook (thread-safe)."""
    with _wb_sessions_lock:
        sid = _WB_SESSIONS.pop(drive_p, None)
    if sid:
        try:
            graph_post(f"{drive_p}/workbook/closeSession", {},
                       extra_headers={"Workbook-Session-Id": sid})
        except Exception:
            pass


def excel_leer_columnas(ruta_origen: str, hoja: str,
                         columnas: list[str],
                         auto_hoja: bool = True) -> list[dict]:
    """
    Lee columnas específicas por nombre de encabezado desde un Excel en OneDrive.

    Parámetros:
        ruta_origen: ruta en OneDrive, ej '/RRHH/Dotacion_Marzo.xlsx'
                     o 'user@emp.cl:/RRHH/Dotacion.xlsx'
        hoja:        nombre de la hoja (ej 'Sheet1', 'Dotación')
        columnas:    lista de encabezados a extraer ['RUT', 'Nombre', 'Sueldo']

    Retorna: lista de dicts con los valores de cada fila
    """
    drive_p = _drive_path(ruta_origen)
    import urllib.parse as _up
    if not hoja:
        _sh = graph_get(f"{drive_p}/workbook/worksheets").get("value", [])
        hoja = _sh[0]["name"] if _sh else "Sheet1"
    hoja_enc = _up.quote(hoja, safe="")
    wks_url  = f"{drive_p}/workbook/worksheets/{hoja_enc}/usedRange"

    rng  = graph_get(wks_url)
    vals = rng.get("values", [])
    if not vals:
        return []

    # Primera fila = encabezados
    headers = [str(h).strip() for h in vals[0]]

    # Mapear columnas pedidas a índices
    idx_map = {}
    for col in columnas:
        # Búsqueda flexible: ignorar mayúsculas y espacios extras
        col_clean = col.strip().lower()
        for i, h in enumerate(headers):
            if h.lower() == col_clean:
                idx_map[col] = i
                break
        if col not in idx_map:
            # Búsqueda parcial
            for i, h in enumerate(headers):
                if col_clean in h.lower() or h.lower() in col_clean:
                    idx_map[col] = i
                    break

    not_found = [c for c in columnas if c not in idx_map]
    if not_found:
        logger.warning("excel_leer_columnas: columnas no encontradas %s | disponibles: %s",
                        not_found, headers)
        return []
        # Retornar lista vacía en vez de lanzar excepción (evita crash en el chat)
        return []

    # Extraer filas de datos (saltando encabezado)
    result = []
    for row in vals[1:]:
        if not any(str(row[idx_map[c]] if idx_map[c] < len(row) else "").strip() for c in columnas):
            continue   # fila vacía
        rec = {}
        for col in columnas:
            idx = idx_map[col]
            rec[col] = row[idx] if idx < len(row) else ""
        result.append(rec)

    logger.info(f"Excel leído: {len(result)} filas desde '{ruta_origen}' → hoja '{hoja}', columnas {columnas}")
    return result


def excel_pegar_columnas(ruta_destino: str, hoja_destino: str,
                          datos: list[dict], columnas: list[str],
                          fila_inicio: int = None,
                          modo: str = "append") -> dict:
    """
    Pega datos en un Excel de OneDrive.

    Parámetros:
        ruta_destino: ruta en OneDrive del informe destino
        hoja_destino: nombre de la hoja donde pegar
        datos:        lista de dicts (resultado de excel_leer_columnas)
        columnas:     orden de columnas a escribir
        fila_inicio:  None = auto-detectar última fila + 1
        modo:         'append' (añadir al final) | 'overwrite' (desde fila_inicio)

    Retorna: dict con info del resultado
    """
    drive_p   = _drive_path(ruta_destino)
    import urllib.parse as _up
    hoja_url   = f"{drive_p}/workbook/worksheets/{_up.quote(hoja_destino, safe='')}"

    # Abrir sesión de workbook para escritura confiable
    sid       = _wb_session(drive_p)
    sess_hdrs = {"Workbook-Session-Id": sid} if sid else {}

    # Detectar última fila usada si modo=append
    if modo == "append" or fila_inicio is None:
        try:
            used = graph_get(f"{hoja_url}/usedRange", extra_headers=sess_hdrs)
            vals = used.get("values", [[]])
            fila_inicio = len(vals) + 1   # siguiente fila libre (1-based)
        except Exception:
            fila_inicio = 2   # si no hay datos, empezar en fila 2

    # Armar la matriz de valores (None → "")
    matrix = [[str(r.get(col, "") if r.get(col) is not None else "")
               for col in columnas] for r in datos]

    # Calcular el rango de destino — soporta >26 columnas (AA, AB…)
    def _col_letter(n: int) -> str:
        s = ""
        while n >= 0:
            s = chr(ord("A") + n % 26) + s
            n = n // 26 - 1
        return s
    col_letra_fin = _col_letter(len(columnas) - 1)
    fila_fin = fila_inicio + len(matrix) - 1
    rango = f"A{fila_inicio}:{col_letra_fin}{fila_fin}"

    try:
        graph_patch(
            f"{hoja_url}/range(address='{rango}')",
            {"values": matrix},
            extra_headers=sess_hdrs
        )
    finally:
        _wb_close_session(drive_p)

    logger.info("Excel pegado: %d filas en '%s' → '%s' rango %s",
                len(matrix), ruta_destino, hoja_destino, rango)
    return {
        "filas_pegadas": len(matrix),
        "rango":         rango,
        "hoja":          hoja_destino,
        "archivo":       ruta_destino,
    }


def excel_copiar_entre_archivos(
    ruta_origen:    str,
    hoja_origen:    str,
    columnas:       list[str],
    ruta_destino:   str,
    hoja_destino:   str,
    modo:           str = "append",
    fila_inicio:    int = None,
) -> dict:
    """
    Copia columnas de un Excel OneDrive a otro en un solo paso.
    """
    # 1. Leer
    datos = excel_leer_columnas(ruta_origen, hoja_origen, columnas)
    if not datos:
        return {"ok": False, "error": "No se encontraron datos en el origen", "filas": 0}

    # 2. Pegar
    result = excel_pegar_columnas(ruta_destino, hoja_destino, datos, columnas,
                                   fila_inicio=fila_inicio, modo=modo)
    return {
        "ok":           True,
        "filas_copiadas": result["filas_pegadas"],
        "rango_destino":  result["rango"],
        "origen":         ruta_origen,
        "destino":        ruta_destino,
        "hoja_origen":    hoja_origen,
        "hoja_destino":   hoja_destino,
        "columnas":       columnas,
        "muestra":        datos[:3],   # primeras 3 filas como preview
    }


def excel_leer_todo(ruta_origen: str, hoja: str = "") -> list[dict]:
    """
    Lee TODAS las columnas de una hoja Excel en OneDrive via Graph API.
    Si hoja="" usa la primera hoja disponible.
    Retorna lista de dicts {encabezado: valor} por cada fila de datos.
    Ignora filas completamente vacías.
    """
    drive_p = _drive_path(ruta_origen)

    # Resolver hoja: si no se especifica, tomar la primera
    if not hoja:
        wks_list = graph_get(f"{drive_p}/workbook/worksheets")
        sheets   = wks_list.get("value", [])
        if not sheets:
            raise ValueError(f"No se encontraron hojas en '{ruta_origen}'")
        hoja = sheets[0]["name"]

    # Encodear nombre de hoja para URL (espacios → %20)
    import urllib.parse as _up
    hoja_enc = _up.quote(hoja, safe="")
    rng  = graph_get(f"{drive_p}/workbook/worksheets/{hoja_enc}/usedRange")
    vals = rng.get("values", [])
    if not vals:
        return []

    headers = [str(h).strip() for h in vals[0]]
    result  = []
    for row in vals[1:]:
        # Saltar filas completamente vacías
        row_padded = list(row) + [""] * (len(headers) - len(row))
        if all(v is None or str(v).strip() == "" for v in row_padded):
            continue  # saltar filas genuinamente vacías (no las que tienen 0)
        rec = {headers[i]: row_padded[i] for i in range(len(headers))}
        result.append(rec)

    logger.info(f"excel_leer_todo: {len(result)} filas · {len(headers)} cols · '{ruta_origen}'")
    return result


def excel_escribir_todo(ruta_destino: str, hoja: str,
                         datos: list[dict], limpiar_primero: bool = True,
                         chunk_size: int = 500) -> dict:
    """
    Escribe un dataset completo en una hoja de OneDrive via Graph API.
    Usa sesión de workbook para escritura transaccional.
    Escribe en chunks de 500 filas para evitar límites de payload.
    """
    if not datos:
        return {"ok": False, "error": "Sin datos para escribir", "filas_escritas": 0}

    drive_p = _drive_path(ruta_destino)
    import urllib.parse as _up
    hoja_enc = _up.quote(hoja, safe="")
    hoja_url = f"{drive_p}/workbook/worksheets/{hoja_enc}"

    headers = list(datos[0].keys())

    def _col_letter(n: int) -> str:
        s = ""
        while n >= 0:
            s = chr(ord("A") + n % 26) + s
            n = n // 26 - 1
        return s

    col_fin = _col_letter(len(headers) - 1)

    # Abrir sesión de workbook (necesaria para escritura confiable)
    sid = _wb_session(drive_p)
    sess_hdrs = {"Workbook-Session-Id": sid} if sid else {}

    try:
        if limpiar_primero:
            try:
                usado = graph_get(f"{hoja_url}/usedRange", extra_headers=sess_hdrs)
                if usado.get("address"):
                    graph_post(f"{hoja_url}/usedRange/clear",
                               {"applyTo": "contents"}, extra_headers=sess_hdrs)
            except Exception:
                pass

        # Escribir encabezados primero
        hdr_rango = f"A1:{col_fin}1"
        graph_patch(f"{hoja_url}/range(address='{hdr_rango}')",
                    {"values": [headers]}, extra_headers=sess_hdrs)

        # Escribir datos en chunks
        total_escritas = 0
        for chunk_start in range(0, len(datos), chunk_size):
            chunk = datos[chunk_start:chunk_start + chunk_size]
            matrix_chunk = [[str(rec.get(h, "") if rec.get(h) is not None else "")
                             for h in headers] for rec in chunk]
            fila_ini = chunk_start + 2   # +2 porque encabezado está en fila 1
            fila_fin_chunk = fila_ini + len(chunk) - 1
            rango_chunk = f"A{fila_ini}:{col_fin}{fila_fin_chunk}"
            graph_patch(f"{hoja_url}/range(address='{rango_chunk}')",
                        {"values": matrix_chunk}, extra_headers=sess_hdrs)
            total_escritas += len(chunk)
            logger.info("excel_escribir_todo: chunk %d-%d (%d filas)",
                        chunk_start, chunk_start+len(chunk)-1, len(chunk))

        rango_total = f"A1:{col_fin}{total_escritas + 1}"
        logger.info("excel_escribir_todo: %d filas · %d cols → '%s' [%s]",
                    total_escritas, len(headers), ruta_destino, rango_total)
        return {
            "ok":            True,
            "filas_escritas": total_escritas,
            "cols":           len(headers),
            "rango":          rango_total,
            "hoja":           hoja,
            "archivo":        ruta_destino,
        }

    finally:
        # Cerrar sesión al terminar (éxito o error)
        _wb_close_session(drive_p)


def graph_listar_archivos(carpeta: str = "/") -> list[dict]:
    """
    Lista archivos en una carpeta de OneDrive/SharePoint via Graph API.
    Usa /users/{email}/drive — correcto para Client Credentials (sin /me/).
    Soporta paginación automática y espacios/caracteres especiales en rutas.
    """
    import urllib.parse as _up
    drv = _drive_root()   # /users/{email}/drive

    if carpeta.strip() in ("/", "", "."):
        ruta = f"{drv}/root/children?$top=200&$select=name,size,lastModifiedDateTime,parentReference,file"
    else:
        segs    = carpeta.strip("/").split("/")
        encoded = "/".join(_up.quote(s, safe="") for s in segs if s)
        ruta    = f"{drv}/root:/{encoded}:/children?$top=200&$select=name,size,lastModifiedDateTime,parentReference,file"

    resp  = graph_get(ruta)
    items = resp.get("value", [])
    return [
        {
            "nombre":     i["name"],
            "ruta":       (i.get("parentReference", {}).get("path", "") + "/" + i["name"]).replace("/drive/root:", ""),
            "tamanio":    i.get("size", 0),
            "modificado": i.get("lastModifiedDateTime", "")[:16],
            "es_excel":   i["name"].lower().endswith((".xlsx", ".xls", ".xlsm")),
            "es_carpeta": "file" not in i,
            "id":         i.get("id",""),
        }
        for i in items
        if i.get("name") and not i["name"].startswith("~$")  # excluir archivos temp
    ]



# ══════════════════════════════════════════════════════════════════════
#  UTILIDADES RUT / SELLO — limpiar_rut() · generar_sello()
# ══════════════════════════════════════════════════════════════════════

def limpiar_rut(rut) -> str:
    """Normaliza RUT chileno: elimina puntos/guiones/espacios, K a mayuscula.
    Ej: '18.570.338-K' -> '18570338K' (compatible con BuscarV / VLOOKUP)
    Maneja None, NaN (float y numpy.float64), y valores vacíos.
    """
    if rut is None: return ""
    try:
        # Captura float nativo Y numpy.float64 / numpy.nan
        import math
        if math.isnan(float(rut)): return ""
    except (ValueError, TypeError):
        pass  # no es numérico — continuar
    s = str(rut).strip()
    if s in ("", "nan", "None", "NaN"): return ""
    return re.sub(r'[^0-9Kk]', '', s).upper()


def generar_sello(analista: str) -> str:
    """Crea sello de auditoria: 'ANALISTA-DD-MM-HH:MM'
    Ej: 'YEIMY AVILA-11-03-13:41'
    """
    import datetime as _dt
    ts = _dt.datetime.now().strftime("%d-%m-%H:%M")
    return str(analista).strip() + "-" + ts


class RemiDaemon:
    """
    Monitorea ~/OneDrive/Remi_Automatizacion/Entrada/ cada N minutos.
    Consolida los .xlsx nuevos en un archivo Maestro y los mueve a Histórico.
    
    Integrado con RemiApp:
      • app._daemon_log(msg)  → burbuja "system" en el chat (thread-safe)
      • notificar(t, m)       → notificación de escritorio (plyer)
      • app._daemon_badge()   → actualiza el badge del sidebar
    """

    def __init__(self, app):
        self.app      = app
        self._stop    = threading.Event()   # permite detener limpiamente
        self._thread  = None
        self._running = False

        # Rutas (se leen de CFG en cada ciclo para que cambios se apliquen sin reiniciar)
        self.base_path = (Path(os.path.expanduser("~"))
                          / "OneDrive" / "Remi_Automatizacion")
        self.entrada   = self.base_path / "Entrada"
        self.historico = self.base_path / "Historico"
        self.maestro   = self.base_path / "Base_Datos" / "Maestro.xlsx"

        # Crear estructura si no existe (silencioso)
        try:
            for p in [self.entrada, self.historico, self.maestro.parent]:
                p.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

    # ── API pública ──────────────────────────────────────────────────
    def iniciar(self):
        if self._running:
            return "ℹ️ El daemon ya está activo."
        self._stop.clear()
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        self._running = True
        self._log("🚀 Consolidación automática ACTIVADA — revisaré la carpeta Entrada cada "
                  f"{CFG.get('daemon_intervalo', 30)} minutos.")
        self.app.after(0, self.app._daemon_badge)
        return True

    def detener(self):
        if not self._running:
            return "ℹ️ El daemon no estaba activo."
        self._stop.set()
        self._running = False
        self._log("🛑 Consolidación automática DETENIDA.")
        self.app.after(0, self.app._daemon_badge)
        return True

    @property
    def activo(self):
        return self._running

    # ── Loop interno ─────────────────────────────────────────────────
    def _loop(self):
        while not self._stop.is_set():
            try:
                self._procesar()
            except Exception as e:
                self._log(f"⚠️ Error en ciclo daemon: {e}")
            # Esperar N minutos en intervalos de 5s (permite detención rápida)
            intervalo = int(CFG.get("daemon_intervalo", 30)) * 60
            for _ in range(intervalo // 5):
                if self._stop.is_set():
                    return
                time.sleep(5)

    # ── Procesamiento ─────────────────────────────────────────────────
    def _procesar(self):
        try:
            import pandas as pd
        except ImportError:
            self._log("❌ Falta pandas. Instala con: pip install pandas openpyxl")
            return

        archivos = sorted(self.entrada.glob("*.xlsx"))
        if not archivos:
            return   # sin ruido si no hay nada

        self._log(f"🔍 {len(archivos)} archivo(s) nuevos en Entrada — procesando…")

        # Cargar (o crear) Maestro
        try:
            df_maestro = pd.read_excel(self.maestro, engine="openpyxl") if self.maestro.exists() else pd.DataFrame()
        except Exception as e:
            self._log(f"⚠️ No pude leer Maestro.xlsx: {e} — se creará uno nuevo.")
            df_maestro = pd.DataFrame()

        procesados = 0
        for arc in archivos:
            if self._stop.is_set():
                break
            try:
                df_nuevo = pd.read_excel(arc, engine="openpyxl")
                if df_nuevo.empty:
                    continue

                # ── Enriquecimiento ───────────────────────────────────
                ts = datetime.datetime.now().strftime("%Y%m%d%H%M")
                df_nuevo["ID_REMI"] = [f"REC-{ts}-{i}" for i in range(len(df_nuevo))]

                col_ref = next(
                    (c for c in ["RUT", "rut", "Rut", "RUT_EMPLEADO"] if c in df_nuevo.columns),
                    df_nuevo.columns[0]
                )
                # Normalizar RUT con limpiar_rut() para búsquedas consistentes
                df_nuevo["RUT_LIMP"] = df_nuevo[col_ref].apply(limpiar_rut)
                df_nuevo["REGISTRO_CONCATENADO"] = (
                    df_nuevo["ID_REMI"].astype(str) + " | "
                    + df_nuevo[col_ref].astype(str) + " | "
                    + datetime.datetime.now().date().isoformat()
                )
                df_nuevo["ARCHIVO_ORIGEN"]  = arc.name
                df_nuevo["FECHA_CARGA"]     = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

                # ── Concatenar al maestro ─────────────────────────────
                df_maestro = pd.concat([df_maestro, df_nuevo], ignore_index=True, sort=False)
                df_maestro = df_maestro.fillna("")

                # ── Mover a Histórico (manejo de conflictos de nombre) ─
                destino = self.historico / f"{datetime.date.today()}_{arc.name}"
                if destino.exists():
                    destino = self.historico / f"{ts}_{arc.name}"
                arc.rename(destino)
                procesados += 1

            except PermissionError:
                self._log(f"⏳ `{arc.name}` está siendo sincronizado por OneDrive — se reintentará en el próximo ciclo.")
            except Exception as e:
                self._log(f"❌ No se pudo procesar `{arc.name}`: {e}")

        if procesados == 0:
            return

        # ── Guardar Maestro ───────────────────────────────────────────
        try:
            df_maestro.to_excel(self.maestro, index=False, engine="openpyxl")
        except Exception as e:
            self._log(f"❌ No pude guardar Maestro.xlsx: {e}")
            return

        msg = f"✅ Maestro actualizado — {procesados} archivo(s) consolidados ({len(df_maestro)} filas totales)."
        self._log(msg)
        notificar("REMI: Consolidación exitosa",
                  f"Se procesaron {procesados} archivo(s) en Entrada.")

    # ── Matriz Finiquitos ───────────────────────────────────────────
    def sincronizar_finiquitos(self):
        """Consolida Finiquitos_*.xlsx de una carpeta OneDrive en Matriz Madre."""
        import threading
        threading.Thread(target=self._sinc_fin, daemon=True).start()

    def _sinc_fin(self):
        """
        Consolida archivos de analistas → MATRIZ GENERAL (archivo Madre).

        MODO 1 — Graph API / Excel 365 (principal, sin localhost):
            • Lista todos los .xlsx en la carpeta OneDrive configurada
            • Lee TODAS las columnas de cada archivo via Graph API (excel_leer_todo)
            • Aplica limpiar_rut() y generar_sello() por analista
            • Sobreescribe el archivo Madre en OneDrive (excel_escribir_todo)

        MODO 2 — Local (fallback automático si Graph sin configurar):
            • Lee desde carpeta local sincronizada por OneDrive Desktop
            • Mismo procesamiento, guarda Madre localmente

        El sistema detecta automáticamente qué modo usar según CFG.
        """
        try: import pandas as pd
        except ImportError:
            self._log("Falta pandas. pip install pandas openpyxl"); return

        # ── Detectar modo ─────────────────────────────────────────────
        tid, cid, sec = _graph_creds()
        ruta_od  = CFG.get("finiquitos_ruta_od","").strip()   # ej: /Finiquitos/2024
        ruta_loc = CFG.get("finiquitos_ruta","").strip()       # ej: C:\Users\...\OneDrive\Finiquitos
        madre    = CFG.get("finiquitos_madre","MATRIZ_GENERAL_FINIQUITOS.xlsx").strip()
        hoja     = CFG.get("finiquitos_hoja","Finiquitos").strip() or "Finiquitos"
        usar_graph = bool(tid and cid and sec and ruta_od)

        self._log(
            ("☁️  Modo Graph API / Excel 365" if usar_graph else "💾  Modo local (OneDrive Desktop)")
            + "  —  iniciando consolidación..."
        )

        # ════════════════════════════════════════════════════════════
        # MODO 1 — GRAPH API
        # ════════════════════════════════════════════════════════════
        if usar_graph:
            try:
                # 1. Listar archivos Excel en la carpeta OneDrive
                archivos_od = graph_listar_archivos(ruta_od)
                archivos_od = [a for a in archivos_od
                               if a["es_excel"]
                               and a["nombre"] != madre
                               and not a["nombre"].startswith("~$")]  # excluir archivos temp
                if not archivos_od:
                    self._log(f"No hay archivos .xlsx en OneDrive: {ruta_od}"); return

                self._log(f"Encontrados {len(archivos_od)} archivo(s) en OneDrive")
                lista = []

                for arch in archivos_od:
                    nombre   = arch["nombre"]
                    ruta_arch = ruta_od.rstrip("/") + "/" + nombre
                    stem     = nombre.rsplit(".",1)[0]
                    analista = (stem.replace("Finiquitos_","")
                                    .replace("Matriz_","")
                                    .replace("_"," ").strip().upper())
                    try:
                        # Leer TODAS las columnas del archivo via Graph API
                        filas = excel_leer_todo(ruta_arch, hoja)
                        if not filas:
                            self._log(f"  {nombre}: sin datos en hoja '{hoja}', intentando primera hoja...")
                            filas = excel_leer_todo(ruta_arch, "")  # primera hoja disponible
                        if not filas:
                            self._log(f"  {nombre}: sin datos — omitido"); continue

                        df = pd.DataFrame(filas)

                        # Columnas de identidad y auditoría
                        df["ANALISTA"]         = analista
                        df["SELLO_AUDITORIA"]  = generar_sello(analista)
                        df["Control Analista"] = df["SELLO_AUDITORIA"]
                        df["ARCHIVO_ORIGEN"]   = nombre
                        df["FECHA_CARGA"]      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
                        df["MODIFICADO_OD"]    = arch.get("modificado","")[:16]

                        # Normalizar RUTs
                        rut_cols = [c for c in df.columns if re.search(r"(?i)^rut", c)]
                        for rc in rut_cols:
                            df[f"RUT_LIMP_{re.sub(r'[^a-zA-Z0-9]','_',rc)}"] = df[rc].apply(limpiar_rut)
                        if rut_cols:
                            df["RUT_LIMP"] = df[rut_cols[0]].apply(limpiar_rut)

                        lista.append(df)
                        self._log(f"  ☁️  {analista}: {len(df)} filas · mod {arch.get('modificado','')[:10]}")

                    except Exception as e:
                        self._log(f"  ❌ {nombre}: {e}")

                if not lista:
                    self._log("No se pudo leer ningún archivo desde OneDrive."); return

                resultado = pd.concat(lista, ignore_index=True, sort=False)
                resultado = resultado.fillna("")  # celdas NaN → "" (seguro para JSON)
                resultado.insert(0, "ID_REMI", range(1, len(resultado)+1))

                # 2. Escribir Matriz Madre en OneDrive (sobrescribe)
                ruta_madre_od = ruta_od.rstrip("/") + "/" + madre
                filas_out = resultado.to_dict(orient="records")
                res = excel_escribir_todo(ruta_madre_od, hoja, filas_out,
                                          limpiar_primero=True)
                n_a = resultado["ANALISTA"].nunique()
                self._log(
                    f"✅ Consolidación OneDrive completa\n"
                    f"  Analistas: {n_a} · Filas: {len(resultado)}\n"
                    f"  Matriz: {ruta_madre_od} [{res.get('rango','')}]"
                )
                # Guardar también copia local para el visor
                import pathlib as _pl, os
                if ruta_loc and os.path.isdir(ruta_loc):
                    dest_loc = os.path.join(ruta_loc, madre)
                    resultado.to_excel(dest_loc, index=False, engine="openpyxl")
                    self._log(f"  Copia local: {dest_loc}")
                notificar("REMI Finiquitos",
                          f"{n_a} analistas · {len(resultado)} filas → {madre} (OneDrive)")

            except Exception as e:
                self._log(f"❌ Error Graph API: {e}"); return

        # ════════════════════════════════════════════════════════════
        # MODO 2 — LOCAL (fallback)
        # ════════════════════════════════════════════════════════════
        else:
            import os, pathlib as _pl
            if not ruta_loc:
                ruta_loc = str(_pl.Path.home() / "OneDrive" / "Finiquitos_2024")
            if not os.path.isdir(ruta_loc):
                self._log(
                    "No hay ruta configurada.\n"
                    "Configura 'Ruta OneDrive (carpeta)' en ⚙️ Finiquitos\n"
                    "o bien 'Ruta local' si aún no tienes Graph API."
                ); return

            import glob as _glob
            todos = [os.path.basename(p) for p in _glob.glob(os.path.join(ruta_loc,"*.xlsx"))
                     if os.path.basename(p) != madre]
            if not todos:
                self._log(f"No hay .xlsx en {ruta_loc}"); return

            self._log(f"Encontrados {len(todos)} archivo(s) locales")
            lista = []
            for nombre in todos:
                stem     = _pl.Path(nombre).stem
                analista = (stem.replace("Finiquitos_","").replace("Matriz_","")
                                .replace("_"," ").strip().upper())
                try:
                    df = pd.read_excel(os.path.join(ruta_loc, nombre), engine="openpyxl")
                    df["ANALISTA"]         = analista
                    df["SELLO_AUDITORIA"]  = generar_sello(analista)
                    df["Control Analista"] = df["SELLO_AUDITORIA"]
                    df["ARCHIVO_ORIGEN"]   = nombre
                    df["FECHA_CARGA"]      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
                    rut_cols = [c for c in df.columns if re.search(r"(?i)^rut", c)]
                    for rc in rut_cols:
                        df[f"RUT_LIMP_{re.sub('[^a-zA-Z0-9]','_',rc)}"] = df[rc].apply(limpiar_rut)
                    if rut_cols: df["RUT_LIMP"] = df[rut_cols[0]].apply(limpiar_rut)
                    lista.append(df)
                    self._log(f"  💾 {analista}: {len(df)} filas")
                except PermissionError:
                    self._log(f"  ⏳ {nombre} bloqueado por OneDrive sync, omitido.")
                except Exception as e:
                    self._log(f"  ❌ {nombre}: {e}")

            if not lista:
                self._log("No se pudo leer ningún archivo."); return

            resultado = pd.concat(lista, ignore_index=True, sort=False)
            resultado = resultado.fillna("")
            resultado.insert(0, "ID_REMI", range(1, len(resultado)+1))
            destino = os.path.join(ruta_loc, madre)
            resultado.to_excel(destino, index=False, engine="openpyxl")
            n_a = resultado["ANALISTA"].nunique()
            self._log(
                f"✅ Consolidación local completa\n"
                f"  Analistas: {n_a} · Filas: {len(resultado)}\n"
                f"  Archivo: {destino}"
            )
            notificar("REMI Finiquitos", f"{n_a} analistas · {len(resultado)} filas → {madre}")

    # ── Helpers ───────────────────────────────────────────────────────
    def _log(self, msg: str):
        """Envía un mensaje al chat de Remi (thread-safe)."""
        try:
            self.app.after(0, lambda m=msg: self.app._daemon_log(m))
        except Exception:
            pass


CHIPS_RAPIDOS = [
    ("📋 Dotación hoy",  "Genera el script Python para el informe de dotación diaria de hoy"),
    ("📊 Cierre comisiones", "Guíame paso a paso en el cierre de comisiones SLD/PDV/EOS de este mes"),
    ("📝 Macro VBA",    "Crea una macro VBA para consolidar los Excel de comisiones SLD del mes"),
    ("⚖️ Finiquito",    "Calcula el finiquito. Pídeme los datos que necesites"),
    ("📂 Ver Excel",    "Muéstrame los archivos Excel disponibles en mis rutas de trabajo"),
    ("📧 Correo DOT",   "Redacta el correo de dotación semanal para gerencia"),
    ("📅 Calendario",   "Muéstrame el calendario de obligaciones de remuneraciones de este mes"),
    ("🔍 SQL Softland", "Ayúdame con una consulta SQL en Softland para remuneraciones"),
]


# ══════════════════════════════════════════════════════════════════════
#  APP PRINCIPAL
# ══════════════════════════════════════════════════════════════════════
class RemiApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Remi · Remuneraciones & Control de Gestión")
        self.geometry("1160x740")
        self.minsize(600, 480)
        self.resizable(True, True)
        self.configure(fg_color=T["bg"])

        # Estado interno
        self._conv      = []   # historial conversación (máx 30 mensajes)
        self._busy      = False
        self._bact      = None  # burbuja activa en stream
        self._img       = None  # imagen adjunta (legacy: 1)
        self._imgs      = []    # lista de imágenes adjuntas (paths)
        self._file      = None  # archivo adjunto
        self._ultimo_excel = None  # último xlsx creado/abierto
        self._ph_active = True  # placeholder activo
        self._cv_sb     = None  # canvas avatar sidebar

        # Performance: debounce/throttle UI
        self._scroll_job = None
        self._stream_job = None
        self._stream_pending = ""
        self._next_blink_at = time.time() + 8.0
        self._next_avatar_at = time.time() + 2.0
        self._avatar_busy_last = None

        self._build_ui()
        GestorAlarmas(self._on_alarma)
        self.after(400,  self._tick_anim)
        self.after(6000, self._tick_prov)
        self.after(1800, self._bienvenida)  # dar tiempo al render completo
        # Auto-update check en segundo plano (no bloquea)
        self.after(3000, self._check_update_bg)

        # ── Daemon de consolidación ──────────────────────────────────
        self._daemon = RemiDaemon(self)
        if CFG.get("daemon_activo"):
            self.after(3000, self._daemon.iniciar)   # iniciar 3s después de la UI

    # ══════════════════════════════════════════════════════════
    # CONSTRUCCIÓN INTERFAZ
    # ══════════════════════════════════════════════════════════
    def _build_ui(self):
        # Columna 0 = sidebar (ancho fijo pero puede colapsar)
        # Columna 1 = chat (se expande)
        self.columnconfigure(0, weight=0, minsize=220)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)
        self._build_sidebar()
        self._build_main()
        # Adaptar sidebar al tamaño de la ventana
        self.bind("<Configure>", self._on_resize)

    def _on_resize(self, e=None):
        """Adapta el layout cuando cambia el tamaño de la ventana."""
        try:
            w = self.winfo_width()
            # Si la ventana es estrecha, colapsar sidebar
            if w < 700:
                self.columnconfigure(0, weight=0, minsize=0)
                try: self._sb_frame.grid_remove()
                except: pass
            else:
                sb_w = min(240, max(200, w // 5))
                self.columnconfigure(0, weight=0, minsize=sb_w)
                try: self._sb_frame.grid()
                except: pass
        except Exception:
            pass

    # ── SIDEBAR ───────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════
    # SIDEBAR — Header fijo + zona scrollable + footer fijo
    # Secciones colapsables con flecha (acordeón)
    # ══════════════════════════════════════════════════════════
    def _build_sidebar(self):
        sb = tk.Frame(self, bg=T["sb"], width=236)
        sb.grid(row=0, column=0, sticky="nsew"); sb.grid_propagate(False)
        sb.rowconfigure(1, weight=1)   # fila 1 = zona scrollable, se expande
        self._sb_frame = sb

        # ╔══════════════════════════════════════════╗
        # ║  HEADER FIJO (logo + badge IA + nueva)   ║
        # ╚══════════════════════════════════════════╝
        hdr = tk.Frame(sb, bg=T["sb"])
        hdr.grid(row=0, column=0, sticky="ew")

        top = tk.Frame(hdr, bg=T["sb"]); top.pack(fill="x", padx=14, pady=(18,8))
        self._cv_sb = tk.Canvas(top, width=58, height=58,
                                bg=T["sb"], highlightthickness=0)
        self._cv_sb.pack(side="left")
        dibujar_remi(self._cv_sb, 58)

        lf = tk.Frame(top, bg=T["sb"]); lf.pack(side="left", padx=10, pady=2)
        tk.Label(lf, text="REMI", font=(FONT, 16, "bold"),
                 bg=T["sb"], fg=T["tx"]).pack(anchor="w")
        tk.Label(lf, text="Remuneraciones · RR.HH.", font=(FONT, 8),
                 bg=T["sb"], fg=T["t2"]).pack(anchor="w")
        tk.Label(lf, text=f"v{REMI_VERSION}  ·  Excel 365 + IA", font=(FONT, 7),
                 bg=T["sb"], fg=T.get("a1","#3fb950")).pack(anchor="w")
        try:
            _ses = getattr(self, "_sesion", _SESION_ACTUAL)
            _u_nom = _ses.get("nombre", _ses.get("usuario",""))
            _u_ic  = "👑" if _ses.get("rol") == "admin" else "👤"
            tk.Label(lf, text=f"{_u_ic} {_u_nom[:18]}", font=(FONT, 8),
                     bg=T["sb"], fg=T.get("t2","#8b949e")).pack(anchor="w")
        except Exception:
            pass

        tk.Frame(hdr, bg=T.get("a1","#3fb950"), height=2).pack(fill="x")

        ia_f = tk.Frame(hdr, bg=T["sb"]); ia_f.pack(fill="x", padx=12, pady=(8,2))
        self._lbl_prov = tk.Label(ia_f, text="—", font=(FONT, 9),
                                   bg=T["sb"], fg=T["a2"])
        self._lbl_prov.pack(side="left")
        self._lbl_web = tk.Label(ia_f, font=(FONT, 8),
                                  text=("· web" if WEB else ""),
                                  bg=T["sb"], fg=T["t2"])
        self._lbl_web.pack(side="left", padx=4)
        self._update_prov_label()

        # Fila: Nueva conv + Historial
        _btn_row = tk.Frame(hdr, bg=T["sb"]); _btn_row.pack(fill="x", padx=10, pady=(6,10))
        _btn_row.columnconfigure(0, weight=1)

        self._mk_sb_btn(_btn_row, "＋  Nueva conversación", self._nueva_conv,
                        bg=T["a1"], fg="white", bold=True, pady=7).grid(
                        row=0, column=0, sticky="ew", padx=(0,4))

        _btn_hist = tk.Button(_btn_row, text="💬", font=(FONT, 13),
                              bg=T["card"], fg=T["t2"],
                              activebackground=T["a2"], activeforeground="white",
                              relief="flat", padx=8, pady=6,
                              cursor="hand2", bd=0,
                              command=self._win_historial_chats)
        _btn_hist.grid(row=0, column=1, sticky="ew")
        _btn_hist.bind("<Enter>", lambda e: _btn_hist.configure(bg=T["a2"], fg="white"))
        _btn_hist.bind("<Leave>", lambda e: _btn_hist.configure(bg=T["card"], fg=T["t2"]))
        # Tooltip
        tk.Label(hdr, text="💬 Ver conversaciones anteriores",
                 font=(FONT, 7), bg=T["sb"], fg=T["t2"]).pack(anchor="w", padx=14, pady=(0,4))

        # ╔══════════════════════════════════════════╗
        # ║  ZONA SCROLLABLE (Canvas + Scrollbar)    ║
        # ╚══════════════════════════════════════════╝
        scroll_wrap = tk.Frame(sb, bg=T["sb"])
        scroll_wrap.grid(row=1, column=0, sticky="nsew")
        scroll_wrap.rowconfigure(0, weight=1)
        scroll_wrap.columnconfigure(0, weight=1)

        _cv = tk.Canvas(scroll_wrap, bg=T["sb"], highlightthickness=0,
                        bd=0, yscrollincrement=1)
        _cv.grid(row=0, column=0, sticky="nsew")

        _vsb = tk.Scrollbar(scroll_wrap, orient="vertical",
                             command=_cv.yview,
                             bg=T["sb"], troughcolor=T["sb"],
                             width=5, relief="flat", bd=0)
        _vsb.grid(row=0, column=1, sticky="ns")
        _cv.configure(yscrollcommand=_vsb.set)

        # Frame interior que va dentro del canvas
        _inner = tk.Frame(_cv, bg=T["sb"])
        _win_id = _cv.create_window((0, 0), window=_inner, anchor="nw")

        def _on_resize(e):
            _cv.itemconfig(_win_id, width=e.width)
        _cv.bind("<Configure>", _on_resize)

        def _on_frame_resize(e):
            _cv.configure(scrollregion=_cv.bbox("all"))
        _inner.bind("<Configure>", _on_frame_resize)

        # Scroll con rueda del ratón
        def _on_wheel(e):
            _cv.yview_scroll(int(-1*(e.delta/120)), "units")
        def _bind_wheel(w):
            w.bind("<MouseWheel>", _on_wheel)
            for ch in w.winfo_children():
                _bind_wheel(ch)
        _inner.bind("<MouseWheel>", _on_wheel)

        # ── Paleta de colores por sección ───────────────────────────
        # Cada grupo tiene su propio color de acento para la barra lateral
        SECTION_COLORS = {
            "rem":   T.get("a1",   "#3fb950"),   # verde — remuneraciones
            "xl":    "#4d8ef0",                   # azul  — excel & datos
            "fin":   "#e3b341",                   # ámbar — finiquitos
            "ges":   "#9b7cf0",                   # violeta — gestión
            "adm":   T.get("err",  "#f85149"),   # rojo  — administración
        }

        # ── Función: grupo colapsable v2 ─────────────────────────────
        def _sb_group(titulo, botones, abierto=True, color_key="rem", n_items=None):
            """
            Grupo colapsable mejorado:
            · Barra de color vertical a la izquierda (identidad visual)
            · Encabezado con fondo diferenciado + badge de cantidad
            · Flecha animada ▶/▼
            · Botones con indentación e indicador de hover suave
            · Línea separadora inferior
            """
            estado  = {"abierto": abierto}
            c_acent = SECTION_COLORS.get(color_key, T.get("a1","#3fb950"))
            n       = n_items if n_items is not None else len(botones)

            # ── Wrapper externo con margen superior ──────────────────
            wrap = tk.Frame(_inner, bg=T["sb"])
            wrap.pack(fill="x", pady=(6,0))

            # ── Barra de color + encabezado en una fila ──────────────
            hdr_row = tk.Frame(wrap, bg=T["sb"]); hdr_row.pack(fill="x")

            # Barra vertical de color (3px) a la izquierda
            bar = tk.Frame(hdr_row, bg=c_acent, width=3)
            bar.pack(side="left", fill="y")
            bar.pack_propagate(False)

            # Fondo del encabezado (ligeramente más claro que sb)
            def _blend(hex_c, factor=0.08):
                """Mezcla un color con blanco para aclarar levemente."""
                try:
                    r = int(hex_c[1:3],16); g = int(hex_c[3:5],16); b = int(hex_c[5:7],16)
                    r2 = int(r + (255-r)*factor); g2 = int(g + (255-g)*factor); b2 = int(b + (255-b)*factor)
                    return f"#{r2:02x}{g2:02x}{b2:02x}"
                except Exception:
                    return hex_c

            bg_hdr = _blend(T["sb"], 0.10)

            hdr_inner = tk.Frame(hdr_row, bg=bg_hdr, cursor="hand2")
            hdr_inner.pack(side="left", fill="both", expand=True)

            # Fila de contenido del encabezado
            hdr_content = tk.Frame(hdr_inner, bg=bg_hdr)
            hdr_content.pack(fill="x", padx=0, pady=0)

            # Icono + texto del título (sin emoji de sección en UPPERCASE)
            # Separar emoji del texto
            parts = titulo.split(" ", 1)
            ico_txt  = parts[0] if len(parts) > 1 else ""
            sec_txt  = parts[1].upper() if len(parts) > 1 else titulo.upper()

            lbl_ico = tk.Label(hdr_content, text=ico_txt, font=(FONT, 11),
                               bg=bg_hdr, fg=c_acent, cursor="hand2")
            lbl_ico.pack(side="left", padx=(10,2), pady=7)

            lbl_titulo = tk.Label(hdr_content, text=sec_txt,
                                  font=(FONT, 8, "bold"),
                                  bg=bg_hdr, fg=T["tx"],
                                  anchor="w", cursor="hand2")
            lbl_titulo.pack(side="left", pady=7)

            # Badge con cantidad de ítems
            badge = tk.Label(hdr_content,
                text=str(n),
                font=(FONT, 7, "bold"),
                bg=c_acent, fg="white",
                padx=5, pady=1,
                cursor="hand2")
            badge.pack(side="left", padx=6)

            # Flecha a la derecha
            lbl_arrow = tk.Label(hdr_content,
                text="▼" if abierto else "▶",
                font=(FONT, 8), bg=bg_hdr,
                fg=T.get("t2","#8b949e"),
                cursor="hand2")
            lbl_arrow.pack(side="right", padx=10)

            # Línea inferior del encabezado (color de acento cuando abierto)
            sep_line = tk.Frame(hdr_inner, height=1,
                                bg=c_acent if abierto else T["br"])
            sep_line.pack(fill="x")

            # ── Cuerpo del grupo (botones) ────────────────────────────
            # Fondo del body ligeramente distinto al sb
            bg_body = T["sb"]

            grp_body = tk.Frame(wrap, bg=bg_body)
            if abierto:
                grp_body.pack(fill="x")

            def _toggle(e=None):
                if estado["abierto"]:
                    grp_body.pack_forget()
                    lbl_arrow.configure(text="▶")
                    sep_line.configure(bg=T["br"])
                    bar.configure(bg=T.get("br","#2a2e35"))
                    estado["abierto"] = False
                else:
                    grp_body.pack(fill="x")
                    lbl_arrow.configure(text="▼")
                    sep_line.configure(bg=c_acent)
                    bar.configure(bg=c_acent)
                    estado["abierto"] = True
                _inner.update_idletasks()
                _cv.configure(scrollregion=_cv.bbox("all"))

            # Si está cerrado inicialmente, ajustar colores de barra
            if not abierto:
                bar.configure(bg=T.get("br","#2a2e35"))

            for widget in (hdr_row, hdr_inner, hdr_content, lbl_ico,
                           lbl_titulo, badge, lbl_arrow):
                widget.bind("<Button-1>", _toggle)

            def _hdr_enter(e=None):
                hdr_inner.configure(bg=_blend(T["sb"], 0.16))
                hdr_content.configure(bg=_blend(T["sb"], 0.16))
                for w in (lbl_ico, lbl_titulo, lbl_arrow, badge):
                    try: w.configure(bg=_blend(T["sb"], 0.16))
                    except Exception: pass
            def _hdr_leave(e=None):
                hdr_inner.configure(bg=bg_hdr)
                hdr_content.configure(bg=bg_hdr)
                for w in (lbl_ico, lbl_titulo, lbl_arrow):
                    try: w.configure(bg=bg_hdr)
                    except Exception: pass

            for widget in (hdr_inner, hdr_content, lbl_ico, lbl_titulo, lbl_arrow):
                widget.bind("<Enter>", _hdr_enter)
                widget.bind("<Leave>", _hdr_leave)

            # ── Botones del grupo ─────────────────────────────────────
            bg_btn_hover = _blend(T["card"], 0.05)
            for idx, (txt, cmd) in enumerate(botones):
                # Separador muy sutil entre botones (excepto el primero)
                if idx > 0:
                    tk.Frame(grp_body, bg=T["br"], height=1).pack(
                        fill="x", padx=16)

                btn_row = tk.Frame(grp_body, bg=bg_body, cursor="hand2")
                btn_row.pack(fill="x")

                # Indicador lateral (fino, del color del grupo)
                btn_ind = tk.Frame(btn_row, bg=bg_body, width=3)
                btn_ind.pack(side="left", fill="y")
                btn_ind.pack_propagate(False)

                b = tk.Button(btn_row,
                    text=f"  {txt}",
                    font=(FONT, 9),
                    bg=bg_body, fg=T.get("t2","#8b949e"),
                    activebackground=c_acent,
                    activeforeground="white",
                    relief="flat", anchor="w",
                    padx=12, pady=6,
                    cursor="hand2", bd=0, command=cmd)
                b.pack(side="left", fill="x", expand=True)

                def _btn_enter(e, w=b, ind=btn_ind, row=btn_row, c=c_acent):
                    w.configure(bg=T["card"], fg=T["tx"])
                    row.configure(bg=T["card"])
                    ind.configure(bg=c)

                def _btn_leave(e, w=b, ind=btn_ind, row=btn_row):
                    w.configure(bg=bg_body, fg=T.get("t2","#8b949e"))
                    row.configure(bg=bg_body)
                    ind.configure(bg=bg_body)

                btn_row.bind("<Enter>", _btn_enter)
                btn_row.bind("<Leave>", _btn_leave)
                b.bind("<Enter>", _btn_enter)
                b.bind("<Leave>", _btn_leave)
                btn_ind.bind("<Button-1>", lambda e, c=cmd: c())
                btn_ind.bind("<Enter>", _btn_enter)
                btn_ind.bind("<Leave>", _btn_leave)
                b.bind("<MouseWheel>", _on_wheel)
                btn_row.bind("<MouseWheel>", _on_wheel)

            # Línea de cierre del grupo
            tk.Frame(wrap, bg=T["br"], height=1).pack(fill="x")

            return grp_body

        # ── SECCIÓN: REMUNERACIONES ───────────────────────────────
        _sb_group("📊 Remuneraciones", [
            ("📂  Carpetas",              self._win_carpetas),
            ("📊  Dotación diaria",       lambda: self._chip_cmd(
                "Genera script Python para el informe de dotación diaria de hoy")),
            ("💰  Comisiones SLD",        lambda: self._chip_cmd(
                "Explica el proceso y genera macro VBA para el cierre de comisiones SLD del mes")),
            ("📧  Correo dotación",       lambda: self._chip_cmd(
                "Redacta el correo de dotación semanal para gerencia")),
        ], abierto=True, color_key="rem")

        _sb_group("📑 Excel & Datos", [
            ("📝  Modificar Excel",       self._win_mod_xl),
            ("✏️  Escribir Excel",        self._win_escribir_xl),
            ("🔄  Comparar Excel",        self._win_cmp_xl),
            ("📂  Cargar carpeta",        self._cargar_carpeta_analistas),
        ], abierto=False, color_key="xl")

        _sb_group("📋 Finiquitos", [
            ("🔄  Sincronizar",           lambda: self._daemon.sincronizar_finiquitos()),
            ("📋  Matriz Finiquitos",     self._abrir_matriz_finiquitos),
            ("📈  Dashboard",             lambda: self._chip_cmd("genera dashboard finiquitos")),
        ], abierto=False, color_key="fin")

        _sb_group("🛠️ Gestión", [
            ("🧠  Procesos",             self._win_procesos),
            ("⏰  Alarmas",              self._win_alarmas),
            ("📝  Nota rápida",          self._win_nota),
            ("📊  Dashboard general",    self._win_dashboard),
            ("📈  KPIs",                 self._win_kpi),
            ("📋  Gantt Check-in",       self._win_gantt),
            ("🗂️  Registro REMI",        self._win_registro),
        ], abierto=False, color_key="ges")

        # Sección Admin solo si rol=admin
        try:
            _ses = getattr(self, "_sesion", _SESION_ACTUAL)
            if _ses.get("rol") == "admin":
                _sb_group("🔐 Administración", [
                    ("👥  Usuarios",         self._win_usuarios),
                    ("🌐  Panel Web Admin",  lambda: webbrowser.open(
                        f"http://localhost:{_WEB_PORT}/admin")),
                ], abierto=False, color_key="adm")
        except Exception:
            pass

        # Activar scroll con rueda en todo el inner
        try:
            _inner.update_idletasks()
            _bind_wheel(_inner)
        except Exception:
            pass

        # ╔══════════════════════════════════════════╗
        # ║  FOOTER FIJO (config + status)           ║
        # ╚══════════════════════════════════════════╝
        ftr = tk.Frame(sb, bg=T["sb"])
        ftr.grid(row=2, column=0, sticky="ew")

        tk.Frame(ftr, bg=T["br"], height=1).pack(fill="x")
        bc = self._mk_sb_btn(ftr, "⚙️  Configuración", self._win_config, fg=T["t2"])
        bc.pack(fill="x", padx=10, pady=(4,2))
        bc.bind("<Enter>", lambda e: bc.configure(bg=T["card"], fg=T["a2"]))
        bc.bind("<Leave>", lambda e: bc.configure(bg=T["sb"], fg=T["t2"]))

        self._lbl_srv = tk.Label(ftr, text="○ Graph API: —", font=(FONT, 7),
                                  bg=T["sb"], fg=T["t2"], cursor="hand2")
        self._lbl_srv.pack(anchor="w", padx=14, pady=(2,0))
        self._lbl_srv.bind("<Button-1>", lambda e: self.after(50, self._srv_reconectar))

        self._lbl_sys = tk.Label(ftr, text="", font=(FONT, 7),
                                  bg=T["sb"], fg=T["t2"])
        self._lbl_sys.pack(anchor="w", padx=14, pady=(0,4))
        self._tick_sys()
        self._tick_srv()

    # ── Helpers de sidebar (para burbujas y otras ventanas) ────────
    def _sb_sep(self, parent, titulo):
        f = tk.Frame(parent, bg=T["sb"]); f.pack(fill="x", padx=10, pady=(10,2))
        tk.Label(f, text=titulo.upper(), font=(FONT, 7),
                 bg=T["sb"], fg=T["t2"]).pack(anchor="w")
        tk.Frame(f, bg=T["br"], height=1).pack(fill="x", pady=(2,0))

    def _sb_btn(self, parent, txt, cmd):
        b = self._mk_sb_btn(parent, txt, cmd)
        b.pack(fill="x", padx=8, pady=2)
        b.bind("<Enter>", lambda e,w=b: w.configure(bg=T["card"], fg=T["tx"]))
        b.bind("<Leave>", lambda e,w=b: w.configure(bg=T["sb"], fg=T["tx"]))
        return b

    def _mk_sb_btn(self, parent, txt, cmd, bg=None, fg=None, bold=False, pady=5):
        return tk.Button(parent, text=f"  {txt}", font=(FONT, 10, "bold" if bold else "normal"),
                         bg=bg or T["sb"], fg=fg or T["tx"],
                         activebackground=T["a1"], activeforeground="white",
                         relief="flat", anchor="w", padx=10, pady=pady,
                         cursor="hand2", bd=0, command=cmd)

    # ── ÁREA PRINCIPAL ────────────────────────────────────────
    def _build_main(self):
        m = tk.Frame(self, bg=T["bg"]); m.grid(row=0, column=1, sticky="nsew")
        m.columnconfigure(0, weight=1)
        self._main = m

        # ── Header mejorado (row 0) ──
        hdr = tk.Frame(m, bg=T["sb"]); hdr.grid(row=0, column=0, sticky="ew")
        # Barra de color izquierda (3px acento)
        tk.Frame(hdr, bg=T["a1"], width=3).pack(side="left", fill="y")

        hdr_l = tk.Frame(hdr, bg=T["sb"]); hdr_l.pack(side="left", padx=18, pady=10)
        now  = datetime.datetime.now()
        dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

        # Fila título + badge estado IA en la misma línea
        titulo_row = tk.Frame(hdr_l, bg=T["sb"]); titulo_row.pack(anchor="w")
        tk.Label(titulo_row, text="Asistente de Remuneraciones",
                 font=(FONT, 13, "bold"), bg=T["sb"], fg=T["tx"]).pack(side="left")
        # Badge "IA activa" inline
        _badge_ia = tk.Label(titulo_row, text=" IA ",
                              font=(FONT, 7, "bold"),
                              bg=T.get("a1","#3fb950"), fg="white",
                              padx=4, pady=1)
        _badge_ia.pack(side="left", padx=(8,0), pady=2)

        alerta_hoy = next((msg for d,msg in CAL_REM.items() if abs(now.day-d)<=0), "")
        sub_txt = f"{dias[now.weekday()]}  {now.strftime('%d/%m/%Y')}"
        if alerta_hoy: sub_txt += f"   ·  {alerta_hoy}"
        tk.Label(hdr_l, text=sub_txt, font=(FONT, 9), bg=T["sb"],
                 fg=T["warn"] if alerta_hoy else T["t2"]).pack(anchor="w")

        hdr_r = tk.Frame(hdr, bg=T["sb"]); hdr_r.pack(side="right", padx=16)
        self._lbl_est = tk.Label(hdr_r, text="● Listo", font=(FONT, 9),
                                  bg=T["sb"], fg=T["ok"])
        self._lbl_est.pack(anchor="e")
        # Botón historial en header principal
        _hbtn = tk.Button(hdr_r, text="💬 Historial", font=(FONT, 8),
                          bg=T["card"], fg=T["t2"],
                          activebackground=T["a1"], activeforeground="white",
                          relief="flat", padx=8, pady=3, cursor="hand2", bd=0,
                          command=self._win_historial_chats)
        _hbtn.pack(anchor="e", pady=(2,0))
        _hbtn.bind("<Enter>", lambda e: _hbtn.configure(bg=T["a1"], fg="white"))
        _hbtn.bind("<Leave>", lambda e: _hbtn.configure(bg=T["card"], fg=T["t2"]))
        self._pv = tk.StringVar(value=CFG.get("proveedor","claude"))
        sel_f = tk.Frame(hdr_r, bg=T["sb"]); sel_f.pack(anchor="e", pady=(2,0))
        tk.Label(sel_f, text="IA:", font=(FONT,8), bg=T["sb"], fg=T["t2"]).pack(side="left")
        prov_menu = tk.OptionMenu(sel_f, self._pv,
                                   "claude","gemini","openai","mistral","ollama",
                                   command=self._cambiar_prov)
        prov_menu.configure(bg=T["card"], fg=T["a2"], activebackground=T["a1"],
                            activeforeground="white", relief="flat",
                            font=(FONT,9), bd=0, highlightthickness=0, padx=4)
        prov_menu["menu"].configure(bg=T["card"], fg=T["tx"],
                                    activebackground=T["a1"], font=(FONT,9))
        prov_menu.pack(side="left", padx=2)

        # ── Chips acceso rápido mejorados (row 1) ──
        chips_row = tk.Frame(m, bg=T["sb"]); chips_row.grid(row=1, column=0, sticky="ew")
        tk.Frame(chips_row, bg=T["br"], height=1).pack(fill="x")
        cf2 = tk.Frame(chips_row, bg=T["sb"]); cf2.pack(fill="x", padx=12, pady=(5,5))
        # Label "Acceso rápido:"
        tk.Label(cf2, text="⚡", font=(FONT, 9),
                 bg=T["sb"], fg=T.get("a1","#3fb950")).pack(side="left", padx=(0,6))
        for etiq, prompt in CHIPS_RAPIDOS:
            b = tk.Button(cf2, text=etiq, font=(FONT, 8),
                          bg=T["card"], fg=T["t2"],
                          activebackground=T["a1"], activeforeground="white",
                          relief="flat", padx=10, pady=4,
                          cursor="hand2", bd=0,
                          highlightthickness=1,
                          highlightbackground=T["br"],
                          command=lambda p=prompt: self._chip_cmd(p))
            b.pack(side="left", padx=2)
            b.bind("<Enter>", lambda e,w=b: w.configure(
                bg=T["a1"], fg="white", highlightbackground=T["a1"]))
            b.bind("<Leave>", lambda e,w=b: w.configure(
                bg=T["card"], fg=T["t2"], highlightbackground=T["br"]))

        # ── Chat scrollable (row 2, expande) ──
        self._sc = ctk.CTkScrollableFrame(m, fg_color=T["bg"],
                                           scrollbar_button_color=T["br"],
                                           scrollbar_button_hover_color=T["a1"],
                                           corner_radius=0)
        self._sc.grid(row=2, column=0, sticky="nsew")
        self._sc.columnconfigure(0, weight=1)

        # ── Progress bar (row 3) ──
        self._pb = ctk.CTkProgressBar(m, mode="indeterminate", height=3,
                                       fg_color=T["bg"], progress_color=T["a1"],
                                       corner_radius=0)
        self._pb.grid(row=3, column=0, sticky="ew"); self._pb.set(0)

        # ── Preview adjunto (row 4) ──
        self._frm_img = tk.Frame(m, bg=T["card"])
        self._frm_img.grid(row=4, column=0, sticky="ew")
        self._lbl_img  = tk.Label(self._frm_img, text="", font=(FONT,9),
                                   bg=T["card"], fg=T["a2"])
        self._lbl_file = tk.Label(self._frm_img, text="", font=(FONT,9),
                                   bg=T["card"], fg=T["a2"])

        # ── Separador + input (rows 5-6) ──────────────────────────────────
        tk.Frame(m, bg=T["br"], height=1).grid(row=5, column=0, sticky="ew")

        inp_area = tk.Frame(m, bg=T["sb"])
        inp_area.grid(row=6, column=0, sticky="ew")
        inp_area.columnconfigure(0, weight=1)

        # ── Caja de texto — LAYOUT CORRECTO ──────────────────────────────
        # row=0: el tk.Text con overlay del hint
        # row=1: barra de botones

        # Contenedor del texto (para poder poner el hint encima)
        inp_box = tk.Frame(inp_area, bg=T["inp"],
                           highlightthickness=1,
                           highlightbackground=T["br"])
        inp_box.grid(row=0, column=0, padx=12, pady=(10,4), sticky="ew")
        inp_box.columnconfigure(0, weight=1)

        # tk.Text dentro del contenedor
        self._inp = tk.Text(inp_box, height=3,
                            font=(FONT, 12),
                            bg=T["inp"], fg=T["tx"],
                            insertbackground=T["tx"],
                            relief="flat", bd=0,
                            wrap="word",
                            padx=10, pady=8,
                            selectbackground=T.get("a1","#3fb950"),
                            selectforeground="white",
                            highlightthickness=0)
        self._inp.pack(fill="x", expand=True)

        # Hint label superpuesto (desaparece al escribir)
        self._inp_hint = tk.Label(inp_box,
            text="  Escribe tu mensaje…  (Enter = Enviar, Shift+Enter = nueva línea)",
            font=(FONT, 10), bg=T["inp"], fg=T["t2"], anchor="w",
            cursor="xterm")
        self._inp_hint.place(x=0, y=0, relwidth=1, height=36)
        self._inp_hint.bind("<Button-1>", lambda e: (
            self._inp.focus_set(),
            self._inp_hint.place_forget()
        ))

        # Al hacer foco en el Text → ocultar hint
        def _inp_focus_in(e=None):
            try: self._inp_hint.place_forget()
            except: pass
        def _inp_focus_out(e=None):
            try:
                if not self._inp.get("1.0","end").strip():
                    self._inp_hint.place(x=0, y=0, relwidth=1, height=36)
            except: pass

        # Foco inicial en el inp
        inp_box.bind("<Button-1>", lambda e: self._inp.focus_set())

        # Binds directos al tk.Text
        self._inp.bind("<Return>",    self._on_enter)
        self._inp.bind("<KeyPress>",  self._on_keypress_inp)
        self._inp.bind("<FocusIn>",   _inp_focus_in)
        self._inp.bind("<FocusOut>",  _inp_focus_out)
        self._inp.bind("<Control-v>", lambda e: self._paste_image())
        self._inp.bind("<Control-V>", lambda e: self._paste_image())

        self._ph_active = False

        # ── Barra inferior: acciones + enviar ─────────────────────────────
        bf = tk.Frame(inp_area, bg=T["sb"])
        bf.grid(row=1, column=0, padx=10, pady=(0,10), sticky="ew")

        def ibtn(parent, t, cmd, tooltip=""):
            b = tk.Button(parent, text=t, font=(FONT, 11),
                          bg=T["card"], fg=T["t2"],
                          activebackground=T["a1"], activeforeground="white",
                          relief="flat", padx=8, pady=5,
                          cursor="hand2", bd=0, command=cmd)
            b.pack(side="left", padx=2)
            b.bind("<Enter>", lambda e: b.configure(bg=T["a1"], fg="white"))
            b.bind("<Leave>", lambda e: b.configure(bg=T["card"], fg=T["t2"]))
            return b

        _acc_frame = tk.Frame(bf, bg=T["sb"]); _acc_frame.pack(side="left")
        ibtn(_acc_frame, "📷", self._adj_img, "Imagen")
        ibtn(_acc_frame, "📎", self._adj_file, "Archivo")
        ibtn(_acc_frame, "🎤", self._voz, "Voz")
        ibtn(_acc_frame, "🌐", self._buscar_manual, "Buscar web")

        # Separador vertical
        tk.Frame(bf, bg=T["br"], width=1).pack(side="left", fill="y", padx=6, pady=4)

        # Campo ruta — ahora más compacto y estilizado
        self._e_ruta = tk.Entry(bf, font=(FMONO, 9),
                                bg=T["card"], fg=T["t2"],
                                insertbackground=T["tx"],
                                relief="flat", bd=0,
                                highlightthickness=1,
                                highlightbackground=T["br"],
                                highlightcolor=T.get("a1","#3fb950"))
        self._e_ruta.pack(side="left", fill="x", expand=True, ipady=5, padx=(4,6))
        self._e_ruta.insert(0, "Pega ruta de Excel…")
        self._e_ruta.configure(fg=T["t2"])

        def _ruta_focus_in(e):
            if self._e_ruta.get().startswith("Pega"):
                self._e_ruta.delete(0, "end")
                self._e_ruta.configure(fg=T["tx"])

        def _ruta_focus_out(e):
            if not self._e_ruta.get().strip():
                self._e_ruta.insert(0, "Pega ruta de Excel…")
                self._e_ruta.configure(fg=T["t2"])

        def _adjuntar_ruta(e=None):
            ruta = self._e_ruta.get().strip().strip('"\'')
            if not ruta or ruta.startswith("Pega"):
                return
            if not os.path.exists(ruta):
                self._e_ruta.configure(fg="#f08080")
                self._e_ruta.delete(0, "end")
                self._e_ruta.insert(0, f"✗ No encontrado: {os.path.basename(ruta)}")
                return
            self._file = ruta
            nombre = os.path.basename(ruta)
            self._lbl_file.configure(text=f"  📎  {nombre}   [doble clic = quitar]")
            self._lbl_file.pack(pady=(0,4))
            self._lbl_file.bind("<Double-Button-1>", lambda _: self._quit_file())
            self._e_ruta.delete(0, "end")
            self._e_ruta.configure(fg=T["t2"])
            self._e_ruta.insert(0, "Pega ruta de Excel…")
            if ruta.lower().endswith((".xlsx",".xlsm",".xls")):
                ctx = leer_excel_contexto(ruta, max_filas=5)
                self._bub_sys(
                    f"✅ **{nombre}** adjuntado.\n\n"
                    f"```\n{ctx[:600]}\n```\n\n"
                    f"¿Qué quieres hacer con este archivo?", "ok")
            else:
                self._bub_sys(f"✅ **{nombre}** adjuntado. ¿Qué quieres hacer?", "ok")

        self._e_ruta.bind("<FocusIn>",  _ruta_focus_in)
        self._e_ruta.bind("<FocusOut>", _ruta_focus_out)
        self._e_ruta.bind("<Return>",   _adjuntar_ruta)

        # Botón Enviar — más prominente
        self._btn_send = tk.Button(bf, text="  Enviar ⏎  ",
                                    font=(FONT, 10, "bold"), bg=T["a1"], fg="white",
                                    activebackground=T["a2"], activeforeground="white",
                                    relief="flat", padx=14, pady=7,
                                    cursor="hand2", bd=0, command=self._enviar)
        self._btn_send.pack(side="right", padx=(6,0))
        self._btn_send.bind("<Enter>", lambda e: self._btn_send.configure(bg=T["a2"]))
        self._btn_send.bind("<Leave>", lambda e: self._btn_send.configure(bg=T["a1"]))

        m.rowconfigure(0, weight=0)   # header
        m.rowconfigure(1, weight=0)   # chips
        m.rowconfigure(2, weight=1)   # chat scrollable ← se expande
        m.rowconfigure(3, weight=0)   # progress bar
        m.rowconfigure(4, weight=0)   # preview adjunto
        m.rowconfigure(5, weight=0)   # separador
        m.rowconfigure(6, weight=0)   # input
    # ══════════════════════════════════════════════════════════
    # HELPERS UI
    # ══════════════════════════════════════════════════════════
    _PH_TXT = ""   # sin placeholder en el widget

    def _set_ph(self):
        """Limpia el input después de enviar y muestra el hint."""
        try:
            self._inp.delete("1.0", "end")
            self._inp_hint.place(x=0, y=0, relwidth=1, height=36)
        except Exception:
            pass
        self._ph_active = False

    def _limpiar_ph(self, _=None):
        """Oculta el hint al hacer foco."""
        try: self._inp_hint.place_forget()
        except: pass
        self._ph_active = False

    def _on_keypress_inp(self, e=None):
        """Oculta el hint al escribir."""
        try: self._inp_hint.grid_remove()
        except: pass

    def _restaurar_ph(self, _=None):
        """Muestra el hint si el input está vacío."""
        try:
            if not self._inp.get("1.0","end").strip():
                self._inp_hint.place(x=0, y=0, relwidth=1, height=36)
        except: pass

    def _scroll(self):
        # Debounce: evita spamear after() en streaming
        if self._scroll_job is not None:
            return
        try:
            self._scroll_job = self.after(60, self._scroll_flush)
        except:
            self._scroll_job = None

    def _scroll_flush(self):
        self._scroll_job = None
        try:
            if hasattr(self._sc, "_parent_canvas") and self._sc.winfo_exists():
                self._sc._parent_canvas.yview_moveto(1.0)
        except:
            pass

    def _est(self, txt, color=None):
        try: self._lbl_est.configure(text=txt, fg=color or T["ok"])
        except: pass

    def _pb_on(self):
        try: self._pb.start(); self._pb.configure(progress_color=T["a1"])
        except: pass

    def _pb_off(self):
        try: self._pb.stop(); self._pb.set(0)
        except: pass

    def _update_prov_label(self):
        try:
            sel = (CFG.get("proveedor", "claude") or "claude").lower().strip()
            nm  = PROV_NOMBRES.get(sel, "—")
            ic  = PROV_ICON.get(sel, "○")
            ok  = _prov_configurado(sel)
            cl  = PROV_COLOR.get(sel, T.get("t2", "#666")) if ok else T.get("t2", "#666")

            if ok:
                suf = ""
            elif sel == "claude" and not CLAUDE:
                suf = " (falta librería)"
            elif sel == "gemini" and not GEMINI:
                suf = " (falta librería)"
            elif sel == "openai" and not OPENAI:
                suf = " (falta librería)"
            elif sel == "mistral" and not MISTRAL:
                suf = " (falta librería)"
            else:
                suf = " (sin key)" if sel in {"claude","gemini","openai","mistral"} else " (offline)"
            self._lbl_prov.configure(text=f"{ic} {nm}{suf}", fg=cl)
            if hasattr(self, "_pv"):
                self._pv.set(sel)
        except Exception:
            pass

    def _tick_prov(self):
        self._update_prov_label()
        if self.winfo_exists(): self.after(20000, self._tick_prov)

    def _tick_sys(self):
        if SYS:
            try:
                cpu = psutil.cpu_percent(interval=None)
                ram = psutil.virtual_memory().percent
                self._lbl_sys.configure(text=f"CPU {cpu:.0f}%  MEM {ram:.0f}%")
            except: pass
        if self.winfo_exists(): self.after(12000, self._tick_sys)


    def _tick_srv(self):
        """Comprueba Graph API directamente cada 30s y actualiza el badge."""
        def _check():
            try:
                if not _graph_creds_ok():
                    email = CFG.get("graph_user_email","").strip()
                    if not email:
                        txt = "○ Graph API: falta Email OD en config"
                    else:
                        txt = "○ Graph API: sin config (Tenant/Client/Secret)"
                    color = T["t2"]
                else:
                    graph_token()
                    email = CFG.get("graph_user_email","").strip()
                    short = email.split("@")[0][:12] if "@" in email else email[:12]
                    txt   = f"● Graph API: {short}@…"
                    color = T["ok"]
            except Exception as _e:
                err_s = str(_e)[:40]
                txt   = f"⊘ Graph: {err_s}"
                color = T["err"]
            try:
                self.after(0, lambda t=txt, c=color: self._lbl_srv.configure(text=t, fg=c))
            except Exception:
                pass
        import threading
        threading.Thread(target=_check, daemon=True).start()
        if self.winfo_exists(): self.after(30_000, self._tick_srv)

    def _srv_reconectar(self):
        """Fuerza comprobación Graph API inmediata. Sin burbuja — solo actualiza el label."""
        now = time.time()
        if hasattr(self, "_srv_last_check") and now - self._srv_last_check < 3:
            return
        self._srv_last_check = now
        # Limpiar caché de token para forzar re-auth
        _graph_token_cache["token"]   = None
        _graph_token_cache["expires"] = 0
        try:
            self.after(0, lambda: self._lbl_srv.configure(text="○ Verificando…", fg=T["t2"]))
        except Exception:
            pass
        self._tick_srv()

    def _daemon_log(self, msg: str):
        """Muestra un mensaje del daemon como burbuja 'system' en el chat."""
        self._bub_sys(msg, "system")

    def _daemon_badge(self):
        """Actualiza el badge de estado del daemon en el sidebar."""
        try:
            activo = self._daemon.activo
            txt    = "● Daemon: activo" if activo else "○ Daemon: inactivo"
            color  = T["ok"] if activo else T["t2"]
            self._lbl_daemon.configure(text=txt, fg=color)
        except Exception:
            pass

    def _tick_anim(self):
        # Evitar loop a 20fps: solo refresca cuando toca
        now = time.time()
        try:
            if self._cv_sb and self._cv_sb.winfo_exists():
                # Redibujo si cambió estado busy (hablando) o cada ~2s
                if (self._avatar_busy_last is None) or (self._avatar_busy_last != self._busy) or (now >= self._next_avatar_at):
                    dibujar_remi(self._cv_sb, 64, habla=self._busy)
                    self._avatar_busy_last = self._busy
                    self._next_avatar_at = now + 2.0

                # Parpadeo aprox cada 8s
                if now >= self._next_blink_at:
                    dibujar_remi(self._cv_sb, 64, blink=True, habla=self._busy)
                    self.after(130, lambda: (
                        dibujar_remi(self._cv_sb, 64, habla=self._busy)
                        if self._cv_sb and self._cv_sb.winfo_exists() else None))
                    self._next_blink_at = now + 8.0
        except:
            pass

        if self.winfo_exists():
            self.after(150, self._tick_anim)

    # ══════════════════════════════════════════════════════════
    # BURBUJAS
    # ══════════════════════════════════════════════════════════
    def _bub_user(self, txt, img=None):
        b = Burbuja(self._sc, txt, "user", img)
        b.pack(fill="x", padx=16, pady=(4,6), anchor="e"); self._scroll(); return b

    def _bub_bot(self, txt=""):
        b = Burbuja(self._sc, txt, "bot")
        b.pack(fill="x", padx=16, pady=(4,6), anchor="w")
        # ── Auto-detectar URLs de Excel 365 en la respuesta ─────────────
        try:
            urls_365 = self._detectar_excel365_urls(txt)
            if urls_365:
                self.after(120, lambda u=urls_365: self._bub_excel365_links(u))
        except Exception:
            pass
        self._scroll()
        return b

    def _bub_sys(self, txt, tipo="system", codigo_pendiente=None, ofrecer_ejecutar=False):
        b = Burbuja(self._sc, txt, tipo)
        b.pack(fill="x", padx=40, pady=(2,4))

        # ── Botones de ejecución si se pide confirmación ─────────────────
        if ofrecer_ejecutar and codigo_pendiente:
            btn_bar = tk.Frame(self._sc, bg=T["bg"])
            btn_bar.pack(anchor="e", padx=44, pady=(0,6))

            def _ejecutar_ahora():
                btn_bar.destroy()
                self._auto_ejecutar_xl(codigo_pendiente)

            def _cancelar():
                btn_bar.destroy()
                self._bub_sys("Entendido, no ejecuto el código.", "system")

            tk.Button(btn_bar, text="▶  Sí, ejecutar",
                      font=(FONT,10,"bold"), bg=T["ok"], fg="white",
                      activebackground="#2d7a4a", activeforeground="white",
                      relief="flat", padx=14, pady=6, cursor="hand2", bd=0,
                      command=_ejecutar_ahora).pack(side="left", padx=(0,6))
            tk.Button(btn_bar, text="✕  No, cancelar",
                      font=(FONT,10), bg=T["card"], fg=T["t2"],
                      activebackground=T["br"], activeforeground=T["tx"],
                      highlightthickness=1, highlightbackground=T["br"],
                      padx=14, pady=6, cursor="hand2", bd=0, relief="flat",
                      command=_cancelar).pack(side="left")

        self._scroll()
        return b

    def _check_update_bg(self):
        """Verifica actualizaciones en segundo plano al arrancar."""
        def _cb(tipo, msg):
            self.after(0, lambda t=tipo, m=msg: self._on_update_result(t, m))
        _remi_check_update(callback_ui=_cb)

    def _on_update_result(self, tipo, msg):
        """Recibe el resultado del check de actualización."""
        if tipo == "update":
            # Mostrar notificación no intrusiva
            # Mostrar notificacion no intrusiva
            self._bub_sys(
                "🔔 **" + msg + "**\n\nDi **'actualizar remi'** para instalar automaticamente.",
                "ok"
            )
        elif tipo == "error":
            pass  # silencioso en errores de red

    def _bienvenida(self):
        hoy    = datetime.datetime.now()
        # Usar el nombre del usuario logueado si está disponible
        try:
            _ses = getattr(self, "_sesion", _SESION_ACTUAL)
            nombre = _ses.get("nombre") or _ses.get("usuario") or CFG.get("nombre_usuario","Yerko")
        except Exception:
            nombre = CFG.get("nombre_usuario", "Yerko")
        p      = prov_activo()
        ia     = PROV_NOMBRES.get(p, "ninguna")
        hora   = hoy.hour
        saludo = ("Buenos días"   if hora < 13
                  else "Buenas tardes" if hora < 20
                  else "Buenas noches")
        alertas   = [msg for d, msg in CAL_REM.items() if abs(hoy.day - d) <= 0]
        alert_str = ("\n\n⚠️  **Hoy en el calendario:** " + " · ".join(alertas)) if alertas else ""
        # Último archivo reciente en Descargas
        ult_str = ""
        try:
            xls = sorted(DESCARGAS.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
            if xls:
                ult_str = f"\n\nÚltimo archivo: **{xls[0].name}** — di *abre el último Excel* para verlo."
        except Exception:
            pass
        # Estado Graph API en la bienvenida
        srv_txt = ""
        try:
            if _graph_creds_ok():
                try:
                    graph_token()   # valida el token
                    email = CFG.get("graph_user_email","").strip()
                    srv_txt = "\n\n☁️  **OneDrive conectado** (" + email + ") — datos Excel 365 disponibles."
                except Exception as _ge:
                    srv_txt = "\n\n⚠️  Graph API: " + str(_ge)[:80] + " — revisa ⚙️ Configuración."
            else:
                email = CFG.get("graph_user_email","").strip()
                if not email:
                    srv_txt = "\n\n○ OneDrive: agrega el **Email OD** en ⚙️ Config → Graph API."
                else:
                    srv_txt = "\n\n○ OneDrive sin configurar — ve a ⚙️ Config → Microsoft Graph API."
        except Exception:
            srv_txt = "\n\n⚠️  Graph API: error de autenticación — revisa ⚙️ Configuración."

        texto = (
            f"{saludo} **{nombre}** 👋\n\n"
            f"Soy **REMI**, tu asistente de Remuneraciones y Control de Gestión.\n"
            f"IA activa: **{ia}**.{alert_str}{ult_str}{srv_txt}\n\n"
            f"Puedo ayudarte con:\n"
            f"  — Macros VBA · Python · Softland SQL · PowerQuery\n"
            f"  — Liquidaciones, finiquitos, contratos (Código del Trabajo)\n"
            f"  — Crear, modificar y analizar archivos Excel\n"
            f"  — Dotación, comisiones SLD/PDV/EOS, libro de remuneraciones\n"
            f"  — Redactar correos y documentos\n"
            f"  — Datos en tiempo real desde el **Remi Server**\n\n"
            f"Escríbeme directamente o usa los **accesos rápidos** de arriba."
        )
        self._bub_bot(texto)
    # ══════════════════════════════════════════════════════════
    # ENVÍO DE MENSAJES
    # ══════════════════════════════════════════════════════════
    def _on_enter(self, e):
        if e.state & 0x1:   # Shift+Enter = nueva línea
            return
        self._enviar()
        return "break"

    def _enviar(self):
        # ── Leer texto ────────────────────────────────────────────────────
        try:
            raw = self._inp.get("1.0", "end").strip()
        except Exception as _ex:
            logger.error("_enviar: error leyendo input: %s", _ex)
            return
        if not raw:
            return
        if self._busy:
            return
        txt = raw
        logger.info("_enviar: ENVIANDO txt=%r", txt[:80])

        img = self._img
        fpath = self._file

        # Procesar imágenes pegadas/adjuntas (self._imgs)
        ocr_combined = ""
        if getattr(self, '_imgs', None):
            parts = []
            try:
                from remi_utils import ocr_image
            except Exception:
                ocr_image = None

            if ocr_image:
                for p in self._imgs:
                    try:
                        t = ocr_image(Path(p))
                    except Exception:
                        t = ''
                    parts.append((os.path.basename(p), t))
            else:
                # sin OCR, sólo listar nombres
                for p in self._imgs:
                    parts.append((os.path.basename(p), ''))

            if parts:
                for name, t in parts:
                    ocr_combined += f"--- {name} ---\n{t}\n\n"

        # Limpiar input y previews
        self._inp.delete("1.0","end"); self._set_ph(); self._quit_img(); self._quit_file()

        # Mostrar adjuntos en la burbuja del usuario (mínimo, sin preview complejo)
        txt_user = txt
        if fpath:
            txt_user = f"{txt}\n\n📎 {os.path.basename(fpath)}"
        if getattr(self, '_imgs', None) and len(self._imgs) > 0:
            if len(self._imgs) == 1:
                txt_user = f"{txt_user}\n\n📷 {os.path.basename(self._imgs[0])}"
            else:
                txt_user = f"{txt_user}\n\n📷 {len(self._imgs)} imágenes adjuntas"

        self._bub_user(txt_user, img)

        # Si hay OCR, agregarlo al texto que se enviará para ayudar a "entender la idea"
        if ocr_combined:
            txt = txt + "\n\n[OCR imágenes]\n" + ocr_combined

            # Análisis automático VBA: detectar y aplicar correcciones simples
            try:
                from remi_utils import analyze_vba, apply_vba_fixes_to_file
                # Usar el texto OCR completo para buscar código VBA
                vres = analyze_vba(ocr_combined)
                if vres.get('issues') or vres.get('missing_vars'):
                    # guardar código corregido en un .bas en la carpeta BASE
                    ts = int(time.time())
                    out = DESCARGAS / f"remi_macro_corregida_{ts}.bas"
                    # si analyze_vba devolvió fixed_code lo guardamos
                    fixed = vres.get('fixed_code', ocr_combined)
                    try:
                        out.write_text(fixed, encoding='utf-8')
                        # ── Registrar macro corregida ────────────────────
                        try:
                            _registro_auto_macro(fixed, f"macro_corregida_OCR")
                        except Exception:
                            pass
                        issues_txt = '\n'.join([f"- {i}" for i in vres.get('issues',[])])
                        vars_txt = '\n'.join([f"- {n} : {t}" for n,t in vres.get('missing_vars',[])])
                        msg = f"REMI detectó problemas en el código (OCR) y generó una corrección automática:\n\n" \
                              f"Archivo guardado: {out}\n\n"
                        if issues_txt:
                            msg += f"Issues:\n{issues_txt}\n\n"
                        if vars_txt:
                            msg += f"Variables agregadas:\n{vars_txt}\n\n"
                        msg += "Puedes abrir el archivo .bas y revisarlo antes de importarlo en Excel."
                        self._bub_sys(msg, 'ok')
                    except Exception as e:
                        self._bub_sys(f"REMI detectó problemas, pero no pudo guardar el archivo: {e}", 'error')
            except Exception:
                # silenciar fallos del analizador para no bloquear la UI
                pass

        # Guardar en historial si parece un logro
        kw_logro = ["terminé","generé","listo","completé","guardé","procesé","cerré el mes"]
        if any(k in txt.lower() for k in kw_logro):
            h = _rl(F["hist"])
            h.append({"fecha":datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                       "resumen":txt[:140]})
            _w(F["hist"], h)

        # Acción directa (sin IA)
        resp_dir = self._accion_directa(txt)
        if resp_dir is not None:
            if isinstance(resp_dir, str) and resp_dir:
                # Respuesta de texto directo → mostrar burbuja
                self._bub_sys(resp_dir, "system")
                self._conv_push(txt, resp_dir)
            # True = acción del servidor ya mostró sus propias burbujas
            return

        # → IA
        self._busy = True; self._pb_on()
        self._est("● Preparando…", T["a2"])
        self._bact = self._bub_bot("▌")
        threading.Thread(target=self._ia_run, args=(txt, img, fpath), daemon=True).start()

    def _conv_push(self, user, asst):
        self._conv.append({"role":"user","content":user})
        self._conv.append({"role":"assistant","content":asst})
        if len(self._conv) > 30: self._conv = self._conv[-30:]

    def _detectar_excel365_urls(self, texto: str) -> list[str]:
        """
        Extrae URLs de Excel / SharePoint / OneDrive en el texto.
        Filtra solo las que apuntan a archivos Excel 365 (sharepoint, onedrive, 1drv).
        """
        todos = re.findall(r'https?://[^\s"]+', texto)
        return [u for u in todos if any(d in u.lower() for d in
                ["sharepoint.com", "onedrive.live", "1drv.ms",
                 "office.com", "excel", "my.sharepoint"])]

    def _bub_excel365_links(self, urls: list[str]):
        """
        Muestra una burbuja compacta con botones de acceso directo a Excel 365.
        Se llama automáticamente cuando la IA responde con URLs de 365.
        """
        if not urls:
            return
        bub = Burbuja(self._sc, tipo="ok")
        bub.pack(fill="x", padx=10, pady=(0,6))
        ctn = bub._ctn

        tk.Label(ctn, text="🔗  Links Excel 365 detectados — abre directo en el navegador:",
                 font=(FONT, 9, "bold"), bg=bub._bg, fg=T.get("ok","#4db877"),
                 anchor="w").pack(anchor="w", pady=(0,4))

        for url in urls[:5]:   # Máximo 5 links por respuesta
            disp = url
            # Extraer nombre del archivo de la URL si es posible
            match = re.search(r'/([^/?#]+\.xlsx?[^/?#]*)', url, re.I)
            if match:
                disp = "📊 " + match.group(1)[:60]
            else:
                disp = "🔗 " + url[:65] + ("…" if len(url) > 65 else "")

            btn_f = tk.Frame(ctn, bg=bub._bg)
            btn_f.pack(anchor="w", fill="x", pady=1)

            btn = tk.Button(btn_f, text=disp,
                            font=(FONT, 9), bg=T.get("card","#1c1f28"),
                            fg=T.get("a2","#6fa3e0"), relief="flat",
                            cursor="hand2", anchor="w", padx=8, pady=4, bd=0,
                            command=lambda u=url: webbrowser.open(u))
            btn.pack(side="left", fill="x", expand=True)
            btn.bind("<Enter>", lambda e, w=btn: w.configure(fg=T.get("a1","#4d85cc"),
                                                              bg=T.get("br","#282e3e")))
            btn.bind("<Leave>", lambda e, w=btn: w.configure(fg=T.get("a2","#6fa3e0"),
                                                              bg=T.get("card","#1c1f28")))

            # Botón copiar URL
            def _copy_url(u=url):
                try:
                    self.clipboard_clear(); self.clipboard_append(u)
                except Exception:
                    pass
            tk.Button(btn_f, text="📋", font=(FONT, 9),
                      bg=T.get("card","#1c1f28"), fg=T.get("t2","#7080a8"),
                      relief="flat", cursor="hand2", padx=6, pady=4, bd=0,
                      command=_copy_url).pack(side="right", padx=2)

        self._scroll()

    def _chip_cmd(self, prompt):
        self._limpiar_ph()
        self._inp.delete("1.0","end")
        self._inp.insert("1.0", prompt)
        self._inp.configure(fg=T["tx"])
        self._ph_active = False
        self._enviar()

    # ══════════════════════════════════════════════════════════
    # ACCIONES DIRECTAS (sin IA)
    # ══════════════════════════════════════════════════════════
    def _accion_directa(self, txt):
        import unicodedata
        def _norm(s):
            """Quita acentos y pasa a minúsculas para comparaciones."""
            return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode().lower()
        t = _norm(txt)

        # ── Actualización automática ────────────────────────────────────
        if re.search(r'(actualiza|update|actualizate|descarga).{0,15}(remi|version|app)', t):
            if not REMI_UPDATE_URL:
                return (
                    "Auto-actualizacion no configurada.\n\n"
                    "Para activarla:\n"
                    "1. Sube pequeno_remi.py a un repo GitHub\n"
                    "2. Copia la URL raw del archivo\n"
                    "3. Pegala en REMI_UPDATE_URL al inicio del codigo\n\n"
                    "Despues REMI se actualiza solo con decir: actualizar remi"
                )
            def _cb_update(tipo, msg):
                self.after(0, lambda t=tipo, m=msg: self._bub_sys(m, "ok" if t=="ok" else "error"))
            _remi_aplicar_update(callback_ui=_cb_update)
            return True

        # ── Ver versión ─────────────────────────────────────────────────
        if re.search(r'(que|qué).{0,10}(version|versión)|version.{0,5}(remi|tienes|eres)', t):
            upd_txt = "Actualizacion activa" if REMI_UPDATE_URL else "No configurada (ver REMI_UPDATE_URL)"
            return "Soy REMI v" + REMI_VERSION + ".\nAuto-actualizacion: " + upd_txt




        # ── Daemon de consolidación (comandos de chat) ────────────────
        if re.search(r'(activa|inicia|enciende|arranca|start).{0,20}(daemon|consolidaci[oó]n|automatiz)', t):
            if self._daemon.activo:
                return "ℹ️ El daemon ya está activo. Di **detén daemon** para pararlo."
            self._daemon.iniciar()
            CFG["daemon_activo"] = True; guardar_cfg(CFG)
            return True

        if re.search(r'(det[eé]n|para|stop|desactiva|apaga).{0,20}(daemon|consolidaci[oó]n|automatiz)', t):
            self._daemon.detener()
            CFG["daemon_activo"] = False; guardar_cfg(CFG)
            return True

        if re.search(r'(abre?|muestra|ver|visualiza).{0,25}(matriz|madre|visor)', t):
            self._abrir_matriz_finiquitos()
            return True

        # Cargar carpeta local de analistas (v10: cargar_carpeta)
        # Listar archivos de finiquitos en OneDrive
        if re.search(r'(lista|muestra|que|ver).{0,20}(archivos?|excel).{0,20}(onedrive|finiquito|carpeta)', t):
            def _listar():
                try:
                    ruta = CFG.get('finiquitos_ruta_od','').strip()
                    if not ruta:
                        self.after(0, lambda: self._bub_sys(
                            'Configura primero la Carpeta OneDrive en ⚙️ Finiquitos.', 'system')); return
                    items = graph_listar_archivos(ruta)
                    excels = [i for i in items if i['es_excel']]
                    if not excels:
                        self.after(0, lambda: self._bub_sys(f'No hay archivos Excel en {ruta}', 'system')); return
                    lines = [f'📂 {ruta}  ({len(excels)} archivos):']
                    for e in excels:
                        mod = e.get('modificado','')[:10]
                        sz  = e.get('tamanio',0)//1024
                        lines.append(f'  📊 {e["nombre"]}  —  {mod}  ({sz} KB)')
                    self.after(0, lambda t='\n'.join(lines): self._bub_sys(t, 'system'))
                except Exception as ex:
                    self.after(0, lambda err=str(ex): self._bub_sys('Error listando OneDrive: '+err, 'error'))
            import threading
            threading.Thread(target=_listar, daemon=True).start()
            return True

        if re.search(r'(carga|selecciona|abre?).{0,20}(carpeta|folder)|carpeta.{0,15}(analistas?|onedrive)', t):
            import threading
            def _pick():
                from tkinter import filedialog
                path = filedialog.askdirectory(title='Carpeta de analistas (OneDrive)')
                if not path: return
                CFG['finiquitos_ruta'] = path
                guardar_cfg(CFG)
                self._daemon.sincronizar_finiquitos()
                self.after(0, lambda: self._bub_sys(
                    'Carpeta configurada: ' + path + '  Consolidando...', 'system'))
            threading.Thread(target=_pick, daemon=True).start()
            return True

        if re.search(r'(sincroniza|actualiza|consolida).{0,25}(finiquito|matriz|madre)', t):
            self._bub_sys("🔄 Consolidando Finiquitos desde OneDrive…", "system")
            self._daemon.sincronizar_finiquitos()
            return True

        # ── Dashboard Finiquitos (módulo crear_dashboard_finiquitos) ──────────
        if re.search(r'(dashboard|visual|gr[aá]fico|reporte).{0,25}(finiquito|costo|cc)|'
                     r'(crea?|genera?|hace?).{0,20}(dashboard|reporte.*finiquito)', t):
            self._bub_sys("📊 Generando Dashboard de Finiquitos…", "system")
            def _gen_dash():
                try:
                    # Intentar leer datos reales de OneDrive primero
                    ruta_od  = CFG.get("finiquitos_ruta_od","").strip()
                    ruta_loc = CFG.get("finiquitos_ruta","").strip()
                    madre    = CFG.get("finiquitos_madre","MATRIZ_GENERAL_FINIQUITOS.xlsx")
                    datos    = []

                    if ruta_od and _graph_creds_ok():
                        try:
                            ruta_madre = ruta_od.rstrip("/") + "/" + madre
                            datos = excel_leer_todo(ruta_madre, "")
                        except Exception:
                            pass

                    if not datos and ruta_loc:
                        try:
                            import pandas as _pd
                            p_m = os.path.join(ruta_loc, madre)
                            if os.path.exists(p_m):
                                df_m = _pd.read_excel(p_m, engine="openpyxl")
                                datos = df_m.to_dict(orient="records")
                        except Exception:
                            pass

                    if not datos:
                        # Datos de ejemplo para demostración
                        datos = [
                            {"Encargado":"Yerko","Centro Costo":"Ventas","Monto":1500000,"Rut":"12345678-9","Nombre":"Juan Pérez","Estado":"Cerrado"},
                            {"Encargado":"Yerko","Centro Costo":"Logística","Monto":980000,"Rut":"98765432-1","Nombre":"María López","Estado":"Cerrado"},
                            {"Encargado":"Ignacio","Centro Costo":"Ventas","Monto":2100000,"Rut":"11111111-1","Nombre":"Pedro Soto","Estado":"En proceso"},
                            {"Encargado":"Ignacio","Centro Costo":"RRHH","Monto":750000,"Rut":"22222222-2","Nombre":"Ana García","Estado":"Cerrado"},
                        ]
                        self.after(0, lambda: self._bub_sys(
                            "⚠️ No encontré datos reales — mostrando dashboard de **ejemplo**.\n"
                            "Configura OneDrive o la carpeta local en ⚙️ Finiquitos para usar datos reales.",
                            "warn"))

                    msg = crear_dashboard_finiquitos(datos)
                    self.after(0, lambda m=msg: self._bub_sys(m, "ok"))

                    # Abrir el archivo si se generó
                    if "Guardado en:" in msg:
                        import re as _re
                        m_ruta = _re.search(r'Guardado en: `([^`]+)`', msg)
                        if m_ruta and os.path.exists(m_ruta.group(1)):
                            ruta_xlsx = m_ruta.group(1)
                            def _open_xlsx(p=ruta_xlsx):
                                try: os.startfile(p)
                                except Exception: pass
                            self.after(500, _open_xlsx)

                except Exception as e:
                    self.after(0, lambda err=str(e): self._bub_sys(
                        f"❌ Error generando dashboard: {err}", "error"))

            import threading
            threading.Thread(target=_gen_dash, daemon=True).start()
            return True


        if re.search(r'estado.{0,15}daemon|daemon.{0,15}(activo|corriendo|estado)', t):
            _d_on   = self._daemon.activo
            _d_ico  = "🟢" if _d_on else "⚫"
            _d_est  = "ACTIVO" if _d_on else "INACTIVO"
            _d_hint = "Di **detén daemon** para pararlo." if _d_on else "Di **activa daemon** para iniciarlo."
            _d_ivl  = CFG.get("daemon_intervalo", 30)
            _d_msg  = (
                _d_ico + " **Daemon de consolidación: " + _d_est + "**\n\n"
                "  · Carpeta vigilada: `" + str(self._daemon.entrada) + "`\n"
                "  · Intervalo: **" + str(_d_ivl) + " minutos**\n"
                "  · Maestro: `" + str(self._daemon.maestro) + "`\n\n"
                + _d_hint
            )
            return _d_msg

        if re.search(r'procesa.{0,20}(ahora|ya|inmediato|manual)|fuerza.{0,15}consolidaci', t):
            self._bub_sys("🔄 Forzando consolidación manual…", "system")
            import threading as _th
            _th.Thread(target=self._daemon._procesar, daemon=True).start()
            return True

        # ── OneDrive / Excel 365 — directo vía Graph API ──────────────────────
        hoy = datetime.datetime.now()

        # ── Diagnóstico Graph API desde chat ──────────────────────────────
        if re.search(r"(verifica|prueba|comprueba|estado).{0,20}(graph|azure|onedrive|365|conexi[oó]n)", t):
            def _diag_graph():
                if not _graph_creds_ok():
                    email = CFG.get("graph_user_email","").strip()
                    tid, cid, sec = _graph_creds()
                    faltantes = []
                    if not tid: faltantes.append("Tenant ID")
                    if not cid: faltantes.append("Client ID")
                    if not sec: faltantes.append("Secret")
                    if not email or "@" not in email: faltantes.append("Email OD")
                    msg = ("⚙️ **Graph API incompleto**\n\n"
                           "Falta configurar: **" + ", ".join(faltantes) + "**\n\n"
                           "Ve a ⚙️ Configuración → Microsoft Graph API y completa todos los campos.")
                    self.after(0, lambda m=msg: self._bub_sys(m, "warn"))
                    return
                try:
                    import urllib.request, urllib.parse, json as _j, time as _t
                    tid, cid, sec = _graph_creds()
                    email = CFG.get("graph_user_email","").strip()

                    # 1. Token
                    token_url = "https://login.microsoftonline.com/" + tid + "/oauth2/v2.0/token"
                    body_d = urllib.parse.urlencode({
                        "grant_type": "client_credentials",
                        "client_id": cid,
                        "client_secret": sec,
                        "scope": "https://graph.microsoft.com/.default",
                    }).encode()
                    req = urllib.request.Request(token_url, data=body_d,
                        headers={"Content-Type": "application/x-www-form-urlencoded"})
                    with urllib.request.urlopen(req, timeout=15) as r:
                        tok = _j.loads(r.read())
                    token = tok.get("access_token","")
                    exp   = tok.get("expires_in", 0)

                    # 2. Organización
                    req2 = urllib.request.Request(
                        "https://graph.microsoft.com/v1.0/organization",
                        headers={"Authorization": "Bearer " + token, "Accept": "application/json"})
                    with urllib.request.urlopen(req2, timeout=15) as r2:
                        org = _j.loads(r2.read())
                    org_name = org.get("value",[{}])[0].get("displayName","—")

                    # 3. Drive del usuario
                    req3 = urllib.request.Request(
                        "https://graph.microsoft.com/v1.0/users/" + email + "/drive",
                        headers={"Authorization": "Bearer " + token, "Accept": "application/json"})
                    try:
                        with urllib.request.urlopen(req3, timeout=15) as r3:
                            drv = _j.loads(r3.read())
                        used_gb = drv.get("quota",{}).get("used",0) / (1024**3)
                        total_gb = drv.get("quota",{}).get("total",1) / (1024**3)
                        drv_info = f"  · Drive: {used_gb:.1f} GB / {total_gb:.0f} GB"
                    except Exception as ed:
                        drv_info = f"  · Drive: error ({str(ed)[:60]})"

                    _graph_token_cache["token"] = token
                    _graph_token_cache["expires"] = _t.time() + exp
                    msg = ("OK Graph API\n\n"
                           "Organizacion: " + org_name + "\n"
                           "Usuario: " + email + "\n"
                           "Token valido: " + str(exp//60) + " min\n"
                           + drv_info + "\n\n"
                           "Puedes usar todos los comandos de OneDrive/Excel 365.")
                except Exception as e:
                    err = str(e)
                    if "AADSTS" in err:
                        hint = "\nRevisión Azure: Permisos de API → Conceder consentimiento de administrador."
                    elif "404" in err:
                        hint = "\nEl Email OD no existe en este tenant. Verifica el email corporativo."
                    elif "401" in err or "403" in err:
                        hint = "\nVerifica que los permisos Files.ReadWrite.All y Sites.Read.All están concedidos."
                    else:
                        hint = ""
                    self.after(0, lambda m="❌ Error Graph API: " + err[:200] + hint:
                               self._bub_sys(m, "error"))

            import threading
            threading.Thread(target=_diag_graph, daemon=True).start()
            return True

        # Dotación diaria
        if re.search(r'dotaci[oó]n.{0,20}(hoy|diaria|activa)?|genera.{0,15}dotaci[oó]n', t):
            self._bub_sys("☁️ Leyendo dotación desde OneDrive…", "system")
            def _dot():
                try:
                    ruta = CFG.get("onedrive_dotacion","").strip()
                    if not ruta:
                        self.after(0, lambda: self._bub_bot(
                            "Configura la ruta del Excel de dotación en ⚙️ Configuración → OneDrive.\n"
                            "O dime: \"genera el script de dotación\" para que lo haga con IA."))
                        return
                    cols = ["RUT","Nombre","Centro de Costo","Cargo","Fecha Ingreso"]
                    datos = excel_leer_columnas(ruta, "", cols)
                    activos = [d for d in datos if d.get("Estado","ACTIVO") != "INACTIVO"]
                    por_cc = {}
                    for d in activos:
                        cc = d.get("Centro de Costo","Sin CC")
                        por_cc[cc] = por_cc.get(cc, 0) + 1
                    txt = ("📋 **Dotación activa — " + hoy.strftime("%d/%m/%Y") + "**\n\n"
                           "Total: **" + str(len(activos)) + " empleados**\n\n")
                    for cc, n in sorted(por_cc.items())[:12]:
                        txt += "  · " + cc + ": " + str(n) + "\n"
                except Exception as e:
                    txt = "❌ Error leyendo OneDrive: " + str(e)
                self.after(0, lambda v=txt: self._bub_bot(v))
            import threading; threading.Thread(target=_dot, daemon=True).start()
            return True

        # Comisiones
        if re.search(r'comisi[oó]n.{0,20}(sld|pdv|eos|mes|calcul)|cierre.{0,15}comisi[oó]n', t):
            tipo = "SLD" if "sld" in t else ("PDV" if "pdv" in t else ("EOS" if "eos" in t else "SLD"))
            self._bub_sys("☁️ Leyendo comisiones " + tipo + " desde OneDrive…", "system")
            def _com(tp=tipo):
                try:
                    ruta = CFG.get("onedrive_comisiones_" + tp.lower(), "").strip()
                    if not ruta:
                        self.after(0, lambda tp2=tp: self._bub_bot(
                            "Configura la ruta del Excel de comisiones " + tp2
                            + " en ⚙️ Configuración → OneDrive.\n"
                            "O dime: \"genera macro VBA de comisiones " + tp2 + "\" para crearla con IA."))
                        return
                    cols = ["RUT","Nombre","Venta","Meta","Comisión"]
                    datos = excel_leer_columnas(ruta, "", cols)
                    total = sum(float(str(d.get("Comisión",0)).replace(",","")) for d in datos if d.get("Comisión"))
                    txt = ("💰 **Comisiones " + tp + " — " + hoy.strftime("%m/%Y") + "**\n\n"
                           "Registros: **" + str(len(datos)) + "**\n"
                           "Total estimado: **$" + f"{total:,.0f}" + "**\n\n")
                    for d in datos[:8]:
                        txt += "  · " + str(d.get("Nombre","")) + ": $" + str(d.get("Comisión","—")) + "\n"
                except Exception as e:
                    txt = "❌ Error leyendo OneDrive: " + str(e)
                self.after(0, lambda v=txt: self._bub_bot(v))
            import threading; threading.Thread(target=_com, daemon=True).start()
            return True

        # Liquidaciones
        if re.search(r'liquidaci[oó]n|remuneraci[oó]n|sueldo.{0,10}(mes|calcul)', t):
            self._bub_sys("☁️ Leyendo liquidaciones desde OneDrive…", "system")
            def _liq():
                try:
                    ruta = CFG.get("onedrive_liquidaciones","").strip()
                    if not ruta:
                        self.after(0, lambda: self._bub_bot(
                            "Configura la ruta del Excel de liquidaciones en ⚙️ Configuración → OneDrive.\n"
                            "O dime: \"genera el script de liquidaciones\" y lo hago con IA."))
                        return
                    cols = ["RUT","Nombre","Total Haberes","Total Descuentos","Total Líquido"]
                    datos = excel_leer_columnas(ruta, "", cols)
                    hab = sum(float(str(d.get("Total Haberes",0)).replace(",","").replace("$","")) for d in datos if d.get("Total Haberes"))
                    liq = sum(float(str(d.get("Total Líquido",0)).replace(",","").replace("$","")) for d in datos if d.get("Total Líquido"))
                    txt = ("📊 **Liquidaciones " + hoy.strftime("%m/%Y") + "**\n\n"
                           "Empleados: **" + str(len(datos)) + "**\n"
                           "Total haberes: **$" + f"{hab:,.0f}" + "**\n"
                           "Total líquido: **$" + f"{liq:,.0f}" + "**")
                except Exception as e:
                    txt = "❌ Error leyendo OneDrive: " + str(e)
                self.after(0, lambda v=txt: self._bub_bot(v))
            import threading; threading.Thread(target=_liq, daemon=True).start()
            return True

        # Captura de pantalla — única que aún necesita herramienta local
        if re.search(r'captura|screenshot|pantallazo|toma.{0,10}pantalla', t):
            def _cap():
                try:
                    import mss, base64 as b64
                    with mss.mss() as sc:
                        mon  = sc.monitors[1]
                        img  = sc.grab(mon)
                        import tempfile
                        tmp  = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        mss.tools.to_png(img.rgb, img.size, output=tmp.name)
                        txt  = "📸 Captura guardada: `" + tmp.name + "`"
                        if PIL_OK:
                            self.after(0, lambda p=tmp.name, v=txt: self._bub_bot(v, img_path=p))
                            return
                except ImportError:
                    txt = "⚠️ Instala mss: `pip install mss`"
                except Exception as e:
                    txt = "❌ Error en captura: " + str(e)
                self.after(0, lambda v=txt: self._bub_bot(v))
            import threading; threading.Thread(target=_cap, daemon=True).start()
            return True

            # Listar Excel en OneDrive — DIRECTO via Graph API
            if re.search(r'(lista|muestra|qu[eé]).{0,20}excel.{0,20}(onedrive|nube|carpeta)|archivos.{0,15}(excel|xlsx).{0,15}(onedrive|nube)', t):
                self._bub_sys("🔄 Consultando OneDrive…", "system")
                def _lst():
                    try:
                        items  = graph_listar_archivos("/")
                        excels = [i for i in items if i["es_excel"]]
                        if excels:
                            txt2 = "📂 **Excel en OneDrive:**\n\n"
                            for e in excels[:15]:
                                txt2 += "  · `" + e["nombre"] + "`\n"
                            txt2 += "\nDi **copia** [columnas] **de** [archivo] **al** [destino]"
                        else:
                            txt2 = "📂 No encontré archivos Excel en la raíz de OneDrive."
                    except Exception as e:
                        txt2 = "❌ " + str(e)
                    self.after(0, lambda v=txt2: self._bub_bot(v))
                import threading; threading.Thread(target=_lst, daemon=True).start()
                return True

            # Copiar columnas de Excel a Excel (OneDrive)
            m_copia = re.search(
                r'(copia|pasa|transfiere|mueve).{0,30}(excel|archivo|planilla|informe|hoja)',
                t
            )
            if m_copia or re.search(r'(copia|pega).{0,20}columna', t):
                # Extraer rutas y columnas del mensaje
                orig_m = re.search(r'de(?:l archivo?)?\s+["\']?(/[^\s"\']+\.xlsx["\']?)', txt, re.I)
                dest_m = re.search(r'a(?:l|l\s+archivo)?\s+["\']?(/[^\s"\']+\.xlsx["\']?)', txt, re.I)
                cols_m = re.search(r'columnas?[:\s]+([A-Za-z\xe1\xe9\xed\xf3\xfa\xc1\xc9\xcd\xd3\xda\xf1\xd1\s,]+?)\s+(?:de|desde)\s', txt, re.I)

                if cols_m and orig_m and dest_m:
                    columnas_raw = [c.strip() for c in re.split(r'[,;]', cols_m.group(1)) if c.strip()]
                    ruta_origen  = orig_m.group(1).strip().strip("'\"")
                    ruta_destino = dest_m.group(1).strip().strip("'\"")

                    self._bub_sys(f"\U0001f504 Copiando {columnas_raw} de `{ruta_origen}` a `{ruta_destino}`\u2026", "system")
                    def _cpy(cols=columnas_raw, ro=ruta_origen, rd=ruta_destino):
                        try:
                            r = excel_copiar_entre_archivos(
                                ruta_origen=ro, hoja_origen="",
                                columnas=cols,
                                ruta_destino=rd, hoja_destino="",
                                modo="append",
                            )
                            if r.get("ok"):
                                muestra = r.get("muestra", [])
                                txt2 = ("✅ **Copia completada**\n\n"
                                        "**" + str(r["filas_copiadas"]) + "** filas copiadas\n"
                                        "De: `" + ro + "`\n"
                                        "A:  `" + rd + "`\n\n")
                                if muestra:
                                    txt2 += "**Primeras filas:**\n```\n"
                                    for fila in muestra:
                                        txt2 += "  " + " | ".join(str(v) for v in fila.values()) + "\n"
                                    txt2 += "```"
                            else:
                                txt2 = "❌ " + r.get("error","Error") + "\n\nVerifica rutas y columnas."
                        except Exception as e:
                            txt2 = "❌ Error Graph API: " + str(e)
                        self.after(0, lambda v=txt2: self._bub_bot(v))
        # Monitor de procesos
        if re.search(r'\b(graba|inicia|comienza|registra)\b.{0,20}proceso', t):
            nombre = re.sub(r'.*proceso\s*[:\-]?\s*','',txt,flags=re.I).strip()[:60] or "Proceso nuevo"
            return MONITOR.iniciar(nombre, list(RUTAS.values()))
        if re.search(r'\b(det[eé]n|para|termina|finaliza)\b.{0,20}(grabaci[oó]n|proceso)', t):
            return MONITOR.detener()
        if re.search(r'(procesos?\s+aprendidos?|qu[eé]\s+aprendiste)', t):
            procs = _rl(F["proc"])
            if not procs: return "No hay procesos guardados aún. Di **'graba proceso Comisiones'** para empezar."
            return "**Procesos aprendidos:**\n" + "\n".join(
                f"  • **{p['nombre']}** ({p.get('fecha','')}) — {p.get('resumen','')[:80]}"
                for p in procs[-8:])

        # Grabando activo
        if MONITOR.grabando and len(t) < 40:
            return f"🔴 Grabando: **{MONITOR.nombre}** — di **'detén grabación'** para terminar."

        # Rutas
        if re.search(r'\b(rutas?|carpetas?|directorio)\b', t):
            lines = ["**Rutas configuradas:**"]
            for n, r in RUTAS.items():
                ok = "✅" if os.path.exists(r) else "❌"
                lines.append(f"  {ok}  **{n}**: `{r}`")
            return "\n".join(lines)

        # Calendario
        if re.search(r'\b(calendario|compromisos?|fechas?)\b', t):
            hoy = datetime.datetime.now()
            lines = [f"**Calendario Remuneraciones — {datetime.datetime.now().strftime('%B %Y')}**"]
            for d, msg in sorted(CAL_REM.items()):
                marker = " ← HOY" if d == hoy.day else (" ← PRÓXIMO" if d == hoy.day+1 else "")
                lines.append(f"  Día {d:2d}: {msg}{marker}")
            return "\n".join(lines)

        # Estado IA
        if re.search(r'\b(qu[eé]\s+ia|ia\s+activa|qu[eé]\s+modelo|estado\s+ia)\b', t):
            p = prov_activo(); cadena = cadena_proveedores()
            return (f"**IA activa:** {PROV_NOMBRES.get(p,'Ninguna')} ({PROV_ICON.get(p,'')})\n"
                    f"**Cadena de fallback:** {' → '.join(PROV_NOMBRES.get(x,x) for x in cadena)}\n"
                    f"**Multi-agente:** {'Activo' if CFG.get('multi_agent') else 'Inactivo'}")

        # Ayuda / capacidades
        if re.search(r'(qu[eé]\s+puedes|para\s+qu[eé]\s+sirves|c[oó]mo\s+te\s+uso|ayuda|que\s+haces|qu[eé]\s+sabes)', t):
            return (
                "**¿Qué puedo hacer por ti?**\n\n"
                "  — *Crea un Excel con [datos]* → genero el archivo\n"
                "  — *Abre el gantt / el SLD / el último Excel* → abro el archivo\n"
                "  — *Genera una macro para [proceso]* → código VBA completo\n"
                "  — *Calcula el finiquito de [datos]* → cálculo legal\n"
                "  — *Muéstrame el calendario* → compromisos del mes\n"
                "  — *Graba proceso [nombre]* → aprendo tu flujo de trabajo\n"
                "  — *Busca [tema]* → busco en la web\n\n"
                "También puedo redactar correos, explicar normativa DT y consultar Softland."
            )

        # Limpiar / reiniciar chat
        if re.search(r'(borra|limpia|vac[ií]a|nuevo\s+chat|reinicia|empieza\s+de\s+cero)', t):
            self._conv.clear()
            return "Chat reiniciado. ¿En qué te ayudo?"

        # Hora / fecha
        if re.search(r'\b(qu[eé]\s+hora|fecha\s+hoy|d[ií]a\s+hoy)\b', t):
            now = datetime.datetime.now()
            dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
            return f"Hoy es **{dias[now.weekday()]}** {now.strftime('%d/%m/%Y')}, {now.strftime('%H:%M')} hrs."

        # Generar macro simple (VBA) si se solicita
        if re.search(r'\b(generar|genera|crear|crea|hacer)\b.*\bmacro\b', t):
            spec = self._parse_macro_spec(txt)
            if not spec:
                return "No entendí la especificación de la macro. Indica: Hoja1: Col1,Col2; Hoja2: ColA,ColB"
            vba = self._build_vba_macro(spec)
            # guardar en carpeta de datos (usar utilitario si está disponible)
            try:
                try:
                    from remi_utils import save_macro_to_file
                except Exception:
                    save_macro_to_file = None

                out = DESCARGAS / f"macro_generada_{int(time.time())}.bas"
                if save_macro_to_file:
                    save_macro_to_file(vba, out)
                else:
                    with open(out, 'w', encoding='utf-8') as fp:
                        fp.write(vba)
                # ── Registrar macro generada ─────────────────────────────
                try:
                    _registro_auto_macro(vba, f"macro_generada")
                except Exception:
                    pass
                return (f"Generé una macro VBA simple y la guardé en {out}.\n\nCódigo:\n{vba}")
            except Exception as e:
                return (f"Generé la macro (no se pudo guardar en disco):\n{vba}\n\nError al guardar: {e}")

        # ── Abrir / leer / mostrar archivo ──────────────────────────────

        # ¿El mensaje contiene una ruta de archivo Excel?
        rutas_msg = re.findall(r'[A-Za-z]:[/\\][^\n"\'<>|*?]{3,}\.xlsx?', txt, re.I)
        rutas_msg = [r.strip("\"' ,;") for r in rutas_msg]
        rutas_exist = [r for r in rutas_msg if os.path.exists(r)]

        # "adjunta esta ruta" / "usa este archivo" / ruta sola en el mensaje
        es_adjuntar = bool(re.search(r'\b(adjunta|adjuntar|usa|utiliza|carga|cargar|agrega|agregar)\b.{0,30}(ruta|archivo|excel|path|xlsx)\b', t))
        solo_ruta   = bool(rutas_exist and len(txt.strip().splitlines()) <= 2 and any(r.lower() in t for r in [r.split("\\")[-1].lower() for r in rutas_exist]))

        if rutas_exist and (es_adjuntar or solo_ruta):
            ruta = rutas_exist[0]
            nombre = os.path.basename(ruta)
            # Adjuntar automáticamente
            def _attach():
                self._file = ruta
                self._lbl_file.configure(text=f"  📎  {nombre}   [doble clic = quitar]")
                self._lbl_file.pack(pady=(0,4))
                self._lbl_file.bind("<Double-Button-1>", lambda _: self._quit_file())
            self.after(0, _attach)
            ctx = leer_excel_contexto(ruta, max_filas=8)
            return (f"✅ **{nombre}** adjuntado.\n\n"
                    f"```\n{ctx[:800]}\n```\n\n"
                    f"¿Qué quieres hacer? Por ejemplo:\n"
                    f"- *Agrégale una fila con los datos X*\n"
                    f"- *Cambia el formato de la columna Sueldo*\n"
                    f"- *Crea un resumen de los datos*")

        if re.search(r'(abre?r?|lee?r?|muestra(?:me)?|mostrar|ver|carga?r?|analiza?r?|ensena(?:me)?|dame|quiero\s+ver|abrelo|abrir?lo)\b.{0,50}(archivo|excel|xlsx|gantt|carta|libro|informe|dotaci)', t):

            # ¿Intención es ABRIR (Excel real) o LEER (preview texto)?
            quiere_abrir = bool(re.search(r'(abre?r?|abrir?lo|abrelo|muestra(?:me)?|ensena(?:me)?|dame|quiero\s+ver|ver|mostrar)', t))

            # — Resolver ruta —
            ruta_enc = None
            # 1) ruta explícita en el mensaje
            for m in re.finditer(r'[A-Za-z]:[/\\][^\n"\'<>|*?]+\.xlsx?', txt, re.I):
                r = m.group(0).strip("\"' ")
                if os.path.exists(r):
                    ruta_enc = r; break
            # 2) por palabra clave conocida
            if not ruta_enc:
                kw_rutas = {
                    "gantt": ["Gantt_Interactiva1.xlsx","Carta_Gantt.xlsx","Gantt.xlsx"],
                    "sld":   ["SLD.xlsx"],
                    "pdv":   ["PDV.xlsx"],
                    "eos":   ["EOS.xlsx"],
                    "dot":   ["Dotacion.xlsx","dotacion_diaria.xlsx"],
                }
                for clave, nombres in kw_rutas.items():
                    if clave in t:
                        base_dir = RUTAS.get(clave.upper(), BASE)
                        for nombre in nombres:
                            c = os.path.join(str(base_dir), nombre)
                            if os.path.exists(c):
                                ruta_enc = c; break
                        if not ruta_enc and os.path.exists(str(base_dir)):
                            xlsxs = glob.glob(os.path.join(str(base_dir), "*.xlsx"))
                            if xlsxs: ruta_enc = xlsxs[0]
                        if ruta_enc: break
                        # 4) Más reciente en Descargas (para "abre el último")
            if not ruta_enc:
                try:
                    xls_desc = sorted(DESCARGAS.glob("*.xlsx"),
                                      key=lambda f: f.stat().st_mtime, reverse=True)
                    if xls_desc:
                        ult_refs2 = ["ultimo","último","recien","recién","el último","el que hiciste"]
                        if any(k in t for k in ult_refs2):
                            ruta_enc = str(xls_desc[0])
                except Exception:
                    pass
            # 5) carpeta BASE de Remi
# 3) carpeta BASE de Remi
            if not ruta_enc:
                for pat in ["*.xlsx","*.xlsm"]:
                    m2 = glob.glob(os.path.join(str(BASE), pat))
                    if m2: ruta_enc = m2[0]; break

            if ruta_enc:
                nombre = os.path.basename(ruta_enc)
                if quiere_abrir:
                    # Abrir con Excel real (os.startfile en Windows)
                    try:
                        os.startfile(ruta_enc)
                        self._ultimo_excel = ruta_enc
                        return (f"✅ Abriendo **{nombre}** en Excel...\n\n"
                                f"`{ruta_enc}`\n\n"
                                f"¿Quieres que también te muestre un preview aquí dentro?")
                    except Exception as e:
                        try:
                            subprocess.Popen(["start", "", ruta_enc], shell=True)
                            return f"✅ Abriendo **{nombre}** en Excel...\n\n`{ruta_enc}`"
                        except Exception as e2:
                            return f"❌ No pude abrir el archivo: {e2}\n\nRuta: `{ruta_enc}`"
                else:
                    # Mostrar preview de texto + botón abrir
                    ctx = leer_excel_contexto(ruta_enc, max_filas=50)
                    return (f"📊 **{nombre}**  —  `{ruta_enc}`\n\n"
                            f"```excel\n{ctx}\n```\n\n"
                            f"Para abrirlo en Excel di: **abre {nombre}**")
            else:
                # Listar disponibles
                encontrados = []
                for n, r in RUTAS.items():
                    if os.path.exists(str(r)):
                        for f2 in glob.glob(os.path.join(str(r),"*.xlsx")) + glob.glob(os.path.join(str(r),"*.xlsm")):
                            encontrados.append(f"  📄 **{n}** → `{os.path.basename(f2)}`  (`{f2}`)")
                if encontrados:
                    return ("No encontré el archivo exacto. Estos están disponibles:\n\n"
                            + "\n".join(encontrados[:15])
                            + "\n\n¿Cuál quieres? Di por ejemplo: **abre Gantt_Interactiva1.xlsx**")
                return ("No encontré el archivo. ¿Está guardado en alguna de las rutas configuradas?\n\n"
                        "Puedes darme la ruta completa, ej:\n`X:\\INFORMES\\Carta_Gantt.xlsx`")

        # ── Crear / escribir Excel con datos directamente ────────────────
        if re.search(r'\b(crea|crear|genera|generar|escribe|escribir|guarda|guardar|haz|hacer|dame|arma|armar)\b.{0,50}(excel|xlsx|hoja|tabla|planilla)', t):

            # Extraer nombre sugerido del archivo
            m_name = re.search(r'(?:llamad[ao]|nombrad[ao]|de\s+nombre|archivo|excel)\s+["\']?([A-Za-záéíóúñÑ0-9_ ]+?)["\']?(?:\s|\.xlsx|$)', txt, re.I)
            nombre_arch = (m_name.group(1).strip() + ".xlsx") if m_name else "datos_remi.xlsx"
            nombre_arch = re.sub(r'[<>:"/\\|?*]', '', nombre_arch)
            if not nombre_arch.endswith(".xlsx"): nombre_arch += ".xlsx"
            ruta_out = str(DESCARGAS / nombre_arch)

            # Detectar tabla inline en el mensaje (líneas con | o tabs)
            lineas = [l.strip() for l in txt.splitlines() if l.strip()]
            tabla_lines = [l for l in lineas if "|" in l or "\t" in l]
            # También detectar listas numeradas o con guiones como datos
            lista_lines = [re.sub(r'^[\d\-\*\.\)]+\s*', '', l)
                           for l in lineas
                           if re.match(r'^[\d\-\*\.\)]+\s+.+', l)]

            if tabla_lines and len(tabla_lines) >= 2:
                # Parsear tabla con separador | o tab
                sep = "|" if "|" in tabla_lines[0] else "\t"
                tabla = [[c.strip().strip("|").strip() for c in l.split(sep) if c.strip()]
                         for l in tabla_lines if any(c.strip() for c in l.split(sep))]
                tabla = [r for r in tabla if r]
                if len(tabla) >= 2:
                    headers, rows = tabla[0], tabla[1:]
                    datos_hojas = {re.sub(r'[^\w\s]','',nombre_arch.replace('.xlsx',''))[:31]:
                                   {"headers": headers, "rows": rows}}
                    res = crear_excel_desde_datos(ruta_out, datos_hojas)
                    if "✅" in res:
                        try:
                            os.startfile(ruta_out)
                            self._ultimo_excel = ruta_out
                        except: pass
                        return (f"{res}\n\n📁 **Guardado en:** `{ruta_out}`\n\n"
                                f"El archivo se abre automáticamente en Excel.")
            elif lista_lines and len(lista_lines) >= 2:
                # Lista → columna única
                datos_hojas = {"Datos": {"headers": ["Ítem"], "rows": [[v] for v in lista_lines]}}
                res = crear_excel_desde_datos(ruta_out, datos_hojas)
                if "✅" in res:
                    try:
                        os.startfile(ruta_out)
                        self._ultimo_excel = ruta_out
                    except: pass
                    return f"{res}\n\n📁 `{ruta_out}`"

            # Sin datos inline → abrir la ventana de escritura directa
            self.after(50, self._win_escribir_xl)
            return ("Abriendo la herramienta de escritura Excel...\n\n"
                    "Puedes pegar tus datos directamente en el área de texto "
                    "(separados por Tab, coma o |) y Remi creará el archivo de inmediato.")

        return None  # → pasar a IA

    # ── MACROS: parsing simple y generación de VBA ─────────────────────
    def _parse_macro_spec(self, txt: str):
        """Intenta extraer una especificación sencilla de macro desde `txt`.

        Formatos soportados (flexible):
        - "Hoja1: ColA, ColB; Hoja2: ColX, ColY"
        - "Hoja1 (ColA,ColB)\nHoja2 (ColX,ColY)"
        Devuelve lista de {'name':..., 'cols':[...]}
        """
        t = txt.strip()
        # tomar lo que sigue a la palabra 'macro' si existe
        m = re.search(r'macro[:\-\s]*(.*)', t, flags=re.I | re.S)
        content = m.group(1).strip() if m else t
        parts = re.split(r'[;\n]+', content)
        spec = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            # intentar formas comunes
            name = None; cols = ''
            if ':' in p:
                name, cols = p.split(':', 1)
            elif '(' in p and ')' in p:
                name = p.split('(', 1)[0]
                cols = re.search(r'\((.*)\)', p).group(1) if re.search(r'\((.*)\)', p) else ''
            else:
                # quizá "Hoja1 ColA,ColB"
                m2 = re.match(r'([^,\(\:]+)\s+([A-Za-z0-9_,\s]+)', p)
                if m2:
                    name = m2.group(1)
                    cols = m2.group(2)
            if not name:
                continue
            cols_list = [c.strip() for c in re.split(r'[,:\-\\/|]+', cols) if c.strip()]
            spec.append({'name': name.strip(), 'cols': cols_list})
        return spec

    def _build_vba_macro(self, spec, macro_name='MacroGeneradaPorRemi'):
        """Genera código VBA simple que crea hojas y escribe cabeceras."""
        lines = []
        lines.append(f'Sub {macro_name}()')
        lines.append('  On Error Resume Next')
        lines.append('  Application.ScreenUpdating = False')
        for s in spec:
            name = s.get('name','').replace('"','')
            lines.append(f'  \' Hoja: {name}')
            lines.append('  Dim ws As Worksheet')
            lines.append('  Set ws = Nothing')
            lines.append(f'  Set ws = ThisWorkbook.Sheets("{name}")')
            lines.append('  If ws Is Nothing Then')
            lines.append('    Set ws = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))')
            lines.append(f'    ws.Name = "{name}"')
            lines.append('  End If')
            cols = s.get('cols',[]) or []
            for i, c in enumerate(cols):
                c_esc = c.replace('"','')
                lines.append(f'  ws.Cells(1, {i+1}).Value = "{c_esc}"')
        lines.append('  Application.ScreenUpdating = True')
        lines.append('End Sub')
        return '\n'.join(lines)


    # ══════════════════════════════════════════════════════════
    # LLAMADA A IA (en hilo)
    # ══════════════════════════════════════════════════════════
    def _ia_run(self, texto, img_path, file_path=None):
        try:
            logger.info("_ia_run INICIO texto=%r img=%s file=%s", texto[:60], bool(img_path), bool(file_path))
            t_low = texto.lower()
            # ── Clasificar intención (movido antes de razon_pfx) ──────────
            KW_CODE   = ["macro","vba","script","código","codigo","función","python",
                         "powershell","sql","batch","genera","escribe","crea","hace un",
                         "automatiza","formula","fórmula"]
            KW_LEGAL  = ["finiquito","liquidaci","contrato","articulo","artículo","dt",
                         "codigo del trabajo","código del trabajo","afc","afp",
                         "previred","remuneracion","remuneración","horas extra"]
            KW_EXCEL  = ["excel","xlsx","gantt","planilla","tabla","hoja","celda",
                         "columna","fila","formato","pivot","tabla din"]
            KW_BUSCAR = ["busca","precio","noticia","reciente","qué es","que es",
                         "legislacion","ley","decreto","circular","normativa"]
            codear    = any(k in t_low for k in KW_CODE)
            es_legal  = any(k in t_low for k in KW_LEGAL)
            es_excel  = any(k in t_low for k in KW_EXCEL)
            simple    = (len(texto.split()) < 7 and not codear and not es_legal)


            # ── Contexto archivo adjunto ──
            ctx_file = ""
            if file_path:
                try:
                    ctx_file = self._archivo_ctx(file_path)
                except Exception:
                    ctx_file = f"\n[ARCHIVO:{os.path.basename(file_path)}]\n(no se pudo leer)\n[/ARCHIVO]"

            # ── Contexto Excel ──
            ctx_xl = ""
            for ruta in extraer_rutas_excel(texto)[:2]:
                self.after(0, lambda: self._est("📊 Leyendo Excel…", T["a2"]))
                ctx_xl += f"\n[EXCEL:{os.path.basename(ruta)}]\n{leer_excel_contexto(ruta)}\n[/EXCEL]"

            # ── Contexto procesos aprendidos ──
            ctx_pr = ""
            procs  = _rl(F["proc"])
            if procs:
                kw_proc = ["sld","pdv","eos","comision","dotacion","finiquito","sueldo","macro","proceso"]
                for p in reversed(procs):
                    n = p.get("nombre","").lower()
                    if any(k in t_low or k in n for k in kw_proc):
                        ctx_pr = (f"\n[PROCESO_APRENDIDO:{p['nombre']}]\n"
                                  f"Resumen:{p.get('resumen','')}\n"
                                  f"Archivos:{', '.join(p.get('archivos',[]))}\n[/PROCESO]")
                        break

            # ── Contexto web ──
            ctx_web = ""
            if CFG.get("web_auto") and WEB and not ctx_xl:
                kw_web = ["busca","precio","noticia","reciente","softland","que es","qué es",
                          "legislacion","ley","decreto","circular"]
                kw_code = ["macro","vba","python","sql","script","función","powershell","batch"]
                if any(k in t_low for k in kw_web):
                    self.after(0, lambda: self._est("🌐 Buscando…", T["a2"]))
                    ctx_web = f"\n[WEB]\n{buscar_web(texto,4)}\n[/WEB]"
                elif any(k in t_low for k in kw_code):
                    self.after(0, lambda: self._est("🔍 Referencias…", T["a2"]))
                    ctx_web = f"\n[WEB]\n{buscar_web(f'ejemplo {texto[:60]} chile remuneraciones',2)}\n[/WEB]"

            # Hint cuando servidor no está activo y se piden datos en tiempo real
            _srv_hint = ""
            _gt, _gc, _gs = _graph_creds()
            if not all([_gt, _gc, _gs]):
                _srv_kw = ["dotacion","dotación","comision","comisión","liquidacion",
                           "viatico","bono","sueldo del mes","finiquito","matriz"]
                if any(k in t_low for k in _srv_kw):
                    _srv_hint = ("\n\n[NOTA SISTEMA] OneDrive/Excel 365 no configurado. "
                                 "Responde como experto en remuneraciones Chile guiando el proceso "
                                 "paso a paso. Menciona que para acceso directo a datos OneDrive, "
                                 "el usuario debe configurar Graph API en ⚙️ Configuración.")
            extra = ctx_xl + ctx_web + ctx_pr + ctx_file + _srv_hint
            # ── Prefijo de razonamiento según intención ───────────────
            _razon_pfx = ""
            if codear:
                _razon_pfx = ("\n\n[MODO CÓDIGO] Antes de escribir el código, indica en una línea"
                              " el enfoque general (qué estructura usarás, qué librerías)."
                              " Luego entrega el código completo y funcional. No truncar.")
            elif es_legal:
                _razon_pfx = ("\n\n[MODO LEGAL] Estructura tu respuesta así: 1) Hechos del caso"
                              " → 2) Artículo aplicable del Código del Trabajo → 3) Fórmula/cálculo"
                              " → 4) Resultado claro. Cita el artículo exacto.")
            elif es_excel:
                _razon_pfx = ("\n\n[MODO EXCEL] Si generarás código: primero menciona qué hace"
                              " cada sección principal, luego el código completo con comentarios.")
            if _razon_pfx:
                extra += _razon_pfx
            sys_p = sistema_prompt(extra)

            # ── Clasificar intención ─────────────────────────────────

            max_t     = (1200 if simple
                         else 5000 if codear
                         else 3500 if es_legal
                         else 2800)
            # Mensaje de estado descriptivo según intención
            _estado_msgs = {
                "code":  ("⚙️ Generando código…",   "🔨 Revisando código…"),
                "legal": ("⚖️ Revisando normativa…", "📖 Consultando código del trabajo…"),
                "excel": ("📊 Procesando Excel…",    "📊 Analizando planilla…"),
                "buscar":("🌐 Buscando…",            "🌐 Consultando fuentes…"),
                "simple":("● Procesando…",           "● Pensando…"),
            }
            _intent_key = ("code" if codear else "legal" if es_legal
                           else "excel" if es_excel else "simple")
            _msg_estado = _estado_msgs[_intent_key][0]

            # Incluir último excel en contexto si el usuario lo referencia
            _ref_ult = any(k in t_low for k in ["ultimo","último","ese excel","el archivo",
                                                   "lo que hiciste","el que creaste","abre el último"])
            if _ref_ult and self._ultimo_excel and os.path.exists(self._ultimo_excel):
                extra += f"\n[ULTIMO_EXCEL]\n{self._ultimo_excel}\n[/ULTIMO_EXCEL]"
            sys_p = sistema_prompt(extra)

            cadena = cadena_proveedores()
            logger.info("_ia_run cadena=%r busy=%s", cadena, self._busy)
            if not cadena:
                self.after(0, lambda: self._st_end(
                           "Sin IA configurada.\n\nUsa ⚙️ **Configuración** para agregar una API Key.\n\n"
                           "Con **una sola** basta (recomendado: **Claude** o **Gemini**).\n"
                           "— Claude: console.anthropic.com\n"
                           "— Gemini: aistudio.google.com"))
                return

            prov_inicial = cadena[0]
            self.after(0, lambda p=prov_inicial, m=_msg_estado:
                       self._est(f"{PROV_ICON.get(p,'●')} {m}",
                                 PROV_COLOR.get(p,T["a2"])))

            # ── Multi-agente (código) ──
            if CFG.get("multi_agent") and codear:
                cl = _get_claude(); gm = _get_gemini()
                if cl and gm:
                    _temp = 0.15
                    self._ia_multi(texto, sys_p, cl, gm, max_t); return

            # ── Cadena de fallback automática ──
            fallas = []  # [(prov, err_str)]
            for prov_try in cadena:
                try:
                    self.after(0, lambda p=prov_try:
                               self._est(f"{PROV_ICON.get(p,'●')} {PROV_NOMBRES.get(p,'IA')}…",
                                         PROV_COLOR.get(p,T["a2"])))
                    _temp = 0.15 if codear else 0.18 if es_legal else 0.30
                    resp = self._llamar(prov_try, texto, img_path, sys_p, max_t, codear, _temp)
                    if resp:
                        self._conv_push(texto, resp)
                        self.after(0, lambda _r=resp: self._st_end(_r)); return
                except Exception as ex:
                    err = str(ex)
                    fallas.append((prov_try, err))
                    _marcar_error(prov_try)
                    self.after(0, lambda p=prov_try, e=err[:50]:
                               self._est(f"↪ {PROV_NOMBRES.get(p,p)} falló ({e[:30]}…), rotando…",
                                         T["warn"]))
                    time.sleep(0.4)
                    continue

            # Mensaje final más explícito
            det = ""
            if fallas:
                p_last, e_last = fallas[-1]
                nm = PROV_NOMBRES.get(p_last, p_last)
                e_low = (e_last or "").lower()

                if "429" in e_low or "resourceexhausted" in e_low or "quota" in e_low or "rate_limit" in e_low:
                    det = (f"\n\nDetalle: **{nm}** respondió con **cuota/rate-limit (429)**.\n"
                           "Solución: Remi intentará automáticamente con **Groq (gratis)** si tienes la key configurada. "
                           "Si no, espera unos minutos o agrega tu key de Groq en ⚙️ Configuración.")
                elif "not found" in e_low and "models/" in e_low:
                    det = (f"\n\nDetalle: **{nm}** indicó **modelo no disponible**.\n"
                           "Solución: cambia el modelo de Gemini (config) o usa otro proveedor.")
                elif "api key not valid" in e_low or "invalid api key" in e_low or ("key" in e_low and ("invalid" in e_low or "not valid" in e_low or "expired" in e_low)):
                    det = (f"\n\nDetalle: **{nm}** rechazó la key — **key inválida o expirada**.\n"
                           "Solución: genera una nueva key o usa Groq (gratis) como alternativa.")
                elif "sin key" in e_low or "no configurado" in e_low or ("falta" in e_low and "key" in e_low):
                    det = (f"\n\nDetalle: **{nm}** no está configurado (falta key).\n"
                           "Solución: ve a ⚙️ Configuración y pega una key válida.\n"
                           "💡 Tip: agrega la key de **Groq** (gratis en console.groq.com) para tener siempre un fallback.")
                else:
                    det = f"\n\nÚltimo error ({nm}): {e_last[:220]}"

            _msg_err = ("⚠️ No se pudo obtener respuesta de la IA." + det + "\n\n"
                        "Tips rápidos:\n"
                        "  — **Groq es gratis** (console.groq.com) y funciona como fallback automático.\n"
                        "  — Si tienes Groq configurado, se activa solo cuando otra IA falla por cuota.\n"
                        "  — Prueba cambiar a Claude si tienes key, o usar Ollama si lo tienes instalado.\n"
                        "  — Revisa conexión a internet.")
            self.after(0, lambda _m=_msg_err: self._st_end(_m))
        except Exception as e:
            try:
                logger.exception("Error interno en _ia_run: %s", e)
            except Exception:
                pass
            self.after(0, lambda _e=str(e): self._st_end("Error interno: " + _e))
    def _llamar(self, prov, texto, img_path, sys_p, max_t, codear, temp=0.25):
        """Llama al proveedor indicado. Retorna texto o lanza excepción."""
        try:
            logger.debug(f"LLAMAR prov={prov} codear={codear} len_text={len(texto) if texto else 0}")
        except Exception:
            pass

        if prov == "claude":
            if not CLAUDE:
                raise ValueError(
                    "Falta la librería de Claude (anthropic).\n\n"
                    "Instala con:\n\npython -m pip install anthropic"
                )
            cl = _get_claude()
            if not cl:
                raise ValueError("Claude no configurado (falta API key en ⚙️ Configuración)")
            # Construir mensajes con historial
            hist = list(self._conv[-16:])
            content: list = []
            if img_path and PIL_OK:
                with open(img_path,"rb") as fh: b64 = base64.b64encode(fh.read()).decode()
                ext = os.path.splitext(img_path)[1].lower().lstrip(".")
                mt  = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png",
                       "gif":"image/gif","webp":"image/webp"}.get(ext,"image/png")
                content.append({"type":"image","source":{"type":"base64","media_type":mt,"data":b64}})
            content.append({"type":"text","text":texto})
            hist.append({"role":"user","content":content})

            modelo = M_SMART if codear else CFG.get("modelo", M_FAST)

            # Doble-pasada Haiku→Sonnet para código
            if codear and CFG.get("modelo",M_FAST) == M_FAST:
                r1 = cl.messages.create(model=M_FAST, max_tokens=2800,
                                         system=sys_p, messages=hist)
                borrador = r1.content[0].text
                self.after(0, lambda _b=borrador: self._st_upd(_b))
                self.after(0, lambda: self._est("🔍 Refinando con Sonnet…", T["a2"]))
                hist2 = hist + [
                    {"role":"assistant","content":borrador},
                    {"role":"user","content":
                     "Revisa el código anterior: corrige errores, mejora comentarios, asegúrate que funcione en el entorno de Remuneraciones con las rutas X:\\. Responde completo en español."}
                ]
                resp = ""
                with cl.messages.stream(model=M_SMART, max_tokens=4000,
                                         system=sistema_prompt(), messages=hist2) as st:
                    for chunk in st.text_stream:
                        resp += chunk; self.after(0, lambda _r=resp: self._st_upd(_r))
            else:
                resp = ""
                with cl.messages.stream(model=modelo, max_tokens=max_t,
                                         system=sys_p, messages=hist) as st:
                    for chunk in st.text_stream:
                        resp += chunk; self.after(0, lambda _r=resp: self._st_upd(_r))
            return resp

        elif prov == "gemini":
            # Gemini via google.generativeai o adaptador google.genai
            try:
                if not GEMINI:
                    raise ValueError(
                        "Falta la librería de Gemini (google-generativeai).\n\n"
                        "Instala con:\n\npython -m pip install google-generativeai"
                    )
                gm = _get_gemini()
                if not gm:
                    raise ValueError("Gemini no configurado (falta API key en ⚙️ Configuración)")
                # Algunos paquetes antiguos (`google.generativeai`) no exponen
                # el método `start_chat`. Si falta, intentar fallback inmediato a Groq.
                if not hasattr(gm, "start_chat"):
                    try:
                        gq = _get_groq()
                        if gq:
                            try:
                                logger.warning("El módulo de Gemini instalado no tiene 'start_chat' — fallback a Groq")
                            except Exception:
                                pass
                            return self._llamar('groq', texto, img_path, sys_p, max_t, codear)
                    except Exception:
                        pass
                    raise ValueError("El módulo de Gemini instalado no soporta 'start_chat'. Instala 'google.genai' o actualiza la librería.")
                # Construir historial en formato Gemini
                gem_hist = []
                for msg in self._conv[-14:]:
                    role = "user" if msg["role"]=="user" else "model"
                    gem_hist.append({"role":role,"parts":[msg["content"]]})
                # Usar el modelo configurado (default: gemini-2.5-flash)
                _gem_model_name = CFG.get("gemini_model", "gemini-2.5-flash")
                try:
                    _gem_model = gm.GenerativeModel(
                        _gem_model_name,
                        system_instruction=sys_p
                    )
                    chat = _gem_model.start_chat(history=gem_hist)
                except Exception:
                    # Fallback: start_chat directo en el módulo (versión antigua)
                    chat = gm.start_chat(history=gem_hist)
                full_prompt = f"[CONTEXTO DEL SISTEMA]\n{sys_p}\n\n[PREGUNTA]\n{texto}"
                resp = ""
                last_chunk = None
                for chunk in chat.send_message(
                    full_prompt,
                    stream=True,
                    generation_config={"max_output_tokens":max_t, "temperature":temp, "top_p":0.9},
                ):
                    last_chunk = chunk
                    try:
                        piece = _gemini_extraer_texto(chunk)
                    except Exception:
                        piece = None
                    if piece:
                        resp += piece
                        self.after(0, lambda _r=resp: self._st_upd(_r))

                # A veces el streaming entrega el texto completo en el último chunk.
                final_full = _gemini_extraer_texto(last_chunk) if last_chunk is not None else ""
                if final_full and len(final_full) >= len(resp):
                    resp = final_full

                if resp:
                    return resp

                fr = _gemini_finish_reason(last_chunk) if last_chunk is not None else None
                fin = f" (finish_reason={fr})" if fr is not None else ""
                raise ValueError(f"Gemini devolvió respuesta vacía{fin}")
            except Exception as ex:
                # Log error
                try:
                    logger.exception(f"Gemini error: {ex}")
                except Exception:
                    pass
                e_low = (str(ex) or "").lower()
                # Detectar claves inválidas / expiradas / auth
                if prov != 'groq' and ("invalid" in e_low and "key" in e_low or "api key" in e_low or "401" in e_low or "403" in e_low or "rejected" in e_low or "invalid api key" in e_low or "key inválida" in e_low):
                    # Intentar fallback inmediato a Groq si está configurado
                    try:
                        gq = _get_groq()
                        if gq:
                            try:
                                logger.info("Gemini key rejected — reintentando inmediatamente con Groq")
                            except Exception:
                                pass
                            return self._llamar('groq', texto, img_path, sys_p, max_t, codear)
                    except Exception:
                        pass
                # Si no se pudo fallback, propagar
                raise

        elif prov == "openai":
            if not OPENAI:
                raise ValueError(
                    "Falta la librería de OpenAI (`openai`).\n\n"
                    "Instala con:\n\n"
                    "```\npython -m pip install openai\n```"
                )
            oa = _get_openai()
            if not oa:
                raise ValueError("OpenAI no configurado (falta API key en ⚙️ Configuración)")
            msgs = ([{"role":"system","content":sys_p}]
                    + self._conv[-14:]
                    + [{"role":"user","content":texto}])
            resp = ""
            for chunk in oa.chat.completions.create(
                    model="gpt-4o-mini", max_tokens=max_t,
                    stream=True, temperature=temp, messages=msgs):
                delta = chunk.choices[0].delta.content or ""
                resp += delta; self.after(0, lambda _r=resp: self._st_upd(_r))
            return resp or None

        elif prov == "mistral":
            if not MISTRAL:
                raise ValueError(
                    "Falta la librería de Mistral (`mistralai`).\n\n"
                    "Instala con:\n\n"
                    "```\npython -m pip install mistralai\n```"
                )
            ms = _get_mistral()
            if not ms:
                raise ValueError("Mistral no configurado (falta API key en ⚙️ Configuración)")
            msgs = ([{"role":"system","content":sys_p}]
                    + self._conv[-14:]
                    + [{"role":"user","content":texto}])
            resp = ""
            # Intentar con complete (más compatible) y fallback a stream
            _mistral_model = CFG.get("mistral_model", "mistral-small-latest")
            try:
                result = ms.chat.complete(model=_mistral_model,
                                           max_tokens=max_t, temperature=temp, messages=msgs)
                resp = result.choices[0].message.content or ""
                self.after(0, lambda _r=resp: self._st_upd(_r))
            except AttributeError:
                # Versión antigua de mistralai
                for ev in ms.chat_stream(model=_mistral_model,
                                          max_tokens=max_t, temperature=temp, messages=msgs):
                    delta = ev.data.choices[0].delta.content or ""
                    resp += delta; self.after(0, lambda _r=resp: self._st_upd(_r))
            return resp or None

        elif prov == "groq":
            if not GROQ:
                raise ValueError(
                    "Falta la librería de Groq (`groq`).\n\n"
                    "Instala con:\n\n"
                    "```\npython -m pip install groq\n```"
                )
            gq = _get_groq()
            if not gq:
                raise ValueError("Groq no configurado (falta API key en ⚙️ Configuración)")

            msgs = ([{"role":"system","content":sys_p}]
                    + self._conv[-14:]
                    + [{"role":"user","content":texto}])

            # Intentar modelos en orden hasta que uno funcione
            ultimo_err = None
            for modelo_groq in GROQ_MODELOS:
                try:
                    resp = ""
                    stream = gq.chat.completions.create(
                        model=modelo_groq,
                        messages=msgs,
                        max_tokens=min(max_t, 8192),   # Groq límite por modelo
                        temperature=0.25,
                        stream=True,
                    )
                    for chunk in stream:
                        delta = chunk.choices[0].delta.content or ""
                        resp += delta
                        self.after(0, lambda _r=resp: self._st_upd(_r))
                    if resp:
                        return resp
                except Exception as ex:
                    ultimo_err = ex
                    continue   # probar siguiente modelo

            raise ValueError(f"Groq falló en todos los modelos. Último error: {ultimo_err}")

        elif prov == "ollama":
            if not REQ:
                raise ValueError(
                    "Falta la librería `requests` (necesaria para Ollama).\n\n"
                    "Instala con:\n\n"
                    "```\npython -m pip install requests\n```"
                )
            url   = CFG.get("ollama_url", OLLAMA_URL)
            model = CFG.get("ollama_model","qwen2.5-coder:7b")
            msgs  = ([{"role":"system","content":sys_p}]
                     + self._conv[-10:]
                     + [{"role":"user","content":texto}])
            # Stream de Ollama
            r = _rq.post(f"{url}/api/chat",
                         json={"model":model,"messages":msgs,"stream":True},
                         stream=True, timeout=180)
            resp = ""
            for line in r.iter_lines():
                if line:
                    try:
                        d = json.loads(line.decode())
                        c = d.get("message",{}).get("content","")
                        resp += c; self.after(0, lambda _r=resp: self._st_upd(_r))
                        if d.get("done"): break
                    except: pass
            return resp or None

        raise ValueError(f"Proveedor desconocido: {prov}")

    def _ia_multi(self, texto, sys_p, cl, gm, max_t):
        """Multi-agente: Haiku + Gemini en paralelo → Sonnet fusiona."""
        bors = {}
        def _h():
            try:
                r = cl.messages.create(model=M_FAST, max_tokens=2800, system=sys_p,
                                        messages=[{"role":"user","content":texto}])
                bors["h"] = r.content[0].text
            except Exception as e: bors["h"] = f"(error Haiku: {e})"
        def _g():
            try:
                _gm2 = gm.GenerativeModel(CFG.get("gemini_model","gemini-2.5-flash"))
                r = _gm2.generate_content(sys_p+f"\n\nPregunta: {texto}",
                                         generation_config={"max_output_tokens":2800,"temperature":0.2})
                bors["g"] = _gemini_extraer_texto(r) or ""
            except Exception as e: bors["g"] = f"(error Gemini: {e})"

        self.after(0, lambda: self._est("🤖×2 Paralelo Haiku+Gemini…", T["a2"]))
        t1=threading.Thread(target=_h,daemon=True); t2=threading.Thread(target=_g,daemon=True)
        t1.start(); t2.start(); t1.join(timeout=30); t2.join(timeout=30)

        self.after(0, lambda: self._est("✦ Sonnet fusionando…", T["a2"]))
        fusion_prompt = (
            f"Tienes dos propuestas para: «{texto}»\n\n"
            f"── Propuesta A (Haiku):\n{bors.get('h','(sin respuesta)')}\n\n"
            f"── Propuesta B (Gemini):\n{bors.get('g','(sin respuesta)')}\n\n"
            f"Fusiona lo mejor de ambas. Da la respuesta final completa y correcta en español."
        )
        resp = ""
        try:
            with cl.messages.stream(model=M_SMART, max_tokens=4000,
                    system=sistema_prompt(), messages=[{"role":"user","content":fusion_prompt}]) as st:
                for chunk in st.text_stream:
                    resp += chunk; self.after(0, lambda _r=resp: self._st_upd(_r))
        except Exception as e:
            resp = bors.get("h") or bors.get("g") or f"Error fusión: {e}"

        self._conv_push(texto, resp)
        self.after(0, lambda _r=resp: self._st_end(_r))

    def _st_upd(self, txt):
        # Throttle UI: el streaming puede mandar decenas de updates por segundo
        self._stream_pending = txt
        if self._stream_job is None:
            try:
                self._stream_job = self.after(50, self._st_flush)
            except:
                self._stream_job = None

    def _st_flush(self):
        self._stream_job = None
        if self._bact:
            try:
                self._bact.stream_upd(self._stream_pending)
            except:
                pass
            self._scroll()

    def _st_end(self, txt):
        # Asegurar que el último update llegue antes del render final
        self._stream_pending = txt
        if self._stream_job is not None:
            try: self.after_cancel(self._stream_job)
            except: pass
            self._stream_job = None
        if self._bact:
            try: self._bact.stream_fin(txt)
            except: pass
        self._bact = None; self._busy = False
        self._pb_off(); self._est("● Listo", T["ok"]); self._scroll()
        self._update_prov_label()
        # ── Resetear cooldown del proveedor que respondió bien ──────────
        try:
            _marcar_ok(CFG.get('proveedor','claude'))
        except Exception:
            pass
        # ── Auto-detectar URLs de Excel 365 en respuesta streaming ──────
        try:
            urls_365 = self._detectar_excel365_urls(txt)
            if urls_365:
                self.after(200, lambda u=urls_365: self._bub_excel365_links(u))
        except Exception:
            pass

        # ── Detectar bloque Python con Excel → pedir permiso antes ───────
        if "```python" in txt.lower() and any(k in txt for k in
                ["to_excel","wb.save","openpyxl","DataFrame","xlsxwriter"]):
            self.after(150, lambda: self._ofrecer_ejecutar(txt))

    def _ofrecer_ejecutar(self, txt):
        """Valida el código y muestra botón de confirmación antes de ejecutar."""
        bloques = re.findall(r'```python\s*\n(.*?)```', txt, re.DOTALL | re.I)
        if not bloques: return
        codigo = bloques[0]
        if not any(k in codigo for k in ["to_excel","wb.save","workbook.save","openpyxl","DataFrame"]):
            return
        errores = _prevalidar_codigo(codigo, sys.executable)
        codigo_fix = _parchear_codigo(codigo)
        if errores:
            msg = (f"⚠️ **Revisé el código y encontré {len(errores)} problema(s) antes de ejecutar:**\n\n"
                   + "\n".join(f"• {e}" for e in errores)
                   + "\n\n¿Quieres que ejecute la versión corregida?")
            self._bub_sys(msg, "warn", codigo_pendiente=codigo_fix, ofrecer_ejecutar=True)
        else:
            self._bub_sys(
                "✅ Revisé el código, todo está correcto. ¿Ejecuto ahora?",
                "ok", codigo_pendiente=codigo_fix, ofrecer_ejecutar=True)

    def _auto_ejecutar_xl(self, codigo):
        """Ejecuta código Python ya validado y parcheado."""
        py = sys.executable
        def _run():
            try:
                def _mod_ok(mod):
                    try:
                        return subprocess.run([py,"-c",f"import {mod}"],
                                              capture_output=True,timeout=15).returncode==0
                    except Exception: return False
                for _pkg in ["openpyxl","pandas"]:
                    if not _mod_ok(_pkg):
                        self.after(0, self._bub_sys,
                                   f"📦 Instalando {_pkg}… (puede tardar 1-2 min)", "system")
                        subprocess.run([py,"-m","pip","install",_pkg,"-q",
                                        "--disable-pip-version-check"],
                                       capture_output=True, timeout=180,
                                       creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0))
                r = subprocess.run([py,"-c",codigo],
                    capture_output=True, text=True, timeout=60,
                    encoding="utf-8", errors="replace",
                    creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0))
                out = r.stdout.strip()
                noise = ["WARNING","DeprecationWarning","FutureWarning","pip","notice","UserWarning"]
                err = "\n".join(l for l in r.stderr.splitlines()
                               if l.strip() and not any(x in l for x in noise))
                if r.returncode == 0:
                    msg = "✅ **Excel creado correctamente.**"
                    if out: msg += f"\n\n{out}"
                    rutas = re.findall(r'[A-Za-z]:[/\\][^\n"\'<>|*?\s]+\.xlsx?', codigo, re.I)
                    for rp in rutas:
                        rp = rp.strip("\"'(),;")
                        if os.path.exists(rp):
                            try:
                                os.startfile(rp)
                                self.after(0, setattr, self, "_ultimo_excel", rp)
                            except: pass
                            break
                else:
                    msg = f"❌ **Error al ejecutar:**\n```\n{err or r.stderr[-400:]}\n```"
                self.after(0, self._bub_sys, msg, "ok" if "✅" in msg else "error")
            except Exception as ex:
                self.after(0, self._bub_sys, f"❌ Error: {ex}", "error")
        threading.Thread(target=_run, daemon=True).start()

    # ══════════════════════════════════════════════════════════
    # IMAGEN / VOZ
    # ══════════════════════════════════════════════════════════
    def _adj_img(self):
        if not PIL_OK:
            self._bub_sys(
                "Para adjuntar imágenes falta la librería **Pillow**.\n\n"
                "Instala con:\n\n"
                "```\npython -m pip install pillow\n```",
                "system",
            )
            return
        paths = filedialog.askopenfilenames(
            filetypes=[("Imágenes","*.png *.jpg *.jpeg *.bmp *.gif *.webp"), ("Todos","*.*")])
        if paths:
            # añadir múltiples rutas
            for p in paths:
                if p and os.path.exists(p):
                    self._imgs.append(p)
            # mostrar resumen
            if len(self._imgs) == 1:
                txt = f"  📷  {os.path.basename(self._imgs[0])}   [doble clic = quitar]"
            else:
                txt = f"  📷  {len(self._imgs)} imágenes seleccionadas   [doble clic = quitar]"
            self._lbl_img.configure(text=txt)
            self._lbl_img.pack(pady=4)
            self._lbl_img.bind("<Double-Button-1>", lambda _: self._quit_img())

    def _adj_file(self):
        r = filedialog.askopenfilename(filetypes=[("Todos","*.*")])
        if r:
            self._file = r
            self._lbl_file.configure(text=f"  📎  {os.path.basename(r)}   [doble clic = quitar]")
            self._lbl_file.pack(pady=(0, 4))
            self._lbl_file.bind("<Double-Button-1>", lambda _: self._quit_file())

    def _paste_image(self):
        # Intentar tomar imagen del portapapeles y guardarla en archivo temporal
        imgs = []
        Image = None
        ImageGrab = None
        try:
            from PIL import Image, ImageGrab
        except Exception:
            try:
                from PIL import Image
            except Exception:
                Image = None

        # Primer intento: ImageGrab (Pillow)
        clip = None
        if ImageGrab is not None:
            try:
                clip = ImageGrab.grabclipboard()
            except Exception:
                clip = None

        # Si ImageGrab no devolvió nada, intentar fallback con pywin32 (CF_DIB)
        if clip is None:
            try:
                import win32clipboard, win32con
                import struct
                win32clipboard.OpenClipboard()
                try:
                    if win32clipboard.IsClipboardFormatAvailable(win32con.CF_DIB):
                        dib = win32clipboard.GetClipboardData(win32con.CF_DIB)
                        # convertir DIB a BMP agregando BITMAPFILEHEADER
                        hdr_size = struct.unpack_from('<I', dib, 0)[0]
                        bfOffBits = 14 + hdr_size
                        file_header = struct.pack('<2sIHHI', b'BM', 14 + len(dib), 0, 0, bfOffBits)
                        bmp = file_header + dib
                        if Image is None:
                            from PIL import Image
                        im = Image.open(io.BytesIO(bmp))
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                        im.save(tmp, format='PNG')
                        tmp.close(); imgs.append(tmp.name)
                    else:
                        # intentar formato lista de rutas (arrastrado)
                        if win32clipboard.IsClipboardFormatAvailable(win32con.CF_HDROP):
                            files = win32clipboard.GetClipboardData(win32con.CF_HDROP)
                            for p in files:
                                if os.path.exists(p): imgs.append(p)
                finally:
                    win32clipboard.CloseClipboard()
            except Exception:
                # no hay fallback disponible
                pass

        # Si ImageGrab devolvió un resultado válido (PIL.Image o lista), manejarlo
        if clip is not None:
            if isinstance(clip, list):
                for p in clip:
                    if os.path.exists(p): imgs.append(p)
            else:
                try:
                    if hasattr(clip, 'save'):
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                        clip.save(tmp, format='PNG')
                        tmp.close(); imgs.append(tmp.name)
                    else:
                        bio = io.BytesIO(clip)
                        if Image is None:
                            from PIL import Image
                        im = Image.open(bio)
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                        im.save(tmp, format='PNG')
                        tmp.close(); imgs.append(tmp.name)
                except Exception:
                    messagebox.showinfo('Portapapeles', 'No se pudo convertir el contenido del portapapeles a imagen.')
                    return

        if not imgs:
            messagebox.showinfo('Portapapeles', 'No se detectó imagen en el portapapeles o no hay soporte en este sistema.')
            return

        for p in imgs:
            self._imgs.append(p)

        if len(self._imgs) == 1:
            txt = f"  📷  {os.path.basename(self._imgs[0])}   [doble clic = quitar]"
        else:
            txt = f"  📷  {len(self._imgs)} imágenes seleccionadas   [doble clic = quitar]"
        self._lbl_img.configure(text=txt)
        self._lbl_img.pack(pady=4)
        self._lbl_img.bind("<Double-Button-1>", lambda _: self._quit_img())

    def _quit_img(self):
        self._img = None
        try:
            # borrar temp files en self._imgs que apunten a temp
            for p in list(self._imgs):
                try:
                    # solo borrar si está en temp
                    if p and os.path.exists(p) and str(tempfile.gettempdir()) in p:
                        os.remove(p)
                except Exception:
                    pass
            self._imgs = []
        except Exception:
            self._imgs = []
        try:
            self._lbl_img.pack_forget()
        except Exception:
            pass

    def _quit_file(self):
        self._file = None
        try:
            self._lbl_file.pack_forget()
        except Exception:
            pass

    def _archivo_ctx(self, path: str) -> str:
        """Devuelve contexto de archivo para la IA (Excel/texto)."""
        p = (path or "").strip()
        if not p:
            return ""
        ext = os.path.splitext(p)[1].lower()
        name = os.path.basename(p)

        if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            if not XL:
                return (
                    f"\n[ARCHIVO_EXCEL:{name}]\n"
                    "Falta la librería **openpyxl** para leer Excel.\n\n"
                    "Instala con:\n\n"
                    "```\npython -m pip install openpyxl\n```\n"
                    "[/ARCHIVO_EXCEL]"
                )
            self.after(0, lambda: self._est("📊 Leyendo archivo Excel…", T["a2"]))
            return f"\n[ARCHIVO_EXCEL:{name}]\n{leer_excel_contexto(p)}\n[/ARCHIVO_EXCEL]"

        # Texto: intentar leer como UTF-8 (con fallback)
        if ext in {".txt", ".csv", ".json", ".md", ".log", ".sql"}:
            self.after(0, lambda: self._est("📎 Leyendo archivo…", T["a2"]))
            try:
                data = Path(p).read_text(encoding="utf-8", errors="replace")
            except Exception:
                try:
                    data = Path(p).read_text(encoding="latin-1", errors="replace")
                except Exception:
                    data = ""
            data = (data or "")[:8000]
            return f"\n[ARCHIVO:{name}]\n{data}\n[/ARCHIVO]"

        # Otros binarios: solo metadata
        try:
            sz = os.path.getsize(p)
        except Exception:
            sz = None
        sz_str = f"{sz} bytes" if isinstance(sz, int) else "tamaño desconocido"
        return f"\n[ARCHIVO:{name}]\n(archivo adjunto: {sz_str})\n[/ARCHIVO]"

    def _voz(self):
        if not VOZ: self._bub_sys("SpeechRecognition y pyaudio no instalados.","error"); return
        self._est("🎤 Escuchando…", T["a2"]); self._pb_on()
        threading.Thread(target=self._voz_h, daemon=True).start()

    def _voz_h(self):
        txt, err = escuchar()
        self.after(0, self._pb_off)
        if txt:
            def _set():
                self._limpiar_ph()
                self._inp.delete("1.0","end")
                self._inp.insert("1.0", txt)
                self._inp.configure(fg=T["tx"])
                self._ph_active = False
            self.after(0, _set)
            self.after(300, self._enviar)
        else:
            self.after(0, lambda: self._bub_sys(err or "No escuché nada.", "error"))
        self.after(0, lambda: self._est("● Listo", T["ok"]))

    def _buscar_manual(self):
        d = ctk.CTkInputDialog(text="¿Qué buscamos?", title="Búsqueda Web")
        q = d.get_input()
        if not q or not q.strip(): return
        b = self._bub_sys(f"Buscando: **{q}**…", "system")
        self._pb_on()
        def _run():
            resultado = buscar_web(q)
            # Actualizar la burbuja existente con resultado
            self.after(0, lambda: (self._pb_off(), self._bub_sys(resultado, "system")))
        threading.Thread(target=_run, daemon=True).start()

    # ─── Historial de sesiones de conversación ────────────────────
    def _conv_guardar_sesion(self):
        """Guarda la conversación actual como sesión persistente."""
        if not self._conv or len(self._conv) < 2:
            return
        try:
            sesiones = _rl(F.get("conv_sessions", "")) or []
            if not isinstance(sesiones, list):
                sesiones = []
            # Generar título automático del primer mensaje del usuario
            primer_user = next((m["content"][:60] for m in self._conv
                                if m.get("role") == "user"), "Sin título")
            sesion = {
                "id":       datetime.datetime.now().strftime("%Y%m%d_%H%M%S"),
                "titulo":   primer_user,
                "fecha":    datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                "mensajes": self._conv[:40],   # máx 40 mensajes
                "n_msgs":   len(self._conv),
            }
            # Mantener máx 50 sesiones
            sesiones.append(sesion)
            sesiones = sesiones[-50:]
            _w(F.get("conv_sessions", str(BASE / "conversaciones.json")), sesiones)
        except Exception:
            pass

    def _conv_cargar_sesion(self, sesion_id: str):
        """Carga y muestra una sesión guardada."""
        try:
            sesiones = _rl(F.get("conv_sessions","")) or []
            sesion = next((s for s in sesiones if s.get("id") == sesion_id), None)
            if not sesion:
                return
            # Limpiar chat actual
            for w in list(self._sc.winfo_children()):
                try: w.destroy()
                except: pass
            self._conv = sesion.get("mensajes", [])
            self._bub_sys(
                "📂 Cargando conversacion del " + sesion.get("fecha","") + "\n"
                + "_" + sesion.get("titulo","")[:80] + "_", "ok")

            # Re-renderizar mensajes
            for msg in self._conv:
                if msg.get("role") == "user":
                    self._bub_user(msg.get("content",""))
                elif msg.get("role") == "assistant":
                    self._bub_bot(msg.get("content",""))
        except Exception as e:
            self._bub_sys(f"Error cargando sesión: {e}", "error")

    def _win_historial_chats(self):
        """Panel flotante con historial de conversaciones anteriores."""
        sesiones = _rl(F.get("conv_sessions", str(BASE / "conversaciones.json"))) or []
        if not isinstance(sesiones, list):
            sesiones = []

        def build(panel, cerrar):
            sc = ctk.CTkScrollableFrame(panel, fg_color=T["bg"],
                                         scrollbar_button_color=T["a1"],
                                         scrollbar_button_hover_color=T["a2"],
                                         corner_radius=0)
            sc.pack(fill="both", expand=True, padx=0, pady=0)

            if not sesiones:
                tk.Label(sc, text="Sin conversaciones guardadas. Se guardan al iniciar nueva conversacion.", font=(FONT, 10), bg=T["bg"], fg=T["t2"], justify="center").pack(pady=40)
                return

            # Mostrar sesiones en orden inverso (más reciente primero)
            for ses in reversed(sesiones[-50:]):
                sid    = ses.get("id","")
                titulo = ses.get("titulo","Sin título")[:65]
                fecha  = ses.get("fecha","")
                n      = ses.get("n_msgs", len(ses.get("mensajes",[])))

                card = tk.Frame(sc, bg=T["card"],
                                highlightthickness=1,
                                highlightbackground=T["br"],
                                cursor="hand2")
                card.pack(fill="x", padx=12, pady=4)

                # Barra de color lateral
                tk.Frame(card, bg=T["a1"], width=3).pack(side="left", fill="y")

                info = tk.Frame(card, bg=T["card"]); info.pack(side="left",
                                fill="both", expand=True, padx=12, pady=8)

                tk.Label(info, text=titulo, font=(FONT, 10, "bold"),
                         bg=T["card"], fg=T["tx"], anchor="w",
                         wraplength=360).pack(anchor="w")

                meta_f = tk.Frame(info, bg=T["card"]); meta_f.pack(anchor="w", pady=(2,0))
                tk.Label(meta_f, text=f"📅 {fecha}",
                         font=(FONT, 8), bg=T["card"], fg=T["t2"]).pack(side="left")
                tk.Label(meta_f, text=f"  ·  {n} mensajes",
                         font=(FONT, 8), bg=T["card"], fg=T["t2"]).pack(side="left")

                btn_f = tk.Frame(card, bg=T["card"]); btn_f.pack(side="right",
                                                                   padx=10, pady=8)
                def _cargar(sid_=sid, c=cerrar):
                    c()
                    self._conv_guardar_sesion()   # guardar actual antes
                    self._conv_cargar_sesion(sid_)

                def _eliminar(sid_=sid):
                    try:
                        ss = _rl(F.get("conv_sessions","")) or []
                        ss = [s for s in ss if s.get("id") != sid_]
                        _w(F.get("conv_sessions", str(BASE / "conversaciones.json")), ss)
                        cerrar()
                        self._win_historial_chats()
                    except Exception:
                        pass

                ctk.CTkButton(btn_f, text="Abrir", width=70, height=28,
                    fg_color=T["a1"], hover_color=T["a2"],
                    font=(FONT,9,"bold"), text_color="white",
                    command=_cargar).pack(pady=2)
                ctk.CTkButton(btn_f, text="🗑", width=32, height=28,
                    fg_color=T["card"], hover_color=T["err"],
                    font=(FONT,9), text_color=T["t2"],
                    command=_eliminar).pack(pady=2)

                # Hover en card
                for w in (card, info):
                    w.bind("<Enter>", lambda e, c=card: c.configure(
                        highlightbackground=T["a1"]))
                    w.bind("<Leave>", lambda e, c=card: c.configure(
                        highlightbackground=T["br"]))
                    w.bind("<Button-1>", lambda e, sid_=sid, c=cerrar: _cargar(sid_, c))

        self._modal("💬  Historial de Conversaciones", build, w=560, h=560)

    def _nueva_conv(self):
        # Guardar la conversación actual antes de limpiar
        self._conv_guardar_sesion()
        for w in list(self._sc.winfo_children()):
            try: w.destroy()
            except: pass
        try:
            self._inp.delete("1.0","end")
            self._inp_hint.place(x=0, y=0, relwidth=1, height=36)
        except: pass
        self._conv        = []
        self._ultimo_excel = None
        self._img          = None
        self._file         = None
        if hasattr(self, "_imgs"): self._imgs = []
        # Limpiar previews visuales
        try: self._quit_img()
        except: pass
        try: self._quit_file()
        except: pass
        self._bub_bot("Conversación reiniciada. ¿En qué te ayudo?")

    # ══════════════════════════════════════════════════════════
    # CAMBIO DE IA
    # ══════════════════════════════════════════════════════════
    def _cambiar_prov(self, prov):
        CFG["proveedor"] = prov
        guardar_cfg(CFG)
        _ERR_PROV.clear(); invalidar_cache_prov()
        self._update_prov_label()
        if not _prov_configurado(prov):
            self._bub_sys(f"⚠️ **{PROV_NOMBRES.get(prov,prov)}** no tiene API Key.\n"
                           "Agrega la clave en ⚙️ Configuración.", "error")
        else:
            self._bub_sys(f"✅ Ahora usando **{PROV_NOMBRES.get(prov,prov)}**", "ok")

    # ══════════════════════════════════════════════════════════
    # ALARMA
    # ══════════════════════════════════════════════════════════
    def _on_alarma(self, msg):
        notificar("⏰ ALARMA — Remi", msg)
        self.after(0, lambda: self._alerta_win(msg))

    def _alerta_win(self, msg):
        v = tk.Toplevel(self); v.title("⏰ ALARMA")
        v.configure(bg="#1a0a0a"); v.attributes("-topmost",True); v.resizable(False,False)
        sw, sh = v.winfo_screenwidth(), v.winfo_screenheight()
        v.geometry(f"440x240+{(sw-440)//2}+{(sh-240)//2}")
        tk.Frame(v, bg=T["err"], height=4).pack(fill="x")
        tk.Label(v, text="⏰  ALARMA", font=(FONT,22,"bold"),
                 bg="#1a0a0a", fg=T["warn"]).pack(pady=(18,8))
        tk.Label(v, text=msg, font=(FONT,13), bg="#1a0a0a", fg=T["tx"],
                 wraplength=400, justify="center").pack(padx=20)
        tk.Button(v, text="  ✔  Entendido  ", font=(FONT,12,"bold"),
                  bg=T["ok"], fg="white", relief="flat", padx=20, pady=8,
                  cursor="hand2", command=v.destroy).pack(pady=20)
        self._blink(v,0)

    def _blink(self, v, n):
        try:
            v.configure(bg="#2a0808" if n%2==0 else "#1a0a0a")
            if n < 8: v.after(350, lambda: self._blink(v, n+1))
        except: pass

    # ══════════════════════════════════════════════════════════
    # VENTANAS MODALES
    # ══════════════════════════════════════════════════════════
    def _modal(self, titulo, fn, w=680, h=560):
        self.update_idletasks()
        rx,ry = self.winfo_rootx(), self.winfo_rooty()
        rw,rh = self.winfo_width(), self.winfo_height()
        px = rx + max(10,(rw-w)//2); py = ry + max(10,(rh-h)//2)
        v = ctk.CTkToplevel(self)
        v.title(titulo); v.geometry(f"{w}x{h}+{px}+{py}")
        v.configure(fg_color=T["bg"])

        # Mantener el modal al frente (Windows a veces lo manda atrás)
        try: v.transient(self)
        except: pass
        def _front_and_grab():
            try:
                if not v.winfo_exists():
                    return
                try: v.deiconify()
                except: pass
                v.lift()
                v.wm_attributes("-topmost", 1)
                v.after(250, lambda: (v.wm_attributes("-topmost", 0) if v.winfo_exists() else None))
                try: v.grab_set()
                except: pass
            except:
                pass

        # Ejecutar al próximo ciclo de UI para asegurar que la ventana ya existe/está mapeada
        try:
            v.after(0, _front_and_grab)
        except:
            pass
        # No forzar focus en FocusIn: puede bloquear escritura en CTkEntry

        # Header del modal
        hdr = ctk.CTkFrame(v, fg_color=T["a1"], corner_radius=0, height=48)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text=f"  {titulo}", font=(FONT,12,"bold"),
                     text_color="white").pack(side="left", padx=12, pady=12)
        def cerrar():
            try: v.grab_release()
            except: pass
            try: v.destroy()
            except: pass
            try:
                # devolver foco al input principal
                if hasattr(self, "_inp") and self._inp.winfo_exists():
                    self._inp.focus_set()
            except:
                pass
        ctk.CTkButton(hdr, text="✕", width=32, height=32, corner_radius=6,
                      fg_color="transparent", hover_color=T["err"],
                      font=(FONT,12,"bold"), text_color="white",
                      command=cerrar).pack(side="right", padx=8, pady=8)
        v.protocol("WM_DELETE_WINDOW", cerrar)

        panel = ctk.CTkFrame(v, fg_color=T["bg"], corner_radius=0)
        panel.pack(fill="both", expand=True)
        fn(panel, cerrar)
        return v

    def _sc_modal(self, panel):
        sc = ctk.CTkScrollableFrame(panel, fg_color=T["bg"],
                                     scrollbar_button_color=T["a1"],
                                     scrollbar_button_hover_color=T["a2"],
                                     corner_radius=0)
        sc.pack(fill="both", expand=True); return sc

    def _sec(self, parent, ico, txt, color=None):
        f = ctk.CTkFrame(parent, fg_color=T["card"], corner_radius=6,
                         border_color=T["br"], border_width=1)
        f.pack(fill="x", padx=14, pady=(8,2))
        ctk.CTkLabel(f, text=f"  {ico}  {txt}", font=(FONT,11,"bold"),
                     text_color=color or T["a2"]).pack(anchor="w", padx=12, pady=(10,8))
        return f

    # ── CONFIGURACIÓN ────────────────────────────────────────
    def _win_config(self):
        def build(panel, cerrar):
            sc = self._sc_modal(panel)

            # Proveedor
            s1 = self._sec(sc,"🤖","Proveedor de IA")
            pv = ctk.StringVar(value=CFG.get("proveedor","claude"))
            pf = ctk.CTkFrame(s1,fg_color="transparent"); pf.pack(fill="x",padx=12,pady=(0,4))
            for nm,val,col in [("Claude","claude",PROV_COLOR["claude"]),
                                ("Gemini","gemini",PROV_COLOR["gemini"]),
                                ("OpenAI","openai",PROV_COLOR["openai"]),
                                ("Mistral","mistral",PROV_COLOR["mistral"]),
                                ("Groq ⚡","groq",PROV_COLOR["groq"]),
                                ("Ollama","ollama",PROV_COLOR["ollama"])]:
                ctk.CTkRadioButton(pf,text=nm,variable=pv,value=val,font=(FONT,10),
                    text_color=col,fg_color=col,border_color=col).pack(side="left",padx=6)

            mav = ctk.BooleanVar(value=CFG.get("multi_agent",False))
            ctk.CTkCheckBox(s1,text="Multi-agente (Haiku + Gemini → Sonnet para código)",
                variable=mav,font=(FONT,10),text_color=T["tx"],
                fg_color=T["a1"]).pack(anchor="w",padx=12,pady=(0,8))

            # Nombre usuario
            s0 = self._sec(sc,"👤","Usuario")
            ef = ctk.CTkFrame(s0,fg_color="transparent"); ef.pack(fill="x",padx=12,pady=(0,8))
            ctk.CTkLabel(ef,text="Tu nombre:",font=(FONT,10),text_color=T["t2"]).pack(side="left")
            en = ctk.CTkEntry(ef,width=200,height=30,font=(FONT,11),
                               fg_color=T["inp"],border_color=T["br"],text_color=T["tx"])
            en.pack(side="left",padx=8); en.insert(0,CFG.get("nombre_usuario","Yerko"))

            def lnk(parent,txt,url):
                ctk.CTkButton(parent,text=f"  {txt}",height=22,fg_color="transparent",
                    hover_color=T["card"],border_width=1,border_color=T["br"],
                    font=(FONT,8),text_color=T["t2"],
                    command=lambda:webbrowser.open(url)).pack(anchor="w",padx=12,pady=(0,6))

            # ── Daemon de Consolidación ──────────────────────────────────
            s_dm = self._sec(sc,"⚙️","Daemon de Consolidación  —  Excel automático")
            dmv = ctk.BooleanVar(value=CFG.get("daemon_activo", False))
            ctk.CTkCheckBox(s_dm, text="Activar al iniciar Remi",
                variable=dmv, font=(FONT,10), text_color=T["tx"],
                fg_color=T["a1"]).pack(anchor="w", padx=12, pady=(0,4))

            rf_dm = ctk.CTkFrame(s_dm, fg_color="transparent"); rf_dm.pack(fill="x", padx=12, pady=(0,4))
            ctk.CTkLabel(rf_dm, text="Intervalo (min):", font=(FONT,10), text_color=T["t2"]).pack(side="left")
            e_dm_ivl = ctk.CTkEntry(rf_dm, width=60, height=28, font=(FONT,10),
                                     fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_dm_ivl.pack(side="left", padx=8)
            e_dm_ivl.insert(0, str(CFG.get("daemon_intervalo", 30)))

            lbl_dm_ruta = ctk.CTkLabel(s_dm, font=(FONT,8), text_color=T["t2"],
                text="Carpeta consolidación: ~/OneDrive/Remi_Automatizacion/Entrada/")
            lbl_dm_ruta.pack(anchor="w", padx=12, pady=(0,2))

            # ── Ruta OneDrive (Graph API) ────────────────────────
            ctk.CTkLabel(s_dm, text="☁️  Ruta OneDrive (Graph API) — principal",
                font=(FONT,9,"bold"), text_color=T["a1"]).pack(anchor="w", padx=12, pady=(4,1))
            ctk.CTkLabel(s_dm,
                text="Carpeta en OneDrive donde están los archivos de analistas. Ej: /Finiquitos/2024",
                font=(FONT,8), text_color=T["t2"]).pack(anchor="w", padx=12, pady=(0,2))
            rf_fin_od = ctk.CTkFrame(s_dm, fg_color="transparent"); rf_fin_od.pack(fill="x", padx=12, pady=(0,4))
            ctk.CTkLabel(rf_fin_od, text="Carpeta OD:", font=(FONT,10), text_color=T["t2"], width=110).pack(side="left")
            e_fin_ruta_od = ctk.CTkEntry(rf_fin_od, height=28, font=(FONT,9),
                placeholder_text="/Finiquitos/2024  o  /RRHH/Finiquitos",
                fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_fin_ruta_od.pack(side="left", fill="x", expand=True, padx=4)
            e_fin_ruta_od.insert(0, CFG.get("finiquitos_ruta_od",""))

            rf_fin_hoja = ctk.CTkFrame(s_dm, fg_color="transparent"); rf_fin_hoja.pack(fill="x", padx=12, pady=(0,4))
            ctk.CTkLabel(rf_fin_hoja, text="Hoja Excel:", font=(FONT,10), text_color=T["t2"], width=110).pack(side="left")
            e_fin_hoja = ctk.CTkEntry(rf_fin_hoja, height=28, font=(FONT,9),
                placeholder_text="Finiquitos  (dejar vacío = primera hoja)",
                fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_fin_hoja.pack(side="left", fill="x", expand=True, padx=4)
            e_fin_hoja.insert(0, CFG.get("finiquitos_hoja","Finiquitos"))

            # ── Ruta local (fallback) ────────────────────────────
            ctk.CTkLabel(s_dm, text="💾  Ruta local (fallback sin Graph API)",
                font=(FONT,9), text_color=T["t2"]).pack(anchor="w", padx=12, pady=(4,1))
            rf_fin = ctk.CTkFrame(s_dm, fg_color="transparent"); rf_fin.pack(fill="x", padx=12, pady=(0,2))
            ctk.CTkLabel(rf_fin, text="Ruta local:", font=(FONT,10), text_color=T["t2"], width=110).pack(side="left")
            e_fin_ruta = ctk.CTkEntry(rf_fin, height=28, font=(FONT,9),
                placeholder_text=r"C:\Users\...\OneDrive\Finiquitos_2024",
                fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_fin_ruta.pack(side="left", fill="x", expand=True, padx=4)
            e_fin_ruta.insert(0, CFG.get("finiquitos_ruta",""))

            rf_fin2 = ctk.CTkFrame(s_dm, fg_color="transparent"); rf_fin2.pack(fill="x", padx=12, pady=(0,4))
            ctk.CTkLabel(rf_fin2, text="Archivo Madre:", font=(FONT,10), text_color=T["t2"], width=110).pack(side="left")
            e_fin_madre = ctk.CTkEntry(rf_fin2, height=28, font=(FONT,9),
                fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_fin_madre.pack(side="left", fill="x", expand=True, padx=4)
            e_fin_madre.insert(0, CFG.get("finiquitos_madre","MATRIZ_GENERAL_FINIQUITOS.xlsx"))

            ctk.CTkButton(s_dm, text="🔄  Sincronizar ahora", width=160, height=28,
                fg_color=T["a1"], hover_color=T["a2"], font=(FONT,9,"bold"), text_color="white",
                command=lambda: self._daemon.sincronizar_finiquitos()).pack(anchor="w", padx=12, pady=(0,8))
            # Botón cargar carpeta (v10)
            def _pick_carpeta():
                import threading
                def _do():
                    from tkinter import filedialog
                    path = filedialog.askdirectory(title="Carpeta de analistas (OneDrive)")
                    if not path: return
                    CFG["finiquitos_ruta"] = path
                    e_fin_ruta.delete(0, "end")
                    e_fin_ruta.insert(0, path)
                    guardar_cfg(CFG)
                    self._daemon.sincronizar_finiquitos()
                threading.Thread(target=_do, daemon=True).start()
            ctk.CTkButton(s_dm, text="📂  Cargar carpeta analistas", width=200, height=28,
                fg_color=T["card"], hover_color=T["a1"], border_width=1, border_color=T["a1"],
                font=(FONT,9,"bold"), text_color=T["a1"],
                command=_pick_carpeta).pack(anchor="w", padx=12, pady=(0,4))
            ctk.CTkButton(s_dm, text="📋  Abrir Matriz", width=160, height=28,
                fg_color=T["card"], hover_color=T["a1"], border_width=1, border_color=T["br"],
                font=(FONT,9), text_color=T["t2"],
                command=self._abrir_matriz_finiquitos).pack(anchor="w", padx=12, pady=(0,8))

            lbl_dm_st = ctk.CTkLabel(s_dm, font=(FONT,9), text_color=T["t2"], text="")
            lbl_dm_st.pack(anchor="w", padx=12, pady=(0,4))

            def _dm_toggle():
                if hasattr(self, "_daemon"):
                    if self._daemon.activo:
                        self._daemon.detener()
                        lbl_dm_st.configure(text="⚫ Detenido", text_color=T["t2"])
                    else:
                        self._daemon.iniciar()
                        lbl_dm_st.configure(text="🟢 Activo", text_color=T["ok"])

            _dm_est = "🟢 Activo" if (hasattr(self,"_daemon") and self._daemon.activo) else "⚫ Inactivo"
            lbl_dm_st.configure(text=_dm_est)
            ctk.CTkButton(s_dm, text="Iniciar / Detener ahora", width=160, height=28,
                fg_color=T["card"], hover_color=T["a1"], border_width=1, border_color=T["br"],
                font=(FONT,9), text_color=T["t2"],
                command=_dm_toggle).pack(anchor="w", padx=12, pady=(0,8))

            # ── Remi Server ───────────────────────────────────────────
            # Servidor local (solo para capturas / SMTP — OPCIONAL)
            s_srv = self._sec(sc,"🖥️","Servidor local  —  solo SMTP · Screenshots (opcional)")
            srv_actv = ctk.BooleanVar(value=CFG.get("server_activo",False))
            ctk.CTkCheckBox(s_srv, text="Activar servidor local",
                variable=srv_actv, font=(FONT,10), text_color=T["tx"],
                fg_color=T["a1"]).pack(anchor="w", padx=12, pady=(0,4))
            rf_srv = ctk.CTkFrame(s_srv, fg_color="transparent"); rf_srv.pack(fill="x", padx=12, pady=(0,2))
            ctk.CTkLabel(rf_srv, text="URL:", font=(FONT,10), text_color=T["t2"]).pack(side="left")
            e_srv_url = ctk.CTkEntry(rf_srv, width=220, height=26, font=(FONT,10),
                fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_srv_url.pack(side="left", padx=4)
            e_srv_url.insert(0, CFG.get("server_url","http://127.0.0.1:8765"))
            ctk.CTkLabel(rf_srv, text="Token:", font=(FONT,10), text_color=T["t2"]).pack(side="left", padx=(6,0))
            e_srv_tok = ctk.CTkEntry(rf_srv, width=140, height=26, font=(FONT,10), show="*",
                fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            e_srv_tok.pack(side="left", padx=4)
            e_srv_tok.insert(0, CFG.get("server_token","remi-piloto-2026"))
            lbl_ping = ctk.CTkLabel(s_srv, text="", font=(FONT,8), text_color=T["t2"])
            lbl_ping.pack(anchor="w", padx=12, pady=(0,4))
            def _test_srv():
                lbl_ping.configure(text="🔄 Comprobando…", text_color=T["t2"])
                import threading, urllib.request, json as _j2
                def _chk():
                    url = e_srv_url.get().rstrip("/") + "/ping"
                    try:
                        req = urllib.request.Request(url, headers={"X-Token": e_srv_tok.get()})
                        d   = _j2.loads(urllib.request.urlopen(req, timeout=4).read())
                        msg = ("✅ Conectado · SMTP=" + ("✓" if d.get("smtp") else "✗")
                               + " · Screenshots=" + ("✓" if d.get("capturas") else "✗"))
                        col = T["ok"]
                    except Exception as e:
                        msg = "❌ " + str(e)[:60]; col = T["err"]
                    lbl_ping.after(0, lambda: lbl_ping.configure(text=msg, text_color=col))
                threading.Thread(target=_chk, daemon=True).start()
            ctk.CTkButton(s_srv, text="Probar", width=100, height=26,
                fg_color=T["card"], hover_color=T["a1"], border_width=1, border_color=T["br"],
                font=(FONT,9), text_color=T["t2"],
                command=_test_srv).pack(anchor="w", padx=12, pady=(0,8))

            # Rutas OneDrive por archivo (corazón del sistema)
            s_od = self._sec(sc,"📂","Rutas OneDrive  —  Excel 365 directo")
            ctk.CTkLabel(s_od, text="Ruta en OneDrive: /Carpeta/Archivo.xlsx  (ej: /RRHH/Dotacion.xlsx)",
                font=(FONT,8), text_color=T["t2"]).pack(anchor="w", padx=12, pady=(0,4))

            od_campos = [
                ("Dotación",         "onedrive_dotacion"),
                ("Liquidaciones",    "onedrive_liquidaciones"),
                ("Comisiones SLD",   "onedrive_comisiones_sld"),
                ("Comisiones PDV",   "onedrive_comisiones_pdv"),
                ("Comisiones EOS",   "onedrive_comisiones_eos"),
            ]
            e_od = {}
            for lbl_txt, cfg_key in od_campos:
                rf = ctk.CTkFrame(s_od, fg_color="transparent"); rf.pack(fill="x", padx=12, pady=1)
                ctk.CTkLabel(rf, text=lbl_txt+":", font=(FONT,9), text_color=T["t2"], width=120).pack(side="left")
                e = ctk.CTkEntry(rf, height=26, font=(FONT,9),
                    placeholder_text="/RRHH/Archivo.xlsx",
                    fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
                e.pack(side="left", fill="x", expand=True, padx=4)
                e.insert(0, CFG.get(cfg_key, ""))
                e_od[cfg_key] = e

            s_graph = self._sec(sc,"☁️","Microsoft Graph API  —  OneDrive / Excel 365")
            ctk.CTkLabel(s_graph, text="🌐 Conexión principal para datos OneDrive/Excel 365. Requiere app en portal.azure.com",
                font=(FONT,9), text_color=T["t2"]).pack(anchor="w", padx=12, pady=(0,4))

            def _graph_row(parent, label, cfg_key, show=None):
                rf = ctk.CTkFrame(parent, fg_color="transparent"); rf.pack(fill="x", padx=12, pady=(0,2))
                ctk.CTkLabel(rf, text=label, font=(FONT,10), text_color=T["t2"], width=90).pack(side="left")
                kw = dict(height=28, font=(FONT,10), fg_color=T["inp"],
                          border_color=T["br"], text_color=T["tx"])
                if show: kw["show"] = show
                e = ctk.CTkEntry(rf, **kw)
                e.pack(side="left", fill="x", expand=True, padx=4)
                e.insert(0, CFG.get(cfg_key,""))
                return e

            e_g_tenant = _graph_row(s_graph, "Tenant ID:", "graph_tenant")
            e_g_client = _graph_row(s_graph, "Client ID:", "graph_client")
            e_g_secret = _graph_row(s_graph, "Secret:", "graph_secret", show="*")

            # ── CAMPO NUEVO: Email usuario OneDrive ─────────────────────────────
            ctk.CTkLabel(s_graph,
                text="📧 Email del propietario del OneDrive (ej: remuneraciones@empresa.cl)",
                font=(FONT,8), text_color=T["warn"]).pack(anchor="w", padx=12, pady=(4,0))
            ctk.CTkLabel(s_graph,
                text="   ⚠ Con Client Credentials NO se puede usar /me/drive — este email es obligatorio.",
                font=(FONT,8), text_color=T["t2"]).pack(anchor="w", padx=12)
            e_g_email = _graph_row(s_graph, "Email OD:", "graph_user_email")

            lbl_graph = ctk.CTkLabel(s_graph, text="", font=(FONT,9), text_color=T["t2"])
            lbl_graph.pack(anchor="w", padx=12, pady=(0,4))
            def _test_graph():
                """Prueba Graph API con diagnóstico detallado. Endpoint correcto para Client Credentials."""
                lbl_graph.configure(text="Probando conexion...", text_color=T["t2"])
                import threading, urllib.request, urllib.error, urllib.parse, json as _jt, re as _re2

                def _chk():
                    tid = e_g_tenant.get().strip()
                    cid = e_g_client.get().strip()
                    sec = e_g_secret.get().strip()

                    # Validar formato UUID
                    uuid_pat = r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"
                    if not _re2.match(uuid_pat, tid, _re2.I):
                        lbl_graph.after(0, lambda: lbl_graph.configure(
                            text="❌ Tenant ID inválido. Debe ser UUID: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx",
                            text_color=T["err"])); return
                    if not _re2.match(uuid_pat, cid, _re2.I):
                        lbl_graph.after(0, lambda: lbl_graph.configure(
                            text="❌ Client ID inválido. Debe ser UUID: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx",
                            text_color=T["err"])); return
                    if len(sec) < 8:
                        lbl_graph.after(0, lambda: lbl_graph.configure(
                            text="❌ Secret muy corto. Verifica que copiaste el Valor (no el ID del secreto).",
                            text_color=T["err"])); return

                    email_od = e_g_email.get().strip()
                    if not email_od or "@" not in email_od:
                        lbl_graph.after(0, lambda: lbl_graph.configure(
                            text="❌ Falta el Email OD. Agrega el email corporativo del propietario del OneDrive.",
                            text_color=T["err"])); return

                    # Guardar temporalmente
                    _orig = (CFG.get("graph_tenant",""), CFG.get("graph_client",""), CFG.get("graph_secret",""))
                    CFG["graph_tenant"] = tid
                    CFG["graph_client"] = cid
                    CFG["graph_secret"] = sec
                    _graph_token_cache["token"]   = None
                    _graph_token_cache["expires"] = 0

                    try:
                        # PASO 1: Token OAuth
                        lbl_graph.after(0, lambda: lbl_graph.configure(
                            text="[1/2] Obteniendo token OAuth...", text_color=T["t2"]))

                        token_url = "https://login.microsoftonline.com/" + tid + "/oauth2/v2.0/token"
                        body_data = urllib.parse.urlencode({
                            "grant_type":    "client_credentials",
                            "client_id":     cid,
                            "client_secret": sec,
                            "scope":         "https://graph.microsoft.com/.default",
                        }).encode("utf-8")

                        try:
                            req_tok = urllib.request.Request(
                                token_url, data=body_data,
                                headers={"Content-Type": "application/x-www-form-urlencoded"})
                            with urllib.request.urlopen(req_tok, timeout=15) as r:
                                tok_data = _jt.loads(r.read())
                        except urllib.error.HTTPError as he:
                            err_body = {}
                            try: err_body = _jt.loads(he.read())
                            except Exception: pass
                            desc = err_body.get("error_description","")[:300]
                            # Diagnóstico por código de error Azure
                            if "AADSTS700016" in desc:
                                hint = "❌ Error 400: Client ID no reconocido en este Tenant. Verifica que la app pertenece a este directorio."
                            elif "AADSTS7000215" in desc:
                                hint = "❌ Error 400: Client Secret inválido o expirado. Crea uno nuevo en Azure > Certificados y secretos."
                            elif "AADSTS90002" in desc:
                                hint = "❌ Error 400: Tenant ID no encontrado. Usa el ID de directorio, no el de suscripción."
                            elif he.code == 400:
                                short = desc[:200] if desc else err_body.get("error","")
                                hint = "❌ Error 400 Bad Request: " + short
                            else:
                                hint = "❌ HTTP " + str(he.code) + ": " + desc[:150]
                            CFG["graph_tenant"], CFG["graph_client"], CFG["graph_secret"] = _orig
                            lbl_graph.after(0, lambda m=hint: lbl_graph.configure(text=m, text_color=T["err"])); return

                        access_token = tok_data.get("access_token","")
                        if not access_token:
                            CFG["graph_tenant"], CFG["graph_client"], CFG["graph_secret"] = _orig
                            lbl_graph.after(0, lambda: lbl_graph.configure(
                                text="❌ Token vacío. Verifica Client ID y Secret.", text_color=T["err"])); return

                        # PASO 2: Verificar permisos con /organization (correcto para app credentials)
                        lbl_graph.after(0, lambda: lbl_graph.configure(
                            text="[2/2] Verificando permisos...", text_color=T["t2"]))

                        try:
                            req_org = urllib.request.Request(
                                "https://graph.microsoft.com/v1.0/organization",
                                headers={"Authorization": "Bearer " + access_token,
                                         "Accept": "application/json"})
                            with urllib.request.urlopen(req_org, timeout=15) as r2:
                                org_data = _jt.loads(r2.read())
                            org_name = (org_data.get("value",[{}])[0].get("displayName","Organización"))
                            exp_min  = tok_data.get("expires_in", 3600) // 60
                            _graph_token_cache["token"]   = access_token
                            _graph_token_cache["expires"] = __import__("time").time() + tok_data.get("expires_in",3600)
                            # Persistir email en CFG
                            CFG["graph_user_email"] = email_od
                            msg = "✅ Graph API conectado  |  Org: " + org_name + "  |  Usuario: " + email_od + "  |  Token " + str(exp_min) + " min"
                            lbl_graph.after(0, lambda m=msg: lbl_graph.configure(text=m, text_color=T["ok"]))
                        except urllib.error.HTTPError as he2:
                            err2 = {}
                            try: err2 = _jt.loads(he2.read())
                            except Exception: pass
                            if he2.code == 403:
                                hint2 = ("❌ Token OK pero sin permisos. "
                                         "Ve a Azure > tu app > Permisos de API > "
                                         "haz clic en 'Conceder consentimiento de administrador'.")
                            else:
                                hint2 = "❌ Token OK pero fallo en Graph: HTTP " + str(he2.code)
                            CFG["graph_tenant"], CFG["graph_client"], CFG["graph_secret"] = _orig
                            lbl_graph.after(0, lambda m=hint2: lbl_graph.configure(text=m, text_color=T["err"]))

                    except Exception as e_gen:
                        CFG["graph_tenant"], CFG["graph_client"], CFG["graph_secret"] = _orig
                        lbl_graph.after(0, lambda m=str(e_gen)[:200]: lbl_graph.configure(
                            text="❌ " + m, text_color=T["err"]))

                threading.Thread(target=_chk, daemon=True).start()
            ctk.CTkButton(s_graph, text="Probar Graph API", width=140, height=28,
                fg_color=T["card"], hover_color=T["a1"], border_width=1, border_color=T["br"],
                font=(FONT,9), text_color=T["t2"],
                command=_test_graph).pack(anchor="w", padx=12, pady=(0,8))




            # Claude
            s2 = self._sec(sc,"◆","Claude  —  console.anthropic.com")
            rk = ctk.CTkFrame(s2, fg_color="transparent")
            rk.pack(fill="x", padx=12, pady=(0,4))
            ek = ctk.CTkEntry(rk, height=32, placeholder_text="sk-ant-api...", show="*",
                               font=(FONT,10), fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            ek.pack(side="left", fill="x", expand=True)
            ctk.CTkButton(rk, text="Borrar", width=76, height=32,
                          fg_color=T["card"], hover_color=T["err"],
                          border_width=1, border_color=T["br"],
                          font=(FONT,9,"bold"), text_color=T["t2"],
                          command=lambda: ek.delete(0, "end")
                          ).pack(side="left", padx=(8,0))
            if CFG.get("api_key"): ek.insert(0,CFG["api_key"])
            rf = ctk.CTkFrame(s2,fg_color="transparent"); rf.pack(fill="x",padx=12,pady=(0,4))
            ctk.CTkLabel(rf,text="Modelo:",font=(FONT,10),text_color=T["t2"]).pack(side="left")
            mv = ctk.StringVar(value=CFG.get("modelo",M_FAST))
            ctk.CTkOptionMenu(rf,values=[M_FAST,M_SMART],variable=mv,width=300,
                fg_color=T["card"],button_color=T["a1"],font=(FONT,10),
                text_color=T["tx"],dropdown_fg_color=T["card"],
                dropdown_text_color=T["tx"]).pack(side="left",padx=6)
            lnk(s2,"Obtener API Key →","https://console.anthropic.com")

            # Gemini
            s3 = self._sec(sc,"◉","Gemini  —  aistudio.google.com  (gratis!)")
            rg = ctk.CTkFrame(s3, fg_color="transparent")
            rg.pack(fill="x", padx=12, pady=(0,4))
            egk = ctk.CTkEntry(rg, height=32, placeholder_text="AIza...", show="*",
                                font=(FONT,10), fg_color=T["inp"], border_color=T["br"], text_color=T["tx"])
            egk.pack(side="left", fill="x", expand=True)
            ctk.CTkButton(rg, text="Borrar", width=76, height=32,
                          fg_color=T["card"], hover_color=T["err"],
                          border_width=1, border_color=T["br"],
                          font=(FONT,9,"bold"), text_color=T["t2"],
                          command=lambda: egk.delete(0, "end")
                          ).pack(side="left", padx=(8,0))
            if CFG.get("gemini_key"): egk.insert(0,CFG["gemini_key"])
            lnk(s3,"Obtener key gratis →","https://aistudio.google.com/app/apikey")

            # OpenAI
            s4 = self._sec(sc,"◈","OpenAI / GPT  —  platform.openai.com")
            eok = ctk.CTkEntry(s4,height=32,placeholder_text="sk-...",show="*",
                                font=(FONT,10),fg_color=T["inp"],border_color=T["br"],text_color=T["tx"])
            eok.pack(fill="x",padx=12,pady=(0,4))
            if CFG.get("openai_key"): eok.insert(0,CFG["openai_key"])
            lnk(s4,"Obtener API Key →","https://platform.openai.com/api-keys")

            # Mistral
            s5 = self._sec(sc,"◇","Mistral AI  —  console.mistral.ai  (plan gratis)")
            emk = ctk.CTkEntry(s5,height=32,placeholder_text="API key...",show="*",
                                font=(FONT,10),fg_color=T["inp"],border_color=T["br"],text_color=T["tx"])
            emk.pack(fill="x",padx=12,pady=(0,4))
            if CFG.get("mistral_key"): emk.insert(0,CFG["mistral_key"])
            lnk(s5,"Obtener API Key →","https://console.mistral.ai/api-keys")

            # Groq (gratis)
            s5b = self._sec(sc,"⚡","Groq  —  console.groq.com  (¡100% GRATIS!)")
            ctk.CTkLabel(s5b,
                text="  Fallback automático cuando tu IA principal se queda sin cuota.",
                font=(FONT,9),text_color=T["warn"]).pack(anchor="w",padx=12,pady=(0,2))
            ctk.CTkLabel(s5b,
                text="  Modelos: Llama 3.3 70B · Llama 3.1 8B · Mixtral 8x7B  (sin costo, sin tarjeta)",
                font=(FONT,9),text_color=T["t2"]).pack(anchor="w",padx=12,pady=(0,4))
            rg2 = ctk.CTkFrame(s5b, fg_color="transparent")
            rg2.pack(fill="x", padx=12, pady=(0,4))
            egro = ctk.CTkEntry(rg2, height=32, placeholder_text="gsk_...", show="*",
                                 font=(FONT,10), fg_color=T["inp"], border_color=T["br"],
                                 text_color=T["tx"])
            egro.pack(side="left", fill="x", expand=True)
            ctk.CTkButton(rg2, text="Borrar", width=76, height=32,
                          fg_color=T["card"], hover_color=T["err"],
                          border_width=1, border_color=T["br"],
                          font=(FONT,9,"bold"), text_color=T["t2"],
                          command=lambda: egro.delete(0, "end")
                          ).pack(side="left", padx=(8,0))
            if CFG.get("groq_key"): egro.insert(0, CFG["groq_key"])
            lnk(s5b,"Obtener key gratis →","https://console.groq.com/keys")

            # Ollama
            s6 = self._sec(sc,"◎","Ollama  —  gratis, 100% local, sin internet")
            rf6=ctk.CTkFrame(s6,fg_color="transparent"); rf6.pack(fill="x",padx=12,pady=(0,4))
            ctk.CTkLabel(rf6,text="URL:",font=(FONT,10),text_color=T["t2"]).pack(side="left")
            eou=ctk.CTkEntry(rf6,width=160,height=28,font=(FONT,10),fg_color=T["inp"],
                              border_color=T["br"],text_color=T["tx"])
            eou.pack(side="left",padx=4); eou.insert(0,CFG.get("ollama_url",OLLAMA_URL))
            ctk.CTkLabel(rf6,text="Modelo:",font=(FONT,10),text_color=T["t2"]).pack(side="left",padx=(8,0))
            eom=ctk.CTkEntry(rf6,width=160,height=28,font=(FONT,10),fg_color=T["inp"],
                              border_color=T["br"],text_color=T["tx"])
            eom.pack(side="left",padx=4); eom.insert(0,CFG.get("ollama_model","qwen2.5-coder:7b"))
            lnk(s6,"Descargar Ollama →","https://ollama.ai")


            # Apariencia
            s7 = self._sec(sc,"🎨","Apariencia")
            tf=ctk.CTkFrame(s7,fg_color="transparent"); tf.pack(padx=12,pady=(0,4))
            for nom,pal in PALETAS.items():
                ctk.CTkButton(tf,text=nom,height=28,width=110,
                    fg_color=pal["card"],hover_color=pal["a1"],border_color=pal["acc"],
                    border_width=2,font=(FONT,9,"bold"),text_color=pal["a2"],
                    command=lambda p=pal:(guardar_tema(p),
                                          lbl_ok.configure(text="Tema guardado — reinicia Remi"))).pack(side="left",padx=2,pady=2)
            wv=ctk.BooleanVar(value=CFG.get("web_auto",True))
            ctk.CTkCheckBox(s7,text="Búsqueda web automática al preguntar",
                variable=wv,font=(FONT,10),text_color=T["tx"],
                fg_color=T["a1"]).pack(anchor="w",padx=12,pady=(4,8))

            # Footer
            ftr=ctk.CTkFrame(panel,fg_color=T["sb"],corner_radius=0,height=56)
            ftr.pack(fill="x",side="bottom"); ftr.pack_propagate(False)
            lbl_ok=ctk.CTkLabel(ftr,text="",font=(FONT,10),text_color=T["ok"])
            lbl_ok.pack(side="left",padx=14,pady=16)

            def save():
                CFG["api_key"]       = ek.get().strip()
                CFG["gemini_key"]    = egk.get().strip()
                CFG["openai_key"]    = eok.get().strip()
                CFG["mistral_key"]   = emk.get().strip()
                CFG["groq_key"]      = egro.get().strip()
                CFG["modelo"]        = mv.get()
                CFG["proveedor"]     = pv.get()
                CFG["multi_agent"]   = mav.get()
                CFG["web_auto"]      = wv.get()
                CFG["ollama_url"]    = eou.get().strip()
                CFG["ollama_model"]  = eom.get().strip()
                CFG["nombre_usuario"]= en.get().strip() or "Yerko"
                CFG["server_url"]    = e_srv_url.get().strip()
                CFG["server_token"]  = e_srv_tok.get().strip()
                CFG["server_activo"] = srv_actv.get()
                for cfg_key, e in e_od.items():
                    CFG[cfg_key] = e.get().strip()
                CFG["daemon_activo"]   = dmv.get()
                try: CFG["daemon_intervalo"] = max(1, int(e_dm_ivl.get()))
                except ValueError: pass
                CFG["finiquitos_ruta"]    = e_fin_ruta.get().strip()
                CFG["finiquitos_madre"]   = e_fin_madre.get().strip() or "MATRIZ_GENERAL_FINIQUITOS.xlsx"
                CFG["finiquitos_ruta_od"] = e_fin_ruta_od.get().strip()
                CFG["finiquitos_hoja"]    = e_fin_hoja.get().strip() or "Finiquitos"
                CFG["graph_tenant"]  = e_g_tenant.get().strip()
                CFG["graph_client"]  = e_g_client.get().strip()
                CFG["graph_secret"]  = e_g_secret.get().strip()
                CFG["graph_user_email"] = e_g_email.get().strip().lower()
                # Limpiar caché token para re-autenticar con nuevas credenciales
                _graph_token_cache["token"]   = None
                _graph_token_cache["expires"] = 0
                _ERR_PROV.clear(); invalidar_cache_prov()
                guardar_cfg(CFG)
                self._pv.set(pv.get()); self._update_prov_label()
                lbl_ok.configure(text="✅ Configuración guardada")

            ctk.CTkButton(ftr,text="  Guardar  ",height=38,fg_color=T["a1"],
                font=(FONT,12,"bold"),text_color="white",
                command=save).pack(side="right",padx=14,pady=9)

        self._modal("⚙️  Configuración", build, w=760, h=820)

    # ── CARPETAS ─────────────────────────────────────────────
    def _win_carpetas(self):
        def build(panel, cerrar):
            sc = self._sc_modal(panel)
            for n, r in RUTAS.items():
                ok = os.path.exists(r)
                card = ctk.CTkFrame(sc,fg_color=T["card"],corner_radius=8,
                                    border_color=T["ok"] if ok else T["err"],border_width=1)
                card.pack(fill="x",padx=14,pady=4)
                ctk.CTkLabel(card,text=f"{'✅' if ok else '❌'}  {n}",
                             font=(FONT,12,"bold"),
                             text_color=T["ok"] if ok else T["err"]).pack(side="left",padx=12,pady=10)
                ctk.CTkLabel(card,text=r,font=(FMONO,8),
                             text_color=T["t2"]).pack(side="left",padx=4)
                if ok:
                    ctk.CTkButton(card,text="Abrir",height=28,width=70,
                        fg_color=T["a1"],font=(FONT,10),text_color="white",
                        command=lambda rr=r:subprocess.Popen(
                            f'explorer "{rr}"',
                            creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0))).pack(side="right",padx=10,pady=8)
        self._modal("📂  Carpetas de trabajo", build, w=660, h=360)

    # ── ESCRIBIR EN EXCEL ─────────────────────────────────────────────
    def _win_escribir_xl(self):
        """Modal para crear/modificar Excel con datos reales directamente."""
        def build(panel, cerrar):
            sc = self._sc_modal(panel)

            # ── Sección: archivo destino ──────────────────────────────────
            s_arch = self._sec(sc, "📁", "Archivo Excel destino")
            rv = [None]  # ruta seleccionada

            row_arch = ctk.CTkFrame(s_arch, fg_color="transparent")
            row_arch.pack(fill="x", padx=12, pady=(0,8))
            lbl_arch = ctk.CTkLabel(row_arch, text="Sin archivo seleccionado",
                                     font=(FONT,10), text_color=T["t2"])
            lbl_arch.pack(side="left", fill="x", expand=True)

            def elegir_existente():
                r = filedialog.askopenfilename(
                    title="Seleccionar Excel existente",
                    filetypes=[("Excel","*.xlsx *.xlsm")])
                if r:
                    rv[0] = r
                    lbl_arch.configure(text=os.path.basename(r), text_color=T["a2"])

            def elegir_nuevo():
                r = filedialog.asksaveasfilename(
                    title="Nuevo archivo Excel",
                    initialdir=str(DESCARGAS),
                    defaultextension=".xlsx",
                    filetypes=[("Excel","*.xlsx")])
                if r:
                    rv[0] = r
                    lbl_arch.configure(text=os.path.basename(r)+" (nuevo)", text_color=T["ok"])

            bf_arch = ctk.CTkFrame(s_arch, fg_color="transparent")
            bf_arch.pack(padx=12, pady=(0,10))
            ctk.CTkButton(bf_arch, text="📂 Abrir existente", height=30, width=160,
                fg_color=T["a1"], font=(FONT,10), text_color="white",
                command=elegir_existente).pack(side="left", padx=4)
            ctk.CTkButton(bf_arch, text="✨ Crear nuevo", height=30, width=140,
                fg_color=T["card"], border_width=1, border_color=T["a1"],
                font=(FONT,10), text_color=T["a2"],
                command=elegir_nuevo).pack(side="left", padx=4)

            # ── Sección: nombre de hoja ───────────────────────────────────
            s_hoja = self._sec(sc, "📋", "Hoja")
            row_hoja = ctk.CTkFrame(s_hoja, fg_color="transparent")
            row_hoja.pack(fill="x", padx=12, pady=(0,8))
            ctk.CTkLabel(row_hoja, text="Nombre hoja:", font=(FONT,10),
                         text_color=T["t2"]).pack(side="left")
            e_hoja = ctk.CTkEntry(row_hoja, width=200, height=28, placeholder_text="Hoja1",
                                   font=(FONT,10), fg_color=T["inp"],
                                   border_color=T["br"], text_color=T["tx"])
            e_hoja.pack(side="left", padx=8)
            ctk.CTkLabel(row_hoja, text="(vacío = hoja activa)",
                         font=(FONT,8), text_color=T["t2"]).pack(side="left")

            # ── Sección: modo de escritura ────────────────────────────────
            s_modo = self._sec(sc, "⚙️", "Modo")
            modo_var = ctk.StringVar(value="append")
            row_modo = ctk.CTkFrame(s_modo, fg_color="transparent")
            row_modo.pack(padx=12, pady=(0,8))
            for lbl_m, val_m, tip_m in [
                ("➕ Agregar al final", "append", "Agrega filas sin borrar existentes"),
                ("🔄 Reemplazar todo",  "overwrite","Borra el contenido actual y escribe desde cero"),
                ("📍 Celda específica", "cell",    "Escribe en una celda exacta (ej: A1, B5)")
            ]:
                ctk.CTkRadioButton(
                    row_modo, text=lbl_m, variable=modo_var, value=val_m,
                    font=(FONT,10), text_color=T["tx"],
                    fg_color=T["a1"], border_color=T["br"]
                ).pack(side="left", padx=8)

            # ── Sección: datos ────────────────────────────────────────────
            s_datos = self._sec(sc, "📊", "Datos a escribir")

            # Sub-frame celda específica
            frame_celda = ctk.CTkFrame(s_datos, fg_color="transparent")
            row_celda = ctk.CTkFrame(frame_celda, fg_color="transparent")
            row_celda.pack(fill="x", padx=12, pady=4)
            ctk.CTkLabel(row_celda, text="Celda (ej: B5):", font=(FONT,10),
                         text_color=T["t2"]).pack(side="left")
            e_ref = ctk.CTkEntry(row_celda, width=80, height=28,
                                  placeholder_text="A1",
                                  font=(FONT,11), fg_color=T["inp"],
                                  border_color=T["a1"], text_color=T["tx"])
            e_ref.pack(side="left", padx=8)
            ctk.CTkLabel(row_celda, text="Valor:", font=(FONT,10),
                         text_color=T["t2"]).pack(side="left", padx=(12,0))
            e_val_celda = ctk.CTkEntry(row_celda, width=200, height=28,
                                        placeholder_text="Texto, número o fecha",
                                        font=(FONT,10), fg_color=T["inp"],
                                        border_color=T["a1"], text_color=T["tx"])
            e_val_celda.pack(side="left", padx=8)

            # Área texto principal
            ctk.CTkLabel(s_datos,
                text="  Pega o escribe los datos (separados por Tab o | entre columnas, Enter entre filas):",
                font=(FONT,9), text_color=T["t2"]).pack(anchor="w", padx=12, pady=(8,2))
            e_datos = ctk.CTkTextbox(s_datos, height=140, font=(FMONO,10),
                                      fg_color=T["inp"], border_color=T["a1"],
                                      border_width=2, text_color=T["tx"])
            e_datos.pack(fill="x", padx=12, pady=(0,4))

            # Placeholder de ejemplo
            EJEMPLO = ("Nombre\tRUT\tSueldo\tEstado\n"
                       "Juan Pérez\t12.345.678-9\t850000\tActivo\n"
                       "María López\t9.876.543-2\t920000\tActivo\n"
                       "Carlos Silva\t15.432.100-K\t780000\tLicencia")
            e_datos.insert("1.0", EJEMPLO)
            e_datos.configure(text_color=T["t2"])

            def limpiar_ph(e=None):
                if e_datos.get("1.0","end").strip() == EJEMPLO.strip():
                    e_datos.delete("1.0","end")
                    e_datos.configure(text_color=T["tx"])
            e_datos.bind("<FocusIn>", limpiar_ph)

            # Opciones extra
            opts_f = ctk.CTkFrame(s_datos, fg_color="transparent")
            opts_f.pack(fill="x", padx=12, pady=(4,8))
            var_header = ctk.BooleanVar(value=True)
            var_estilo  = ctk.BooleanVar(value=True)
            ctk.CTkCheckBox(opts_f, text="Primera fila = encabezados",
                variable=var_header, font=(FONT,9),
                text_color=T["tx"], fg_color=T["a1"]).pack(side="left", padx=4)
            ctk.CTkCheckBox(opts_f, text="Aplicar estilo tabla",
                variable=var_estilo, font=(FONT,9),
                text_color=T["tx"], fg_color=T["a1"]).pack(side="left", padx=4)

            # Mostrar/ocultar frame celda según modo
            def on_modo(*_):
                if modo_var.get() == "cell":
                    frame_celda.pack(fill="x", padx=12, pady=4)
                else:
                    frame_celda.pack_forget()
            modo_var.trace_add("write", on_modo)

            # ── Resultado ─────────────────────────────────────────────────
            lbl_res = ctk.CTkLabel(sc, text="", font=(FONT,10), text_color=T["ok"],
                                    wraplength=620)
            lbl_res.pack(anchor="w", padx=16, pady=4)

            # ── Botón principal ────────────────────────────────────────────
            def _parsear_datos(texto):
                """Convierte texto pegado en headers + rows."""
                lineas = [l for l in texto.strip().splitlines() if l.strip()]
                if not lineas: return [], []
                # Detectar separador: tab, |, ;, coma
                sep = "\t"
                for s in ["\t", "|", ";", ","]:
                    if s in lineas[0]:
                        sep = s; break
                tabla = [[c.strip().strip("|").strip() for c in l.split(sep)] for l in lineas]
                if var_header.get() and len(tabla) > 1:
                    return tabla[0], tabla[1:]
                return [], tabla

            def ejecutar():
                if not rv[0]:
                    lbl_res.configure(text="⚠️ Selecciona un archivo primero.", text_color=T["warn"])
                    return
                ruta = rv[0]
                modo = modo_var.get()
                nombre_hoja = e_hoja.get().strip() or None
                lbl_res.configure(text="⏳ Escribiendo...", text_color=T["a2"])
                panel.update()

                try:
                    if modo == "cell":
                        ref = e_ref.get().strip().upper() or "A1"
                        val = e_val_celda.get().strip()
                        if not val:
                            lbl_res.configure(text="⚠️ Ingresa un valor para la celda.", text_color=T["warn"])
                            return
                        # Crear archivo si no existe
                        if not os.path.exists(ruta):
                            if not XL:
                                lbl_res.configure(text="❌ openpyxl no instalado.", text_color=T["err"])
                                return
                            wb_n = openpyxl.Workbook()
                            if nombre_hoja: wb_n.active.title = nombre_hoja
                            wb_n.save(ruta)
                        res = escribir_celdas(ruta, {ref: val}, hoja=nombre_hoja)
                    else:
                        raw = e_datos.get("1.0","end").strip()
                        if not raw or raw == EJEMPLO.strip():
                            lbl_res.configure(text="⚠️ Ingresa datos en el área de texto.", text_color=T["warn"])
                            return
                        headers, rows = _parsear_datos(raw)
                        if not rows:
                            lbl_res.configure(text="⚠️ No se detectaron datos válidos.", text_color=T["warn"])
                            return

                        if modo == "overwrite" or not os.path.exists(ruta):
                            # Crear nuevo o sobreescribir hoja
                            datos_hojas = {nombre_hoja or "Datos": {"headers": headers, "rows": rows}}
                            res = crear_excel_desde_datos(ruta, datos_hojas, estilo=var_estilo.get())
                        else:
                            # Append: agregar filas al final de hoja existente
                            if not XL:
                                lbl_res.configure(text="❌ openpyxl no instalado.", text_color=T["err"])
                                return
                            wb_a = openpyxl.load_workbook(ruta)
                            ws_a = wb_a[nombre_hoja] if nombre_hoja and nombre_hoja in wb_a.sheetnames else wb_a.active
                            es_vacio = ws_a.max_row is None or ws_a.max_row == 0 or (ws_a.max_row == 1 and ws_a.cell(1,1).value is None)
                            if es_vacio and headers:
                                ws_a.append([str(h) for h in headers])
                            for row in rows:
                                ws_a.append([_xl_valor(v) for v in row])
                            if var_estilo.get():
                                _xl_estilo_tabla(ws_a)
                            wb_a.save(ruta)
                            res = f"✅ {len(rows)} filas agregadas a {os.path.basename(ruta)} — hoja '{ws_a.title}'"

                    lbl_res.configure(text=res, text_color=T["ok"] if "✅" in res else T["err"])
                    self._bub_sys(res, "ok" if "✅" in res else "error")

                    # Ofrecer abrir el archivo
                    if "✅" in res:
                        try: os.startfile(ruta)
                        except: pass

                except Exception as ex:
                    msg = f"❌ Error: {ex}"
                    lbl_res.configure(text=msg, text_color=T["err"])
                    self._bub_sys(msg, "error")

            # Footer
            ftr = ctk.CTkFrame(panel, fg_color=T["sb"], corner_radius=0, height=56)
            ftr.pack(fill="x", side="bottom"); ftr.pack_propagate(False)
            ctk.CTkButton(ftr, text="  ✏️  Escribir en Excel  ", height=38,
                fg_color=T["a1"], font=(FONT,12,"bold"), text_color="white",
                command=ejecutar).pack(side="right", padx=14, pady=9)
            ctk.CTkButton(ftr, text="Cancelar", height=38, width=100,
                fg_color=T["card"], border_width=1, border_color=T["br"],
                font=(FONT,10), text_color=T["t2"],
                command=cerrar).pack(side="right", padx=4, pady=9)

        self._modal("✏️  Escribir en Excel", build, w=740, h=700)

    # ── MODIFICAR EXCEL ──────────────────────────────────────
    def _win_mod_xl(self):
        def build(panel, cerrar):
            rv = [None]
            card = ctk.CTkFrame(panel,fg_color=T["card"],corner_radius=8)
            card.pack(fill="x",padx=14,pady=14)
            row1 = ctk.CTkFrame(card,fg_color="transparent"); row1.pack(fill="x",padx=12,pady=10)
            la = ctk.CTkLabel(row1,text="Sin archivo seleccionado",
                               font=(FONT,10),text_color=T["t2"])
            la.pack(side="left")
            def elegir():
                r=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xlsm *.xls")])
                if r: rv[0]=r; la.configure(text=os.path.basename(r),text_color=T["a2"])
            ctk.CTkButton(row1,text="Elegir…",height=30,width=100,
                fg_color=T["a1"],font=(FONT,10),text_color="white",
                command=elegir).pack(side="right")

            ctk.CTkLabel(panel,text="¿Qué modificar?",font=(FONT,10,"bold"),
                         text_color=T["t2"]).pack(anchor="w",padx=16,pady=(4,2))
            e = ctk.CTkTextbox(panel,height=55,font=(FONT,11),
                               fg_color=T["inp"],border_color=T["a1"],
                               border_width=2,text_color=T["tx"])
            e.pack(fill="x",padx=14,pady=4)

            sf=ctk.CTkFrame(panel,fg_color="transparent"); sf.pack(padx=14,pady=2)
            for sug in ["color encabezados azul","autoajusta columnas",
                        "congela fila 1","negrita encabezados",
                        "aplica bordes","activa filtros"]:
                ctk.CTkButton(sf,text=sug,height=24,fg_color="transparent",
                    hover_color=T["card"],border_width=1,border_color=T["br"],
                    font=(FONT,8),text_color=T["t2"],
                    command=lambda x=sug:(e.delete("1.0","end"),e.insert("1.0",x))).pack(side="left",padx=2,pady=2)

            lbl_r=ctk.CTkLabel(panel,text="",font=(FONT,10),text_color=T["ok"],wraplength=600)
            lbl_r.pack(anchor="w",padx=16,pady=4)

            def aplicar():
                if rv[0]:
                    r=modificar_excel(rv[0],e.get("1.0","end").strip())
                    lbl_r.configure(text=r,text_color=T["ok"] if "✅" in r else T["err"])
                    self._bub_sys(r,"ok" if "✅" in r else "error")

            ctk.CTkButton(panel,text="  Aplicar cambios  ",height=38,fg_color=T["a1"],
                font=(FONT,11,"bold"),text_color="white",command=aplicar).pack(pady=12)

        self._modal("✏️  Modificar Excel", build, w=580, h=420)

    # ── COMPARAR EXCEL ───────────────────────────────────────
    def _win_cmp_xl(self):
        r1=filedialog.askopenfilename(title="Archivo BASE",filetypes=[("Excel","*.xlsx *.xlsm *.xls")])
        if not r1: return
        r2=filedialog.askopenfilename(title="Archivo COMPARAR",filetypes=[("Excel","*.xlsx *.xlsm *.xls")])
        if not r2: return
        res = comparar_excel(r1,r2)
        self._bub_sys(res,"ok" if "✅" in res else "system")

    # ── ALARMAS ──────────────────────────────────────────────
    def _win_alarmas(self):
        def build(panel, cerrar):
            sc=self._sc_modal(panel)
            alarmas=_rl(F["alarm"])
            if not alarmas:
                ctk.CTkLabel(sc,text="No hay alarmas.",font=(FONT,11),
                             text_color=T["t2"]).pack(pady=20)
            for i,a in enumerate(reversed(alarmas[-15:])):
                activa=a.get("activa",False)
                card=ctk.CTkFrame(sc,fg_color=T["card"],corner_radius=6,
                                   border_color=T["warn"] if activa else T["br"],border_width=1)
                card.pack(fill="x",padx=12,pady=3)
                ico="🔔" if activa else "✓"
                ctk.CTkLabel(card,text=f"{ico}  {a.get('hora','')}  —  {a.get('mensaje','')[:65]}",
                             font=(FONT,10,"bold" if activa else "normal"),
                             text_color=T["warn"] if activa else T["t2"]).pack(side="left",pady=8,padx=12)
                idx_real=len(alarmas)-1-i
                def toggle(ix=idx_real,al=a):
                    als=_rl(F["alarm"]); als[ix]["activa"]=not als[ix].get("activa",False)
                    _w(F["alarm"],als); cerrar(); self._win_alarmas()
                ctk.CTkButton(card,text="On/Off",height=24,width=60,
                    fg_color=T["card"],hover_color=T["a1"],
                    font=(FONT,8),text_color=T["t2"],command=toggle).pack(side="right",padx=8)

            fila=ctk.CTkFrame(panel,fg_color="transparent"); fila.pack(pady=10)
            eh=ctk.CTkEntry(fila,width=72,height=32,placeholder_text="HH:MM",
                             font=(FONT,11),fg_color=T["inp"],border_color=T["a1"],text_color=T["tx"])
            eh.pack(side="left",padx=4)
            em=ctk.CTkEntry(fila,width=300,height=32,placeholder_text="Mensaje de alarma…",
                             font=(FONT,11),fg_color=T["inp"],border_color=T["a1"],text_color=T["tx"])
            em.pack(side="left",padx=4)
            def add():
                h=eh.get().strip(); m=em.get().strip()
                if h and m:
                    al=_rl(F["alarm"]); al.append({"hora":h,"mensaje":m,"activa":True})
                    _w(F["alarm"],al); cerrar(); self._win_alarmas()
            ctk.CTkButton(fila,text="+",width=34,height=32,fg_color=T["a1"],
                font=(FONT,13,"bold"),text_color="white",command=add).pack(side="left",padx=2)
        self._modal("⏰  Alarmas", build, w=620, h=440)

    # ── NOTA STICKY ──────────────────────────────────────────
    def _win_nota(self):
        def build(panel, cerrar):
            e=ctk.CTkTextbox(panel,height=100,font=(FONT,12),
                              fg_color=T["inp"],border_color=T["a1"],
                              border_width=2,text_color=T["tx"])
            e.pack(fill="x",padx=14,pady=14)
            e.focus()
            ctk.CTkButton(panel,text="  Crear nota  ",height=38,fg_color=T["a1"],
                font=(FONT,11,"bold"),text_color="white",
                command=lambda:(StickyNote(self,e.get("1.0","end").strip()),cerrar())
                ).pack(pady=8)
        self._modal("📝  Nueva nota sticky", build, w=380, h=240)

    # ── DASHBOARD ─────────────────────────────────────────────
    def _win_dashboard(self):
        def build(panel, cerrar):
            sc=self._sc_modal(panel)
            hoy=datetime.datetime.now()

            def lsec(ico,txt,color=None):
                ctk.CTkLabel(sc,text=f"  {ico}  {txt}",font=(FONT,10,"bold"),
                             text_color=color or T["a2"]).pack(anchor="w",padx=12,pady=(14,0))
                tk.Frame(sc,bg=T["br"],height=1).pack(fill="x",padx=12,pady=(2,4))

            lsec("📅","Calendario del mes")
            for d,msg in sorted(CAL_REM.items()):
                mark = " ← HOY" if d==hoy.day else ("  ← PRÓXIMO" if d==hoy.day+1 else "")
                col = T["warn"] if mark else T["t2"]
                ctk.CTkLabel(sc,text=f"  Día {d:2d}:  {msg}{mark}",
                             font=(FONT,9,"bold" if mark else "normal"),
                             text_color=col).pack(anchor="w",padx=22,pady=1)

            lsec("🤖","Estado IA")
            p=prov_activo(); cadena=cadena_proveedores()
            ctk.CTkLabel(sc,
                text=f"  Activo: {PROV_NOMBRES.get(p,'Ninguno')}  |  Cadena: {' → '.join(PROV_NOMBRES.get(x,x) for x in cadena) or 'ninguna'}",
                font=(FONT,10),text_color=PROV_COLOR.get(p,T["t2"])).pack(anchor="w",padx=22,pady=2)

            if _ERR_PROV:
                lsec("⚠️","Proveedores con error (cooldown 45 min)",T["warn"])
                for p2,ts in _ERR_PROV.items():
                    mins = int((time.time()-ts)/60)
                    ctk.CTkLabel(sc,text=f"  {PROV_NOMBRES.get(p2,p2)}: falló hace {mins} min",
                                 font=(FONT,9),text_color=T["warn"]).pack(anchor="w",padx=22,pady=1)

            lsec("🕐","Historial reciente")
            for h in reversed(_rl(F["hist"])[-6:]):
                ctk.CTkLabel(sc,text=f"  [{h.get('fecha','')}]  {h.get('resumen','')[:85]}",
                             font=(FONT,9),text_color=T["t2"]).pack(anchor="w",padx=22,pady=1)

            lsec("⏰","Alarmas activas")
            act=[x for x in _rl(F["alarm"]) if x.get("activa")]
            if act:
                for a in act:
                    ctk.CTkLabel(sc,text=f"  🔔  {a['hora']}  —  {a.get('mensaje','')[:60]}",
                                 font=(FONT,10),text_color=T["warn"]).pack(anchor="w",padx=22,pady=1)
            else:
                ctk.CTkLabel(sc,text="  Sin alarmas activas",font=(FONT,9),
                             text_color=T["t2"]).pack(anchor="w",padx=22,pady=1)

            if SYS:
                lsec("🌡️","Sistema")
                cpu=psutil.cpu_percent(interval=None); ram=psutil.virtual_memory().percent
                col=T["err"] if cpu>85 or ram>85 else T["ok"]
                ctk.CTkLabel(sc,text=f"  CPU {cpu:.0f}%   RAM {ram:.0f}%",
                             font=(FONT,11,"bold"),text_color=col).pack(anchor="w",padx=22,pady=2)

        self._modal("📊  Dashboard", build, w=680, h=580)

    # ── KPIs ─────────────────────────────────────────────────
    def _win_kpi(self):
        def build(panel, cerrar):
            kpis = _r(F["kpi"])
            if not isinstance(kpis, dict): kpis = {}

            sc = self._sc_modal(panel)
            ctk.CTkLabel(sc,text="Indicadores clave — actualiza manualmente",
                         font=(FONT,9),text_color=T["t2"]).pack(anchor="w",padx=14,pady=(4,6))

            ITEMS = [
                ("dot_total",  "👥 Dotación total"),
                ("dot_activos","✅ Activos"),
                ("dot_ausentes","⚠️ Ausentes hoy"),
                ("comision_sld","💰 Comisiones SLD ($)"),
                ("comision_pdv","💰 Comisiones PDV ($)"),
                ("horas_extra", "⏱ Horas extra acumuladas"),
                ("finiquitos",  "📋 Finiquitos pendientes"),
                ("contratos_vencer","⚠️ Contratos a vencer"),
            ]
            entries = {}
            for key, lbl in ITEMS:
                row = ctk.CTkFrame(sc,fg_color=T["card"],corner_radius=6)
                row.pack(fill="x",padx=12,pady=3)
                ctk.CTkLabel(row,text=f"  {lbl}",font=(FONT,10),
                             text_color=T["tx"],width=260,anchor="w").pack(side="left",pady=8)
                ev = ctk.StringVar(value=str(kpis.get(key,"")))
                e  = ctk.CTkEntry(row,textvariable=ev,width=150,height=28,
                                   font=(FONT,10),fg_color=T["inp"],
                                   border_color=T["br"],text_color=T["a2"])
                e.pack(side="right",padx=10)
                entries[key] = ev

            ftr=ctk.CTkFrame(panel,fg_color=T["sb"],corner_radius=0,height=52)
            ftr.pack(fill="x",side="bottom"); ftr.pack_propagate(False)
            lbl_ok=ctk.CTkLabel(ftr,text="",font=(FONT,10),text_color=T["ok"])
            lbl_ok.pack(side="left",padx=14)
            def guardar():
                nuevo = {k:v.get() for k,v in entries.items()}
                _w(F["kpi"],nuevo); lbl_ok.configure(text="✅ KPIs guardados")
            def analizar():
                d={k:v.get() for k,v in entries.items()}
                self._chip_cmd(
                    f"Analiza estos KPIs de Remuneraciones: {json.dumps(d,ensure_ascii=False)}. "
                    f"Identifica alertas, tendencias y recomendaciones.")
                cerrar()
            ctk.CTkButton(ftr,text="  Guardar  ",height=36,width=110,fg_color=T["a1"],
                font=(FONT,11),text_color="white",command=guardar).pack(side="right",padx=4,pady=8)
            ctk.CTkButton(ftr,text="  Analizar con IA  ",height=36,fg_color=T["a2"],
                font=(FONT,11),text_color="white",command=analizar).pack(side="right",padx=4,pady=8)

        self._modal("📈  KPIs — Control de Gestión", build, w=580, h=560)

    # ── GANTT CHECK-IN ───────────────────────────────────────
    def _win_gantt(self):
        """Ventana de Gantt: check-in diario + vista de todas las tareas."""
        data = gantt_load()
        tasks = data.get("tasks", [])
        checkins = data.get("checkin", {})
        today = datetime.date.today().isoformat()
        dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
        hoy_obj = datetime.date.today()
        dia_label = f"{dias[hoy_obj.weekday()]} {hoy_obj.strftime('%d/%m/%Y')}"

        def get_active():
            return [t for t in tasks if t["inicio"] <= today <= t["fin"] and t["avance"] < 1.0]

        def status_color(t):
            if t["avance"] >= 1.0: return T["ok"]
            if t["fin"] < today:   return T["err"]
            if t["avance"] >= 0.5: return T["warn"]
            if t["avance"] > 0:    return T["a1"]
            return T["t2"]

        def status_label(t):
            if t["avance"] >= 1.0: return "Completado"
            if t["fin"] < today:   return "Vencido"
            if t["avance"] >= 0.5: return "En progreso"
            if t["avance"] > 0:    return "Iniciado"
            return "Pendiente"

        def build(panel, cerrar):
            # Tab bar
            tab_var = tk.StringVar(value="checkin")
            tab_f = ctk.CTkFrame(panel, fg_color=T["sb"], corner_radius=0, height=44)
            tab_f.pack(fill="x"); tab_f.pack_propagate(False)

            overall = round(sum(t["avance"] for t in tasks) / len(tasks) * 100) if tasks else 0
            ctk.CTkLabel(tab_f, text=f"  {dia_label}",
                         font=(FONT,10), text_color=T["t2"]).pack(side="left", padx=8, pady=10)
            ctk.CTkLabel(tab_f, text=f"Avance general: {overall}%",
                         font=(FONT,10,"bold"), text_color=T["ok"]).pack(side="right", padx=12)

            btab_f = ctk.CTkFrame(panel, fg_color=T["card"], corner_radius=0, height=38)
            btab_f.pack(fill="x"); btab_f.pack_propagate(False)

            content_f = ctk.CTkFrame(panel, fg_color=T["bg"], corner_radius=0)
            content_f.pack(fill="both", expand=True)

            # ── Helpers para renderizar vistas ────────────────
            def clear_content():
                for w in content_f.winfo_children():
                    try: w.destroy()
                    except: pass

            def show_checkin():
                tab_var.set("checkin")
                clear_content()
                actives = get_active()
                sc2 = ctk.CTkScrollableFrame(content_f, fg_color=T["bg"],
                    scrollbar_button_color=T["a1"], corner_radius=0)
                sc2.pack(fill="both", expand=True)

                done_today = today in checkins
                if not actives:
                    ctk.CTkLabel(sc2, text="✅  Sin tareas activas hoy",
                                 font=(FONT,14,"bold"), text_color=T["ok"]).pack(pady=40)
                    return

                if done_today:
                    ctk.CTkLabel(sc2, text=f"✅  Check-in completado para hoy ({len(actives)} tareas)",
                                 font=(FONT,12,"bold"), text_color=T["ok"]).pack(pady=14, padx=16)
                    ctk.CTkButton(sc2, text="Rehacer check-in", height=30, width=160,
                        fg_color=T["card"], border_width=1, border_color=T["br"],
                        font=(FONT,10), text_color=T["t2"],
                        command=lambda:(checkins.pop(today, None),
                                        gantt_save({"tasks":tasks,"checkin":checkins}),
                                        show_checkin())
                    ).pack(pady=4, padx=16, anchor="w")
                    # Resumen tareas activas
                    for t in actives:
                        col = status_color(t)
                        card = ctk.CTkFrame(sc2, fg_color=T["card"], corner_radius=8,
                                            border_color=col, border_width=1)
                        card.pack(fill="x", padx=14, pady=3)
                        ctk.CTkLabel(card, text=f"  {t['name']}",
                                     font=(FONT,11,"bold"), text_color=T["tx"]).pack(side="left", padx=4, pady=10)
                        ctk.CTkLabel(card, text=f"{round(t['avance']*100)}%  {status_label(t)}",
                                     font=(FONT,10,"bold"), text_color=col).pack(side="right", padx=12)
                    return

                # Check-in interactivo
                state = {"idx": 0, "frame": None}

                def render_question():
                    if state["frame"] and state["frame"].winfo_exists():
                        try: state["frame"].destroy()
                        except: pass

                    if state["idx"] >= len(actives):
                        # Finalizar
                        checkins[today] = True
                        gantt_save({"tasks": tasks, "checkin": checkins})
                        clear_content(); show_checkin()
                        return

                    t = actives[state["idx"]]
                    pct_actual = round(t["avance"] * 100)

                    qf = ctk.CTkFrame(sc2, fg_color=T["card"], corner_radius=10,
                                      border_color=T["a1"], border_width=1)
                    qf.pack(fill="x", padx=14, pady=8)
                    state["frame"] = qf

                    # Progress
                    ctk.CTkLabel(qf, text=f"  Pregunta {state['idx']+1} de {len(actives)}",
                                 font=(FONT,9), text_color=T["t2"]).pack(anchor="w", padx=12, pady=(10,2))
                    pb_bar = ctk.CTkProgressBar(qf, height=4, fg_color=T["br"], progress_color=T["a1"])
                    pb_bar.pack(fill="x", padx=12, pady=(0,8))
                    pb_bar.set(state["idx"] / len(actives))

                    # Responsable + fecha
                    inf_f = ctk.CTkFrame(qf, fg_color="transparent"); inf_f.pack(anchor="w", padx=12, pady=2)
                    ctk.CTkLabel(inf_f, text=f"👤 {t['resp']}",
                                 font=(FONT,9), text_color=T["a2"]).pack(side="left")
                    ctk.CTkLabel(inf_f, text=f"  hasta {t['fin']}",
                                 font=(FONT,9), text_color=T["t2"]).pack(side="left", padx=8)

                    ctk.CTkLabel(qf, text=f"¿Avanzaste en",
                                 font=(FONT,13), text_color=T["t2"]).pack(anchor="w", padx=14, pady=(6,0))
                    ctk.CTkLabel(qf, text=f"  {t['name']}?",
                                 font=(FONT,15,"bold"), text_color=T["a2"]).pack(anchor="w", padx=14, pady=(0,4))
                    if pct_actual > 0:
                        ctk.CTkLabel(qf, text=f"  Avance actual: {pct_actual}%",
                                     font=(FONT,9), text_color=T["t2"]).pack(anchor="w", padx=14, pady=(0,4))

                    # Barra avance actual
                    pb2 = ctk.CTkProgressBar(qf, height=6, fg_color=T["br"],
                                              progress_color=status_color(t))
                    pb2.pack(fill="x", padx=14, pady=(0,10))
                    pb2.set(t["avance"])

                    # Botones respuesta
                    btn_f = ctk.CTkFrame(qf, fg_color="transparent")
                    btn_f.pack(fill="x", padx=12, pady=(0,10))

                    partial_f = ctk.CTkFrame(qf, fg_color="transparent")  # se muestra al click

                    def ans_si():
                        t["avance"] = 1.0
                        gantt_save({"tasks": tasks, "checkin": checkins})
                        state["idx"] += 1; render_question()

                    def ans_no():
                        state["idx"] += 1; render_question()

                    def ans_parcial():
                        for w in btn_f.winfo_children():
                            try: w.pack_forget()
                            except: pass
                        partial_f.pack(fill="x", padx=12, pady=(0,10))
                        ctk.CTkLabel(partial_f, text="¿Cuánto % tiene ahora?",
                                     font=(FONT,10), text_color=T["t2"]).pack(anchor="w", pady=4)
                        pf2 = ctk.CTkFrame(partial_f, fg_color="transparent"); pf2.pack(fill="x")
                        epc = ctk.CTkEntry(pf2, width=90, height=32, placeholder_text="ej: 75",
                                           font=(FONT,11), fg_color=T["inp"],
                                           border_color=T["a1"], text_color=T["tx"])
                        epc.pack(side="left"); epc.focus()
                        def guardar_parcial():
                            try:
                                v = float(epc.get())
                                if 0 <= v <= 100:
                                    t["avance"] = round(v / 100, 2)
                                    gantt_save({"tasks": tasks, "checkin": checkins})
                                    state["idx"] += 1; render_question()
                            except: pass
                        ctk.CTkButton(pf2, text="Guardar", height=32, width=90,
                            fg_color=T["a1"], font=(FONT,10,"bold"), text_color="white",
                            command=guardar_parcial).pack(side="left", padx=8)

                    ctk.CTkButton(btn_f, text="✅  Sí, completado", height=36,
                        fg_color="#0d2018", border_color=T["ok"], border_width=1,
                        font=(FONT,10,"bold"), text_color=T["ok"],
                        command=ans_si).pack(side="left", padx=3, fill="x", expand=True)
                    ctk.CTkButton(btn_f, text="🔄  Parcialmente", height=36,
                        fg_color="#1c1509", border_color=T["warn"], border_width=1,
                        font=(FONT,10,"bold"), text_color=T["warn"],
                        command=ans_parcial).pack(side="left", padx=3, fill="x", expand=True)
                    ctk.CTkButton(btn_f, text="❌  No avancé", height=36,
                        fg_color="#1c0909", border_color=T["err"], border_width=1,
                        font=(FONT,10,"bold"), text_color=T["err"],
                        command=ans_no).pack(side="left", padx=3, fill="x", expand=True)

                render_question()

            def show_tasks():
                tab_var.set("tasks")
                clear_content()
                sc2 = ctk.CTkScrollableFrame(content_f, fg_color=T["bg"],
                    scrollbar_button_color=T["a1"], corner_radius=0)
                sc2.pack(fill="both", expand=True)

                hdr = ctk.CTkFrame(sc2, fg_color="transparent")
                hdr.pack(fill="x", padx=14, pady=(8,4))
                completas = sum(1 for t in tasks if t["avance"] >= 1.0)
                ctk.CTkLabel(hdr, text=f"{completas}/{len(tasks)} completadas",
                             font=(FONT,9), text_color=T["t2"]).pack(side="left")

                for t in tasks:
                    col = status_color(t)
                    pct = round(t["avance"] * 100)
                    card = ctk.CTkFrame(sc2, fg_color=T["card"], corner_radius=8,
                                        border_color=T["br"], border_width=1)
                    card.pack(fill="x", padx=14, pady=3)

                    top_row = ctk.CTkFrame(card, fg_color="transparent")
                    top_row.pack(fill="x", padx=12, pady=(8,2))
                    ctk.CTkLabel(top_row, text=t["name"], font=(FONT,11,"bold"),
                                 text_color=T["tx"]).pack(side="left")
                    ctk.CTkLabel(top_row, text=f"{pct}%  {status_label(t)}",
                                 font=(FONT,10,"bold"), text_color=col).pack(side="right")

                    ctk.CTkLabel(card, text=f"  {t['resp']}  ·  {t['inicio']} → {t['fin']}",
                                 font=(FONT,8), text_color=T["t2"]).pack(anchor="w", padx=12)

                    pb = ctk.CTkProgressBar(card, height=6, fg_color=T["br"], progress_color=col)
                    pb.pack(fill="x", padx=12, pady=(4,6))
                    pb.set(t["avance"])

                    if t["avance"] < 1.0:
                        btn_row = ctk.CTkFrame(card, fg_color="transparent")
                        btn_row.pack(anchor="w", padx=12, pady=(0,8))
                        for v in [25, 50, 75, 100]:
                            is_curr = pct == v
                            def mk_cmd(tt=t, vv=v):
                                tt["avance"] = round(vv / 100, 2)
                                gantt_save({"tasks": tasks, "checkin": checkins})
                                show_tasks()
                            ctk.CTkButton(btn_row, text=f"{v}%", height=24, width=52,
                                fg_color=T["a1"] if is_curr else T["inp"],
                                border_color=T["a1"] if is_curr else T["br"],
                                border_width=1, font=(FONT,8), text_color="white" if is_curr else T["t2"],
                                command=mk_cmd).pack(side="left", padx=2)

            # Tab buttons
            for lbl, key, fn in [("📋 Check-in diario","checkin",show_checkin),
                                   ("📊 Todas las tareas","tasks",show_tasks)]:
                is_active = tab_var.get() == key
                ctk.CTkButton(btab_f, text=lbl, height=38, fg_color="transparent",
                    border_width=0,
                    font=(FONT,10,"bold" if is_active else "normal"),
                    text_color=T["a2"] if is_active else T["t2"],
                    command=fn).pack(side="left", padx=4)

            show_checkin()

        self._modal("📋  Gantt — Check-in diario", build, w=700, h=600)

    # ── PROCESOS ─────────────────────────────────────────────
    def _win_procesos(self):
        def build(panel, cerrar):
            ctrl=ctk.CTkFrame(panel,fg_color=T["card"],corner_radius=8)
            ctrl.pack(fill="x",padx=14,pady=12)
            cf=ctk.CTkFrame(ctrl,fg_color="transparent"); cf.pack(fill="x",padx=12,pady=10)
            ctk.CTkLabel(cf,text="Nombre del proceso:",font=(FONT,10),text_color=T["t2"]).pack(side="left")
            en=ctk.CTkEntry(cf,width=260,height=30,placeholder_text="ej: Cierre Comisiones SLD Feb",
                             font=(FONT,11),fg_color=T["inp"],border_color=T["a1"],text_color=T["tx"])
            en.pack(side="left",padx=8)
            if MONITOR.grabando: en.insert(0,MONITOR.nombre)

            if MONITOR.grabando:
                ctk.CTkLabel(ctrl,text=f"⏺ Grabando: {MONITOR.nombre}",
                             font=(FONT,9,"bold"),text_color=T["err"]).pack(anchor="w",padx=12,pady=(0,4))

            bf=ctk.CTkFrame(ctrl,fg_color="transparent"); bf.pack(padx=12,pady=(0,10))
            if not MONITOR.grabando:
                def iniciar():
                    nb=en.get().strip() or "Proceso"
                    m=MONITOR.iniciar(nb,list(RUTAS.values()))
                    self._bub_sys(m,"ok"); cerrar()
                ctk.CTkButton(bf,text="⏺  Iniciar grabación",height=32,
                    fg_color=T["err"],hover_color="#e05070",
                    font=(FONT,10,"bold"),text_color="white",
                    command=iniciar).pack(side="left",padx=4)
            else:
                def detener():
                    m=MONITOR.detener(); self._bub_sys(m,"ok"); cerrar(); self._win_procesos()
                ctk.CTkButton(bf,text="⏹  Detener grabación",height=32,
                    fg_color=T["ok"],hover_color="#5abd7a",
                    font=(FONT,10,"bold"),text_color="white",
                    command=detener).pack(side="left",padx=4)

            ctk.CTkLabel(panel,text="Procesos guardados:",font=(FONT,10,"bold"),
                         text_color=T["a2"]).pack(anchor="w",padx=14,pady=(4,2))
            sc=self._sc_modal(panel)
            procs=_rl(F["proc"])
            if not procs:
                ctk.CTkLabel(sc,text="No hay procesos aún.\nInicia una grabación para aprender un proceso.",
                             font=(FONT,11),text_color=T["t2"],justify="center").pack(pady=30)
            else:
                for i,p in enumerate(reversed(procs)):
                    card=ctk.CTkFrame(sc,fg_color=T["card"],corner_radius=8,
                                      border_color=T["br"],border_width=1)
                    card.pack(fill="x",padx=12,pady=4)
                    t=ctk.CTkFrame(card,fg_color="transparent"); t.pack(fill="x",padx=12,pady=(8,2))
                    ctk.CTkLabel(t,text=f"🧠 {p.get('nombre','')}",
                                 font=(FONT,11,"bold"),text_color=T["a2"]).pack(side="left")
                    ctk.CTkLabel(t,text=p.get("fecha",""),
                                 font=(FONT,9),text_color=T["t2"]).pack(side="right")
                    resumen=p.get("resumen","")
                    if resumen:
                        ctk.CTkLabel(card,text=f"  {resumen}",font=(FONT,9),
                                     text_color=T["tx"],wraplength=580).pack(anchor="w",padx=12,pady=(0,4))
                    archs=", ".join(p.get("archivos",[])[:5])
                    if archs:
                        ctk.CTkLabel(card,text=f"  📄 {archs}",font=(FONT,8),
                                     text_color=T["t2"]).pack(anchor="w",padx=12,pady=(0,4))
                    bf2=ctk.CTkFrame(card,fg_color="transparent"); bf2.pack(anchor="e",padx=12,pady=(0,6))
                    def aplicar(pp=p):
                        self._chip_cmd(
                            f"Aplica el proceso '{pp['nombre']}'. "
                            f"Archivos: {', '.join(pp.get('archivos',[]))}. "
                            f"Genera el código o pasos completos.")
                        cerrar()
                    idx_r=len(procs)-1-i
                    def eliminar(ix=idx_r):
                        ps=_rl(F["proc"]); ps.pop(ix); _w(F["proc"],ps)
                        cerrar(); self._win_procesos()
                    ctk.CTkButton(bf2,text="▶ Aplicar",height=26,width=100,
                        fg_color=T["a1"],font=(FONT,9,"bold"),text_color="white",
                        command=aplicar).pack(side="left",padx=3)
                    ctk.CTkButton(bf2,text="🗑",width=30,height=26,
                        fg_color=T["card"],hover_color=T["err"],
                        font=(FONT,10),text_color=T["t2"],command=eliminar).pack(side="left",padx=2)

        self._modal("🧠  Procesos Aprendidos", build, w=720, h=600)


    # ══════════════════════════════════════════════════════════════
    #  CARGAR CARPETA LOCAL DE ANALISTAS (v10)
    # ══════════════════════════════════════════════════════════════
    def _cargar_carpeta_analistas(self):
        """Abre dialogo de carpeta, configura ruta y consolida de inmediato."""
        import threading
        def _pick():
            from tkinter import filedialog
            path = filedialog.askdirectory(
                title="Selecciona carpeta de analistas (OneDrive local)")
            if not path: return
            CFG["finiquitos_ruta"] = path
            guardar_cfg(CFG)
            self.after(0, lambda: self._bub_sys(
                "Carpeta de analistas configurada:\n`" + path + "`\n"
                "Iniciando consolidacion...", "system"))
            self._daemon.sincronizar_finiquitos()
        threading.Thread(target=_pick, daemon=True).start()

    # ══════════════════════════════════════════════════════════════
    #  MATRIZ MADRE — Visor React en browser
    # ══════════════════════════════════════════════════════════════
    def _abrir_matriz_finiquitos(self):
        """Abre la Matriz Madre de Finiquitos en el navegador como HTML interactivo."""
        import threading
        threading.Thread(target=self._abrir_matriz_run, daemon=True).start()

    def _abrir_matriz_run(self):
        import os, json as _j, pathlib as _pl
        madre    = CFG.get("finiquitos_madre","MATRIZ_GENERAL_FINIQUITOS.xlsx").strip()
        hoja     = CFG.get("finiquitos_hoja","Finiquitos").strip() or "Finiquitos"
        ruta_od  = CFG.get("finiquitos_ruta_od","").strip()
        ruta_loc = CFG.get("finiquitos_ruta","").strip()
        tid, cid, sec = _graph_creds()
        usar_graph = bool(tid and cid and sec and ruta_od)
        data, meta = [], {}

        if usar_graph:
            # ── Leer Madre directamente desde OneDrive via Graph API ──
            self.after(0, lambda: self._bub_sys("☁️ Leyendo Matriz Madre desde OneDrive...", "system"))
            try:
                ruta_madre_od = ruta_od.rstrip("/") + "/" + madre
                data = excel_leer_todo(ruta_madre_od, hoja)
                if not data:
                    data = excel_leer_todo(ruta_madre_od, "")  # fallback primera hoja
                if not data:
                    self.after(0, lambda: self._bub_sys(
                        "El archivo Madre existe pero está vacío. Sincroniza primero.", "system"))
                    return
                meta["fuente"] = madre + " (OneDrive)"
                meta["modo"]   = "Graph API"
            except Exception as e:
                self.after(0, lambda err=str(e): self._bub_sys(
                    "Error leyendo Madre desde OneDrive: " + err, "error"))
                return
        else:
            # ── Fallback: leer desde ruta local ──────────────────────
            if ruta_loc:
                maestro_path = str(_pl.Path(ruta_loc) / madre)
            else:
                maestro_path = str(self._daemon.maestro)
            if maestro_path and os.path.exists(maestro_path):
                self.after(0, lambda: self._bub_sys("📊 Generando vista de Matriz Madre...", "system"))
                try:
                    import pandas as _pd
                    df = _pd.read_excel(maestro_path, engine="openpyxl")
                    for _, row in df.iterrows():
                        rec = {}
                        for col in df.columns:
                            val = row[col]
                            if hasattr(val, "item"): val = val.item()
                            if hasattr(val, "isoformat"): val = str(val)[:10]
                            elif val is None or (hasattr(val,"__class__") and val.__class__.__name__ in ("float","float64") and val != val): val = ""
                            rec[col] = val
                        data.append(rec)
                    meta["fuente"] = os.path.basename(maestro_path)
                    meta["modo"]   = "Local"
                except Exception as e:
                    self.after(0, lambda err=str(e): self._bub_sys("Error leyendo Maestro: " + err, "error"))
            else:
                self.after(0, lambda: self._bub_sys(
                    "No encontré el Maestro. Sincroniza primero en ⚙️ Finiquitos.", "system"))
        try:
            import pandas as _pd2
            _df_meta = _pd2.DataFrame(data)
            _analistas = sorted(_df_meta["ANALISTA"].dropna().unique().tolist()) \
                         if "ANALISTA" in _df_meta.columns else []
        except Exception:
            _analistas = []
        meta.update({
            "empresa":    CFG.get("nombre_usuario","Salesland Chile"),
            "generado":   datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "version":    "Remi", "region": "METROPOLITANA / RM",
            "anio":       str(datetime.datetime.now().year),
            "total":      len(data),
            "analistas":  _analistas,
            "n_analistas": len(_analistas),
        })
        html = MATRIZ_HTML_TEMPLATE.replace(
            "__REMI_DATA_PLACEHOLDER__", json.dumps(data, ensure_ascii=False, default=str)
        ).replace(
            "__REMI_META_PLACEHOLDER__", json.dumps(meta, ensure_ascii=False)
        )
        import tempfile
        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix=".html", prefix="remi_matriz_",
            mode="w", encoding="utf-8"
        )
        tmp.write(html); tmp.close()
        webbrowser.open("file:///" + tmp.name.replace("\\", "/"))
        n = len(data)
        self.after(0, lambda: self._bub_sys(
            "Matriz Madre abierta en el navegador. "
            + str(n) + " registros, columnas A:AU activas. "
            + "Busca, filtra y ordena cualquier columna.",
            "ok"
        ))

    # ══════════════════════════════════════════════════════════
    #  VENTANA REGISTRO REMI — historial de archivos generados
    # ══════════════════════════════════════════════════════════
    def _win_usuarios(self):
        """Abre el panel de administración de usuarios (solo admin)."""
        try:
            ses = getattr(self, "_sesion", _SESION_ACTUAL)
            if ses.get("rol") != "admin":
                self._bub_sys("🔒 Acceso restringido — solo administradores.", "warn")
                return
        except Exception:
            pass
        PanelUsuarios(self)

    def _win_registro(self):
        """Ventana de registro: muestra todos los archivos generados por REMI."""
        top = tk.Toplevel(self)
        top.title("🗂️ Registro REMI — Archivos generados")
        top.geometry("820x560")
        top.configure(bg=T["bg"])
        top.resizable(True, True)

        # ── Barra de título ──────────────────────────────────────────────
        hdr = tk.Frame(top, bg=T["sb"], height=44)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="🗂️  Registro de archivos generados por REMI",
                 font=(FONT, 12, "bold"), bg=T["sb"], fg=T["tx"]).pack(
                 side="left", padx=14, pady=10)
        tk.Button(hdr, text="📁 Abrir carpeta Registro", font=(FONT, 9),
                  bg=T["a1"], fg="white", relief="flat", padx=10, pady=4,
                  cursor="hand2", bd=0,
                  command=lambda: REGISTRO.abrir_carpeta()
                  ).pack(side="right", padx=10, pady=8)

        # ── Filtros ──────────────────────────────────────────────────────
        fil_f = tk.Frame(top, bg=T["card"])
        fil_f.pack(fill="x", padx=0)
        tk.Label(fil_f, text="  Filtrar:", font=(FONT, 9),
                 bg=T["card"], fg=T["t2"]).pack(side="left", padx=(12, 4), pady=8)

        _tipo_var = tk.StringVar(value="todos")
        _tipo_map = {
            "Todos":           "todos",
            "📊 Excel":        "excel",
            "📜 Macros":       "macro",
            "🐍 Scripts":      "script",
            "🔒 Backups":      "backup",
            "📄 Informes":     "informe",
            "📋 Finiquitos":   "finiquito",
        }
        btn_refs = {}

        def _cargar(tipo_key="todos"):
            _tipo_var.set(tipo_key)
            # Resaltar botón activo
            for k, b in btn_refs.items():
                b.configure(bg=T["a1"] if k == tipo_key else T["card"],
                            fg="white"  if k == tipo_key else T["t2"])
            # Limpiar lista
            for w in lista_f.winfo_children():
                try: w.destroy()
                except: pass
            entradas = REGISTRO.listar(
                tipo=None if tipo_key == "todos" else tipo_key, n=200)
            if not entradas:
                tk.Label(lista_f, text="  Sin registros para este filtro.",
                         font=(FONT, 10), bg=T["bg"], fg=T["t2"],
                         anchor="w").pack(fill="x", padx=16, pady=20)
                return
            iconos = {"macro":"📜","excel":"📊","script":"🐍",
                      "informe":"📄","backup":"🔒","finiquito":"📋"}
            for e in entradas:
                ic  = iconos.get(e.get("tipo",""), "📁")
                row = tk.Frame(lista_f, bg=T["card"],
                               highlightthickness=1, highlightbackground=T["br"])
                row.pack(fill="x", padx=10, pady=2)
                # Icono + nombre
                info_f = tk.Frame(row, bg=T["card"])
                info_f.pack(side="left", fill="x", expand=True, padx=8, pady=6)
                tk.Label(info_f,
                         text=f"{ic}  {e['archivo']}",
                         font=(FONT, 10, "bold"),
                         bg=T["card"], fg=T["tx"],
                         anchor="w").pack(anchor="w")
                desc_txt = e.get("descripcion","") or ""
                tk.Label(info_f,
                         text=f"   {e['fecha']}  ·  {desc_txt[:60]}",
                         font=(FONT, 8),
                         bg=T["card"], fg=T["t2"],
                         anchor="w").pack(anchor="w")
                # Botones: abrir + abrir carpeta
                btn_f = tk.Frame(row, bg=T["card"])
                btn_f.pack(side="right", padx=6, pady=4)
                ruta = e.get("ruta","")
                def _abrir(r=ruta):
                    try:
                        os.startfile(r)
                    except Exception:
                        try:
                            import subprocess as _sp
                            _sp.Popen(f'explorer /select,"{r}"',
                                      creationflags=getattr(_sp,"CREATE_NO_WINDOW",0))
                        except Exception:
                            pass
                def _abrir_dir(r=ruta):
                    try:
                        import subprocess as _sp
                        _sp.Popen(f'explorer "{os.path.dirname(r)}"',
                                  creationflags=getattr(_sp,"CREATE_NO_WINDOW",0))
                    except Exception:
                        pass
                tk.Button(btn_f, text="Abrir", font=(FONT, 8),
                          bg=T["a1"], fg="white", relief="flat",
                          padx=8, pady=3, cursor="hand2", bd=0,
                          command=_abrir).pack(side="left", padx=2)
                tk.Button(btn_f, text="Carpeta", font=(FONT, 8),
                          bg=T["br"], fg=T["t2"], relief="flat",
                          padx=8, pady=3, cursor="hand2", bd=0,
                          command=_abrir_dir).pack(side="left", padx=2)

        for label, key in _tipo_map.items():
            b = tk.Button(fil_f, text=label, font=(FONT, 8),
                          bg=T["card"], fg=T["t2"],
                          relief="flat", padx=8, pady=4,
                          cursor="hand2", bd=0,
                          command=lambda k=key: _cargar(k))
            b.pack(side="left", padx=2, pady=4)
            btn_refs[key] = b

        # ── Separador ────────────────────────────────────────────────────
        tk.Frame(top, bg=T["br"], height=1).pack(fill="x")

        # ── Lista scrollable ─────────────────────────────────────────────
        sc = ctk.CTkScrollableFrame(top, fg_color=T["bg"],
                                     scrollbar_button_color=T["a1"],
                                     corner_radius=0)
        sc.pack(fill="both", expand=True)
        lista_f = sc

        # ── Pie con ruta de la carpeta ───────────────────────────────────
        pie_f = tk.Frame(top, bg=T["sb"])
        pie_f.pack(fill="x")
        tk.Label(pie_f,
                 text=f"  📁 {REGISTRO._raiz}",
                 font=(FMONO, 8), bg=T["sb"], fg=T["t2"],
                 anchor="w").pack(side="left", padx=8, pady=6)
        tk.Button(pie_f, text="Actualizar ↺", font=(FONT, 8),
                  bg=T["card"], fg=T["t2"], relief="flat",
                  padx=8, pady=3, cursor="hand2", bd=0,
                  command=lambda: _cargar(_tipo_var.get())).pack(
                  side="right", padx=10, pady=4)

        _cargar("todos")


class StickyNote(tk.Toplevel):
    def __init__(self, master, texto):
        super().__init__(master)
        self.overrideredirect(True); self.attributes("-topmost",True)
        self.configure(bg=T["card"]); self.geometry("260x170+130+130")
        self._dx = self._dy = 0
        bar=tk.Frame(self,bg=T["a1"],height=26); bar.pack(fill="x")
        tk.Label(bar,text="  📝 Nota  (doble clic = cerrar)",
                 font=(FONT,8),bg=T["a1"],fg="white").pack(side="left")
        bar.bind("<ButtonPress-1>",   lambda e:(setattr(self,"_dx",e.x),setattr(self,"_dy",e.y)))
        bar.bind("<B1-Motion>",       lambda e:self.geometry(
            f"+{self.winfo_x()+e.x-self._dx}+{self.winfo_y()+e.y-self._dy}"))
        bar.bind("<Double-Button-1>", lambda _: self.destroy())
        tk.Frame(self,bg=T["a1"],height=1).pack(fill="x")
        t=tk.Text(self,font=(FONT,10),bg=T["card"],fg=T["tx"],
                  wrap="word",relief="flat",padx=10,pady=8,bd=0)
        t.insert("1.0",texto); t.pack(fill="both",expand=True)




MATRIZ_HTML_TEMPLATE = '<!DOCTYPE html>\n<html lang="es">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>Remi — Control Finiquitos</title>\n<script src="https://cdnjs.cloudflare.com/ajax/libs/react/18.2.0/umd/react.production.min.js"></script>\n<script src="https://cdnjs.cloudflare.com/ajax/libs/react-dom/18.2.0/umd/react-dom.production.min.js"></script>\n<script src="https://cdnjs.cloudflare.com/ajax/libs/babel-standalone/7.23.2/babel.min.js"></script>\n<script src="https://cdn.tailwindcss.com"></script>\n<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;700&family=IBM+Plex+Sans:wght@400;600;700;900&display=swap" rel="stylesheet">\n<style>\n  * { box-sizing: border-box; }\n  body { margin: 0; font-family: \'IBM Plex Sans\', sans-serif; background: #f0f3f7; }\n  ::-webkit-scrollbar { width: 5px; height: 5px; }\n  ::-webkit-scrollbar-track { background: #e2e8f0; }\n  ::-webkit-scrollbar-thumb { background: #94a3b8; border-radius: 3px; }\n  .mono { font-family: \'IBM Plex Mono\', monospace; }\n  .tab-active  { border-bottom: 3px solid #107c41; color: #107c41; background: white; }\n  .tab-passive { border-bottom: 3px solid transparent; color: #64748b; background: #f8fafc; }\n  .tab-passive:hover { background: white; color: #374151; }\n  .truncate { overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }\n  @keyframes pulse-green { 0%,100%{opacity:1} 50%{opacity:0.4} }\n  .live-dot { animation: pulse-green 2s infinite; }\n  @keyframes slide-in { from{opacity:0;transform:translateY(4px)} to{opacity:1;transform:none} }\n  .slide-in { animation: slide-in 0.25s ease-out; }\n</style>\n</head>\n<body>\n<div id="root"></div>\n<script>\n  window.__REMI_DATA__ = __REMI_DATA_PLACEHOLDER__;\n  window.__REMI_META__ = __REMI_META_PLACEHOLDER__;\n</script>\n<script type="text/babel">\nconst { useState, useMemo, useCallback } = React;\n\n// ─── Formato moneda CLP ────────────────────────────────────────────────\nconst f = (n) => {\n  if (n === null || n === undefined || n === \'\') return \'—\';\n  const num = typeof n === \'number\' ? n : parseFloat(String(n).replace(/[^0-9.-]/g,\'\'));\n  if (isNaN(num)) return String(n);\n  return new Intl.NumberFormat(\'es-CL\').format(Math.round(num));\n};\nconst toNum = (v) => {\n  const n = parseFloat(String(v ?? \'\').replace(/[^0-9.-]/g,\'\'));\n  return isNaN(n) ? 0 : n;\n};\n\n// ─── Helpers UI ────────────────────────────────────────────────────────\nconst Pill = ({ ok, yes=\'SI\', no=\'NO\', val }) => {\n  const isOk = String(val).toUpperCase().trim() === yes.toUpperCase();\n  return (\n    <span className={`px-2 py-0.5 rounded text-[9px] font-bold mono\n      ${isOk ? \'bg-emerald-100 text-emerald-700\' : \'bg-rose-100 text-rose-600\'}`}>\n      {val || no}\n    </span>\n  );\n};\n\nconst AuditBadge = ({ legalizado, finanzas }) => {\n  const ok = String(legalizado).toUpperCase().trim() === \'SI\'\n          && String(finanzas).toUpperCase().trim() === \'SI\';\n  return ok\n    ? <span className="inline-flex items-center gap-1 text-[9px] font-bold text-emerald-600">\n        <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2.5"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/></svg>OK\n      </span>\n    : <span className="inline-flex items-center gap-1 text-[9px] font-bold text-amber-500">\n        <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2.5"><circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="12"/><line x1="12" y1="16" x2="12.01" y2="16"/></svg>PENDIENTE\n      </span>;\n};\n\n// ─── Cabecera ordenable para tabla A:AU ───────────────────────────────\nconst Th = ({ col: c, label, cls=\'\', sortCol, setSortCol, sortDir, setSortDir }) => (\n  <th className={`px-2 py-2.5 border-r border-slate-600 cursor-pointer select-none whitespace-nowrap text-[9px] ${cls}`}\n      onClick={() => { setSortCol(c); setSortDir(p => c === sortCol && p === \'asc\' ? \'desc\' : \'asc\'); }}>\n    {label}{sortCol === c ? (sortDir === \'asc\' ? \' ▲\' : \' ▼\') : \'\'}\n  </th>\n);\n\n// ─── Mapeador flexible de columnas ────────────────────────────────────\nconst col = (row, keys) => {\n  for (const k of keys) if (row[k] !== undefined && row[k] !== null && row[k] !== \'\') return row[k];\n  return \'\';\n};\n\n// ══════════════════════════════════════════════════════════════════════\n//  VISTA 1 — REMI SYNC (Motor de cruce por analista)\n// ══════════════════════════════════════════════════════════════════════\nconst RemiSync = ({ data, meta }) => {\n  const [search, setSearch]     = useState(\'\');\n  const [analista, setAnalista] = useState(\'TODOS\');\n  const [sortCol, setSortCol]   = useState(null);\n  const [sortDir, setSortDir]   = useState(\'asc\');\n\n  // Normalizar registros para esta vista\n  const rows = useMemo(() => data.map(r => ({\n    cc:         col(r, [\'colG_CC\',\'CC\',\'Centro de Costo\',\'cc\']),\n    analista:   col(r, [\'colD_Analista\',\'Analista\',\'ANALISTA\',\'Control Analista\',\'analista\']),\n    rut_pc:     col(r, [\'RUT_LIMP\',\'colAM_RutPC\',\'RUT_PC\',\'Rut_PC\',\'rut_pc\',\'RUT PC\']),\n    nombre:     col(r, [\'colI_Nombre\',\'Nombre\',\'NOMBRE\',\'nombre\']),\n    total:      toNum(col(r, [\'colAJ_Total\',\'Total_Liquido\',\'Total Líquido\',\'Liquido\',\'total\',\'liquido\'])),\n    haberes:    toNum(col(r, [\'colAC_TotalHaberes\',\'Total_Haberes\',\'Total Haberes\',\'haberes\'])),\n    legalizado: col(r, [\'colAR_Legalizado\',\'Legalizado\',\'LEGALIZADO\',\'legalizado\']),\n    finanzas:   col(r, [\'colAS_Finanzas\',\'Finanzas\',\'FINANZAS\',\'finanzas\']),\n    estado:     col(r, [\'colA_Estado\',\'Estado\',\'ESTADO\',\'estado\']),\n    ctrl:       col(r, [\'SELLO_AUDITORIA\',\'REGISTRO_CONCATENADO\',\'Control Analista\',\'ID_REMI\']),\n    _r: r,\n  })), [data]);\n\n  const analistas = useMemo(() =>\n    [\'TODOS\', ...Array.from(new Set(rows.map(r => r.analista).filter(Boolean))).sort()],\n    [rows]);\n\n  const filtered = useMemo(() => {\n    const s = search.toLowerCase();\n    return rows.filter(r =>\n      (analista === \'TODOS\' || r.analista.toUpperCase() === analista.toUpperCase()) &&\n      (!s || r.nombre.toLowerCase().includes(s) || String(r.rut_pc).includes(s) || r.cc.toLowerCase().includes(s))\n    );\n  }, [rows, search, analista]);\n\n  const sorted = useMemo(() => {\n    if (!sortCol) return filtered;\n    return [...filtered].sort((a, b) => {\n      const va = a[sortCol], vb = b[sortCol];\n      const cmp = (typeof va === \'number\' && typeof vb === \'number\')\n        ? va - vb : String(va).localeCompare(String(vb));\n      return sortDir === \'asc\' ? cmp : -cmp;\n    });\n  }, [filtered, sortCol, sortDir]);\n\n  const totalLiq  = sorted.reduce((a, r) => a + (r.total   || 0), 0);\n  const totalHab  = sorted.reduce((a, r) => a + (r.haberes || 0), 0);\n  const pendientes = sorted.filter(r =>\n    String(r.legalizado || \'\').toUpperCase() !== \'SI\' || String(r.finanzas || \'\').toUpperCase() !== \'SI\').length;\n  const listos     = sorted.length - pendientes;\n\n  const SortTh = ({ col: c, label, cls=\'\' }) => (\n    <th className={`p-3 text-left cursor-pointer select-none whitespace-nowrap ${cls}`}\n        onClick={() => { setSortCol(c); setSortDir(p => c === sortCol && p === \'asc\' ? \'desc\' : \'asc\'); }}>\n      {label}{sortCol === c ? (sortDir === \'asc\' ? \' ▲\' : \' ▼\') : \'\'}\n    </th>\n  );\n\n  return (\n    <div className="slide-in flex flex-col h-full">\n      {/* Header SYNC */}\n      <div className="bg-[#107c41] px-6 py-4 flex justify-between items-end shrink-0">\n        <div>\n          <div className="flex items-center gap-2">\n            <div className="live-dot w-2 h-2 rounded-full bg-white opacity-80"></div>\n            <h2 className="text-xl font-black italic tracking-tight text-white">REMI SYNC v21</h2>\n          </div>\n          <p className="text-[10px] font-semibold opacity-70 uppercase tracking-widest text-white mt-0.5">\n            Extracción directa · OneDrive Excel 365\n          </p>\n        </div>\n        <div className="flex gap-6 text-right text-white">\n          <div>\n            <p className="text-[8px] font-bold opacity-60 uppercase">Total Haberes</p>\n            <p className="text-lg font-black mono">${f(totalHab)}</p>\n          </div>\n          <div>\n            <p className="text-[8px] font-bold opacity-60 uppercase">Total Líquido</p>\n            <p className="text-xl font-black mono">${f(totalLiq)}</p>\n          </div>\n          <div>\n            <p className="text-[8px] font-bold opacity-60 uppercase">Auditados</p>\n            <p className="text-xl font-black">{listos}<span className="text-sm opacity-60">/{sorted.length}</span></p>\n          </div>\n        </div>\n      </div>\n\n      {/* Filtros */}\n      <div className="bg-white border-b border-slate-200 px-4 py-2.5 flex gap-3 items-center shrink-0">\n        <div className="relative flex-1 max-w-xs">\n          <svg className="absolute left-3 top-1/2 -translate-y-1/2 text-slate-400" width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>\n          <input type="text" placeholder="Nombre, RUT PC, CC…"\n            className="pl-9 pr-3 py-1.5 bg-slate-50 border border-slate-200 rounded-lg text-[11px] w-full outline-none focus:ring-2 focus:ring-emerald-400/30"\n            onChange={e => setSearch(e.target.value)} />\n        </div>\n        <select className="bg-slate-100 border border-slate-200 rounded-lg px-3 py-1.5 text-[10px] font-bold uppercase text-slate-600 outline-none cursor-pointer"\n          onChange={e => setAnalista(e.target.value)}>\n          {analistas.map(a => <option key={a} value={a}>{a === \'TODOS\' ? \'Todos los analistas\' : a}</option>)}\n        </select>\n        {pendientes > 0 && (\n          <span className="ml-auto text-[10px] font-bold text-amber-600 bg-amber-50 px-3 py-1.5 rounded-lg border border-amber-200">\n            ⚠ {pendientes} pendiente{pendientes !== 1 ? \'s\' : \'\'}\n          </span>\n        )}\n        <span className="text-[10px] font-bold text-slate-400 ml-auto">\n          {analistas.length - 1} analistas · {sorted.length} registros\n        </span>\n      </div>\n\n      {/* Tabla */}\n      <div className="flex-1 overflow-auto">\n        <table className="w-full text-left">\n          <thead className="sticky top-0 bg-slate-800 text-white text-[9px] font-black uppercase z-10">\n            <tr>\n              <SortTh col="analista" label="Analista (D)" />\n              <SortTh col="cc"       label="CC (G)" />\n              <SortTh col="rut_pc"   label="RUT PC (AM)" cls="mono" />\n              <SortTh col="nombre"   label="Nombre (I)" />\n              <SortTh col="haberes"  label="Haberes (AC)" cls="text-right" />\n              <SortTh col="total"    label="Líquido (AJ)" cls="text-right" />\n              <th className="p-3 text-center">Legalizado (AR)</th>\n              <th className="p-3 text-center">Finanzas (AS)</th>\n              <th className="p-3 text-center">Sello Auditoría</th>\n              <th className="p-3">Control Remi</th>\n            </tr>\n          </thead>\n          <tbody className="text-[11px]">\n            {sorted.length === 0 && (\n              <tr><td colSpan="10" className="text-center py-20 text-slate-400">\n                Sin registros. Ajusta filtros o sincroniza desde Remi.\n              </td></tr>\n            )}\n            {sorted.map((row, i) => (\n              <tr key={i} className={`border-b border-slate-100 hover:bg-emerald-50/40 transition-colors\n                ${i % 2 === 0 ? \'\' : \'bg-slate-50/50\'}`}>\n                <td className="p-3 font-bold text-slate-500 truncate max-w-[9rem]">{row.analista}</td>\n                <td className="p-3 text-slate-500 text-[10px] truncate max-w-[8rem]">{row.cc}</td>\n                <td className="p-3 mono font-bold text-slate-400 text-[10px]">{row.rut_pc}</td>\n                <td className="p-3 font-black text-slate-800 uppercase truncate max-w-[14rem]">{row.nombre}</td>\n                <td className="p-3 text-right mono text-slate-600">${f(row.haberes||0)}</td>\n                <td className="p-3 text-right mono font-black text-slate-800 bg-slate-50">${f(row.total||0)}</td>\n                <td className="p-3 text-center"><Pill val={row.legalizado} /></td>\n                <td className="p-3 text-center"><Pill val={row.finanzas} /></td>\n                <td className="p-3 text-center"><AuditBadge legalizado={row.legalizado} finanzas={row.finanzas} /></td>\n                <td className="p-3 text-[9px] mono text-slate-400 truncate max-w-[10rem]">\n                  {row.ctrl || (String(row.analista).split(\' \')[0] + \'-\' + new Date().toLocaleDateString(\'es-CL\'))}\n                </td>\n              </tr>\n            ))}\n          </tbody>\n        </table>\n      </div>\n\n      {/* Footer SYNC */}\n      <div className="bg-white border-t border-slate-200 px-6 py-1.5 flex justify-between items-center text-[9px] font-bold text-slate-400 uppercase shrink-0">\n        <div className="flex gap-5 items-center">\n          <span className="flex items-center gap-1 text-emerald-600">\n            <svg width="10" height="10" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2.5"><path d="M10 13a5 5 0 0 0 7.54.54l3-3a5 5 0 0 0-7.07-7.07l-1.72 1.71"/><path d="M14 11a5 5 0 0 0-7.54-.54l-3 3a5 5 0 0 0 7.07 7.07l1.71-1.71"/></svg>\n            SharePoint · OneDrive 365\n          </span>\n          <span>{analistas.length - 1} analistas activos</span>\n          <span className="text-slate-300">·</span>\n          <span>Generado: {meta.generado || \'—\'}</span>\n        </div>\n        <span className="text-[8px] italic text-slate-300">REMI ENGINE CHILE · v21</span>\n      </div>\n    </div>\n  );\n};\n\n// ══════════════════════════════════════════════════════════════════════\n//  VISTA 2 — MATRIZ A:AU (Auditoría completa)\n// ══════════════════════════════════════════════════════════════════════\nconst MatrizAU = ({ data, meta }) => {\n  const [search, setSearch]             = useState(\'\');\n  const [filterAnalyst, setFilterAnalyst] = useState(\'Todos\');\n  const [filterEstado, setFilterEstado]   = useState(\'Todos\');\n  const [sortCol, setSortCol]             = useState(null);\n  const [sortDir, setSortDir]             = useState(\'asc\');\n\n  const analysts = useMemo(() => {\n    const s = new Set(data.map(r => col(r, [\'ANALISTA\',\'colD_Analista\',\'Analista\',\'Control Analista\']) || \'\').filter(Boolean));\n    return [\'Todos\', ...Array.from(s).sort()];\n  }, [data]);\n\n  const rows = useMemo(() => data.map(r => ({\n    Estado:      col(r, [\'colA_Estado\',\'Estado\',\'ESTADO\']),\n    Analista:    col(r, [\'ANALISTA\',\'colD_Analista\',\'Analista\',\'Control Analista\']),\n    CC:          col(r, [\'colG_CC\',\'CC\',\'Centro de Costo\']),\n    Ficha:       col(r, [\'colH_Ficha\',\'Ficha\',\'FICHA\']),\n    Nombre:      col(r, [\'colI_Nombre\',\'Nombre\',\'NOMBRE\']),\n    RUT:         col(r, [\'colK_Rut\',\'RUT\',\'Rut\',\'rut\']),\n    FIngreso:    col(r, [\'colL_FIngreso\',\'Fecha Ingreso\',\'F_Ingreso\']),\n    FTermino:    col(r, [\'colM_FTermino\',\'Fecha Termino\',\'F_Termino\']),\n    DiasVac:     col(r, [\'colX_DiasVac\',\'Dias_Vac\']),\n    Vacaciones:  toNum(col(r, [\'colZ_Vacaciones\',\'Vacaciones\'])),\n    Aviso:       toNum(col(r, [\'colAA_Aviso\',\'Aviso Previo\',\'Aviso\'])),\n    TotHaberes:  toNum(col(r, [\'colAC_TotalHaberes\',\'Total_Haberes\',\'Total Haberes\'])),\n    AFC:         toNum(col(r, [\'colAG_AFC\',\'AFC\'])),\n    OtrosDctos:  toNum(col(r, [\'colAH_Otros\',\'Otros_Dctos\'])),\n    TotDctos:    toNum(col(r, [\'colAI_TotalDctos\',\'Total_Dctos\',\'Total Descuentos\'])),\n    Liquido:     toNum(col(r, [\'colAJ_Total\',\'Total_Liquido\',\'Total Líquido\',\'Liquido\'])),\n    Pagado:      col(r, [\'colAK_Pagado\',\'Pagado\']),\n    RutPC:       col(r, [\'RUT_LIMP\',\'colAM_RutPC\',\'Rut_PC\',\'RUT_PC\']),\n    PosiblePago: col(r, [\'colAQ_PosiblePago\',\'Posible_Pago\']),\n    Legalizado:  col(r, [\'colAR_Legalizado\',\'Legalizado\']),\n    Finanzas:    col(r, [\'colAS_Finanzas\',\'Finanzas\']),\n    Obs:         col(r, [\'colAU_Obs\',\'Obs\',\'Observacion\']),\n    ControlRemi: col(r, [\'SELLO_AUDITORIA\',\'REGISTRO_CONCATENADO\',\'Control Analista\',\'ID_REMI\']),\n  })), [data]);\n\n  const filtered = useMemo(() => {\n    const s = search.toLowerCase();\n    return rows.filter(r =>\n      (filterAnalyst === \'Todos\' || r.Analista.toUpperCase() === filterAnalyst.toUpperCase()) &&\n      (filterEstado  === \'Todos\' || r.Estado.toUpperCase()   === filterEstado.toUpperCase()) &&\n      (!s || r.Nombre.toLowerCase().includes(s) || String(r.RUT).toLowerCase().includes(s) || String(r.RutPC).includes(s))\n    );\n  }, [rows, search, filterAnalyst, filterEstado]);\n\n  const sorted = useMemo(() => {\n    if (!sortCol) return filtered;\n    return [...filtered].sort((a, b) => {\n      const va = a[sortCol], vb = b[sortCol];\n      const cmp = (typeof va === \'number\' && typeof vb === \'number\') ? va - vb : String(va).localeCompare(String(vb));\n      return sortDir === \'asc\' ? cmp : -cmp;\n    });\n  }, [filtered, sortCol, sortDir]);\n\n  const totalLiq  = sorted.reduce((a, r) => a + (r.Liquido  || 0), 0);\n  const totalHab  = sorted.reduce((a, r) => a + (r.TotHaberes || 0), 0);\n  const pendientes = sorted.filter(r => String(r.Estado || \'\').toUpperCase().includes(\'PENDIENTE\')).length;\n\n  const thProps = { sortCol, setSortCol, sortDir, setSortDir };\n\n  return (\n    <div className="slide-in flex flex-col h-full">\n      {/* Header A:AU */}\n      <div className="bg-white border-b border-slate-200 px-6 py-2.5 flex items-center justify-between shrink-0">\n        <div className="flex items-center gap-3">\n          <div className="bg-[#107c41] p-1.5 rounded">\n            <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="white" strokeWidth="2"><rect x="3" y="3" width="18" height="18" rx="2"/><path d="M3 9h18M3 15h18M9 3v18"/></svg>\n          </div>\n          <div>\n            <h2 className="text-sm font-black uppercase tracking-tight text-slate-800">\n              Matriz Madre {meta.empresa || \'Salesland\'} {meta.anio || \'\'}\n            </h2>\n            <p className="text-[8px] font-bold uppercase tracking-widest text-slate-400">\n              Auditoría columnas A:AU · {sorted.length} registros visibles\n            </p>\n          </div>\n        </div>\n        <div className="flex gap-4 text-[10px] font-black">\n          <div className="bg-emerald-50 text-emerald-700 px-3 py-1.5 rounded-lg border border-emerald-200">\n            T.Haberes: <span className="mono">${f(totalHab)}</span>\n          </div>\n          <div className="bg-blue-50 text-blue-700 px-3 py-1.5 rounded-lg border border-blue-200">\n            T.Líquido: <span className="mono">${f(totalLiq)}</span>\n          </div>\n          {pendientes > 0 && (\n            <div className="bg-amber-50 text-amber-600 px-3 py-1.5 rounded-lg border border-amber-200">\n              ⚠ {pendientes} pendientes\n            </div>\n          )}\n        </div>\n      </div>\n\n      {/* Filtros */}\n      <div className="bg-white/90 border-b border-slate-200 px-4 py-2 flex items-center gap-3 shrink-0 overflow-x-auto">\n        <div className="relative min-w-[240px]">\n          <svg className="absolute left-3 top-1/2 -translate-y-1/2 text-slate-400" width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>\n          <input type="text" placeholder="Nombre, RUT o RUT PC…"\n            className="pl-9 pr-3 py-1.5 bg-slate-50 border border-slate-200 rounded-lg text-[11px] w-full outline-none focus:ring-2 focus:ring-emerald-400/30"\n            onChange={e => setSearch(e.target.value)} />\n        </div>\n        <select className="bg-slate-100 border border-slate-200 rounded-lg px-3 py-1.5 text-[10px] font-bold uppercase text-slate-600 outline-none cursor-pointer"\n          onChange={e => setFilterAnalyst(e.target.value)}>\n          {analysts.map(a => <option key={a} value={a}>{a === \'Todos\' ? \'Todos los analistas\' : a}</option>)}\n        </select>\n        <select className="bg-slate-100 border border-slate-200 rounded-lg px-3 py-1.5 text-[10px] font-bold uppercase text-slate-600 outline-none cursor-pointer"\n          onChange={e => setFilterEstado(e.target.value)}>\n          {[\'Todos\',\'FINIQUITO NORMAL\',\'PENDIENTE\'].map(e =>\n            <option key={e} value={e}>{e === \'Todos\' ? \'Todos los estados\' : e}</option>)}\n        </select>\n      </div>\n\n      {/* Tabla A:AU */}\n      <div className="flex-1 overflow-auto bg-slate-100">\n        <div className="inline-block min-w-full">\n          <table className="border-separate border-spacing-0 min-w-full">\n            <thead className="sticky top-0 z-30">\n              <tr className="bg-[#1e293b] text-white uppercase font-black text-center tracking-tight">\n                <th className="sticky left-0 z-40 bg-[#0f172a] px-3 py-2.5 border-r border-slate-600 w-28 text-[9px]">Estado (A)</th>\n                <th className="sticky left-28 z-40 bg-[#0f172a] px-4 py-2.5 border-r border-slate-600 w-36 text-[9px]">Analista (D)</th>\n                <Th col="CC"       label="CC (G)"           cls="w-44 bg-slate-800" {...thProps} />\n                <Th col="Ficha"    label="Ficha (H)"        cls="w-20 bg-slate-800" {...thProps} />\n                <Th col="Nombre"   label="Nombre (I)"       cls="w-64 text-left bg-slate-800" {...thProps} />\n                <Th col="RUT"      label="RUT (K)"          cls="w-32 bg-slate-800" {...thProps} />\n                <Th col="FIngreso" label="F.Inicio (L)"     cls="w-28 bg-slate-800" {...thProps} />\n                <Th col="FTermino" label="F.Término (M)"    cls="w-28 bg-slate-800" {...thProps} />\n                <Th col="DiasVac"  label="Días Vac (X)"     cls="w-20 bg-slate-700" {...thProps} />\n                <Th col="Vacaciones" label="Vacac $ (Z)"    cls="w-32 bg-slate-700 text-right" {...thProps} />\n                <Th col="Aviso"    label="Aviso (AA)"       cls="w-28 bg-slate-700 text-right" {...thProps} />\n                <Th col="TotHaberes" label="T.Haberes (AC)" cls="w-36 bg-emerald-900 text-right" {...thProps} />\n                <Th col="AFC"      label="AFC (AG)"         cls="w-28 bg-rose-900 text-right" {...thProps} />\n                <Th col="OtrosDctos" label="Otros (AH)"     cls="w-28 bg-rose-900 text-right" {...thProps} />\n                <Th col="TotDctos" label="T.Dctos (AI)"     cls="w-32 bg-rose-950 text-right" {...thProps} />\n                <Th col="Liquido"  label="LÍQUIDO (AJ)"     cls="w-40 bg-blue-900 text-right text-[10px] underline" {...thProps} />\n                <Th col="Pagado"   label="Pagado (AK)"      cls="w-20 bg-slate-800" {...thProps} />\n                <Th col="RutPC"    label="Rut PC (AM)"      cls="w-32 bg-slate-800" {...thProps} />\n                <Th col="PosiblePago" label="Pos Pago (AQ)" cls="w-28 bg-slate-800" {...thProps} />\n                <Th col="Legalizado" label="Legal (AR)"     cls="w-24 bg-slate-800" {...thProps} />\n                <Th col="Finanzas"   label="Finanzas (AS)"  cls="w-24 bg-slate-800" {...thProps} />\n                <th className="px-4 py-2.5 border-r border-slate-600 w-48 bg-slate-800 text-left text-[9px]">Obs (AU)</th>\n                <th className="sticky right-0 z-40 bg-[#1d4ed8] px-3 py-2.5 border-l border-blue-400 w-36 text-[9px]">Sello Remi</th>\n              </tr>\n            </thead>\n            <tbody className="bg-white text-[10px]">\n              {sorted.length === 0 && (\n                <tr><td colSpan="23" className="text-center py-16 text-slate-400 text-sm">\n                  Sin registros. Ajusta los filtros o sincroniza desde Remi.\n                </td></tr>\n              )}\n              {sorted.map((row, idx) => {\n                const isPend = String(row.Estado).toUpperCase().includes(\'PENDIENTE\');\n                return (\n                  <tr key={idx} className="hover:bg-emerald-50/40 border-b border-slate-100 group">\n                    <td className="sticky left-0 z-20 bg-white px-2 py-1.5 border-r border-slate-100 text-center group-hover:bg-emerald-50/40">\n                      <span className={`px-1.5 py-0.5 rounded text-[8px] font-bold ${isPend ? \'bg-amber-100 text-amber-700\' : \'bg-emerald-100 text-emerald-700\'}`}>\n                        {isPend ? \'🏳\' : \'✔\'} {row.Estado || \'NORMAL\'}\n                      </span>\n                    </td>\n                    <td className="sticky left-28 z-20 bg-white px-3 py-1.5 border-r border-slate-100 font-black text-slate-700 uppercase group-hover:bg-emerald-50/40 truncate max-w-[9rem]">\n                      {row.Analista}\n                    </td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-slate-500 truncate uppercase text-[9px]">{row.CC}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center mono text-slate-400">{row.Ficha}</td>\n                    <td className="px-4 py-1.5 border-r border-slate-100 font-black text-slate-900 uppercase truncate max-w-[16rem]">{row.Nombre}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 mono text-slate-600">{row.RUT}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center text-slate-400">{row.FIngreso}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center font-bold text-slate-500">{row.FTermino}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center bg-slate-50">{row.DiasVac}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-right bg-slate-50 mono">${f(row.Vacaciones)}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-right bg-slate-50 mono text-amber-600 font-bold">${f(row.Aviso)}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-right bg-emerald-50 font-black text-emerald-800 mono">${f(row.TotHaberes)}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-right text-rose-600 mono">${f(row.AFC)}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-right text-rose-600 mono">${f(row.OtrosDctos)}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 text-right bg-rose-50 font-black text-rose-900 mono">${f(row.TotDctos)}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-200 text-right bg-blue-50 font-black text-blue-900 text-[11px] mono">${f(row.Liquido)}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center font-bold text-slate-700">{row.Pagado}</td>\n                    <td className="px-3 py-1.5 border-r border-slate-100 mono text-[9px] text-slate-400">{row.RutPC}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center font-bold text-blue-600">{row.PosiblePago}</td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center"><Pill val={row.Legalizado} /></td>\n                    <td className="px-2 py-1.5 border-r border-slate-100 text-center"><Pill val={row.Finanzas} /></td>\n                    <td className="px-4 py-1.5 border-r border-slate-100 text-slate-400 italic truncate max-w-[12rem]">{row.Obs}</td>\n                    <td className="sticky right-0 z-20 bg-blue-50 px-3 py-1.5 border-l border-blue-200 mono font-bold text-[9px] text-blue-700">\n                      <div className="flex items-center gap-1">\n                        <svg width="10" height="10" viewBox="0 0 24 24" fill="none" stroke="#3b82f6" strokeWidth="2.5"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/></svg>\n                        {row.ControlRemi || (String(row.Analista).split(\' \')[0])}\n                      </div>\n                    </td>\n                  </tr>\n                );\n              })}\n            </tbody>\n          </table>\n        </div>\n      </div>\n\n      {/* Footer A:AU */}\n      <div className="bg-white border-t border-slate-200 px-6 py-1.5 flex justify-between items-center text-[9px] font-bold text-slate-400 uppercase shrink-0">\n        <div className="flex gap-4">\n          <span>🏢 {meta.empresa || \'SALESLAND CHILE\'}</span>\n          <span className="text-slate-200">|</span>\n          <span className="text-emerald-600">Columnas A:AU</span>\n          <span className="text-slate-200">|</span>\n          <span>{data.length} registros totales</span>\n          <span className="text-slate-200">|</span>\n          <span className="text-blue-500 italic">Pequeño Remi — Sistema de Auditoría</span>\n        </div>\n        <span>📍 {meta.region || \'RM\'}</span>\n      </div>\n    </div>\n  );\n};\n\n// ══════════════════════════════════════════════════════════════════════\n//  APP RAÍZ — Pestañas\n// ══════════════════════════════════════════════════════════════════════\nconst App = () => {\n  const data = window.__REMI_DATA__ || [];\n  const meta = window.__REMI_META__ || {};\n  const [tab, setTab] = useState(\'sync\');\n\n  return (\n    <div style={{display:\'flex\',flexDirection:\'column\',height:\'100vh\',background:\'#f0f3f7\'}}\n         className="border-t-4 border-[#107c41]">\n\n      {/* Barra de navegación */}\n      <div className="bg-white border-b border-slate-200 flex items-stretch shrink-0 shadow-sm z-50">\n        <div className="flex items-center gap-3 px-5 border-r border-slate-200">\n          <div className="bg-[#107c41] p-1.5 rounded">\n            <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="white" strokeWidth="2.5"><path d="M22 12h-4l-3 9L9 3l-3 9H2"/></svg>\n          </div>\n          <span className="text-[11px] font-black uppercase tracking-tight text-slate-800">Remi Control</span>\n        </div>\n        <button onClick={() => setTab(\'sync\')}\n          className={`px-6 py-3 text-[11px] font-black uppercase tracking-wide transition-all ${tab===\'sync\' ? \'tab-active\' : \'tab-passive\'}`}>\n          ⚡ SYNC v21\n        </button>\n        <button onClick={() => setTab(\'matriz\')}\n          className={`px-6 py-3 text-[11px] font-black uppercase tracking-wide transition-all ${tab===\'matriz\' ? \'tab-active\' : \'tab-passive\'}`}>\n          📊 Matriz A:AU\n        </button>\n        <div className="ml-auto flex items-center gap-3 px-5 text-[9px] font-bold text-slate-400">\n          <span className="live-dot inline-block w-1.5 h-1.5 rounded-full bg-emerald-500"></span>\n          <span>OneDrive 365{(meta.n_analistas || 0) > 0 ? \' · \' + meta.n_analistas + \' analistas\' : \'\'}{meta.modo ? \' · \' + meta.modo : \'\'}</span>\n          <span className="text-slate-300">·</span>\n          <span>{data.length} registros</span>\n          <span className="text-slate-300">·</span>\n          <span>{meta.generado || \'\'}</span>\n        </div>\n      </div>\n\n      {/* Contenido de pestaña */}\n      <div style={{flex:1,minHeight:0,overflow:\'hidden\',display:\'flex\',flexDirection:\'column\'}}>\n        {tab === \'sync\'   && <RemiSync   data={data} meta={meta} />}\n        {tab === \'matriz\' && <MatrizAU   data={data} meta={meta} />}\n      </div>\n    </div>\n  );\n};\n\nReactDOM.createRoot(document.getElementById(\'root\')).render(<App />);\n</script>\n</body>\n</html>'


# ══════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════
#  SISTEMA DE AUTENTICACIÓN REMI
#  · Usuarios en BASE/usuarios.json (hash SHA-256 de contraseña)
#  · Log de accesos en BASE/acceso_log.json
#  · Roles: "admin" (acceso total) | "usuario" (solo chat)
#  · El administrador puede aprobar/denegar/eliminar usuarios
# ══════════════════════════════════════════════════════════════════════
import hashlib as _hashlib
import os as _os_auth
import secrets as _secrets_mod


# ── Seguridad de contraseñas v2 — PBKDF2-HMAC-SHA256 con salt único ──
_PBKDF2_ITERS = 260_000   # NIST 2024 recomendación mínima para SHA-256

def _hash_pwd(pwd: str, salt: str | None = None) -> str:
    """
    Genera un hash seguro PBKDF2-HMAC-SHA256 con salt aleatorio.
    Formato retornado: "pbkdf2$<salt_hex>$<hash_hex>"
    Backward-compatible: si recibe un hash viejo (SHA-256 simple de 64 chars),
    lo verifica con SHA-256 para migración transparente.
    """
    if salt is None:
        salt = _secrets_mod.token_hex(32)   # 256-bit salt único
    dk = _hashlib.pbkdf2_hmac(
        "sha256",
        pwd.encode("utf-8"),
        salt.encode("utf-8"),
        _PBKDF2_ITERS,
    )
    return f"pbkdf2${salt}${dk.hex()}"


def _verificar_pwd(pwd: str, stored_hash: str) -> bool:
    """
    Verifica una contraseña contra el hash almacenado.
    Maneja automáticamente hashes viejos (SHA-256 puro) para migración.
    """
    if not stored_hash:
        return False
    # Hash viejo: SHA-256 puro (64 chars hexadecimales, sin $)
    if stored_hash.count("$") == 0 and len(stored_hash) == 64:
        return _hashlib.sha256(pwd.encode("utf-8")).hexdigest() == stored_hash
    # Hash nuevo: pbkdf2$<salt>$<hash>
    try:
        _, salt, _ = stored_hash.split("$", 2)
        nuevo = _hash_pwd(pwd, salt=salt)
        return _secrets_mod.compare_digest(stored_hash, nuevo)
    except Exception:
        return False


def _validar_fortaleza_pwd(pwd: str) -> tuple[bool, str]:
    """
    Valida que la contraseña cumpla los requisitos mínimos.
    Retorna: (ok, mensaje_error)
    """
    if len(pwd) < 8:
        return False, "Mínimo 8 caracteres."
    if not any(c.isupper() for c in pwd):
        return False, "Debe tener al menos una mayúscula."
    if not any(c.islower() for c in pwd):
        return False, "Debe tener al menos una minúscula."
    if not any(c.isdigit() for c in pwd):
        return False, "Debe tener al menos un número."
    comunes = {"password","123456","remi2026","qwerty","admin123",
               "12345678","contraseña","remuneraciones"}
    if pwd.lower() in comunes:
        return False, "Contraseña demasiado común. Elige otra."
    return True, ""


def _migrar_hash_si_necesario(usr_key: str, datos: dict, pwd: str):
    """
    Si el usuario tiene un hash viejo (SHA-256), lo migra a PBKDF2 en el acto.
    Se llama silenciosamente al hacer login exitoso.
    """
    h = datos.get("pwd_hash","")
    if h.count("$") == 0 and len(h) == 64:
        try:
            usuarios = _usuarios_cargar()
            usuarios[usr_key]["pwd_hash"] = _hash_pwd(pwd)
            _usuarios_guardar(usuarios)
            logger.info("Auth: hash migrado a PBKDF2 para '%s'", usr_key)
        except Exception:
            pass


def _usuarios_cargar() -> dict:
    d = _r(F["usuarios"])
    if not isinstance(d, dict):
        d = {}
    # Crear admin por defecto si no existe
    if not d:
        d["admin"] = {
            "nombre":    "Administrador",
            "pwd_hash":  _hash_pwd("remi2026"),
            "rol":       "admin",
            "activo":    True,
            "creado":    datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        _w(F["usuarios"], d)
    return d


def _usuarios_guardar(d: dict):
    _w(F["usuarios"], d)


def _acceso_log(usuario: str, resultado: str, ip: str = "local"):
    """Registra un intento de acceso en el log."""
    try:
        entradas = _rl(F["acceso_log"])
        entradas.append({
            "fecha":    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "usuario":  usuario,
            "resultado": resultado,   # "ok" | "denegado" | "inactivo" | "no_existe"
            "ip":       ip,
        })
        entradas = entradas[-500:]   # mantener últimos 500
        _w(F["acceso_log"], entradas)
    except Exception:
        pass


def _validar_login(usuario: str, pwd: str) -> tuple[bool, str, dict]:
    """
    Valida credenciales con PBKDF2. Migra hashes viejos automáticamente.
    Retorna: (ok, mensaje, datos_usuario)
    """
    usuarios = _usuarios_cargar()
    u = usuario.strip().lower()
    if u not in usuarios:
        _acceso_log(u, "no_existe")
        return False, "Usuario no encontrado.", {}
    datos = usuarios[u]
    if not datos.get("activo", False):
        _acceso_log(u, "inactivo")
        return False, "Acceso denegado por el administrador.", datos
    if not _verificar_pwd(pwd, datos.get("pwd_hash","")):
        _acceso_log(u, "contraseña_incorrecta")
        return False, "Contraseña incorrecta.", {}
    # Migración transparente de hash viejo → PBKDF2
    _migrar_hash_si_necesario(u, datos, pwd)
    _acceso_log(u, "ok")
    return True, "ok", datos


# ─── Ventana de Login ──────────────────────────────────────────────────
class LoginWindow:
    """
    Ventana de inicio de sesión de REMI.
    Bloquea el acceso hasta que se autentique un usuario válido.
    """

    def __init__(self):
        self.resultado   = None   # dict del usuario autenticado, o None
        self.intentos    = 0
        self._root       = tk.Tk()
        self._root.title("REMI — Acceso")
        self._root.geometry("440x520")
        self._root.resizable(False, False)
        self._root.configure(bg="#101214")
        self._root.protocol("WM_DELETE_WINDOW", self._cancelar)
        self._root.eval("tk::PlaceWindow . center")
        self._build()
        self._root.mainloop()

    def _build(self):
        r = self._root
        BG   = "#101214"
        CARD = "#1c2027"
        A1   = T.get("a1", "#3fb950")
        TX   = T.get("tx", "#eaecef")
        T2   = T.get("t2", "#8b949e")
        BR   = T.get("br", "#2a2e35")

        # ── Logo + título ──────────────────────────────────────────────
        top = tk.Frame(r, bg=BG); top.pack(pady=(32, 0))
        cv = tk.Canvas(top, width=72, height=72, bg=BG, highlightthickness=0)
        cv.pack()
        dibujar_remi(cv, 72)

        tk.Label(r, text="REMI", font=(FONT, 22, "bold"),
                 bg=BG, fg=TX).pack(pady=(10, 0))
        tk.Label(r, text="Remuneraciones · Control de Gestión",
                 font=(FONT, 9), bg=BG, fg=T2).pack()
        tk.Label(r, text="v10  ·  Acceso autorizado", font=(FONT, 8),
                 bg=BG, fg=A1).pack(pady=(2, 0))

        # ── Formulario ────────────────────────────────────────────────
        frm = tk.Frame(r, bg=CARD, highlightthickness=1,
                       highlightbackground=BR, bd=0)
        frm.pack(fill="x", padx=40, pady=24)

        def _entry(lbl_txt, show=None):
            tk.Label(frm, text=lbl_txt, font=(FONT, 9),
                     bg=CARD, fg=T2, anchor="w").pack(
                         fill="x", padx=16, pady=(14, 2))
            e = tk.Entry(frm, font=(FONT, 11), bg="#141619", fg=TX,
                         insertbackground=TX, relief="flat", bd=0,
                         highlightthickness=1, highlightbackground=BR,
                         highlightcolor=A1, show=show)
            e.pack(fill="x", padx=16, ipady=8)
            return e

        self._e_usr = _entry("Usuario")
        self._e_pwd = _entry("Contraseña", show="•")

        # Enter = login
        self._e_usr.bind("<Return>", lambda e: self._e_pwd.focus())
        self._e_pwd.bind("<Return>", lambda e: self._login())

        self._lbl_err = tk.Label(frm, text="", font=(FONT, 9),
                                  bg=CARD, fg=T.get("err", "#f85149"),
                                  wraplength=330, justify="left")
        self._lbl_err.pack(fill="x", padx=16, pady=(6, 0))

        # ── Botón Ingresar ─────────────────────────────────────────────
        btn = tk.Button(frm, text="Ingresar", font=(FONT, 11, "bold"),
                        bg=A1, fg="white", relief="flat",
                        cursor="hand2", bd=0, pady=10,
                        command=self._login)
        btn.pack(fill="x", padx=16, pady=(12, 16))
        btn.bind("<Enter>", lambda e: btn.configure(bg=T.get("a2", "#52c768")))
        btn.bind("<Leave>", lambda e: btn.configure(bg=A1))

        # ── Link registro ──────────────────────────────────────────────
        tk.Label(r, text="¿No tienes acceso? Solicítalo al administrador.",
                 font=(FONT, 8), bg=BG, fg=T2).pack()

        # ── Solicitar acceso (crea solicitud pendiente) ────────────────
        def _solicitar():
            self._win_solicitud()

        tk.Button(r, text="Solicitar acceso →", font=(FONT, 8),
                  bg=BG, fg=A1, relief="flat", cursor="hand2", bd=0,
                  command=_solicitar).pack(pady=(2, 0))

        # Foco inicial
        self._e_usr.focus()

    def _login(self):
        usr = self._e_usr.get().strip().lower()
        pwd = self._e_pwd.get()
        if not usr or not pwd:
            self._lbl_err.configure(text="Completa usuario y contraseña.")
            return
        ok, msg, datos = _validar_login(usr, pwd)
        if ok:
            self.resultado = {"usuario": usr, **datos}
            self._root.destroy()
        else:
            self.intentos += 1
            self._lbl_err.configure(text=f"⚠ {msg}")
            self._e_pwd.delete(0, "end")
            if self.intentos >= 5:
                self._lbl_err.configure(
                    text="Demasiados intentos fallidos. Contacta al administrador.")
                self._e_usr.configure(state="disabled")
                self._e_pwd.configure(state="disabled")

    def _cancelar(self):
        self.resultado = None
        try: self._root.destroy()
        except Exception: pass

    def _win_solicitud(self):
        """Mini-ventana para solicitar acceso (crea usuario inactivo)."""
        w = tk.Toplevel(self._root)
        w.title("Solicitar acceso a REMI")
        w.geometry("380x320")
        w.configure(bg="#101214")
        w.resizable(False, False)

        BG = "#101214"; CARD = "#1c2027"
        A1 = T.get("a1","#3fb950"); TX = T.get("tx","#eaecef")
        T2 = T.get("t2","#8b949e"); BR = T.get("br","#2a2e35")
        ERR = T.get("err","#f85149")

        tk.Label(w, text="Solicitar acceso a REMI", font=(FONT,13,"bold"),
                 bg=BG, fg=TX).pack(pady=(18,4))
        tk.Label(w, text="El administrador revisará tu solicitud.",
                 font=(FONT,9), bg=BG, fg=T2).pack()

        frm = tk.Frame(w, bg=CARD, highlightthickness=1,
                       highlightbackground=BR); frm.pack(fill="x", padx=24, pady=16)

        def _fld(txt, show=None):
            tk.Label(frm, text=txt, font=(FONT,9), bg=CARD, fg=T2,
                     anchor="w").pack(fill="x", padx=14, pady=(10,2))
            e = tk.Entry(frm, font=(FONT,10), bg="#141619", fg=TX,
                         insertbackground=TX, relief="flat",
                         highlightthickness=1, highlightbackground=BR,
                         highlightcolor=A1, show=show)
            e.pack(fill="x", padx=14, ipady=7)
            return e

        e_usr  = _fld("Usuario (sin espacios)")
        e_nom  = _fld("Nombre completo")
        e_pwd  = _fld("Contraseña", show="•")
        lbl_r  = tk.Label(frm, text="", font=(FONT,9), bg=CARD, fg=ERR,
                           wraplength=300)
        lbl_r.pack(fill="x", padx=14, pady=(4,0))

        def _enviar():
            usr = e_usr.get().strip().lower().replace(" ","_")
            nom = e_nom.get().strip()
            pwd = e_pwd.get()
            if not usr or not nom or not pwd:
                lbl_r.configure(text="Completa todos los campos."); return
            ok_f, msg_f = _validar_fortaleza_pwd(pwd)
            if not ok_f:
                lbl_r.configure(text=msg_f); return
            usuarios = _usuarios_cargar()
            if usr in usuarios:
                lbl_r.configure(text="Ese usuario ya existe."); return
            usuarios[usr] = {
                "nombre":   nom,
                "pwd_hash": _hash_pwd(pwd),
                "rol":      "usuario",
                "activo":   False,     # pendiente de aprobación
                "creado":   datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                "solicitud": True,
            }
            _usuarios_guardar(usuarios)
            _acceso_log(usr, "solicitud_enviada")
            lbl_r.configure(fg=A1,
                text="✅ Solicitud enviada. El administrador te habilitará el acceso.")
            w.after(2500, w.destroy)

        tk.Button(frm, text="Enviar solicitud", font=(FONT,10,"bold"),
                  bg=A1, fg="white", relief="flat", cursor="hand2",
                  bd=0, pady=8, command=_enviar).pack(
                  fill="x", padx=14, pady=(10,14))


# ─── Panel de Administración de Usuarios ───────────────────────────────
class PanelUsuarios(tk.Toplevel):
    """
    Ventana de administración de usuarios (solo para rol=admin).
    Accesible desde ⚙️ Configuración o desde el sidebar.
    """

    def __init__(self, master):
        super().__init__(master)
        self.title("👥 Administración de Usuarios — REMI")
        self.geometry("860x600")
        self.configure(bg=T["bg"])
        self.resizable(True, True)
        self._build()

    def _build(self):
        BG   = T["bg"]; CARD = T["card"]; SB = T["sb"]
        A1   = T["a1"]; TX   = T["tx"]; T2 = T["t2"]
        BR   = T["br"]; OK   = T.get("ok","#3fb950")
        ERR  = T.get("err","#f85149")
        WARN = T.get("warn","#e3b341")

        # ── Header ─────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=SB, height=50)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="👥  Administración de Usuarios",
                 font=(FONT,13,"bold"), bg=SB, fg=TX).pack(
                 side="left", padx=14, pady=12)
        tk.Button(hdr, text="➕ Nuevo usuario", font=(FONT,9),
                  bg=A1, fg="white", relief="flat", cursor="hand2",
                  bd=0, padx=12, pady=5,
                  command=self._win_nuevo).pack(side="right", padx=10, pady=8)

        # ── Tabs: Usuarios | Log de accesos ────────────────────────────
        tab_f = tk.Frame(self, bg=CARD); tab_f.pack(fill="x")
        self._tab = tk.StringVar(value="usuarios")

        def _tab_btn(txt, val):
            def _cmd():
                self._tab.set(val)
                _refresca_tab()
            b = tk.Button(tab_f, text=txt, font=(FONT,10),
                          relief="flat", cursor="hand2", bd=0,
                          padx=16, pady=8,
                          command=_cmd)
            b.pack(side="left")
            return b

        btn_u = _tab_btn("👤  Usuarios", "usuarios")
        btn_l = _tab_btn("📋  Log de accesos", "log")
        tk.Frame(tab_f, bg=BR, height=2).pack(fill="x", side="bottom")

        # ── Contenido dinámico ─────────────────────────────────────────
        self._contenido = tk.Frame(self, bg=BG)
        self._contenido.pack(fill="both", expand=True)

        def _refresca_tab():
            # Limpiar
            for w in self._contenido.winfo_children():
                try: w.destroy()
                except: pass
            # Actualizar colores tabs
            for b, v in [(btn_u,"usuarios"),(btn_l,"log")]:
                activo = self._tab.get() == v
                b.configure(bg=BG if activo else CARD,
                            fg=A1 if activo else T2,
                            font=(FONT,10,"bold") if activo else (FONT,10))
            if self._tab.get() == "usuarios":
                self._tab_usuarios()
            else:
                self._tab_log()

        self._refresca = _refresca_tab
        _refresca_tab()

    def _tab_usuarios(self):
        BG  = T["bg"]; CARD = T["card"]
        A1  = T["a1"]; TX   = T["tx"]; T2 = T["t2"]
        BR  = T["br"]; OK   = T.get("ok","#3fb950")
        ERR = T.get("err","#f85149")
        WARN = T.get("warn","#e3b341")

        sc = ctk.CTkScrollableFrame(self._contenido, fg_color=BG,
                                     scrollbar_button_color=A1, corner_radius=0)
        sc.pack(fill="both", expand=True)

        usuarios = _usuarios_cargar()

        # ── Leyenda de estados ─────────────────────────────────────────
        leg_f = tk.Frame(sc, bg=BG); leg_f.pack(fill="x", padx=14, pady=(10,4))
        for txt, fg in [("🟢 Activo",OK), ("🔴 Inactivo/Denegado",ERR),
                         ("🟡 Pendiente aprobación",WARN)]:
            tk.Label(leg_f, text=txt, font=(FONT,8), bg=BG, fg=fg).pack(
                side="left", padx=8)

        # ── Tabla de usuarios ──────────────────────────────────────────
        # Encabezado
        hf = tk.Frame(sc, bg=T.get("sb","#11131a"))
        hf.pack(fill="x", padx=10, pady=(0,2))
        for txt, w in [("Usuario",120),("Nombre",160),("Rol",80),
                        ("Estado",90),("Creado",140),("Acciones",160)]:
            tk.Label(hf, text=txt, font=(FONT,9,"bold"), bg=T.get("sb","#11131a"),
                     fg=T2, width=w//7, anchor="w").pack(side="left", padx=6, pady=6)

        def _fila(usr_key, datos):
            activo   = datos.get("activo", False)
            pendiente = datos.get("solicitud", False) and not activo
            rol      = datos.get("rol","usuario")

            color_est = OK if activo else (WARN if pendiente else ERR)
            est_txt   = "🟢 Activo" if activo else ("🟡 Pendiente" if pendiente else "🔴 Inactivo")

            row = tk.Frame(sc, bg=CARD, highlightthickness=1,
                           highlightbackground=BR)
            row.pack(fill="x", padx=10, pady=2)

            def _lbl(txt, w=None, fg=TX, bold=False):
                kw = dict(text=str(txt)[:22], font=(FONT,9,"bold" if bold else "normal"),
                          bg=CARD, fg=fg, anchor="w")
                if w: kw["width"] = w//7
                tk.Label(row, **kw).pack(side="left", padx=6, pady=8)

            _lbl(usr_key, 120, bold=(rol=="admin"))
            _lbl(datos.get("nombre",""), 160)
            rol_color = A1 if rol=="admin" else T2
            _lbl(rol.upper(), 80, fg=rol_color)
            _lbl(est_txt, 90, fg=color_est)
            _lbl(datos.get("creado","")[:16], 140)

            # Acciones
            acc_f = tk.Frame(row, bg=CARD); acc_f.pack(side="left", padx=4)

            def _btn_acc(txt, cmd, fg_c):
                b = tk.Button(acc_f, text=txt, font=(FONT,8),
                              bg=T.get("br","#2a2e35"), fg=fg_c,
                              relief="flat", cursor="hand2", bd=0,
                              padx=6, pady=3, command=cmd)
                b.pack(side="left", padx=2)
                return b

            def _toggle_activo(uk=usr_key):
                us = _usuarios_cargar()
                if uk not in us: return
                us[uk]["activo"] = not us[uk].get("activo", False)
                if us[uk]["activo"]:
                    us[uk].pop("solicitud", None)
                _usuarios_guardar(us)
                self._refresca()

            def _toggle_rol(uk=usr_key):
                us = _usuarios_cargar()
                if uk not in us: return
                us[uk]["rol"] = "usuario" if us[uk].get("rol") == "admin" else "admin"
                _usuarios_guardar(us)
                self._refresca()

            def _eliminar(uk=usr_key):
                if uk == "admin":
                    return  # No se puede eliminar el admin principal
                import tkinter.messagebox as _mb
                if _mb.askyesno("Eliminar usuario",
                                "¿Eliminar al usuario '" + uk + "'?\nEsta acción no se puede deshacer.",
                                parent=self):
                    us = _usuarios_cargar()
                    us.pop(uk, None)
                    _usuarios_guardar(us)
                    _acceso_log(uk, "eliminado_por_admin")
                    self._refresca()

            if activo:
                _btn_acc("🔒 Denegar", _toggle_activo, ERR)
            else:
                _btn_acc("✅ Aprobar", _toggle_activo, OK)

            if usr_key != "admin":
                lbl_rol = "👑 Admin" if rol != "admin" else "👤 Usuario"
                _btn_acc(lbl_rol, _toggle_rol, A1)
                _btn_acc("🗑", _eliminar, ERR)

        # Primero admin, luego el resto ordenado
        admin_d = usuarios.pop("admin", None)
        if admin_d:
            _fila("admin", admin_d)
            usuarios["admin"] = admin_d

        for uk, datos in sorted(usuarios.items()):
            if uk != "admin":
                _fila(uk, datos)

        # Stats rápidas
        total  = len(usuarios)
        activos = sum(1 for d in usuarios.values() if d.get("activo"))
        pend   = sum(1 for d in usuarios.values() if d.get("solicitud") and not d.get("activo"))
        sf = tk.Frame(sc, bg=BG); sf.pack(fill="x", padx=14, pady=10)
        tk.Label(sf, text=f"Total: {total}  ·  Activos: {activos}  ·  Pendientes: {pend}",
                 font=(FONT,9), bg=BG, fg=T2).pack(side="left")

    def _tab_log(self):
        BG = T["bg"]; CARD = T["card"]; A1 = T["a1"]
        TX = T["tx"]; T2 = T["t2"]; BR = T["br"]
        OK = T.get("ok","#3fb950"); ERR = T.get("err","#f85149")
        WARN = T.get("warn","#e3b341")

        sc = ctk.CTkScrollableFrame(self._contenido, fg_color=BG,
                                     scrollbar_button_color=A1, corner_radius=0)
        sc.pack(fill="both", expand=True)

        log = list(reversed(_rl(F["acceso_log"])))  # más reciente primero

        if not log:
            tk.Label(sc, text="Sin registros de acceso todavía.",
                     font=(FONT,10), bg=BG, fg=T2).pack(pady=30)
            return

        # Encabezado
        hf = tk.Frame(sc, bg=T.get("sb","#11131a"))
        hf.pack(fill="x", padx=10, pady=(8,2))
        for txt, w in [("Fecha",150),("Usuario",130),("Resultado",160),("IP",100)]:
            tk.Label(hf, text=txt, font=(FONT,9,"bold"),
                     bg=T.get("sb","#11131a"), fg=T2,
                     width=w//7, anchor="w").pack(side="left", padx=6, pady=6)

        RESULT_COLOR = {
            "ok":                    OK,
            "no_existe":             ERR,
            "contraseña_incorrecta": ERR,
            "inactivo":              WARN,
            "solicitud_enviada":     WARN,
            "eliminado_por_admin":   ERR,
        }
        RESULT_ICON = {
            "ok":                    "✅",
            "no_existe":             "❌",
            "contraseña_incorrecta": "🔑",
            "inactivo":              "🔒",
            "solicitud_enviada":     "📩",
            "eliminado_por_admin":   "🗑",
        }

        for i, e in enumerate(log[:200]):
            bg_row = CARD if i % 2 == 0 else BG
            row = tk.Frame(sc, bg=bg_row)
            row.pack(fill="x", padx=10, pady=1)
            res = e.get("resultado","")
            col = RESULT_COLOR.get(res, T2)
            ico = RESULT_ICON.get(res, "•")

            def _lbl(txt, fg=TX, w=None, bold=False):
                kw = dict(text=str(txt), font=(FONT,9,"bold" if bold else "normal"),
                          bg=bg_row, fg=fg, anchor="w")
                if w: kw["width"] = w//7
                tk.Label(row, **kw).pack(side="left", padx=6, pady=5)

            _lbl(e.get("fecha",""), w=150)
            _lbl(e.get("usuario",""), w=130, bold=True)
            _lbl(f"{ico} {res}", fg=col, w=160)
            _lbl(e.get("ip","local"), w=100, fg=T2)

        # Botón limpiar log (solo admin)
        def _limpiar():
            import tkinter.messagebox as _mb
            if _mb.askyesno("Limpiar log", "¿Borrar todo el log de accesos?",
                            parent=self):
                _w(F["acceso_log"], [])
                self._refresca()

        tk.Button(sc, text="🗑 Limpiar log completo", font=(FONT,9),
                  bg=T.get("br","#2a2e35"), fg=ERR,
                  relief="flat", cursor="hand2", bd=0, padx=10, pady=5,
                  command=_limpiar).pack(pady=(8,14))

    def _win_nuevo(self):
        """Ventana para crear un nuevo usuario desde el panel admin."""
        w = tk.Toplevel(self)
        w.title("Nuevo usuario")
        w.geometry("380x360")
        w.configure(bg=T["bg"])
        w.resizable(False, False)

        BG = T["bg"]; CARD = T["card"]; A1 = T["a1"]
        TX = T["tx"]; T2 = T["t2"]; BR = T["br"]
        ERR = T.get("err","#f85149")

        tk.Label(w, text="Crear nuevo usuario", font=(FONT,13,"bold"),
                 bg=BG, fg=TX).pack(pady=(16,4))
        frm = tk.Frame(w, bg=CARD, highlightthickness=1,
                       highlightbackground=BR); frm.pack(fill="x", padx=20, pady=12)

        def _fld(lbl, show=None):
            tk.Label(frm, text=lbl, font=(FONT,9), bg=CARD, fg=T2,
                     anchor="w").pack(fill="x", padx=14, pady=(10,2))
            e = tk.Entry(frm, font=(FONT,10), bg="#141619", fg=TX,
                         insertbackground=TX, relief="flat",
                         highlightthickness=1, highlightbackground=BR,
                         highlightcolor=A1, show=show)
            e.pack(fill="x", padx=14, ipady=7)
            return e

        e_usr = _fld("Usuario (sin espacios)")
        e_nom = _fld("Nombre completo")
        e_pwd = _fld("Contraseña", show="•")

        rol_var = tk.StringVar(value="usuario")
        rf = tk.Frame(frm, bg=CARD); rf.pack(fill="x", padx=14, pady=(10,4))
        tk.Label(rf, text="Rol:", font=(FONT,9), bg=CARD, fg=T2).pack(side="left")
        for v, lbl in [("usuario","Usuario"),("admin","Administrador")]:
            tk.Radiobutton(rf, text=lbl, variable=rol_var, value=v,
                           font=(FONT,9), bg=CARD, fg=TX,
                           selectcolor=CARD, activebackground=CARD).pack(
                           side="left", padx=8)

        lbl_r = tk.Label(frm, text="", font=(FONT,9), bg=CARD, fg=ERR,
                          wraplength=320)
        lbl_r.pack(fill="x", padx=14, pady=(4,0))

        def _crear():
            usr = e_usr.get().strip().lower().replace(" ","_")
            nom = e_nom.get().strip()
            pwd = e_pwd.get()
            if not usr or not nom or not pwd:
                lbl_r.configure(text="Completa todos los campos."); return
            ok_f, msg_f = _validar_fortaleza_pwd(pwd)
            if not ok_f:
                lbl_r.configure(text=msg_f); return
            usuarios = _usuarios_cargar()
            if usr in usuarios:
                lbl_r.configure(text="Ese usuario ya existe."); return
            usuarios[usr] = {
                "nombre":  nom,
                "pwd_hash": _hash_pwd(pwd),
                "rol":     rol_var.get(),
                "activo":  True,
                "creado":  datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
            _usuarios_guardar(usuarios)
            lbl_r.configure(fg=A1, text=f"✅ Usuario '{usr}' creado.")
            w.after(1500, lambda: (w.destroy(), self._refresca()))

        tk.Button(frm, text="Crear usuario", font=(FONT,10,"bold"),
                  bg=A1, fg="white", relief="flat", cursor="hand2",
                  bd=0, pady=9, command=_crear).pack(
                  fill="x", padx=14, pady=(10,14))


# Variable global con el usuario de la sesión actual
# ══════════════════════════════════════════════════════════════════════
#  PANEL WEB ADMIN REMI — servidor Flask local
#  Acceso: http://localhost:7265 (solo desde la misma máquina)
#  Autenticación: usuario admin de REMI
#  Arranca automáticamente al iniciar REMI (hilo daemon)
# ══════════════════════════════════════════════════════════════════════

_WEB_PORT    = 7265
_WEB_TOKEN   = None   # token de sesión web, se genera al arrancar
_WEB_THREAD  = None

_HTML_ADMIN = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>REMI · Panel Admin</title>
<style>
  :root{--bg:#101214;--sb:#141619;--card:#1c2027;--inp:#191d24;
        --a1:#3fb950;--a2:#52c768;--tx:#eaecef;--t2:#8b949e;
        --br:#2a2e35;--err:#f85149;--warn:#e3b341;--ok:#3fb950}
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);color:var(--tx);font-family:"Segoe UI",sans-serif;min-height:100vh}
  a{color:var(--a1);text-decoration:none}

  /* ── Nav ── */
  nav{background:var(--sb);border-bottom:2px solid var(--a1);
      display:flex;align-items:center;padding:0 24px;height:56px;gap:16px}
  .logo{display:flex;align-items:center;gap:10px}
  .logo svg{width:32px;height:32px}
  .logo span{font-size:1.2rem;font-weight:700;color:var(--tx)}
  .logo small{font-size:.72rem;color:var(--t2)}
  nav a.tab{color:var(--t2);font-size:.9rem;padding:6px 14px;border-radius:6px;
            transition:.15s;border:1px solid transparent}
  nav a.tab:hover,nav a.tab.active{color:var(--a1);border-color:var(--br);
                                    background:var(--card)}
  .spacer{flex:1}
  .badge{background:var(--card);border:1px solid var(--br);
         padding:4px 12px;border-radius:20px;font-size:.8rem;color:var(--t2)}
  .badge b{color:var(--a1)}

  /* ── Layout ── */
  .wrap{max-width:1100px;margin:0 auto;padding:28px 20px}
  h1{font-size:1.3rem;font-weight:700;margin-bottom:4px}
  .sub{color:var(--t2);font-size:.85rem;margin-bottom:24px}

  /* ── Cards de stats ── */
  .stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin-bottom:28px}
  .stat{background:var(--card);border:1px solid var(--br);border-radius:10px;
        padding:18px 20px;text-align:center}
  .stat .n{font-size:2rem;font-weight:800;color:var(--a1)}
  .stat .l{font-size:.8rem;color:var(--t2);margin-top:4px}

  /* ── Tabla ── */
  .tbl-wrap{background:var(--card);border:1px solid var(--br);border-radius:10px;overflow:hidden;margin-bottom:28px}
  .tbl-hdr{background:var(--sb);padding:14px 20px;display:flex;align-items:center;
           justify-content:space-between;border-bottom:1px solid var(--br)}
  .tbl-hdr h2{font-size:1rem}
  table{width:100%;border-collapse:collapse}
  th{background:var(--sb);color:var(--t2);font-size:.78rem;text-transform:uppercase;
     letter-spacing:.06em;padding:10px 16px;text-align:left;border-bottom:1px solid var(--br)}
  td{padding:11px 16px;border-bottom:1px solid var(--br);font-size:.88rem;vertical-align:middle}
  tr:last-child td{border-bottom:none}
  tr:hover td{background:#ffffff06}

  /* ── Badges estado ── */
  .pill{display:inline-block;padding:3px 10px;border-radius:20px;font-size:.75rem;font-weight:600}
  .pill-ok{background:#1a3d22;color:#3fb950}
  .pill-err{background:#3d1a1a;color:#f85149}
  .pill-warn{background:#3d3000;color:#e3b341}

  /* ── Botones acción ── */
  .btn{display:inline-block;padding:5px 14px;border-radius:6px;font-size:.8rem;
       cursor:pointer;border:none;font-weight:600;text-decoration:none;transition:.15s}
  .btn-ok{background:#1a3d22;color:#3fb950}
  .btn-ok:hover{background:#224d2a}
  .btn-err{background:#3d1a1a;color:#f85149}
  .btn-err:hover{background:#4d2020}
  .btn-warn{background:#3d3000;color:#e3b341}
  .btn-warn:hover{background:#4d3c00}
  .btn-blue{background:#1a2a3d;color:#58a6ff}
  .btn-blue:hover{background:#1e3352}
  .btn-ghost{background:var(--br);color:var(--t2)}
  .btn-ghost:hover{background:#343c4a;color:var(--tx)}
  form.inline{display:inline}

  /* ── Log ── */
  .log-ok   td:nth-child(3){color:var(--ok)}
  .log-err  td:nth-child(3){color:var(--err)}
  .log-warn td:nth-child(3){color:var(--warn)}

  /* ── Modal nueva contraseña / usuario ── */
  .modal-bg{display:none;position:fixed;inset:0;background:#000a;z-index:100;
            align-items:center;justify-content:center}
  .modal-bg.open{display:flex}
  .modal{background:var(--card);border:1px solid var(--br);border-radius:14px;
         padding:28px;width:420px;max-width:95vw}
  .modal h3{margin-bottom:16px;font-size:1.05rem}
  .fld{margin-bottom:14px}
  .fld label{display:block;font-size:.8rem;color:var(--t2);margin-bottom:5px}
  .fld input,.fld select{width:100%;background:var(--inp);color:var(--tx);border:1px solid var(--br);
              border-radius:7px;padding:9px 12px;font-size:.9rem;outline:none}
  .fld input:focus,.fld select:focus{border-color:var(--a1)}
  .pwd-meter{height:4px;border-radius:4px;margin-top:5px;background:var(--br);overflow:hidden}
  .pwd-bar{height:100%;width:0;transition:.3s;border-radius:4px}
  .modal-btns{display:flex;gap:10px;margin-top:20px;justify-content:flex-end}
  .err-msg{color:var(--err);font-size:.8rem;margin-top:4px;min-height:16px}

  /* ── Reglas contraseña ── */
  .pwd-rules{font-size:.75rem;color:var(--t2);margin-top:6px;line-height:1.7}
  .rule{display:flex;align-items:center;gap:5px}
  .rule.ok{color:var(--ok)} .rule.fail{color:var(--err)}
  .rule svg{width:12px;height:12px}

  /* ── Footer ── */
  footer{text-align:center;color:var(--t2);font-size:.75rem;padding:24px;
         border-top:1px solid var(--br);margin-top:20px}

  /* ── Login page ── */
  .login-wrap{display:flex;align-items:center;justify-content:center;min-height:100vh}
  .login-box{background:var(--card);border:1px solid var(--br);border-radius:14px;
             padding:36px;width:360px}
  .login-box .logo{justify-content:center;margin-bottom:20px}
  .login-box h2{text-align:center;margin-bottom:6px}
  .login-box .sub{text-align:center;margin-bottom:24px}
  .login-btn{width:100%;padding:11px;background:var(--a1);color:#fff;border:none;
             border-radius:8px;font-weight:700;font-size:1rem;cursor:pointer;margin-top:6px}
  .login-btn:hover{background:var(--a2)}
  .login-err{color:var(--err);font-size:.85rem;text-align:center;margin-top:10px;min-height:20px}
</style>
</head>
<body>
<!--NAV-->
<nav>
  <div class="logo">
    <svg viewBox="0 0 80 80" fill="none">
      <polygon points="40,4 72,22 72,58 40,76 8,58 8,22" stroke="#3fb950" stroke-width="3" fill="#1c2027"/>
      <line x1="40" y1="14" x2="40" y2="32" stroke="#3fb950" stroke-width="2"/>
      <line x1="40" y1="48" x2="40" y2="66" stroke="#3fb950" stroke-width="2"/>
      <line x1="26" y1="20" x2="40" y2="32" stroke="#3fb950" stroke-width="1.5"/>
      <line x1="54" y1="20" x2="40" y2="32" stroke="#3fb950" stroke-width="1.5"/>
      <circle cx="40" cy="40" r="10" fill="#3fb950" opacity=".15" stroke="#3fb950" stroke-width="2"/>
      <text x="40" y="45" text-anchor="middle" font-size="12" font-weight="bold" fill="#3fb950" font-family="Segoe UI">R</text>
      <circle cx="40" cy="14" r="3" fill="#52c768"/>
      <circle cx="26" cy="20" r="2.5" fill="#52c768"/>
      <circle cx="54" cy="20" r="2.5" fill="#52c768"/>
    </svg>
    <div><span>REMI</span><br><small>Panel de Administración</small></div>
  </div>
  <a href="/admin" class="tab {{tab_u}}">👤 Usuarios</a>
  <a href="/admin/log" class="tab {{tab_l}}">📋 Log de accesos</a>
  <div class="spacer"></div>
  <span class="badge">Sesión: <b>{{admin_user}}</b></span>
  <a href="/admin/logout" class="btn btn-ghost" style="margin-left:8px">Cerrar sesión</a>
</nav>

<div class="wrap">
{{CONTENT}}
</div>

<footer>REMI v10 · Panel Admin · Solo acceso local · {{now}}</footer>
<script>
// ── Modal nueva contraseña ──────────────────────────────────────────
function openPwd(usr){
  document.getElementById('pwd-usr').value=usr;
  document.getElementById('pwd-modal').classList.add('open');
  document.getElementById('npwd').value='';
  document.getElementById('npwd2').value='';
  updatePwdMeter('');
  document.getElementById('npwd').focus();
}
function closePwd(){document.getElementById('pwd-modal').classList.remove('open')}

// ── Modal nuevo usuario ─────────────────────────────────────────────
function openNew(){document.getElementById('new-modal').classList.add('open')}
function closeNew(){document.getElementById('new-modal').classList.remove('open')}

// ── Medidor de fortaleza de contraseña ──────────────────────────────
function updatePwdMeter(v){
  let score=0,color='#f85149';
  const rules={
    'r-len':v.length>=8,
    'r-up':/[A-Z]/.test(v),
    'r-lo':/[a-z]/.test(v),
    'r-num':/[0-9]/.test(v),
  };
  Object.values(rules).forEach(ok=>{if(ok)score++});
  if(score==4)color='#3fb950';
  else if(score>=2)color='#e3b341';
  const bar=document.getElementById('pwd-bar');
  if(bar){bar.style.width=(score*25)+'%';bar.style.background=color}
  Object.entries(rules).forEach(([id,ok])=>{
    const el=document.getElementById(id);
    if(el){el.className='rule '+(ok?'ok':'fail');
           el.querySelector('svg').innerHTML=ok?
           '<polyline points="2,6 5,9 10,3" stroke="currentColor" stroke-width="1.5" fill="none"/>':
           '<line x1="2" y1="2" x2="10" y2="10" stroke="currentColor" stroke-width="1.5"/><line x1="10" y1="2" x2="2" y2="10" stroke="currentColor" stroke-width="1.5"/>';}
  });
}
document.addEventListener('DOMContentLoaded',()=>{
  const inp=document.getElementById('npwd');
  if(inp)inp.addEventListener('input',e=>updatePwdMeter(e.target.value));
});
</script>

<!-- Modal cambiar contraseña -->
<div class="modal-bg" id="pwd-modal" onclick="if(event.target===this)closePwd()">
  <div class="modal">
    <h3>🔑 Cambiar contraseña</h3>
    <form method="post" action="/admin/cambiar_pwd">
      <input type="hidden" name="usuario" id="pwd-usr">
      <div class="fld"><label>Nueva contraseña</label>
        <input type="password" name="nueva" id="npwd" autocomplete="new-password">
        <div class="pwd-meter"><div class="pwd-bar" id="pwd-bar"></div></div>
        <div class="pwd-rules">
          <div class="rule fail" id="r-len"><svg viewBox="0 0 12 12"></svg> Mínimo 8 caracteres</div>
          <div class="rule fail" id="r-up"><svg viewBox="0 0 12 12"></svg> Una mayúscula</div>
          <div class="rule fail" id="r-lo"><svg viewBox="0 0 12 12"></svg> Una minúscula</div>
          <div class="rule fail" id="r-num"><svg viewBox="0 0 12 12"></svg> Un número</div>
        </div>
      </div>
      <div class="fld"><label>Confirmar contraseña</label>
        <input type="password" name="nueva2" id="npwd2" autocomplete="new-password">
      </div>
      <p class="err-msg" id="pwd-err">{{pwd_flash}}</p>
      <div class="modal-btns">
        <button type="button" class="btn btn-ghost" onclick="closePwd()">Cancelar</button>
        <button type="submit" class="btn btn-ok">Guardar</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal nuevo usuario -->
<div class="modal-bg" id="new-modal" onclick="if(event.target===this)closeNew()">
  <div class="modal">
    <h3>➕ Nuevo usuario</h3>
    <form method="post" action="/admin/nuevo_usuario">
      <div class="fld"><label>Usuario (sin espacios)</label><input type="text" name="usr" autocomplete="off"></div>
      <div class="fld"><label>Nombre completo</label><input type="text" name="nom"></div>
      <div class="fld"><label>Contraseña</label>
        <input type="password" name="pwd" id="npwd" autocomplete="new-password">
        <div class="pwd-meter"><div class="pwd-bar" id="pwd-bar"></div></div>
        <div class="pwd-rules">
          <div class="rule fail" id="r-len"><svg viewBox="0 0 12 12"></svg> Mínimo 8 caracteres</div>
          <div class="rule fail" id="r-up"><svg viewBox="0 0 12 12"></svg> Una mayúscula</div>
          <div class="rule fail" id="r-lo"><svg viewBox="0 0 12 12"></svg> Una minúscula</div>
          <div class="rule fail" id="r-num"><svg viewBox="0 0 12 12"></svg> Un número</div>
        </div>
      </div>
      <div class="fld"><label>Rol</label>
        <select name="rol">
          <option value="usuario">Usuario</option>
          <option value="admin">Administrador</option>
        </select>
      </div>
      <p class="err-msg">{{new_flash}}</p>
      <div class="modal-btns">
        <button type="button" class="btn btn-ghost" onclick="closeNew()">Cancelar</button>
        <button type="submit" class="btn btn-ok">Crear</button>
      </div>
    </form>
  </div>
</div>
</body>
</html>
"""

_HTML_LOGIN = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>REMI · Admin Login</title>
<style>
  :root{--bg:#101214;--card:#1c2027;--inp:#191d24;--a1:#3fb950;--a2:#52c768;
        --tx:#eaecef;--t2:#8b949e;--br:#2a2e35;--err:#f85149}
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);color:var(--tx);font-family:"Segoe UI",sans-serif;
       display:flex;align-items:center;justify-content:center;min-height:100vh}
  .box{background:var(--card);border:1px solid var(--br);border-radius:14px;
       padding:36px;width:360px}
  .logo{display:flex;align-items:center;gap:10px;justify-content:center;margin-bottom:20px}
  .logo svg{width:38px;height:38px}
  .logo span{font-size:1.4rem;font-weight:800}
  h2{text-align:center;font-size:1.1rem;margin-bottom:4px}
  .sub{text-align:center;color:var(--t2);font-size:.82rem;margin-bottom:24px}
  .fld{margin-bottom:14px}
  .fld label{display:block;font-size:.8rem;color:var(--t2);margin-bottom:5px}
  .fld input{width:100%;background:var(--inp);color:var(--tx);border:1px solid var(--br);
             border-radius:7px;padding:10px 12px;font-size:.9rem;outline:none}
  .fld input:focus{border-color:var(--a1)}
  button{width:100%;padding:11px;background:var(--a1);color:#fff;border:none;
         border-radius:8px;font-weight:700;font-size:1rem;cursor:pointer;margin-top:6px}
  button:hover{background:var(--a2)}
  .err{color:var(--err);font-size:.85rem;text-align:center;margin-top:10px;min-height:18px}
  footer{text-align:center;color:var(--t2);font-size:.72rem;margin-top:18px}
</style>
</head>
<body>
<div class="box">
  <div class="logo">
    <svg viewBox="0 0 80 80" fill="none">
      <polygon points="40,4 72,22 72,58 40,76 8,58 8,22" stroke="#3fb950" stroke-width="3" fill="#1c2027"/>
      <text x="40" y="47" text-anchor="middle" font-size="20" font-weight="bold" fill="#3fb950" font-family="Segoe UI">R</text>
    </svg>
    <span>REMI</span>
  </div>
  <h2>Panel de Administración</h2>
  <p class="sub">Solo administradores autorizados</p>
  <form method="post">
    <div class="fld"><label>Usuario</label><input type="text" name="usr" autofocus autocomplete="username"></div>
    <div class="fld"><label>Contraseña</label><input type="password" name="pwd" autocomplete="current-password"></div>
    <button type="submit">Ingresar</button>
    <p class="err">{{error}}</p>
  </form>
  <footer>Acceso solo desde localhost · REMI v10</footer>
</div>
</body>
</html>
"""


def _iniciar_servidor_web():
    """
    Lanza el servidor web Flask en un hilo daemon.
    Solo escucha en 127.0.0.1 (no expuesto a la red).
    """
    global _WEB_TOKEN

    try:
        from flask import Flask, request, redirect, session, url_for
    except ImportError:
        try:
            import subprocess as _sp
            _sp.run([sys.executable, "-m", "pip", "install", "flask", "-q",
                     "--disable-pip-version-check"],
                    capture_output=True, timeout=120,
                    creationflags=getattr(_sp, "CREATE_NO_WINDOW", 0))
            from flask import Flask, request, redirect, session, url_for
        except Exception as e:
            logger.warning("Panel web: no se pudo instalar Flask: %s", e)
            return

    import secrets as _sec
    _WEB_TOKEN = _sec.token_hex(32)

    app_web = Flask(__name__)
    app_web.secret_key = _sec.token_hex(32)

    def _render(template: str, content: str = "", **kwargs) -> str:
        import datetime as _dt
        ses_usr = kwargs.pop("admin_user", session.get("admin_user","—"))
        tab_u   = "active" if kwargs.pop("tab","u") == "u" else ""
        tab_l   = "active" if not tab_u else ""
        return (template
                .replace("{{CONTENT}}", content)
                .replace("{{admin_user}}", ses_usr)
                .replace("{{tab_u}}", tab_u)
                .replace("{{tab_l}}", tab_l)
                .replace("{{now}}", _dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
                .replace("{{pwd_flash}}", kwargs.get("pwd_flash",""))
                .replace("{{new_flash}}", kwargs.get("new_flash",""))
                )

    def _requiere_admin():
        return session.get("admin_ok") and session.get("admin_rol") == "admin"

    # ── Login web ──────────────────────────────────────────────────────
    @app_web.route("/admin/login", methods=["GET","POST"])
    def web_login():
        err = ""
        if request.method == "POST":
            usr = request.form.get("usr","").strip().lower()
            pwd = request.form.get("pwd","")
            ok, msg, datos = _validar_login(usr, pwd)
            if ok and datos.get("rol") == "admin":
                session["admin_ok"]   = True
                session["admin_user"] = usr
                session["admin_rol"]  = datos.get("rol","usuario")
                return redirect("/admin")
            elif ok:
                err = "Acceso restringido a administradores."
                _acceso_log(usr, "web_no_admin")
            else:
                err = msg
        return _HTML_LOGIN.replace("{{error}}", err)

    @app_web.route("/admin/logout")
    def web_logout():
        session.clear()
        return redirect("/admin/login")

    @app_web.route("/admin")
    def web_usuarios():
        if not _requiere_admin():
            return redirect("/admin/login")
        usuarios = _usuarios_cargar()
        total   = len(usuarios)
        activos = sum(1 for d in usuarios.values() if d.get("activo"))
        pend    = sum(1 for d in usuarios.values() if d.get("solicitud") and not d.get("activo"))

        def _pill(datos):
            if datos.get("activo"):
                return '<span class="pill pill-ok">🟢 Activo</span>'
            if datos.get("solicitud"):
                return '<span class="pill pill-warn">🟡 Pendiente</span>'
            return '<span class="pill pill-err">🔴 Inactivo</span>'

        def _acciones(uk, datos):
            btns = ""
            if datos.get("activo"):
                btns += f'<form class="inline" method="post" action="/admin/toggle_activo"><input type="hidden" name="usr" value="{uk}"><button class="btn btn-err">🔒 Denegar</button></form> '
            else:
                btns += f'<form class="inline" method="post" action="/admin/toggle_activo"><input type="hidden" name="usr" value="{uk}"><button class="btn btn-ok">✅ Aprobar</button></form> '
            if uk != "admin":
                rol_txt = "👑 Admin" if datos.get("rol") != "admin" else "👤 Usuario"
                btns += f'<form class="inline" method="post" action="/admin/toggle_rol"><input type="hidden" name="usr" value="{uk}"><button class="btn btn-blue">{rol_txt}</button></form> '
                btns += f'<button class="btn btn-warn" onclick="openPwd(\'{uk}\')">🔑 Clave</button> '
                btns += f'<form class="inline" method="post" action="/admin/eliminar" onsubmit="return confirm(\'¿Eliminar al usuario {uk}?\')"><input type="hidden" name="usr" value="{uk}"><button class="btn btn-err">🗑</button></form>'
            else:
                btns += f'<button class="btn btn-warn" onclick="openPwd(\'admin\')">🔑 Clave</button>'
            return btns

        filas = ""
        # Admin primero
        for uk in (["admin"] + [k for k in sorted(usuarios) if k != "admin"]):
            if uk not in usuarios: continue
            d = usuarios[uk]
            rol_badge = '<span style="color:#3fb950;font-weight:700">ADMIN</span>' if d.get("rol")=="admin" else d.get("rol","usuario").upper()
            filas += f"""<tr>
              <td><b>{uk}</b></td>
              <td>{d.get("nombre","")}</td>
              <td>{rol_badge}</td>
              <td>{_pill(d)}</td>
              <td style="color:var(--t2);font-size:.8rem">{d.get("creado","")[:16]}</td>
              <td>{_acciones(uk, d)}</td>
            </tr>"""

        content = f"""
<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px">
  <div><h1>👤 Usuarios de REMI</h1><p class="sub">Gestión de acceso al sistema</p></div>
  <button class="btn btn-ok" onclick="openNew()" style="margin-top:4px">➕ Nuevo usuario</button>
</div>
<div class="stats">
  <div class="stat"><div class="n">{total}</div><div class="l">Total usuarios</div></div>
  <div class="stat"><div class="n" style="color:var(--ok)">{activos}</div><div class="l">Activos</div></div>
  <div class="stat"><div class="n" style="color:var(--warn)">{pend}</div><div class="l">Pendientes</div></div>
  <div class="stat"><div class="n" style="color:var(--err)">{total-activos-pend}</div><div class="l">Inactivos</div></div>
</div>
<div class="tbl-wrap">
  <div class="tbl-hdr"><h2>Lista de usuarios</h2>
    <span style="color:var(--t2);font-size:.8rem">Los cambios se aplican de inmediato</span></div>
  <table>
    <thead><tr><th>Usuario</th><th>Nombre</th><th>Rol</th><th>Estado</th><th>Creado</th><th>Acciones</th></tr></thead>
    <tbody>{filas}</tbody>
  </table>
</div>"""
        return _render(_HTML_ADMIN, content, tab="u")

    @app_web.route("/admin/log")
    def web_log():
        if not _requiere_admin():
            return redirect("/admin/login")
        log = list(reversed(_rl(F["acceso_log"])))

        ICONOS = {
            "ok":"✅","no_existe":"❌","contraseña_incorrecta":"🔑",
            "inactivo":"🔒","solicitud_enviada":"📩","eliminado_por_admin":"🗑",
            "web_no_admin":"⛔",
        }
        CLASES = {
            "ok":"log-ok","no_existe":"log-err","contraseña_incorrecta":"log-err",
            "inactivo":"log-warn","solicitud_enviada":"log-warn",
            "eliminado_por_admin":"log-err","web_no_admin":"log-err",
        }

        filas = ""
        for e in log[:300]:
            res = e.get("resultado","")
            cls = CLASES.get(res,"")
            ico = ICONOS.get(res,"•")
            filas += f"""<tr class="{cls}">
              <td style="color:var(--t2);font-size:.8rem">{e.get("fecha","")}</td>
              <td><b>{e.get("usuario","")}</b></td>
              <td>{ico} {res}</td>
              <td style="color:var(--t2)">{e.get("ip","local")}</td>
            </tr>"""

        ok_c  = sum(1 for e in log if e.get("resultado")=="ok")
        err_c = sum(1 for e in log if e.get("resultado") in ("contraseña_incorrecta","no_existe"))
        pend_c= sum(1 for e in log if e.get("resultado")=="solicitud_enviada")

        content = f"""
<h1>📋 Log de accesos</h1>
<p class="sub">Historial completo de intentos de autenticación</p>
<div class="stats">
  <div class="stat"><div class="n" style="color:var(--ok)">{ok_c}</div><div class="l">Accesos exitosos</div></div>
  <div class="stat"><div class="n" style="color:var(--err)">{err_c}</div><div class="l">Intentos fallidos</div></div>
  <div class="stat"><div class="n" style="color:var(--warn)">{pend_c}</div><div class="l">Solicitudes</div></div>
  <div class="stat"><div class="n">{len(log)}</div><div class="l">Total registros</div></div>
</div>
<div class="tbl-wrap">
  <div class="tbl-hdr"><h2>Registro de eventos</h2>
    <form method="post" action="/admin/limpiar_log" onsubmit="return confirm('¿Limpiar todo el log?')">
      <button class="btn btn-err">🗑 Limpiar log</button>
    </form>
  </div>
  <table>
    <thead><tr><th>Fecha</th><th>Usuario</th><th>Resultado</th><th>IP</th></tr></thead>
    <tbody>{filas if filas else "<tr><td colspan='4' style='text-align:center;color:var(--t2);padding:24px'>Sin registros</td></tr>"}</tbody>
  </table>
</div>"""
        return _render(_HTML_ADMIN, content, tab="l")

    # ── Acciones POST ──────────────────────────────────────────────────
    @app_web.route("/admin/toggle_activo", methods=["POST"])
    def web_toggle_activo():
        if not _requiere_admin(): return redirect("/admin/login")
        uk = request.form.get("usr","").strip().lower()
        us = _usuarios_cargar()
        if uk in us:
            us[uk]["activo"] = not us[uk].get("activo", False)
            if us[uk]["activo"]: us[uk].pop("solicitud", None)
            _usuarios_guardar(us)
        return redirect("/admin")

    @app_web.route("/admin/toggle_rol", methods=["POST"])
    def web_toggle_rol():
        if not _requiere_admin(): return redirect("/admin/login")
        uk = request.form.get("usr","").strip().lower()
        if uk == "admin": return redirect("/admin")
        us = _usuarios_cargar()
        if uk in us:
            us[uk]["rol"] = "usuario" if us[uk].get("rol") == "admin" else "admin"
            _usuarios_guardar(us)
        return redirect("/admin")

    @app_web.route("/admin/eliminar", methods=["POST"])
    def web_eliminar():
        if not _requiere_admin(): return redirect("/admin/login")
        uk = request.form.get("usr","").strip().lower()
        if uk == "admin": return redirect("/admin")
        us = _usuarios_cargar()
        if uk in us:
            us.pop(uk)
            _usuarios_guardar(us)
            _acceso_log(uk, "eliminado_por_admin")
        return redirect("/admin")

    @app_web.route("/admin/cambiar_pwd", methods=["POST"])
    def web_cambiar_pwd():
        if not _requiere_admin(): return redirect("/admin/login")
        uk   = request.form.get("usuario","").strip().lower()
        pwd  = request.form.get("nueva","")
        pwd2 = request.form.get("nueva2","")
        if pwd != pwd2:
            return _render(_HTML_ADMIN, "", pwd_flash="Las contraseñas no coinciden.", tab="u")
        ok_f, msg_f = _validar_fortaleza_pwd(pwd)
        if not ok_f:
            return _render(_HTML_ADMIN, "", pwd_flash=msg_f, tab="u")
        us = _usuarios_cargar()
        if uk in us:
            us[uk]["pwd_hash"] = _hash_pwd(pwd)
            _usuarios_guardar(us)
            _acceso_log(uk, "contrasena_cambiada_admin_web")
        return redirect("/admin")

    @app_web.route("/admin/nuevo_usuario", methods=["POST"])
    def web_nuevo_usuario():
        if not _requiere_admin(): return redirect("/admin/login")
        usr = request.form.get("usr","").strip().lower().replace(" ","_")
        nom = request.form.get("nom","").strip()
        pwd = request.form.get("pwd","")
        rol = request.form.get("rol","usuario")
        ok_f, msg_f = _validar_fortaleza_pwd(pwd)
        if not usr or not nom:
            return _render(_HTML_ADMIN, "", new_flash="Completa todos los campos.", tab="u")
        if not ok_f:
            return _render(_HTML_ADMIN, "", new_flash=msg_f, tab="u")
        us = _usuarios_cargar()
        if usr in us:
            return _render(_HTML_ADMIN, "", new_flash="Usuario ya existe.", tab="u")
        us[usr] = {
            "nombre":  nom, "pwd_hash": _hash_pwd(pwd),
            "rol": rol, "activo": True,
            "creado": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        _usuarios_guardar(us)
        return redirect("/admin")

    @app_web.route("/admin/limpiar_log", methods=["POST"])
    def web_limpiar_log():
        if not _requiere_admin(): return redirect("/admin/login")
        _w(F["acceso_log"], [])
        return redirect("/admin/log")

    @app_web.route("/")
    def web_root():
        return redirect("/admin")

    # ── Arrancar Flask en hilo daemon ──────────────────────────────────
    import threading as _thr_web
    def _run_flask():
        import logging as _lg
        _lg.getLogger("werkzeug").setLevel(_lg.ERROR)  # silenciar logs de Flask
        app_web.run(host="127.0.0.1", port=_WEB_PORT, debug=False, use_reloader=False)

    _t = _thr_web.Thread(target=_run_flask, daemon=True, name="RemiWebAdmin")
    _t.start()
    logger.info("Panel web admin: http://localhost:%d/admin", _WEB_PORT)


_SESION_ACTUAL: dict = {}



if __name__ == "__main__":
    # ── Ocultar ventana CMD en Windows ──────────────────────────────────
    try:
        import ctypes as _ctypes
        _ctypes.windll.user32.ShowWindow(
            _ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except Exception:
        pass

    # ── Redirigir stderr/stdout si no hay consola ────────────────────────
    import sys as _sys
    import os  as _os
    if _sys.stderr is None or not hasattr(_sys.stderr, "write"):
        _sys.stderr = open(_os.devnull, "w")
    if _sys.stdout is None or not hasattr(_sys.stdout, "write"):
        _sys.stdout = open(_os.devnull, "w")

    # ── Panel web admin (Flask, solo localhost) ──────────────────────────
    try:
        _iniciar_servidor_web()
    except Exception as _we:
        logger.warning("Panel web no iniciado: %s", _we)

    # ── Login — pantalla de acceso antes de cargar la app ──────────────
    login = LoginWindow()
    if login.resultado is None:
        # Cerró la ventana sin autenticarse → salir
        sys.exit(0)
    _SESION_ACTUAL.update(login.resultado)

    # ── Arrancar REMI con captura de errores ─────────────────────────────
    try:
        app = RemiApp()
        # Guardar referencia de sesión en la app para el panel de admin
        app._sesion = _SESION_ACTUAL
        app.mainloop()
    except Exception as _startup_err:
        import traceback as _tb
        _err_txt = _tb.format_exc()
        # Guardar el error en un archivo de log legible
        try:
            _log_path = Path.home() / "Desktop" / "remi_error.log"
            _log_path.write_text(
                "REMI Error de arranque — " + str(datetime.datetime.now()) + "\n\n" + _err_txt,
                encoding="utf-8"
            )
        except Exception:
            pass
        # Mostrar ventana de error con el traceback completo
        try:
            import tkinter as _tk
            _r = _tk.Tk()
            _r.title("REMI — Error de arranque")
            _r.geometry("780x460")
            _r.configure(bg="#111111")
            _tk.Label(_r, text="❌  REMI no pudo iniciar — copia el error y envíalo para soporte",
                      font=("Segoe UI", 11, "bold"), bg="#111111", fg="#f08080",
                      wraplength=740, justify="left").pack(padx=14, pady=(14,4), anchor="w")
            _txt = _tk.Text(_r, wrap="word", bg="#0d0d0d", fg="#e0e0e0",
                            font=("Consolas", 9), relief="flat", padx=10, pady=8)
            _txt.pack(fill="both", expand=True, padx=12, pady=4)
            _txt.insert("1.0", _err_txt)
            _txt.configure(state="disabled")
            _bf = _tk.Frame(_r, bg="#111111")
            _bf.pack(fill="x", padx=12, pady=(0,10))
            def _copy_err():
                _r.clipboard_clear(); _r.clipboard_append(_err_txt)
                _cb.configure(text="Copiado ✓")
            _cb = _tk.Button(_bf, text="Copiar error", command=_copy_err,
                             bg="#2d6cdf", fg="white", relief="flat",
                             padx=12, pady=6, font=("Segoe UI", 10))
            _cb.pack(side="left")
            try:
                _tk.Label(_bf, text=f"  Log guardado en: {_log_path}",
                          font=("Consolas", 8), bg="#111111", fg="#888888").pack(side="left", padx=8)
            except Exception:
                pass
            _tk.Button(_bf, text="Cerrar", command=_r.destroy,
                       bg="#333333", fg="white", relief="flat",
                       padx=12, pady=6, font=("Segoe UI", 10)).pack(side="right")
            _r.mainloop()
        except Exception:
            pass
